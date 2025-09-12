from fastapi import APIRouter, Request, Query, Form
from fastapi.responses import HTMLResponse, RedirectResponse, FileResponse, Response
from fastapi.templating import Jinja2Templates
from datetime import date, datetime, timedelta
from typing import Optional, List
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, or_, func, desc, String
from sqlalchemy.orm import selectinload
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from core.database.session import get_async_session
from core.auth.user_manager import UserManager
from apps.web.middleware.auth_middleware import require_owner_or_superadmin
from domain.entities.shift import Shift
from domain.entities.shift_schedule import ShiftSchedule
from domain.entities.object import Object
from domain.entities.user import User

router = APIRouter()
templates = Jinja2Templates(directory="apps/web/templates")
user_manager = UserManager()


async def get_user_id_from_current_user(current_user, session):
    """Получает внутренний ID пользователя из current_user"""
    if isinstance(current_user, dict):
        # current_user - это словарь из JWT payload
        telegram_id = current_user.get("id")
        user_query = select(User).where(User.telegram_id == telegram_id)
        user_result = await session.execute(user_query)
        user_obj = user_result.scalar_one_or_none()
        return user_obj.id if user_obj else None
    else:
        # current_user - это объект User
        return current_user.id


@router.get("/", response_class=HTMLResponse)
async def reports_index(request: Request):
    """Главная страница отчетов"""
    current_user = await require_owner_or_superadmin(request)
    if isinstance(current_user, RedirectResponse):
        return current_user
    
    user_role = current_user.get("role") if isinstance(current_user, dict) else current_user.role
    
    async with get_async_session() as session:
        # Получаем внутренний user_id
        user_id = await get_user_id_from_current_user(current_user, session)
        
        # Получаем объекты владельца
        objects_query = select(Object).where(Object.owner_id == user_id)
        objects_result = await session.execute(objects_query)
        objects = objects_result.scalars().all()
        
        # Получаем всех пользователей, которые работали на объектах владельца (не только employee)
        employees_query = select(User.id, User.telegram_id, User.username, User.first_name, User.last_name, User.phone, User.role, User.is_active, User.created_at, User.updated_at).distinct().join(Shift, User.id == Shift.user_id).where(
            Shift.object_id.in_([obj.id for obj in objects])
        )
        employees_result = await session.execute(employees_query)
        employees = employees_result.all()
        
        # Если нет сотрудников из смен, показываем всех пользователей кроме текущего владельца
        if not employees:
            all_employees_query = select(User.id, User.telegram_id, User.username, User.first_name, User.last_name, User.phone, User.role, User.is_active, User.created_at, User.updated_at).where(User.id != user_id)
            all_employees_result = await session.execute(all_employees_query)
            employees = all_employees_result.all()
        
        # Статистика за последний месяц
        month_ago = datetime.now() - timedelta(days=30)
        
        shifts_query = select(Shift).options(
            selectinload(Shift.object),
            selectinload(Shift.user)
        ).where(
            and_(
                Shift.object_id.in_([obj.id for obj in objects]),
                Shift.start_time >= month_ago
            )
        )
        shifts_result = await session.execute(shifts_query)
        recent_shifts = shifts_result.scalars().all()
        
        stats = {
            "total_shifts": len(recent_shifts),
            "total_hours": sum(s.total_hours or 0 for s in recent_shifts if s.total_hours),
            "total_payment": sum(s.total_payment or 0 for s in recent_shifts if s.total_payment),
            "active_objects": len(objects),
            "employees": len(employees)
        }
        
        return templates.TemplateResponse("reports/index.html", {
            "request": request,
            "current_user": current_user,
            "objects": objects,
            "employees": employees,
            "stats": stats
        })


@router.post("/generate")
async def generate_report(
    request: Request,
    report_type: str = Form(...),
    date_from: str = Form(...),
    date_to: str = Form(...),
    object_id: Optional[int] = Form(None),
    employee_id: Optional[int] = Form(None),
    format: str = Form("excel")
):
    """Генерация отчета"""
    current_user = await require_owner_or_superadmin(request)
    if isinstance(current_user, RedirectResponse):
        return current_user
    
    # Парсинг дат
    try:
        start_date = datetime.strptime(date_from, "%Y-%m-%d").date()
        end_date = datetime.strptime(date_to, "%Y-%m-%d").date()
    except ValueError:
        return {"error": "Неверный формат даты"}
    
    async with get_async_session() as session:
        # Получаем внутренний user_id
        user_id = await get_user_id_from_current_user(current_user, session)
        
        # Получаем объекты владельца
        owner_objects = select(Object.id).where(Object.owner_id == user_id)
        objects_result = await session.execute(owner_objects)
        owner_object_ids = [obj[0] for obj in objects_result.all()]
        
        # Базовый запрос для смен
        shifts_query = select(Shift).options(
            selectinload(Shift.object),
            selectinload(Shift.user)
        ).where(
            and_(
                Shift.object_id.in_(owner_object_ids),
                Shift.start_time >= start_date,
                Shift.start_time <= end_date + timedelta(days=1)
            )
        )
        
        # Применение фильтров
        if object_id and object_id in owner_object_ids:
            shifts_query = shifts_query.where(Shift.object_id == object_id)
        
        if employee_id:
            shifts_query = shifts_query.where(Shift.user_id == employee_id)
        
        # Выполнение запроса
        shifts_result = await session.execute(shifts_query.order_by(desc(Shift.start_time)))
        shifts = shifts_result.scalars().all()
        
        # Генерация отчета в зависимости от типа
        if report_type == "shifts":
            return await _generate_shifts_report(shifts, format, start_date, end_date)
        elif report_type == "employees":
            return await _generate_employees_report(shifts, format, start_date, end_date)
        elif report_type == "objects":
            return await _generate_objects_report(shifts, format, start_date, end_date)
        else:
            return {"error": "Неизвестный тип отчета"}


async def _generate_shifts_report(shifts: List[Shift], format: str, start_date: date, end_date: date):
    """Генерация отчета по сменам"""
    data = []
    
    for shift in shifts:
        data.append({
            "ID": shift.id,
            "Сотрудник": f"{shift.user.first_name} {shift.user.last_name or ''}".strip(),
            "Объект": shift.object.name,
            "Дата начала": shift.start_time.strftime("%d.%m.%Y %H:%M"),
            "Дата окончания": shift.end_time.strftime("%d.%m.%Y %H:%M") if shift.end_time else "Не завершена",
            "Статус": shift.status,
            "Часов": shift.total_hours or 0,
            "Ставка": shift.hourly_rate or 0,
            "Сумма": shift.total_payment or 0,
            "Заметки": shift.notes or ""
        })
    
    if format == "excel":
        return await _create_excel_file(data, f"shifts_report_{start_date}_{end_date}")
    else:
        return {"data": data, "total": len(data)}


async def _generate_employees_report(shifts: List[Shift], format: str, start_date: date, end_date: date):
    """Генерация отчета по сотрудникам"""
    # Группировка по сотрудникам
    employees_data = {}
    
    for shift in shifts:
        employee_id = shift.user_id
        if employee_id not in employees_data:
            employees_data[employee_id] = {
                "employee": shift.user,
                "shifts": [],
                "total_hours": 0,
                "total_payment": 0
            }
        
        employees_data[employee_id]["shifts"].append(shift)
        employees_data[employee_id]["total_hours"] += shift.total_hours or 0
        employees_data[employee_id]["total_payment"] += shift.total_payment or 0
    
    data = []
    for emp_data in employees_data.values():
        data.append({
            "Сотрудник": f"{emp_data['employee'].first_name} {emp_data['employee'].last_name or ''}".strip(),
            "Количество смен": len(emp_data["shifts"]),
            "Общее время": emp_data["total_hours"],
            "Общая сумма": emp_data["total_payment"],
            "Средняя ставка": emp_data["total_payment"] / emp_data["total_hours"] if emp_data["total_hours"] > 0 else 0
        })
    
    if format == "excel":
        return await _create_excel_file(data, f"employees_report_{start_date}_{end_date}")
    else:
        return {"data": data, "total": len(data)}


async def _generate_objects_report(shifts: List[Shift], format: str, start_date: date, end_date: date):
    """Генерация отчета по объектам"""
    # Группировка по объектам
    objects_data = {}
    
    for shift in shifts:
        object_id = shift.object_id
        if object_id not in objects_data:
            objects_data[object_id] = {
                "object": shift.object,
                "shifts": [],
                "total_hours": 0,
                "total_payment": 0,
                "employees": set()
            }
        
        objects_data[object_id]["shifts"].append(shift)
        objects_data[object_id]["total_hours"] += shift.total_hours or 0
        objects_data[object_id]["total_payment"] += shift.total_payment or 0
        objects_data[object_id]["employees"].add(shift.user_id)
    
    data = []
    for obj_data in objects_data.values():
        data.append({
            "Объект": obj_data["object"].name,
            "Адрес": obj_data["object"].address or "",
            "Количество смен": len(obj_data["shifts"]),
            "Количество сотрудников": len(obj_data["employees"]),
            "Общее время": obj_data["total_hours"],
            "Общая сумма": obj_data["total_payment"],
            "Средняя ставка": obj_data["total_payment"] / obj_data["total_hours"] if obj_data["total_hours"] > 0 else 0
        })
    
    if format == "excel":
        return await _create_excel_file(data, f"objects_report_{start_date}_{end_date}")
    else:
        return {"data": data, "total": len(data)}


async def _create_excel_file(data: List[dict], filename: str):
    """Создание Excel файла"""
    if not data:
        return {"error": "Нет данных для отчета"}
    
    # Создание DataFrame
    df = pd.DataFrame(data)
    
    # Создание Excel файла
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет"
    
    # Добавление данных
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # Стилизация
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Автоширина колонок
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохранение в память
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"}
    )


@router.get("/stats/period")
async def period_stats(
    request: Request,
    date_from: str = Query(...),
    date_to: str = Query(...),
    object_id: Optional[int] = Query(None)
):
    """Статистика за период"""
    current_user = await require_owner_or_superadmin(request)
    if isinstance(current_user, RedirectResponse):
        return current_user
    
    try:
        start_date = datetime.strptime(date_from, "%Y-%m-%d").date()
        end_date = datetime.strptime(date_to, "%Y-%m-%d").date()
    except ValueError:
        return {"error": "Неверный формат даты"}
    
    async with get_async_session() as session:
        # Получаем внутренний user_id
        user_id = await get_user_id_from_current_user(current_user, session)
        
        # Получаем объекты владельца
        owner_objects = select(Object.id).where(Object.owner_id == user_id)
        objects_result = await session.execute(owner_objects)
        owner_object_ids = [obj[0] for obj in objects_result.all()]
        
        # Запрос смен за период
        shifts_query = select(Shift).options(
            selectinload(Shift.object),
            selectinload(Shift.user)
        ).where(
            and_(
                Shift.object_id.in_(owner_object_ids),
                Shift.start_time >= start_date,
                Shift.start_time <= end_date + timedelta(days=1)
            )
        )
        
        if object_id and object_id in owner_object_ids:
            shifts_query = shifts_query.where(Shift.object_id == object_id)
        
        shifts_result = await session.execute(shifts_query)
        shifts = shifts_result.scalars().all()
        
        # Расчет статистики
        stats = {
            "period": {
                "from": start_date.strftime("%d.%m.%Y"),
                "to": end_date.strftime("%d.%m.%Y")
            },
            "total_shifts": len(shifts),
            "total_hours": sum(s.total_hours or 0 for s in shifts if s.total_hours),
            "total_payment": sum(s.total_payment or 0 for s in shifts if s.total_payment),
            "avg_hours_per_shift": 0,
            "avg_payment_per_shift": 0,
            "by_status": {},
            "by_object": {},
            "by_employee": {}
        }
        
        if stats["total_shifts"] > 0:
            stats["avg_hours_per_shift"] = stats["total_hours"] / stats["total_shifts"]
            stats["avg_payment_per_shift"] = stats["total_payment"] / stats["total_shifts"]
        
        # Статистика по статусам
        status_counts = {}
        for shift in shifts:
            status = shift.status
            status_counts[status] = status_counts.get(status, 0) + 1
        stats["by_status"] = status_counts
        
        # Статистика по объектам
        object_stats = {}
        for shift in shifts:
            obj_name = shift.object.name
            if obj_name not in object_stats:
                object_stats[obj_name] = {
                    "shifts": 0,
                    "hours": 0,
                    "payment": 0
                }
            object_stats[obj_name]["shifts"] += 1
            object_stats[obj_name]["hours"] += shift.total_hours or 0
            object_stats[obj_name]["payment"] += shift.total_payment or 0
        stats["by_object"] = object_stats
        
        # Статистика по сотрудникам
        employee_stats = {}
        for shift in shifts:
            emp_name = f"{shift.user.first_name} {shift.user.last_name or ''}".strip()
            if emp_name not in employee_stats:
                employee_stats[emp_name] = {
                    "shifts": 0,
                    "hours": 0,
                    "payment": 0
                }
            employee_stats[emp_name]["shifts"] += 1
            employee_stats[emp_name]["hours"] += shift.total_hours or 0
            employee_stats[emp_name]["payment"] += shift.total_payment or 0
        stats["by_employee"] = employee_stats
        
        return stats