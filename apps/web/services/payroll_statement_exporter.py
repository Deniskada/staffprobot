"""Экспорт расчётного листа в Excel."""

from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def build_statement_workbook(statement: Dict[str, Any]) -> bytes:
    """Формирует Excel-файл по данным расчётного листа."""
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Расчётный лист"

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="E5E5E5", end_color="E5E5E5", fill_type="solid")

    summary_headers = [
        "Период",
        "Объект",
        "Начислено",
        "Доплаты",
        "Удержания",
        "К выплате",
        "Выплачено",
        "Остаток",
    ]
    ws_summary.append(summary_headers)
    for cell in ws_summary[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for block in statement["entries"]:
        entry = block.entry
        period = f"{entry.period_start.strftime('%d.%m.%Y')} — {entry.period_end.strftime('%d.%m.%Y')}"
        object_name = entry.object_.name if entry.object_ else "—"
        gross = float(entry.gross_amount or 0)
        bonus = float(entry.total_bonuses or 0)
        deduction = float(entry.total_deductions or 0)
        net = float(entry.net_amount or 0)
        paid = float(block.paid_amount)
        balance = net - paid

        ws_summary.append([period, object_name, gross, bonus, deduction, net, paid, balance])

    totals = statement["totals"]
    ws_summary.append([])
    ws_summary.append(
        [
            "ИТОГО",
            "",
            float(totals.gross),
            float(totals.bonuses),
            float(totals.deductions),
            float(totals.net),
            float(totals.paid),
            float(totals.balance),
        ]
    )

    _auto_fit_columns(ws_summary)

    _add_adjustments_sheet(wb, statement["entries"])
    _add_payments_sheet(wb, statement["entries"])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _auto_fit_columns(worksheet) -> None:
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 60)


def _add_adjustments_sheet(workbook: Workbook, entries: List[Any]) -> None:
    ws = workbook.create_sheet("Корректировки")
    headers = ["Период", "Тип", "Сумма", "Описание", "Привязка"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E5E5E5", end_color="E5E5E5", fill_type="solid")

    for block in entries:
        period = f"{block.entry.period_start.strftime('%d.%m.%Y')} — {block.entry.period_end.strftime('%d.%m.%Y')}"
        for adj in block.adjustments:
            adj_type = adj.get_type_label()
            amount = float(adj.amount or 0)
            description = adj.description or ""
            link = f"Смена #{adj.shift_id}" if adj.shift_id else ""
            ws.append([period, adj_type, amount, description, link])

    _auto_fit_columns(ws)


def _add_payments_sheet(workbook: Workbook, entries: List[Any]) -> None:
    ws = workbook.create_sheet("Выплаты")
    headers = ["Период", "Дата", "Сумма", "Способ", "Статус", "Комментарий"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E5E5E5", end_color="E5E5E5", fill_type="solid")

    for block in entries:
        period = f"{block.entry.period_start.strftime('%d.%m.%Y')} — {block.entry.period_end.strftime('%d.%m.%Y')}"
        for payment in block.payments:
            ws.append(
                [
                    period,
                    payment.payment_date.strftime("%d.%m.%Y"),
                    float(payment.amount or 0),
                    payment.payment_method,
                    payment.status,
                    payment.notes or "",
                ]
            )

    _auto_fit_columns(ws)

