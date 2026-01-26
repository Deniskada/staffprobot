"""
Роуты для владельцев объектов
URL-префикс: /owner/*
"""

from fastapi import APIRouter, Request, Depends, HTTPException, status, Form, Query, Body
from fastapi.responses import RedirectResponse, JSONResponse
from typing import List, Optional, Dict, Any
from fastapi.responses import HTMLResponse, RedirectResponse, Response
from fastapi.templating import Jinja2Templates
from sqlalchemy.ext.asyncio import AsyncSession
from core.database.session import get_db_session
from apps.web.middleware.auth_middleware import require_owner_or_superadmin
from sqlalchemy import select, func, desc, and_, or_
from sqlalchemy.orm import selectinload
from datetime import datetime, timedelta, date, time, timezone
from typing import Optional, List, Dict, Any
from decimal import Decimal
import calendar
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from core.database.session import get_async_session, get_db_session
from apps.web.middleware.auth_middleware import get_current_user
from apps.web.dependencies import get_current_user_dependency, require_role
from apps.web.services.object_service import ObjectService, TimeSlotService
from apps.web.utils.timezone_utils import web_timezone_helper
from shared.services.system_features_service import SystemFeaturesService
from shared.services.calendar_filter_service import CalendarFilterService
from shared.services.shift_history_service import ShiftHistoryService
from shared.services.cancellation_policy_service import CancellationPolicyService
from shared.services.shift_cancellation_service import ShiftCancellationService
from shared.services.shift_status_sync_service import ShiftStatusSyncService
from shared.services.shift_history_service import ShiftHistoryService
from core.cache.redis_cache import cache
from apps.web.utils.shift_history_utils import build_shift_history_items
from shared.models.calendar_data import TimeslotStatus
from domain.entities.user import User, UserRole
from domain.entities.object import Object
from domain.entities.contract import Contract
from domain.entities.shift import Shift
from domain.entities.shift_schedule import ShiftSchedule
from domain.entities.application import Application, ApplicationStatus
from core.logging.logger import logger
from apps.web.services.template_service import TemplateService
import pytz

router = APIRouter()
from apps.web.jinja import templates


async def get_available_interfaces_for_user(user_id: int):
    """Получает доступные интерфейсы для пользователя"""
    from shared.services.role_based_login_service import RoleBasedLoginService
    async with get_async_session() as session:
        login_service = RoleBasedLoginService(session)
        return await login_service.get_available_interfaces(user_id)

async def get_owner_context(user_id: int, session: AsyncSession):
    """Получает общий контекст для страниц владельца"""
    from apps.web.utils.applications_utils import get_new_applications_count
    
    available_interfaces = await get_available_interfaces_for_user(user_id)
    new_applications_count = await get_new_applications_count(user_id, session, "owner")
    
    return {
        "available_interfaces": available_interfaces,
        "new_applications_count": new_applications_count
    }


async def get_user_id_from_current_user(current_user, session):
    """Получает внутренний ID пользователя из current_user"""
    if isinstance(current_user, dict):
        # current_user - это словарь из JWT payload
        telegram_id = current_user.get("telegram_id") or current_user.get("id")
        user_query = select(User).where(User.telegram_id == telegram_id)
        user_result = await session.execute(user_query)
        user_obj = user_result.scalar_one_or_none()
        return user_obj.id if user_obj else None
    else:
        # current_user - это объект User
        return current_user.id


@router.get("/", response_class=HTMLResponse, name="owner_dashboard")
async def owner_dashboard(request: Request):
    """Дашборд владельца"""
    # Проверяем авторизацию и роль владельца
    current_user = await get_current_user(request)
    if not current_user:
        return RedirectResponse(url="/auth/login", status_code=status.HTTP_302_FOUND)
    
    user_role = current_user.get("role", "employee")
    if user_role != "owner":
        return RedirectResponse(url="/dashboard", status_code=status.HTTP_302_FOUND)
    
    try:
        async with get_async_session() as session:
            user_id = await get_user_id_from_current_user(current_user, session)
            
            # Получаем статистику владельца
            objects_count = await session.execute(
                select(func.count(Object.id)).where(Object.owner_id == user_id)
            )
            total_objects = objects_count.scalar()
            
            shifts_count = await session.execute(
                select(func.count(Shift.id)).where(
                    Shift.object_id.in_(
                        select(Object.id).where(Object.owner_id == user_id)
                    )
                )
            )
            total_shifts = shifts_count.scalar()
            
            # Активные смены
            active_shifts_count = await session.execute(
                select(func.count(Shift.id)).where(
                    and_(
                        Shift.status == 'active',
                        Shift.object_id.in_(
                            select(Object.id).where(Object.owner_id == user_id)
                        )
                    )
                )
            )
            active_shifts = active_shifts_count.scalar()
            
            # Все объекты владельца с вычислением статуса и времени последнего изменения статуса
            from domain.entities.object_opening import ObjectOpening
            from core.utils.timezone_helper import timezone_helper
            
            # Загружаем объекты с org_unit и всей цепочкой родителей (до 5 уровней иерархии)
            from domain.entities.org_structure import OrgStructureUnit
            objects_result = await session.execute(
                select(Object)
                .where(
                    Object.owner_id == user_id,
                    Object.is_active == True  # Только активные объекты
                )
                .order_by(desc(Object.created_at))
                .options(
                    selectinload(Object.org_unit).selectinload(OrgStructureUnit.parent)
                    .selectinload(OrgStructureUnit.parent)
                    .selectinload(OrgStructureUnit.parent)
                    .selectinload(OrgStructureUnit.parent)
                    .selectinload(OrgStructureUnit.parent)
                )
            )
            objects_list = objects_result.scalars().all()
            
            # Сегодняшняя дата
            today_local = timezone_helper.utc_to_local(datetime.now(timezone.utc)).date()
            
            # Для каждого объекта берём последнюю запись открытия/закрытия + статус работы
            all_objects = []
            for obj in objects_list:
                last_opening_res = await session.execute(
                    select(ObjectOpening)
                    .where(ObjectOpening.object_id == obj.id)
                    .order_by(desc(ObjectOpening.opened_at))
                    .limit(1)
                )
                last_opening = last_opening_res.scalar_one_or_none()
                status = "Закрыт"
                status_time = None
                if last_opening:
                    if last_opening.closed_at is None:
                        status = "Открыт"
                        status_time = last_opening.opened_at
                    else:
                        status = "Закрыт"
                        status_time = last_opening.closed_at
                
                # Вычислить статус работы (только для сегодняшних смен)
                work_status = None
                work_employee = None
                work_delay = None
                work_early = None
                
                if obj.opening_time and obj.closing_time:
                    # Получить смены объекта на сегодня
                    # Учитываем как запланированные смены (по planned_start), так и спонтанные (по start_time/actual_start)
                    start_of_day_utc = timezone_helper.start_of_day_utc(today_local)
                    end_of_day_utc = timezone_helper.end_of_day_utc(today_local)
                    shifts_query = select(Shift).where(
                        and_(
                            Shift.object_id == obj.id,
                            or_(
                                # Запланированные смены
                                and_(
                                    Shift.planned_start.isnot(None),
                                    Shift.planned_start >= start_of_day_utc,
                                    Shift.planned_start < end_of_day_utc
                                ),
                                # Спонтанные смены (без planned_start, но с start_time или actual_start)
                                and_(
                                    Shift.planned_start.is_(None),
                                    or_(
                                        and_(
                                            Shift.start_time.isnot(None),
                                            Shift.start_time >= start_of_day_utc,
                                            Shift.start_time < end_of_day_utc
                                        ),
                                        and_(
                                            Shift.actual_start.isnot(None),
                                            Shift.actual_start >= start_of_day_utc,
                                            Shift.actual_start < end_of_day_utc
                                        )
                                    )
                                )
                            )
                        )
                    ).options(selectinload(Shift.user))
                    
                    shifts_result = await session.execute(shifts_query)
                    shifts_today = list(shifts_result.scalars().all())
                    
                    # Если смен нет, но объект должен работать - проверяем статус "нет смен"
                    if not shifts_today:
                        # Проверяем, что текущее время находится в рабочем времени объекта (как в Celery)
                        obj_timezone = obj.timezone or "Europe/Moscow"
                        naive_expected_open = datetime.combine(today_local, obj.opening_time)
                        naive_expected_close = datetime.combine(today_local, obj.closing_time)
                        obj_tz = pytz.timezone(obj_timezone)
                        expected_open = obj_tz.localize(naive_expected_open)
                        expected_close = obj_tz.localize(naive_expected_close)
                        now_local = timezone_helper.utc_to_local(datetime.now(timezone.utc), timezone_str=obj_timezone)
                        
                        # Если время открытия прошло и мы еще в рабочем времени - статус "нет смен"
                        if now_local >= expected_open and now_local <= expected_close:
                            work_status = 'no_shifts_today'
                            work_employee = None
                            logger.debug(
                                f"Dashboard: object {obj.id} ({obj.name}) - no shifts today (within working hours), "
                                f"expected_open={expected_open.isoformat()}, expected_close={expected_close.isoformat()}, "
                                f"now={now_local.isoformat()}"
                            )
                    elif shifts_today:
                        # Логика: учитываем все смены, которые фактически открыли объект
                        # Берём первую смену, у которой actual_start или start_time >= expected_open
                        naive_expected = datetime.combine(today_local, obj.opening_time)
                        expected_open = timezone_helper.local_tz.localize(naive_expected)
                        logger.debug(
                            f"Dashboard: object {obj.id} ({obj.name}) expected_open={expected_open.isoformat()}"
                        )

                        # Ищем все смены, которые фактически открыли объект (actual_start или start_time >= expected_open)
                        candidates_after: list[tuple[datetime, object]] = []
                        for s in shifts_today:
                            start_local = None
                            if s.actual_start:
                                start_local = timezone_helper.utc_to_local(s.actual_start)
                            elif s.start_time:
                                start_local = timezone_helper.utc_to_local(s.start_time)
                            
                            if start_local:
                                logger.debug(
                                    f"Dashboard: shift {s.id}: start_local={start_local.isoformat()}, "
                                    f"planned_start={s.planned_start}, "
                                    f"start >= expected_open? {start_local >= expected_open}"
                                )
                                if start_local >= expected_open:
                                    candidates_after.append((start_local, s))
                        
                        logger.debug(
                            f"Dashboard: candidates_after={len(candidates_after)}, "
                            f"candidates={[(c[0].isoformat(), c[1].id) for c in candidates_after] if candidates_after else []}"
                        )

                        if candidates_after:
                            # Берём первое фактическое открытие (минимальный start_local >= expected_open)
                            earliest_open_local, opener_shift = min(candidates_after, key=lambda t: t[0])
                            delay_minutes = int((earliest_open_local - expected_open).total_seconds() / 60)
                            logger.debug(
                                f"Dashboard: found earliest_open_local={earliest_open_local.isoformat()}, "
                                f"delay_minutes={delay_minutes}, opener_shift_id={opener_shift.id}"
                            )
                        else:
                            # Нет открытия после времени начала работы — берём первую смену дня как открывающую
                            # Это может быть случай, когда смена началась до expected_open (раньше времени открытия объекта)
                            if shifts_today:
                                # Берём первую смену по времени начала (actual_start или start_time)
                                first_shift = None
                                first_start_local = None
                                for s in shifts_today:
                                    start_local = None
                                    if s.actual_start:
                                        start_local = timezone_helper.utc_to_local(s.actual_start)
                                    elif s.start_time:
                                        start_local = timezone_helper.utc_to_local(s.start_time)
                                    
                                    if start_local and (first_start_local is None or start_local < first_start_local):
                                        first_start_local = start_local
                                        first_shift = s
                                
                                if first_shift:
                                    opener_shift = first_shift
                                    earliest_open_local = first_start_local
                                    # Если смена началась до expected_open, считаем без опоздания (отрицательная задержка = 0)
                                    if earliest_open_local < expected_open:
                                        delay_minutes = 0
                                    else:
                                        delay_minutes = int((earliest_open_local - expected_open).total_seconds() / 60)
                                    logger.debug(
                                        f"Dashboard: no candidates_after, using first shift {first_shift.id}, "
                                        f"start_local={earliest_open_local.isoformat()}, delay_minutes={delay_minutes}"
                                    )
                                else:
                                    delay_minutes = 0
                                    earliest_open_local = None
                                    opener_shift = None
                                    logger.debug(f"Dashboard: no candidates_after, no shifts with start_time, setting delay_minutes=0")
                            else:
                                # Нет смен вообще
                                delay_minutes = 0
                                earliest_open_local = None
                                opener_shift = None
                                logger.debug(f"Dashboard: no candidates_after, no shifts today, setting delay_minutes=0")

                        logger.debug(
                            f"Dashboard: object {obj.id} ({obj.name}) expected_open={expected_open.isoformat()}, "
                            f"earliest_open_local={earliest_open_local.isoformat() if earliest_open_local else None}, "
                            f"delay_minutes={delay_minutes}"
                        )
                        
                        # Последняя смена для расчёта закрытия (как было)
                        last_shift = max(shifts_today, key=lambda s: s.end_time or s.start_time or datetime.min.replace(tzinfo=timezone.utc))
                        
                        # Логирование для диагностики
                        logger.debug(
                            f"Dashboard: object {obj.id} ({obj.name}) - "
                            f"opener_shift={opener_shift.id if opener_shift else None}, "
                            f"opener_shift.user={opener_shift.user.id if opener_shift and opener_shift.user else None}, "
                            f"last_shift={last_shift.id if last_shift else None}, "
                            f"last_shift.user={last_shift.user.id if last_shift and last_shift.user else None}"
                        )
                        
                        # Проверяем наличие активных смен
                        active_shifts_on_object = [s for s in shifts_today if s.status == 'active']
                        
                        # Получить эффективные настройки опоздания (с учетом наследования)
                        async def get_effective_late_settings_for_object(obj: Object) -> dict:
                            if not obj.inherit_late_settings and obj.late_threshold_minutes is not None:
                                return {
                                    'threshold_minutes': obj.late_threshold_minutes,
                                    'penalty_per_minute': obj.late_penalty_per_minute
                                }
                            if obj.org_unit_id:
                                from domain.entities.org_structure import OrgStructureUnit
                                current_unit_id = obj.org_unit_id
                                while current_unit_id:
                                    unit_query = select(OrgStructureUnit).where(OrgStructureUnit.id == current_unit_id)
                                    unit_result = await session.execute(unit_query)
                                    unit = unit_result.scalar_one_or_none()
                                    if not unit:
                                        break
                                    if not unit.inherit_late_settings and unit.late_threshold_minutes is not None:
                                        return {
                                            'threshold_minutes': unit.late_threshold_minutes,
                                            'penalty_per_minute': unit.late_penalty_per_minute
                                        }
                                    current_unit_id = unit.parent_id
                            return {'threshold_minutes': None, 'penalty_per_minute': None}
                        
                        late_settings = await get_effective_late_settings_for_object(obj)
                        threshold_minutes = late_settings.get('threshold_minutes') or 0
                        
                        # Сначала определяем сотрудника для открытия (если есть opener_shift)
                        opener_employee = None
                        if opener_shift:
                            if opener_shift.user:
                                opener_employee = f"{opener_shift.user.last_name} {opener_shift.user.first_name}".strip()
                            else:
                                # Если смена есть, но user не загружен, загружаем его
                                await session.refresh(opener_shift, ['user'])
                                if opener_shift.user:
                                    opener_employee = f"{opener_shift.user.last_name} {opener_shift.user.first_name}".strip()
                        
                        if delay_minutes > threshold_minutes:
                            work_status = 'late_opening'
                            work_delay = delay_minutes
                            work_employee = opener_employee
                        else:
                            work_status = 'timely_opening'
                            work_employee = opener_employee
                        
                        # Проверка закрытия (как было)
                        # Только если объект не открыт (нет активных смен) и последняя смена завершена
                        if last_shift.end_time and last_shift.status == 'completed' and not active_shifts_on_object:
                            # Используем timezone объекта, а не default_timezone
                            obj_timezone = obj.timezone or "Europe/Moscow"
                            naive_expected_close = datetime.combine(today_local, obj.closing_time)
                            obj_tz = pytz.timezone(obj_timezone)
                            expected_close = obj_tz.localize(naive_expected_close)
                            actual_close_local = timezone_helper.utc_to_local(last_shift.end_time, timezone_str=obj_timezone)
                            early_minutes = int((expected_close - actual_close_local).total_seconds() / 60)
                            
                            # Проверка: если все смены завершены и прошло 10 минут после закрытия - статус "нет смен"
                            # Только до времени закрытия объекта (closing_time)
                            all_shifts_completed = all(s.status in ('completed', 'closed') for s in shifts_today)
                            close_time_utc = last_shift.end_time
                            delay_minutes = 10
                            notification_time_utc = close_time_utc + timedelta(minutes=delay_minutes)
                            now_utc = datetime.now(timezone.utc)
                            now_local = timezone_helper.utc_to_local(now_utc, timezone_str=obj_timezone)
                            
                            # Максимальное время для установки статуса "нет смен" - до времени закрытия объекта
                            max_status_time = expected_close
                            
                            if all_shifts_completed and now_utc >= notification_time_utc and now_local <= max_status_time:
                                work_status = 'no_shifts_today'
                                work_employee = None
                                logger.debug(
                                    f"Dashboard: object {obj.id} ({obj.name}) - all shifts completed, "
                                    f"close_time={close_time_utc.isoformat()}, "
                                    f"notification_time={notification_time_utc.isoformat()}, "
                                    f"now={now_utc.isoformat()}, max_status_time={max_status_time.isoformat()}"
                                )
                            # Порог раннего закрытия: 5 минут (хардкод, можно вынести в настройки объекта)
                            elif early_minutes > 5:
                                work_status = 'early_closing'
                                work_early = early_minutes
                                # Используем сотрудника из last_shift, если он есть, иначе из opener_shift
                                closer_employee = None
                                if last_shift.user:
                                    closer_employee = f"{last_shift.user.last_name} {last_shift.user.first_name}".strip()
                                elif last_shift:
                                    # Если смена есть, но user не загружен, загружаем его
                                    await session.refresh(last_shift, ['user'])
                                    if last_shift.user:
                                        closer_employee = f"{last_shift.user.last_name} {last_shift.user.first_name}".strip()
                                
                                # Если не удалось получить из last_shift, используем opener_employee
                                # НЕ перезаписываем work_employee, если он уже установлен (для своевременного открытия/закрытия)
                                if closer_employee:
                                    work_employee = closer_employee
                                elif not work_employee:
                                    # Только если work_employee еще не установлен, используем opener_employee
                                    work_employee = opener_employee or "Неизвестный"
                            else:
                                work_status = 'closed'
                                # Используем сотрудника из last_shift (для своевременного закрытия всегда показываем closer_employee)
                                closer_employee = None
                                if last_shift.user:
                                    closer_employee = f"{last_shift.user.last_name} {last_shift.user.first_name}".strip()
                                elif last_shift:
                                    # Если смена есть, но user не загружен, загружаем его
                                    await session.refresh(last_shift, ['user'])
                                    if last_shift.user:
                                        closer_employee = f"{last_shift.user.last_name} {last_shift.user.first_name}".strip()
                                
                                # Для своевременного закрытия всегда используем closer_employee, если он есть
                                if closer_employee:
                                    work_employee = closer_employee
                                elif not work_employee:
                                    # Только если work_employee еще не установлен, используем opener_employee
                                    work_employee = opener_employee or "Неизвестный"
                
                all_objects.append(
                    type("OwnerObjectRow", (), {
                        "id": obj.id,
                        "name": obj.name,
                        "address": obj.address,
                        "status": status,
                        "status_time": status_time,
                        "work_status": work_status,
                        "work_employee": work_employee,
                        "work_delay": work_delay,
                        "work_early": work_early,
                    })
                )
            
            # Получаем данные для переключения интерфейсов
            available_interfaces = await get_available_interfaces_for_user(user_id)
            
            # Получаем enabled_features для меню
            from shared.services.system_features_service import SystemFeaturesService
            features_service = SystemFeaturesService()
            enabled_features = await features_service.get_enabled_features(session, user_id)
        
        stats = {
            'total_objects': total_objects,
            'total_shifts': total_shifts,
            'active_shifts': active_shifts,
        }
        
        return templates.TemplateResponse("owner/dashboard.html", {
            "request": request,
            "current_user": current_user,
            "title": "Дашборд владельца",
            "stats": stats,
            "all_objects": all_objects,
            "available_interfaces": available_interfaces,
            "enabled_features": enabled_features,
        })
    except Exception as e:
        logger.error(f"Error loading owner dashboard: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка загрузки дашборда: {str(e)}")


@router.get("/dashboard", response_class=HTMLResponse)
async def owner_dashboard_redirect(request: Request):
    """Редирект с /owner/dashboard на /owner/"""
    return RedirectResponse(url="/owner/", status_code=status.HTTP_302_FOUND)


@router.get("/features", response_class=HTMLResponse, name="owner_features")
async def owner_features(request: Request, session: AsyncSession = Depends(get_db_session)):
    """Страница управления функциями системы"""
    current_user = await get_current_user(request)
    if not current_user:
        return RedirectResponse(url="/auth/login", status_code=status.HTTP_302_FOUND)
    
    user_role = current_user.get("role", "employee")
    if user_role != "owner":
        return RedirectResponse(url="/dashboard", status_code=status.HTTP_302_FOUND)
    
    telegram_id = current_user.get("id")
    # Получить User.id
    from sqlalchemy import select
    from domain.entities.user import User
    user_result = await session.execute(
        select(User.id).where(User.telegram_id == telegram_id)
    )
    user_id = user_result.scalar_one_or_none()
    if not user_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    
    owner_context = await get_owner_context(user_id, session)
    
    return templates.TemplateResponse("owner/features.html", {
        "request": request,
        "current_user": current_user,
        **owner_context
    })


@router.get("/notifications/center", response_class=HTMLResponse, name="owner_notifications_center")
async def owner_notifications_center(request: Request, session: AsyncSession = Depends(get_db_session)):
    """Центр уведомлений владельца"""
    current_user = await get_current_user(request)
    if not current_user:
        return RedirectResponse(url="/auth/login", status_code=status.HTTP_302_FOUND)
    
    user_role = current_user.get("role", "employee")
    if user_role != "owner":
        return RedirectResponse(url="/dashboard", status_code=status.HTTP_302_FOUND)
    
    telegram_id = current_user.get("id")
    
    # Получить User.id для контекста
    from sqlalchemy import select
    from domain.entities.user import User
    user_result = await session.execute(
        select(User.id).where(User.telegram_id == telegram_id)
    )
    user_id = user_result.scalar_one_or_none()
    
    if not user_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    
    owner_context = await get_owner_context(user_id, session)
    
    return templates.TemplateResponse("owner/notifications/center.html", {
        "request": request,
        "current_user": current_user,
        **owner_context
    })


@router.get("/notifications", response_class=HTMLResponse, name="owner_notifications")
async def owner_notifications(request: Request, session: AsyncSession = Depends(get_db_session)):
    """Страница настроек уведомлений владельца"""
    current_user = await get_current_user(request)
    if not current_user:
        return RedirectResponse(url="/auth/login", status_code=status.HTTP_302_FOUND)
    
    user_role = current_user.get("role", "employee")
    if user_role != "owner":
        return RedirectResponse(url="/dashboard", status_code=status.HTTP_302_FOUND)
    
    telegram_id = current_user.get("id")
    
    # Получить User.id
    from sqlalchemy import select
    from domain.entities.user import User
    user_result = await session.execute(
        select(User.id).where(User.telegram_id == telegram_id)
    )
    user_id = user_result.scalar_one_or_none()
    if not user_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    
    # Получить типы уведомлений из мета-таблицы
    from shared.services.notification_service import NotificationService
    from shared.services.notification_type_meta_service import NotificationTypeMetaService
    
    notification_service = NotificationService()
    meta_service = NotificationTypeMetaService()
    
    # Получить настройки пользователя
    # Получить User для доступа к JSON-полю notification_preferences
    user_result = await session.execute(
        select(User).where(User.id == user_id)
    )
    user = user_result.scalar_one_or_none()
    if not user:
        user_settings = {}
    else:
        # notification_preferences - это JSON поле: {type_code: {"telegram": bool, "inapp": bool}}
        user_settings = user.notification_preferences or {}
    
    # Получить типы, сгруппированные по категориям (только для настройки владельцем)
    types_grouped = await meta_service.get_types_grouped_by_category(
        session,
        user_configurable_only=True,
        active_only=True
    )
    
    # Названия категорий на русском
    category_names = {
        "shifts": "Смены",
        "objects": "Объекты",
        "contracts": "Договоры",
        "reviews": "Отзывы",
        "payments": "Платежи",
        "system": "Системные",
        "tasks": "Задачи",
        "applications": "Заявки",
        "incidents": "Инциденты"
    }
    
    # Формирование данных для шаблона
    priority_labels = {
        "critical": "Критический",
        "high": "Высокий",
        "normal": "Обычный",
        "low": "Низкий"
    }
    
    categories_data = []
    for category_code in ["shifts", "objects", "contracts", "reviews", "payments", "tasks", "applications", "incidents", "system"]:
        if category_code not in types_grouped:
            continue
        
        types_in_category = []
        for type_meta in types_grouped[category_code]:
            user_pref = user_settings.get(type_meta.type_code, {})
            types_in_category.append({
                "type_code": type_meta.type_code,
                "title": type_meta.title,
                "description": type_meta.description,
                "priority": type_meta.default_priority,
                "priority_label": priority_labels.get(type_meta.default_priority, type_meta.default_priority),
                "telegram_enabled": user_pref.get("telegram", True),
                "inapp_enabled": user_pref.get("inapp", True),
            })
        
        if types_in_category:
            categories_data.append({
                "code": category_code,
                "name": category_names.get(category_code, category_code),
                "types": types_in_category
            })
    
    owner_context = await get_owner_context(user_id, session)
    
    return templates.TemplateResponse("owner/notifications.html", {
        "request": request,
        "current_user": current_user,
        "categories": categories_data,
        **owner_context
    })


@router.post("/notifications")
async def owner_notifications_save(
    request: Request,
    session: AsyncSession = Depends(get_db_session)
):
    """Сохранение настроек уведомлений владельца"""
    current_user = await get_current_user(request)
    if not current_user or current_user.get("role") != "owner":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Access denied")
    
    telegram_id = current_user.get("id")
    
    # Получить User.id
    from sqlalchemy import select
    from domain.entities.user import User
    user_result = await session.execute(
        select(User.id).where(User.telegram_id == telegram_id)
    )
    user_id = user_result.scalar_one_or_none()
    if not user_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    
    # Парсинг формы
    form_data = await request.form()
    
    # Получить список доступных типов из мета-таблицы
    from shared.services.notification_service import NotificationService
    from shared.services.notification_type_meta_service import NotificationTypeMetaService
    
    notification_service = NotificationService()
    meta_service = NotificationTypeMetaService()
    
    # Получить типы, доступные для настройки владельцем
    user_configurable_types = await meta_service.get_user_configurable_types(
        session,
        active_only=True
    )
    
    # Сохранить настройки для каждого типа
    for type_meta in user_configurable_types:
        type_code = type_meta.type_code
        telegram_key = f"telegram_{type_code}"
        inapp_key = f"inapp_{type_code}"
        
        telegram_enabled = telegram_key in form_data
        inapp_enabled = inapp_key in form_data
        
        await notification_service.set_user_notification_preference(
            user_id=user_id,
            notification_type=type_code,
            channel_telegram=telegram_enabled,
            channel_inapp=inapp_enabled
        )
    
    logger.info(
        "Настройки уведомлений обновлены",
        user_id=user_id,
        telegram_id=telegram_id
    )
    
    return RedirectResponse(url="/owner/notifications", status_code=status.HTTP_303_SEE_OTHER)


@router.get("/objects", response_class=HTMLResponse, name="owner_objects")
async def owner_objects(
    request: Request,
    show_inactive: bool = Query(False),
    view_mode: str = Query("cards")
):
    """Список объектов владельца"""
    # Проверяем авторизацию и роль владельца
    current_user = await get_current_user(request)
    user_role = current_user.get("role", "employee")
    if user_role != "owner":
        return RedirectResponse(url="/auth/login", status_code=status.HTTP_302_FOUND)
    
    try:
        from apps.web.services.object_service import ObjectService
        
        async with get_async_session() as session:
            # Получение объектов владельца из базы данных
            object_service = ObjectService(session)
            # Получаем telegram_id пользователя
            telegram_id = current_user.get("telegram_id") or current_user.get("telegram_id") or current_user.get("id")
            if not telegram_id:
                raise HTTPException(status_code=400, detail="Пользователь не найден")
            
            objects = await object_service.get_objects_by_owner(telegram_id, include_inactive=show_inactive)
            
            # Преобразуем в формат для шаблона
            objects_data = []
            for obj in objects:
                objects_data.append({
                    "id": obj.id,
                    "name": obj.name,
                    "address": obj.address or "",
                    "hourly_rate": float(obj.hourly_rate),
                    "opening_time": obj.opening_time.strftime("%H:%M"),
                    "closing_time": obj.closing_time.strftime("%H:%M"),
                    "max_distance": obj.max_distance_meters,
                    "is_active": obj.is_active,
                    "available_for_applicants": obj.available_for_applicants,
                    "created_at": obj.created_at.strftime("%Y-%m-%d"),
                    "owner_id": obj.owner_id,
                    "work_conditions": obj.work_conditions or "",
                    "employee_position": obj.employee_position or "",
                    "shift_tasks": obj.shift_tasks or []
                })
            
            # Получаем данные для переключения интерфейсов
            user_id = await get_user_id_from_current_user(current_user, session)
            available_interfaces = await get_available_interfaces_for_user(user_id)
            
            return templates.TemplateResponse("owner/objects/list.html", {
                "request": request,
                "title": "Управление объектами",
                "objects": objects_data,
                "current_user": current_user,
                "available_interfaces": available_interfaces,
                "show_inactive": show_inactive,
                "view_mode": view_mode
            })
            
    except Exception as e:
        logger.error(f"Error loading objects list: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки списка объектов")


@router.get("/objects/create", response_class=HTMLResponse, name="owner_objects_create")
async def owner_objects_create(request: Request):
    """Форма создания объекта"""
    # Проверяем авторизацию и роль владельца
    current_user = await get_current_user(request)
    user_role = current_user.get("role", "employee")
    if user_role != "owner":
        return RedirectResponse(url="/auth/login", status_code=status.HTTP_302_FOUND)
    
    # Получаем данные для переключения интерфейсов, графики выплат и подразделения
    async with get_async_session() as session:
        user_id = await get_user_id_from_current_user(current_user, session)
        available_interfaces = await get_available_interfaces_for_user(user_id)
        
        # Загрузить графики выплат (системные + кастомные владельца)
        from domain.entities.payment_schedule import PaymentSchedule
        schedules_query = select(PaymentSchedule).where(
            PaymentSchedule.is_active == True
        ).where(
            (PaymentSchedule.owner_id == None) |  # Системные
            (PaymentSchedule.owner_id == user_id)  # Кастомные владельца
        ).order_by(PaymentSchedule.is_custom.asc(), PaymentSchedule.id.asc())
        schedules_result = await session.execute(schedules_query)
        payment_schedules = schedules_result.scalars().all()
        
        # Загрузить подразделения владельца
        from apps.web.services.org_structure_service import OrgStructureService
        org_service = OrgStructureService(session)
        org_units = await org_service.get_units_by_owner(user_id)
        org_tree = await org_service.get_org_tree(user_id)
    
    return templates.TemplateResponse("owner/objects/create.html", {
        "request": request,
        "title": "Создание объекта",
        "current_user": current_user,
        "available_interfaces": available_interfaces,
        "payment_schedules": payment_schedules,
        "org_units": org_units,
        "org_tree": org_tree
    })


@router.post("/objects/create")
async def owner_objects_create_post(
    request: Request,
    current_user: dict = Depends(require_owner_or_superadmin),
    db: AsyncSession = Depends(get_db_session)
):
    """Создание нового объекта"""
    try:
        from apps.web.services.object_service import ObjectService
        
        # Получение данных формы
        form_data = await request.form()
        
        name = form_data.get("name", "").strip()
        address = form_data.get("address", "").strip()
        hourly_rate_str = form_data.get("hourly_rate", "0").strip()
        opening_time = form_data.get("opening_time", "").strip()
        closing_time = form_data.get("closing_time", "").strip()
        max_distance_str = form_data.get("max_distance", "500").strip()
        latitude_str = form_data.get("latitude", "").strip()
        longitude_str = form_data.get("longitude", "").strip()
        
        logger.info(f"Creating object '{name}' for user {current_user['id']} (type: {type(current_user['id'])})")
        
        # Валидация обязательных полей
        if not name:
            raise HTTPException(status_code=400, detail="Название объекта обязательно")
        if not address:
            raise HTTPException(status_code=400, detail="Адрес объекта обязателен")
        
        # Валидация и преобразование числовых полей
        try:
            # Поддержка запятой как десятичного разделителя ("500,00")
            normalized_rate = hourly_rate_str.replace(",", ".") if hourly_rate_str else "0"
            hourly_rate = int(float(normalized_rate)) if normalized_rate else 0
        except ValueError:
            raise HTTPException(status_code=400, detail="Неверный формат ставки")
        
        try:
            max_distance = int(max_distance_str) if max_distance_str else 500
        except ValueError:
            raise HTTPException(status_code=400, detail="Неверный формат максимального расстояния")
        
        if hourly_rate <= 0:
            raise HTTPException(status_code=400, detail="Ставка должна быть больше 0")
        
        if max_distance <= 0:
            raise HTTPException(status_code=400, detail="Максимальное расстояние должно быть больше 0")
        
        # Обработка координат
        coordinates = None
        if latitude_str and longitude_str:
            try:
                lat = float(latitude_str)
                lon = float(longitude_str)
                # Проверяем диапазон координат
                if -90 <= lat <= 90 and -180 <= lon <= 180:
                    coordinates = f"{lat},{lon}"
                else:
                    raise HTTPException(status_code=400, detail="Координаты вне допустимого диапазона")
            except ValueError:
                raise HTTPException(status_code=400, detail="Неверный формат координат")
        
        # Обработка чекбоксов (они не отправляются, если не отмечены)
        available_for_applicants = "available_for_applicants" in form_data
        auto_create_timeslots = "auto_create_timeslots" in form_data
        
        # Обработка новых полей
        work_conditions = form_data.get("work_conditions", "").strip()
        employee_position = form_data.get("employee_position", "").strip()
        payment_system_id_str = form_data.get("payment_system_id", "").strip()
        payment_system_id = int(payment_system_id_str) if payment_system_id_str else None
        payment_schedule_id_str = form_data.get("payment_schedule_id", "").strip()
        payment_schedule_id = int(payment_schedule_id_str) if payment_schedule_id_str else None
        
        # Обработка настроек штрафов за опоздание
        # JavaScript создает скрытое поле со значением 'false' при снятии галочки
        inherit_late_settings_value = form_data.get("inherit_late_settings", "false")
        inherit_late_settings = inherit_late_settings_value not in ["false", ""]
        late_threshold_minutes_str = form_data.get("late_threshold_minutes", "").strip()
        late_threshold_minutes = int(late_threshold_minutes_str) if late_threshold_minutes_str else None
        late_penalty_per_minute_str = form_data.get("late_penalty_per_minute", "").strip()
        late_penalty_per_minute = float(late_penalty_per_minute_str.replace(",", ".")) if late_penalty_per_minute_str else None
        
        # Обработка настроек штрафов за отмену смены
        inherit_cancellation_settings_value = form_data.get("inherit_cancellation_settings", "false")
        inherit_cancellation_settings = inherit_cancellation_settings_value not in ["false", ""]
        cancellation_short_notice_hours_str = form_data.get("cancellation_short_notice_hours", "").strip()
        cancellation_short_notice_hours = int(cancellation_short_notice_hours_str) if cancellation_short_notice_hours_str else None
        cancellation_short_notice_fine_str = form_data.get("cancellation_short_notice_fine", "").strip()
        cancellation_short_notice_fine = float(cancellation_short_notice_fine_str.replace(",", ".")) if cancellation_short_notice_fine_str else None
        cancellation_invalid_reason_fine_str = form_data.get("cancellation_invalid_reason_fine", "").strip()
        cancellation_invalid_reason_fine = float(cancellation_invalid_reason_fine_str.replace(",", ".")) if cancellation_invalid_reason_fine_str else None
        
        # Обработка Telegram группы для отчетов
        # JavaScript создает скрытое поле со значением 'false' при снятии галочки
        inherit_telegram_chat_value = form_data.get("inherit_telegram_chat", "false")
        inherit_telegram_chat = inherit_telegram_chat_value not in ["false", ""]
        telegram_report_chat_id = form_data.get("telegram_report_chat_id", "").strip()
        telegram_report_chat_id = telegram_report_chat_id if telegram_report_chat_id else None
        
        # Обработка подразделения
        org_unit_id_str = form_data.get("org_unit_id", "").strip()
        org_unit_id = int(org_unit_id_str) if org_unit_id_str else None
        
        # Парсинг задач с новой структурой
        task_texts = form_data.getlist("task_texts[]")
        task_deductions = form_data.getlist("task_deductions[]")
        task_mandatory = form_data.getlist("task_mandatory[]")
        task_requires_media = form_data.getlist("task_requires_media[]")
        
        logger.info(f"Task parsing - texts: {task_texts}, deductions: {task_deductions}, mandatory: {task_mandatory}")
        
        shift_tasks = []
        for idx, text in enumerate(task_texts):
            if text.strip():
                is_mandatory = str(idx) in task_mandatory
                logger.info(f"Task {idx}: text='{text}', is_mandatory={is_mandatory} (checking '{idx}' in {task_mandatory})")
                shift_tasks.append({
                    "text": text.strip(),
                    "is_mandatory": is_mandatory,
                    "deduction_amount": float(task_deductions[idx]) if idx < len(task_deductions) else 100.0
                })
        
        # Обработка графика работы
        work_days = form_data.getlist("work_days")
        work_days_mask = 0
        for day in work_days:
            work_days_mask += int(day)
        
        schedule_repeat_weeks_str = form_data.get("schedule_repeat_weeks", "1").strip()
        try:
            schedule_repeat_weeks = int(schedule_repeat_weeks_str) if schedule_repeat_weeks_str else 1
        except ValueError:
            schedule_repeat_weeks = 1
        
        # Создание объекта в базе данных
        object_service = ObjectService(db)
        object_data = {
            "name": name,
            "address": address,
            "hourly_rate": hourly_rate,
            "opening_time": opening_time,
            "closing_time": closing_time,
            "max_distance": max_distance,
            "available_for_applicants": available_for_applicants,
            "auto_create_timeslots": auto_create_timeslots,
            "payment_system_id": payment_system_id,
            "payment_schedule_id": payment_schedule_id,
            "org_unit_id": org_unit_id,
            "is_active": True,
            "coordinates": coordinates,
            "work_days_mask": work_days_mask,
            "schedule_repeat_weeks": schedule_repeat_weeks,
            "work_conditions": work_conditions if work_conditions else None,
            "employee_position": employee_position if employee_position else None,
            "shift_tasks": shift_tasks if shift_tasks else None,
            "inherit_late_settings": inherit_late_settings,
            "late_threshold_minutes": late_threshold_minutes,
            "late_penalty_per_minute": late_penalty_per_minute,
            "inherit_cancellation_settings": inherit_cancellation_settings,
            "cancellation_short_notice_hours": cancellation_short_notice_hours,
            "cancellation_short_notice_fine": cancellation_short_notice_fine,
            "cancellation_invalid_reason_fine": cancellation_invalid_reason_fine
        }
        
        # Передаем telegram_id в create_object (метод ожидает telegram_id)
        telegram_id = current_user['id']  # Это telegram_id из JWT payload
        
        new_object = await object_service.create_object(object_data, telegram_id)
        logger.info(f"Object {new_object.id} created successfully")
            
        return RedirectResponse(url="/owner/objects", status_code=status.HTTP_302_FOUND)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating object: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка создания объекта: {str(e)}")


@router.get("/objects/{object_id}", response_class=HTMLResponse, name="owner_objects_detail")
async def owner_objects_detail(request: Request, object_id: int):
    """Детальная информация об объекте"""
    # Проверяем авторизацию и роль владельца
    current_user = await get_current_user(request)
    user_role = current_user.get("role", "employee")
    if user_role != "owner":
        return RedirectResponse(url="/auth/login", status_code=status.HTTP_302_FOUND)
    
    try:
        from apps.web.services.object_service import ObjectService, TimeSlotService
        
        async with get_async_session() as session:
            # Получение данных объекта из базы данных с проверкой владельца
            object_service = ObjectService(session)
            timeslot_service = TimeSlotService(session)
            
            # Получаем telegram_id пользователя
            telegram_id = current_user.get("telegram_id") or current_user.get("telegram_id") or current_user.get("id")
            if not telegram_id:
                raise HTTPException(status_code=400, detail="Пользователь не найден")
            
            obj = await object_service.get_object_by_id(object_id, telegram_id)
            if not obj:
                raise HTTPException(status_code=404, detail="Объект не найден")
            
            # Получаем тайм-слоты
            timeslots = await timeslot_service.get_timeslots_by_object(object_id, telegram_id)
            
            # Преобразуем в формат для шаблона
            object_data = {
                "id": obj.id,
                "name": obj.name,
                "address": obj.address or "",
                "hourly_rate": float(obj.hourly_rate),
                "opening_time": obj.opening_time.strftime("%H:%M"),
                "closing_time": obj.closing_time.strftime("%H:%M"),
                "max_distance": obj.max_distance_meters,
                "is_active": obj.is_active,
                "available_for_applicants": obj.available_for_applicants,
                "created_at": obj.created_at.strftime("%Y-%m-%d"),
                "owner_id": obj.owner_id,
                "work_days_mask": obj.work_days_mask,
                "schedule_repeat_weeks": obj.schedule_repeat_weeks,
                "shift_tasks": obj.shift_tasks or [],
                "payment_system_id": obj.payment_system_id,
                "payment_schedule_id": obj.payment_schedule_id,
                "timeslots": [
                    {
                        "id": slot.id,
                        "start_time": slot.start_time.strftime("%H:%M"),
                        "end_time": slot.end_time.strftime("%H:%M"),
                        "hourly_rate": float(slot.hourly_rate) if slot.hourly_rate else float(obj.hourly_rate),
                        "is_active": slot.is_active
                    }
                    for slot in timeslots
                ]
            }
            
            # Получаем данные для переключения интерфейсов
            user_id = await get_user_id_from_current_user(current_user, session)
            available_interfaces = await get_available_interfaces_for_user(user_id)
            
            return templates.TemplateResponse("owner/objects/detail.html", {
                "request": request,
                "title": f"Объект: {object_data['name']}",
                "object": object_data,
                "available_interfaces": available_interfaces,
                "current_user": current_user
            })
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error loading object detail: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки информации об объекте")


@router.get("/objects/{object_id}/edit", response_class=HTMLResponse, name="owner_objects_edit")
async def owner_objects_edit(request: Request, object_id: int):
    """Форма редактирования объекта"""
    # Проверяем авторизацию и роль владельца
    current_user = await get_current_user(request)
    user_role = current_user.get("role", "employee")
    if user_role != "owner":
        return RedirectResponse(url="/auth/login", status_code=status.HTTP_302_FOUND)
    
    try:
        from apps.web.services.object_service import ObjectService
        
        async with get_async_session() as session:
            # Получение данных объекта из базы данных с проверкой владельца
            object_service = ObjectService(session)
            # Получаем telegram_id пользователя
            telegram_id = current_user.get("telegram_id") or current_user.get("telegram_id") or current_user.get("id")
            if not telegram_id:
                raise HTTPException(status_code=400, detail="Пользователь не найден")
            
            obj = await object_service.get_object_by_id(object_id, telegram_id)
            if not obj:
                raise HTTPException(status_code=404, detail="Объект не найден")
            
            # Загрузить графики выплат (системные + кастомные владельца + кастомные объекта)
            from domain.entities.payment_schedule import PaymentSchedule
            user_id = await get_user_id_from_current_user(current_user, session)
            
            schedules_query = select(PaymentSchedule).where(
                PaymentSchedule.is_active == True
            ).where(
                (PaymentSchedule.owner_id == None) |  # Системные
                (PaymentSchedule.owner_id == user_id) |  # Кастомные владельца
                (PaymentSchedule.object_id == object_id)  # Кастомные объекта
            ).order_by(PaymentSchedule.is_custom.asc(), PaymentSchedule.id.asc())
            schedules_result = await session.execute(schedules_query)
            payment_schedules = schedules_result.scalars().all()
            
            # Загрузить подразделения владельца
            from apps.web.services.org_structure_service import OrgStructureService
            org_service = OrgStructureService(session)
            org_units = await org_service.get_units_by_owner(user_id)
            org_tree = await org_service.get_org_tree(user_id)
            
            # Преобразуем в формат для шаблона
            object_data = {
                "id": obj.id,
                "name": obj.name,
                "address": obj.address or "",
                "coordinates": obj.coordinates or "",
                "hourly_rate": obj.hourly_rate,
                "opening_time": obj.opening_time.strftime("%H:%M") if obj.opening_time else "",
                "closing_time": obj.closing_time.strftime("%H:%M") if obj.closing_time else "",
                "max_distance": obj.max_distance_meters or 500,
                "available_for_applicants": obj.available_for_applicants,
                "is_active": obj.is_active,
                "work_days_mask": obj.work_days_mask,
                "schedule_repeat_weeks": obj.schedule_repeat_weeks,
                "work_conditions": obj.work_conditions or "",
                "employee_position": obj.employee_position or "",
                "shift_tasks": obj.shift_tasks or [],
                "payment_system_id": obj.payment_system_id,
                "payment_schedule_id": obj.payment_schedule_id,
                "org_unit_id": obj.org_unit_id if hasattr(obj, 'org_unit_id') else None,
                "inherit_late_settings": obj.inherit_late_settings if hasattr(obj, 'inherit_late_settings') else True,
                "late_threshold_minutes": obj.late_threshold_minutes if hasattr(obj, 'late_threshold_minutes') else None,
                "late_penalty_per_minute": obj.late_penalty_per_minute if hasattr(obj, 'late_penalty_per_minute') else None,
                "inherit_cancellation_settings": obj.inherit_cancellation_settings if hasattr(obj, 'inherit_cancellation_settings') else True,
                "cancellation_short_notice_hours": obj.cancellation_short_notice_hours if hasattr(obj, 'cancellation_short_notice_hours') else None,
                "cancellation_short_notice_fine": obj.cancellation_short_notice_fine if hasattr(obj, 'cancellation_short_notice_fine') else None,
                "cancellation_invalid_reason_fine": obj.cancellation_invalid_reason_fine if hasattr(obj, 'cancellation_invalid_reason_fine') else None,
                "inherit_telegram_chat": obj.inherit_telegram_chat if hasattr(obj, 'inherit_telegram_chat') else True,
                "telegram_report_chat_id": obj.telegram_report_chat_id if hasattr(obj, 'telegram_report_chat_id') else None
            }
            
            # Получаем данные для переключения интерфейсов
            user_id = await get_user_id_from_current_user(current_user, session)
            available_interfaces = await get_available_interfaces_for_user(user_id)
            
            # Загрузить подразделения (уже получили user_id и org_service выше, повторно создаем)
            from apps.web.services.org_structure_service import OrgStructureService
            org_service = OrgStructureService(session)
            org_units = await org_service.get_units_by_owner(user_id)
            org_tree = await org_service.get_org_tree(user_id)
            
            return templates.TemplateResponse("owner/objects/edit.html", {
                "request": request,
                "title": f"Редактирование: {object_data['name']}",
                "object": object_data,
                "available_interfaces": available_interfaces,
                "current_user": current_user,
                "payment_schedules": payment_schedules,
                "org_units": org_units,
                "org_tree": org_tree
            })
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error loading edit form: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки формы редактирования")
@router.post("/objects/{object_id}/edit")
async def owner_objects_edit_post(request: Request, object_id: int):
    """Обновление объекта"""
    # Проверяем авторизацию и роль владельца
    current_user = await get_current_user(request)
    user_role = current_user.get("role", "employee")
    if user_role != "owner":
        return RedirectResponse(url="/auth/login", status_code=status.HTTP_302_FOUND)
    
    try:
        from apps.web.services.object_service import ObjectService
        
        # Получение данных формы
        form_data = await request.form()
        
        # Логируем все данные формы
        logger.info(f"Form data keys: {list(form_data.keys())}")
        logger.info(f"Form data values: {dict(form_data)}")
        
        name = form_data.get("name", "").strip()
        address = form_data.get("address", "").strip()
        hourly_rate_str = form_data.get("hourly_rate", "0").strip()
        opening_time = form_data.get("opening_time", "").strip()
        closing_time = form_data.get("closing_time", "").strip()
        timezone = form_data.get("timezone", "Europe/Moscow").strip()
        max_distance_str = form_data.get("max_distance", "500").strip()
        latitude_str = form_data.get("latitude", "").strip()
        longitude_str = form_data.get("longitude", "").strip()
        
        logger.info(f"Updating object {object_id} for user {current_user['id']}")
        
        # Валидация обязательных полей
        if not name:
            raise HTTPException(status_code=400, detail="Название объекта обязательно")
        if not address:
            raise HTTPException(status_code=400, detail="Адрес объекта обязателен")
        
        # Валидация и преобразование числовых полей
        try:
            # Поддержка запятой как десятичного разделителя ("500,00")
            normalized_rate = hourly_rate_str.replace(",", ".") if hourly_rate_str else "0"
            hourly_rate = int(float(normalized_rate)) if normalized_rate else 0
        except ValueError:
            raise HTTPException(status_code=400, detail="Неверный формат ставки")
        
        try:
            max_distance = int(max_distance_str) if max_distance_str else 500
        except ValueError:
            raise HTTPException(status_code=400, detail="Неверный формат максимального расстояния")
        
        if hourly_rate <= 0:
            raise HTTPException(status_code=400, detail="Ставка должна быть больше 0")
        
        if max_distance <= 0:
            raise HTTPException(status_code=400, detail="Максимальное расстояние должно быть больше 0")
        
        # Обработка координат
        coordinates = None
        if latitude_str and longitude_str:
            try:
                lat = float(latitude_str)
                lon = float(longitude_str)
                # Проверяем диапазон координат
                if -90 <= lat <= 90 and -180 <= lon <= 180:
                    coordinates = f"{lat},{lon}"
                else:
                    raise HTTPException(status_code=400, detail="Координаты вне допустимого диапазона")
            except ValueError:
                raise HTTPException(status_code=400, detail="Неверный формат координат")
        
        # Обработка чекбоксов/скрытых значений
        def to_bool(value: Optional[str]) -> bool:
            if value is None:
                return False
            return value.lower() in ("true", "on", "1", "yes")

        available_for_applicants = to_bool(form_data.get("available_for_applicants"))
        is_active = to_bool(form_data.get("is_active"))
        
        # Обработка новых полей
        work_conditions = form_data.get("work_conditions", "").strip()
        employee_position = form_data.get("employee_position", "").strip()
        payment_system_id_str = form_data.get("payment_system_id", "").strip()
        payment_system_id = int(payment_system_id_str) if payment_system_id_str else None
        payment_schedule_id_str = form_data.get("payment_schedule_id", "").strip()
        payment_schedule_id = int(payment_schedule_id_str) if payment_schedule_id_str else None
        
        # Обработка настроек штрафов за опоздание
        # JavaScript создает скрытое поле со значением 'false' при снятии галочки
        inherit_late_settings_value = form_data.get("inherit_late_settings", "false")
        inherit_late_settings = inherit_late_settings_value not in ["false", ""]
        late_threshold_minutes_str = form_data.get("late_threshold_minutes", "").strip()
        late_threshold_minutes = int(late_threshold_minutes_str) if late_threshold_minutes_str else None
        late_penalty_per_minute_str = form_data.get("late_penalty_per_minute", "").strip()
        late_penalty_per_minute = float(late_penalty_per_minute_str.replace(",", ".")) if late_penalty_per_minute_str else None
        
        # Обработка настроек штрафов за отмену смены
        inherit_cancellation_settings_value = form_data.get("inherit_cancellation_settings", "false")
        inherit_cancellation_settings = inherit_cancellation_settings_value not in ["false", ""]
        cancellation_short_notice_hours_str = form_data.get("cancellation_short_notice_hours", "").strip()
        cancellation_short_notice_hours = int(cancellation_short_notice_hours_str) if cancellation_short_notice_hours_str else None
        cancellation_short_notice_fine_str = form_data.get("cancellation_short_notice_fine", "").strip()
        cancellation_short_notice_fine = float(cancellation_short_notice_fine_str.replace(",", ".")) if cancellation_short_notice_fine_str else None
        cancellation_invalid_reason_fine_str = form_data.get("cancellation_invalid_reason_fine", "").strip()
        cancellation_invalid_reason_fine = float(cancellation_invalid_reason_fine_str.replace(",", ".")) if cancellation_invalid_reason_fine_str else None
        
        # Обработка Telegram группы для отчетов
        # JavaScript создает скрытое поле со значением 'false' при снятии галочки
        inherit_telegram_chat_value = form_data.get("inherit_telegram_chat", "false")
        inherit_telegram_chat = inherit_telegram_chat_value not in ["false", ""]
        telegram_report_chat_id = form_data.get("telegram_report_chat_id", "").strip()
        telegram_report_chat_id = telegram_report_chat_id if telegram_report_chat_id else None
        
        # Обработка подразделения
        org_unit_id_str = form_data.get("org_unit_id", "").strip()
        org_unit_id = int(org_unit_id_str) if org_unit_id_str else None
        
        # Парсинг задач с новой структурой
        task_texts = form_data.getlist("task_texts[]")
        task_deductions = form_data.getlist("task_deductions[]")
        task_mandatory = form_data.getlist("task_mandatory[]")
        task_requires_media = form_data.getlist("task_requires_media[]")
        
        logger.info(f"Task parsing (edit) - texts: {task_texts}, deductions: {task_deductions}, mandatory: {task_mandatory}, requires_media: {task_requires_media}")
        
        shift_tasks = []
        for idx, text in enumerate(task_texts):
            if text.strip():
                is_mandatory = str(idx) in task_mandatory
                requires_media = str(idx) in task_requires_media
                logger.info(f"Task {idx}: text='{text}', is_mandatory={is_mandatory}, requires_media={requires_media}")
                shift_tasks.append({
                    "text": text.strip(),
                    "is_mandatory": is_mandatory,
                    "deduction_amount": float(task_deductions[idx]) if idx < len(task_deductions) else 100.0,
                    "requires_media": requires_media
                })
        
        logger.info(f"Form data - work_conditions: '{work_conditions}', shift_tasks: {shift_tasks}")
        
        # Обработка графика работы
        work_days_mask_str = form_data.get("work_days_mask", "0").strip()
        schedule_repeat_weeks_str = form_data.get("schedule_repeat_weeks", "1").strip()
        
        try:
            work_days_mask = int(work_days_mask_str) if work_days_mask_str else 0
        except ValueError:
            work_days_mask = 0
            
        try:
            schedule_repeat_weeks = int(schedule_repeat_weeks_str) if schedule_repeat_weeks_str else 1
        except ValueError:
            schedule_repeat_weeks = 1
        
        logger.info(f"Work days mask: {work_days_mask}, Schedule repeat weeks: {schedule_repeat_weeks}")
        
        # Обновление объекта в базе данных
        async with get_async_session() as db:
            object_service = ObjectService(db)
            object_data = {
                "name": name,
                "address": address,
                "hourly_rate": hourly_rate,
                "opening_time": opening_time,
                "closing_time": closing_time,
                "timezone": timezone,
                "max_distance": max_distance,
                "available_for_applicants": available_for_applicants,
                "payment_system_id": payment_system_id,
                "payment_schedule_id": payment_schedule_id,
                "is_active": is_active,
                "coordinates": coordinates,
                "work_days_mask": work_days_mask,
                "schedule_repeat_weeks": schedule_repeat_weeks,
                "work_conditions": work_conditions if work_conditions else None,
                "employee_position": employee_position if employee_position else None,
                "shift_tasks": shift_tasks if shift_tasks else None,
                "inherit_late_settings": inherit_late_settings,
                "late_threshold_minutes": late_threshold_minutes,
                "late_penalty_per_minute": late_penalty_per_minute,
                "inherit_cancellation_settings": inherit_cancellation_settings,
                "cancellation_short_notice_hours": cancellation_short_notice_hours,
                "cancellation_short_notice_fine": cancellation_short_notice_fine,
                "cancellation_invalid_reason_fine": cancellation_invalid_reason_fine,
                "inherit_telegram_chat": inherit_telegram_chat,
                "telegram_report_chat_id": telegram_report_chat_id,
                "org_unit_id": org_unit_id
            }
            
            # Получаем внутренний ID пользователя
            user_id = await get_user_id_from_current_user(current_user, db)
            if not user_id:
                raise HTTPException(status_code=400, detail="Пользователь не найден")
            
            updated_object = await object_service.update_object(object_id, object_data, user_id)
            if not updated_object:
                raise HTTPException(status_code=404, detail="Объект не найден или нет доступа")
            
            logger.info(f"Object {object_id} updated successfully")
            
        return RedirectResponse(url=f"/owner/objects/{object_id}", status_code=status.HTTP_302_FOUND)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating object: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка обновления объекта: {str(e)}")


@router.post("/objects/{object_id}/delete")
async def owner_objects_delete(
    object_id: int, 
    request: Request,
    current_user: dict = Depends(require_owner_or_superadmin),
    db: AsyncSession = Depends(get_db_session)
):
    """Полное удаление объекта из базы данных"""
    try:
        from apps.web.services.object_service import ObjectService
        
        logger.info(f"Hard deleting object {object_id} for user {current_user['id']}")

        object_service = ObjectService(db)
        # Получаем внутренний ID пользователя
        user_id = await get_user_id_from_current_user(current_user, db)
        if not user_id:
            raise HTTPException(status_code=400, detail="Пользователь не найден")
        
        success = await object_service.hard_delete_object(object_id, user_id)
        if not success:
            raise HTTPException(status_code=404, detail="Объект не найден или нет доступа")

        return RedirectResponse(url="/owner/objects", status_code=status.HTTP_302_FOUND)
        
    except Exception as e:
        logger.error(f"Error hard deleting object: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка удаления объекта: {str(e)}")


# ===============================
# КАЛЕНДАРЬ
# ===============================


@router.get("/calendar", response_class=HTMLResponse, name="owner_calendar")
async def owner_calendar(
    request: Request,
    year: int = Query(None),
    month: int = Query(None),
    object_id: int = Query(None),
    org_unit_id: int = Query(None)
):
    """Календарный вид планирования"""
    current_user = await get_current_user(request)
    if not current_user or current_user.get("role") != "owner":
        return RedirectResponse(url="/auth/login", status_code=status.HTTP_302_FOUND)
    
    try:
        # Определяем текущую дату или переданные параметры
        today = date.today()
        if year is None:
            year = today.year
        if month is None:
            month = today.month
        
        # Валидация даты
        if not (1 <= month <= 12):
            month = today.month
        if year < 2020 or year > 2030:
            year = today.year
        
        # Получаем объекты пользователя
        async with get_async_session() as session:
            object_service = ObjectService(session)
            timeslot_service = TimeSlotService(session)
            
            # ОРИГИНАЛ: используем telegram_id владельца (а не внутренний id)
            owner_telegram_id = current_user.get("telegram_id") or current_user.get("telegram_id") or current_user.get("id")
            
            # Получаем подразделения владельца
            from apps.web.services.org_structure_service import OrgStructureService
            org_service = OrgStructureService(session)
            user_id = await get_user_id_from_current_user(current_user, session)
            org_units_raw = await org_service.get_units_by_owner(user_id)
            # Преобразуем в список словарей для шаблона
            org_units = [{"id": u.id, "name": u.name, "level": u.level} for u in org_units_raw]
            
            # Получаем ВСЕ объекты владельца (для JS фильтра)
            all_objects = await object_service.get_objects_by_owner(owner_telegram_id)
            
            # Фильтрация по подразделению если выбрано (для календаря)
            if org_unit_id:
                filtered_objects = [obj for obj in all_objects if obj.org_unit_id == org_unit_id]
                logger.info(f"Filtering by org_unit_id={org_unit_id}: {len(filtered_objects)}/{len(all_objects)} objects")
                logger.info(f"Filtered object IDs: {[obj.id for obj in filtered_objects]}")
            else:
                filtered_objects = all_objects
                logger.info(f"No org_unit filter, loading all {len(all_objects)} objects")
                logger.info(f"All object IDs: {[obj.id for obj in all_objects]}")
            
            # Если выбран конкретный объект, проверяем доступ
            selected_object = None
            if object_id:
                for obj in filtered_objects:
                    if obj.id == object_id:
                        selected_object = obj
                        break
                if not selected_object:
                    raise HTTPException(status_code=404, detail="Объект не найден")
            
            # Определяем объекты для загрузки данных (используем отфильтрованные)
            objects_to_load = [selected_object] if selected_object else filtered_objects
            
            # Получаем тайм-слоты для выбранного объекта или всех объектов
            timeslots_data = []
            logger.info(f"Selected object: {selected_object}, Objects to load count: {len(objects_to_load)}")
            
            # Загружаем тайм-слоты для выбранных объектов
            from sqlalchemy import select
            from domain.entities.time_slot import TimeSlot
            
            for obj in objects_to_load:
                query = select(TimeSlot).where(
                    TimeSlot.object_id == obj.id,
                    TimeSlot.is_active == True
                ).order_by(TimeSlot.slot_date, TimeSlot.start_time)
                
                result = await session.execute(query)
                timeslots = result.scalars().all()
                logger.info(f"Found {len(timeslots)} timeslots for object {obj.id} ({obj.name})")
                
                for slot in timeslots:
                    timeslots_data.append({
                        "id": slot.id,
                        "object_id": slot.object_id,
                        "object_name": obj.name,
                        "date": slot.slot_date,
                        "start_time": slot.start_time.strftime("%H:%M"),
                        "end_time": slot.end_time.strftime("%H:%M"),
                        "hourly_rate": float(slot.hourly_rate) if slot.hourly_rate else float(obj.hourly_rate),
                        "max_employees": slot.max_employees if slot.max_employees is not None else 1,
                        "is_active": slot.is_active,
                        "notes": slot.notes or ""
                    })
                    # Debug для тайм-слота 508
                    if slot.id == 508:
                        logger.info(f"Found timeslot 508: object_id={slot.object_id}, max_employees={slot.max_employees}, is_active={slot.is_active}")
            
            logger.info(f"Total timeslots_data: {len(timeslots_data)}")
            
            # Debug: проверим, есть ли тайм-слот 508 в данных
            timeslot_508 = next((ts for ts in timeslots_data if ts["id"] == 508), None)
            if timeslot_508:
                logger.info(f"Timeslot 508 in data: {timeslot_508}")
            else:
                logger.warning("Timeslot 508 NOT found in timeslots_data")
            
            # Получаем смены для выбранного объекта или всех объектов
            shifts_data = []
            try:
                from sqlalchemy import select, and_
                from domain.entities.shift import Shift
                from domain.entities.shift_schedule import ShiftSchedule
                from domain.entities.user import User
                
                # Получаем владельца по telegram_id
                owner_q = select(User).where(User.telegram_id == owner_telegram_id)
                owner = (await session.execute(owner_q)).scalar_one_or_none()
                if owner:
                    # Получаем смены за месяц
                    start_date = date(year, month, 1)
                    end_date = date(year, month, 28) + timedelta(days=4)  # До конца месяца
                    
                    from sqlalchemy.orm import selectinload
                    
                    # Загружаем активные и завершенные смены
                    shifts_query = select(Shift).where(
                        and_(
                            Shift.object_id.in_([obj.id for obj in objects_to_load]),
                            Shift.start_time >= datetime.combine(start_date, time.min),
                            Shift.start_time <= datetime.combine(end_date, time.max)
                        )
                    ).options(selectinload(Shift.user))
                    shifts = (await session.execute(shifts_query)).scalars().all()
                    logger.info(f"Found {len(shifts)} active/completed shifts in database for objects: {[obj.id for obj in objects_to_load]}")
                    
                    for shift in shifts:
                        # Находим объект для смены
                        shift_object = next((obj for obj in all_objects if obj.id == shift.object_id), None)
                        object_name = shift_object.name if shift_object else "Неизвестный объект"
                        
                        shifts_data.append({
                            "id": shift.id,
                            "object_id": shift.object_id,
                            "time_slot_id": getattr(shift, "time_slot_id", None),
                            "object_name": object_name,
                            "date": shift.start_time.date(),
                            "start_time": web_timezone_helper.format_time_with_timezone(shift.start_time, shift.object.timezone if shift.object else 'Europe/Moscow'),
                            "end_time": web_timezone_helper.format_time_with_timezone(shift.end_time, shift.object.timezone if shift.object else 'Europe/Moscow') if shift.end_time else "",
                            "employee_name": f"{shift.user.first_name} {shift.user.last_name}".strip() if shift.user else "Неизвестно",
                            "status": shift.status,
                            "total_hours": float(shift.total_hours) if shift.total_hours else 0,
                            "total_payment": float(shift.total_payment) if shift.total_payment else 0,
                            "status_label": shift.status_label
                        })
                    
                    # Загружаем запланированные смены (исключаем отмененные)
                    schedules_query = select(ShiftSchedule).where(
                        and_(
                            ShiftSchedule.object_id.in_([obj.id for obj in objects_to_load]),
                            ShiftSchedule.planned_start >= datetime.combine(start_date, time.min),
                            ShiftSchedule.planned_start <= datetime.combine(end_date, time.max),
                            ShiftSchedule.status != 'cancelled'  # Исключаем отмененные
                        )
                    ).options(selectinload(ShiftSchedule.user))
                    schedules = (await session.execute(schedules_query)).scalars().all()
                    logger.info(f"Found {len(schedules)} planned shifts in database for objects: {[obj.id for obj in objects_to_load]}")
                    
                    for schedule in schedules:
                        # Находим объект для запланированной смены
                        schedule_object = next((obj for obj in all_objects if obj.id == schedule.object_id), None)
                        object_name = schedule_object.name if schedule_object else "Неизвестный объект"
                        
                        shifts_data.append({
                            "id": f"schedule_{schedule.id}",  # Префикс для отличия от обычных смен
                            "object_id": schedule.object_id,
                            "time_slot_id": getattr(schedule, "time_slot_id", None),
                            "object_name": object_name,
                            "date": schedule.planned_start.date(),
                            "start_time": web_timezone_helper.format_time_with_timezone(schedule.planned_start, schedule_object.timezone if schedule_object else 'Europe/Moscow'),
                            "end_time": web_timezone_helper.format_time_with_timezone(schedule.planned_end, schedule_object.timezone if schedule_object else 'Europe/Moscow') if schedule.planned_end else "",
                            "employee_name": f"{schedule.user.first_name} {schedule.user.last_name}".strip() if schedule.user else "Неизвестно",
                            "status": schedule.status or "planned",
                            "total_hours": 0,
                            "total_payment": 0,
                            "status_label": schedule.status_label
                        })
                    
                    
                    logger.info(f"Loaded {len(shifts_data)} total shifts (active + planned) for calendar")
            except Exception as e:
                logger.warning(f"Could not load shifts for calendar: {e}")
                shifts_data = []
            
            # Создаем календарную сетку
            logger.info(f"Creating calendar grid with {len(timeslots_data)} timeslots and {len(shifts_data)} shifts")
            calendar_data = _create_calendar_grid(year, month, timeslots_data, shifts_data)
            logger.info(f"Calendar grid created with {len(calendar_data)} weeks")
            
            # Подготавливаем данные для шаблона
            # ВСЕ объекты для JS фильтра (не отфильтрованные)
            objects_list = [{"id": obj.id, "name": obj.name, "org_unit_id": obj.org_unit_id} for obj in all_objects]
            org_units_list = org_units
            
            # Навигация по месяцам
            prev_month = month - 1 if month > 1 else 12
            prev_year = year if month > 1 else year - 1
            next_month = month + 1 if month < 12 else 1
            next_year = year if month < 12 else year + 1
            
            # Получаем данные для переключения интерфейсов
            user_id = await get_user_id_from_current_user(current_user, session)
            available_interfaces = await get_available_interfaces_for_user(user_id)
            
            
            # Подготавливаем данные для shared компонентов календаря
            calendar_title = f"{RU_MONTHS[month]} {year}"
            current_date = f"{year}-{month:02d}-01"
            
            # Преобразуем calendar_data в формат, ожидаемый shared компонентами
            calendar_weeks = []
            for week in calendar_data:
                week_data = []
                for day in week:
                    logger.info(f"Processing day {day.get('date')}: timeslots={len(day.get('timeslots', []))}, shifts={len(day.get('shifts', []))}")
                    # Обрабатываем смены
                    shifts = []
                    for shift in day.get("shifts", []):
                        shifts.append({
                            "id": shift.get("id"),
                            "object_id": shift.get("object_id"),
                            "time_slot_id": shift.get("time_slot_id"),
                            "start_time": shift.get("start_time", ""),
                            "end_time": shift.get("end_time", ""),
                            "employee_name": shift.get("employee_name", "Неизвестно"),
                            "object_name": shift.get("object_name", ""),
                            "status": shift.get("status", "pending"),
                            "status_label": shift.get("status_label", "")
                        })
                    
                    # Обрабатываем тайм-слоты
                    timeslots = []
                    for timeslot in day.get("timeslots", []):
                        timeslots.append({
                            "id": timeslot.get("id"),
                            "object_id": timeslot.get("object_id"),
                            "start_time": timeslot.get("start_time", ""),
                            "end_time": timeslot.get("end_time", ""),
                            "object_name": timeslot.get("object_name", ""),
                            "max_employees": timeslot.get("max_employees", 1),
                            "employee_count": timeslot.get("employee_count", 0),
                            "status": timeslot.get("status", "available")
                        })
                    
                    week_data.append({
                        "date": day["date"].strftime("%Y-%m-%d"),
                        "day": day["day"],
                        "is_other_month": day["is_other_month"],
                        "is_today": day["is_today"],
                        "shifts": shifts,
                        "timeslots": timeslots
                    })
                calendar_weeks.append(week_data)
            
            return templates.TemplateResponse("owner/calendar/index.html", {
                "request": request,
                "title": "Календарное планирование",
                "current_user": current_user,
                "year": year,
                "month": month,
                "month_name": RU_MONTHS[month],
                "calendar_title": calendar_title,
                "current_date": current_date,
                "view_type": "month",
                "show_today_button": True,
                "calendar_weeks": calendar_weeks,
                "available_interfaces": available_interfaces,
                "objects": objects_list,
                "org_units": org_units_list,
                "selected_object_id": object_id,
                "selected_org_unit_id": org_unit_id,
                "selected_object": selected_object,
                "timeslots": timeslots_data,
                "prev_month": prev_month,
                "prev_year": prev_year,
                "next_month": next_month,
                "next_year": next_year,
                "today": today
            })
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error loading calendar: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки календаря")


@router.get("/calendar/week", response_class=HTMLResponse, name="owner_calendar_week")
async def owner_calendar_week(
    request: Request,
    year: int = Query(None),
    week: int = Query(None),
    object_id: int = Query(None),
    current_user: dict = Depends(require_owner_or_superadmin)
):
    """Недельный вид календаря владельца"""
    try:
        # Определяем текущую неделю или переданные параметры
        today = date.today()
        if year is None:
            year = today.year
        if week is None:
            # Получаем номер недели для текущей даты
            week = today.isocalendar()[1]
        
        # Получаем первый день недели (понедельник)
        first_day_of_year = date(year, 1, 1)
        first_monday = first_day_of_year - timedelta(days=first_day_of_year.weekday())
        week_start = first_monday + timedelta(weeks=week-1)
        
        # Создаем список дней недели
        week_days = []
        for i in range(7):
            current_date = week_start + timedelta(days=i)
            week_days.append(current_date)
        
        # Получаем объекты и тайм-слоты (как в месячном календаре)
        async with get_async_session() as db:
            object_service = ObjectService(db)
            timeslot_service = TimeSlotService(db)
            
            # ОРИГИНАЛ: используем telegram_id владельца (как в месячном календаре)
            owner_telegram_id = current_user.get("telegram_id") or current_user.get("telegram_id") or current_user.get("id")
            objects = await object_service.get_objects_by_owner(owner_telegram_id)
            
            selected_object = None
            if object_id:
                for obj in objects:
                    if obj.id == object_id:
                        selected_object = obj
                        break
                if not selected_object:
                    raise HTTPException(status_code=404, detail="Объект не найден")
            
            # Получаем тайм-слоты для недели (как в месячном календаре)
            timeslots_data = []
            if selected_object:
                timeslots = await timeslot_service.get_timeslots_by_object(selected_object.id, owner_telegram_id)
                for slot in timeslots:
                    if week_start <= slot.slot_date <= week_days[-1]:
                        timeslots_data.append({
                            "id": slot.id,
                            "object_id": slot.object_id,
                            "object_name": selected_object.name,
                            "date": slot.slot_date,
                            "start_time": slot.start_time.strftime("%H:%M"),
                            "end_time": slot.end_time.strftime("%H:%M"),
                            "hourly_rate": float(slot.hourly_rate) if slot.hourly_rate else float(selected_object.hourly_rate),
                            "max_employees": slot.max_employees if slot.max_employees is not None else 1,
                            "is_active": slot.is_active,
                            "notes": slot.notes or ""
                        })
            else:
                for obj in objects:
                    timeslots = await timeslot_service.get_timeslots_by_object(obj.id, owner_telegram_id)
                    for slot in timeslots:
                        if week_start <= slot.slot_date <= week_days[-1]:
                            timeslots_data.append({
                                "id": slot.id,
                                "object_id": slot.object_id,
                                "object_name": obj.name,
                                "date": slot.slot_date,
                                "start_time": slot.start_time.strftime("%H:%M"),
                                "end_time": slot.end_time.strftime("%H:%M"),
                                "hourly_rate": float(slot.hourly_rate) if slot.hourly_rate else float(obj.hourly_rate),
                                "max_employees": slot.max_employees if slot.max_employees is not None else 1,
                                "is_active": slot.is_active,
                                "notes": slot.notes or ""
                            })
            
            # Группируем тайм-слоты по дням
            week_data = []
            for day_date in week_days:
                day_timeslots = [
                    slot for slot in timeslots_data 
                    if slot["date"] == day_date and slot["is_active"]
                ]
                week_data.append({
                    "date": day_date,
                    "is_today": day_date == today,
                    "timeslots": day_timeslots,
                    "timeslots_count": len(day_timeslots)
                })
            
            # Навигация по неделям
            prev_week = week - 1 if week > 1 else 52
            prev_year = year if week > 1 else year - 1
            next_week = week + 1 if week < 52 else 1
            next_year = year if week < 52 else year + 1
            
            objects_list = [{"id": obj.id, "name": obj.name} for obj in objects]
            
            # Получаем данные для переключения интерфейсов
            user_id = await get_user_id_from_current_user(current_user, db)
            available_interfaces = await get_available_interfaces_for_user(user_id)
            
            return templates.TemplateResponse("owner/calendar/week.html", {
                "request": request,
                "title": "Недельное планирование",
                "current_user": current_user,
                "year": year,
                "week": week,
                "week_data": week_data,
                "week_start": week_start,
                "week_end": week_days[-1],
                "objects": objects_list,
                "available_interfaces": available_interfaces,
                "selected_object_id": object_id,
                "selected_object": selected_object,
                "prev_week": prev_week,
                "prev_year": prev_year,
                "next_week": next_week,
                "next_year": next_year,
                "today": today
            })
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error loading week view: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки недельного вида")


@router.get("/calendar/analysis", response_class=HTMLResponse, name="owner_calendar_analysis")
async def owner_calendar_analysis(
    request: Request,
    object_id: int = Query(None),
    days: int = Query(30),
    message: str = Query(None),
    created: int = Query(None),
    current_user: dict = Depends(require_owner_or_superadmin),
    db: AsyncSession = Depends(get_db_session)
):
    """Анализ пробелов в планировании"""
    try:
        # Получаем объекты и тайм-слоты (как в месячном календаре)
        object_service = ObjectService(db)
        timeslot_service = TimeSlotService(db)
        
        # ОРИГИНАЛ: используем telegram_id владельца (как в месячном календаре)
        owner_telegram_id = current_user.get("telegram_id") or current_user.get("telegram_id") or current_user.get("id")
        objects = await object_service.get_objects_by_owner(owner_telegram_id)
        
        selected_object = None
        if object_id:
            for obj in objects:
                if obj.id == object_id:
                    selected_object = obj
                    break
            if not selected_object:
                raise HTTPException(status_code=404, detail="Объект не найден")
        
        # Анализируем пробелы (как в оригинале)
        analysis_data = await _analyze_gaps(
            timeslot_service, 
            objects if not selected_object else [selected_object], 
            owner_telegram_id, 
            days
        )
        
        objects_list = [{"id": obj.id, "name": obj.name} for obj in objects]
        
        # Получаем данные для переключения интерфейсов
        available_interfaces = await get_available_interfaces_for_user(owner_telegram_id)
        
        return templates.TemplateResponse("owner/calendar/analysis.html", {
            "request": request,
            "title": "Анализ пробелов в планировании",
            "current_user": current_user,
            "objects": objects_list,
            "selected_object_id": object_id,
            "selected_object": selected_object,
            "analysis_data": analysis_data,
            "available_interfaces": available_interfaces,
            "days": days,
            "message": message,
            "created": created
        })
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error loading gap analysis: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки анализа пробелов")


def _is_working_day(obj, check_date: date) -> bool:
    """Проверяет, является ли день рабочим для объекта."""
    # Получаем день недели (0=Понедельник, 6=Воскресенье)
    weekday = check_date.weekday()  # 0=Пн, 1=Вт, ..., 6=Вс
    
    # Конвертируем в битовую позицию (1=Пн, 2=Вт, ..., 64=Вс)
    bit_position = 1 << weekday
    
    # Проверяем, установлен ли бит для этого дня
    return (obj.work_days_mask & bit_position) != 0


def _check_timeslot_coverage(obj, slots, check_date: date) -> List[Dict[str, Any]]:
    """Проверяет покрытие рабочего времени тайм-слотами."""
    gaps = []
    
    if not slots:
        return [{
            "date": check_date,
            "type": "no_slots",
            "message": "Нет тайм-слотов на рабочий день"
        }]
    
    # Сортируем слоты по времени начала
    slots.sort(key=lambda x: x.start_time)
    
    # Проверяем покрытие от времени открытия до времени закрытия
    opening_time = obj.opening_time
    closing_time = obj.closing_time
    
    current_time = opening_time
    slot_index = 0
    
    while current_time < closing_time and slot_index < len(slots):
        slot = slots[slot_index]
        
        # Если есть пробел между текущим временем и началом слота
        if current_time < slot.start_time:
            gap_duration = (datetime.combine(check_date, slot.start_time) - 
                          datetime.combine(check_date, current_time)).total_seconds() / 60
            if gap_duration >= 30:  # Пробелы меньше 30 минут не считаем
                gaps.append({
                    "date": check_date,
                    "type": "time_gap",
                    "message": f"Пробел в расписании: {current_time.strftime('%H:%M')} - {slot.start_time.strftime('%H:%M')} ({gap_duration:.0f} мин)"
                })
            current_time = slot.start_time
        
        # Переходим к концу текущего слота
        if slot.end_time:
            current_time = slot.end_time
        else:
            # Если слот без времени окончания, считаем до закрытия
            current_time = closing_time
        
        slot_index += 1
    
    # Проверяем, есть ли пробел в конце дня
    if current_time < closing_time:
        gap_duration = (datetime.combine(check_date, closing_time) - 
                      datetime.combine(check_date, current_time)).total_seconds() / 60
        if gap_duration >= 30:
            gaps.append({
                "date": check_date,
                "type": "time_gap",
                "message": f"Пробел в конце дня: {current_time.strftime('%H:%M')} - {closing_time.strftime('%H:%M')} ({gap_duration:.0f} мин)"
            })
    
    return gaps


async def _analyze_gaps(
    timeslot_service: TimeSlotService, 
    objects: List, 
    telegram_id: int, 
    days: int
) -> Dict[str, Any]:
    """Анализирует пробелы в планировании с учетом расписания работы объектов."""
    from datetime import datetime
    
    today = date.today()
    end_date = today + timedelta(days=days)
    
    total_gaps = 0
    object_gaps = {}
    
    for obj in objects:
        timeslots = await timeslot_service.get_timeslots_by_object(obj.id, telegram_id)
        
        # Группируем тайм-слоты по дням
        daily_slots = {}
        for slot in timeslots:
            if today <= slot.slot_date <= end_date and slot.is_active:
                if slot.slot_date not in daily_slots:
                    daily_slots[slot.slot_date] = []
                daily_slots[slot.slot_date].append(slot)
        
        # Анализируем пробелы для каждого дня
        gaps = []
        current_date = today
        while current_date <= end_date:
            # Проверяем, является ли день рабочим
            if _is_working_day(obj, current_date):
                if current_date not in daily_slots:
                    # Рабочий день без тайм-слотов
                    gaps.append({
                        "date": current_date,
                        "type": "no_slots",
                        "message": "Нет тайм-слотов на рабочий день"
                    })
                    total_gaps += 1
                else:
                    # Проверяем покрытие рабочего времени
                    day_gaps = _check_timeslot_coverage(obj, daily_slots[current_date], current_date)
                    gaps.extend(day_gaps)
                    total_gaps += len(day_gaps)
            
            current_date += timedelta(days=1)
        
        object_gaps[obj.id] = {
            "object_name": obj.name,
            "gaps": gaps,
            "gaps_count": len(gaps),
            "work_schedule": f"{obj.opening_time.strftime('%H:%M')} - {obj.closing_time.strftime('%H:%M')}",
            "work_days": _get_work_days_text(obj.work_days_mask)
        }
    
    return {
        "total_gaps": total_gaps,
        "object_gaps": object_gaps,
        "period": f"{today.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"
    }


def _get_work_days_text(work_days_mask: int) -> str:
    """Преобразует битовую маску дней в читаемый текст."""
    days = []
    day_names = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
    
    for i in range(7):
        if work_days_mask & (1 << i):
            days.append(day_names[i])
    
    return ", ".join(days) if days else "Нет рабочих дней"


@router.post("/calendar/analysis/fill-gaps/{object_id}")
async def owner_fill_gaps(
    object_id: int,
    request: Request,
    days: int = Form(30),
    current_user: dict = Depends(require_owner_or_superadmin),
    db: AsyncSession = Depends(get_db_session)
):
    """Автоматическое заполнение пробелов в планировании для объекта."""
    try:
        from apps.web.services.object_service import ObjectService, TimeSlotService
        
        object_service = ObjectService(db)
        timeslot_service = TimeSlotService(db)
        
        # Получаем объект
        owner_telegram_id = current_user.get("telegram_id") or current_user.get("telegram_id") or current_user.get("id")
        objects = await object_service.get_objects_by_owner(owner_telegram_id)
        
        target_object = None
        for obj in objects:
            if obj.id == object_id:
                target_object = obj
                break
        
        if not target_object:
            raise HTTPException(status_code=404, detail="Объект не найден")
        
        # Анализируем пробелы для объекта
        analysis_data = await _analyze_gaps(
            timeslot_service, 
            [target_object], 
            owner_telegram_id, 
            days
        )
        
        object_data = analysis_data["object_gaps"].get(object_id, {})
        gaps = object_data.get("gaps", [])
        
        if not gaps:
            return RedirectResponse(
                url=f"/owner/calendar/analysis?object_id={object_id}&days={days}&message=no_gaps",
                status_code=303
            )
        
        # Создаем тайм-слоты для пробелов
        created_slots = 0
        for gap in gaps:
            if gap["type"] == "no_slots":
                # Создаем тайм-слот на весь рабочий день
                slot_data = {
                    "object_id": object_id,
                    "slot_date": gap["date"],
                    "start_time": target_object.opening_time.strftime("%H:%M"),
                    "end_time": target_object.closing_time.strftime("%H:%M"),
                    "hourly_rate": float(target_object.hourly_rate),
                    "max_employees": 1,
                    "is_active": True,
                    "notes": "Автоматически создан для заполнения пробела"
                }
                
                await timeslot_service.create_timeslot(slot_data, object_id, owner_telegram_id)
                created_slots += 1
            elif gap["type"] == "time_gap":
                # Для пробелов во времени создаем тайм-слот только для конкретного интервала
                # Парсим время из сообщения "Пробел в расписании: 12:00 - 21:00 (540 мин)"
                message = gap["message"]
                if "Пробел в расписании:" in message:
                    # Извлекаем время из сообщения
                    time_part = message.split("Пробел в расписании: ")[1].split(" (")[0]
                    start_time_str, end_time_str = time_part.split(" - ")
                    
                    slot_data = {
                        "object_id": object_id,
                        "slot_date": gap["date"],
                        "start_time": start_time_str,
                        "end_time": end_time_str,
                        "hourly_rate": float(target_object.hourly_rate),
                        "max_employees": 1,
                        "is_active": True,
                        "notes": "Автоматически создан для заполнения пробела во времени"
                    }
                    
                    await timeslot_service.create_timeslot(slot_data, object_id, owner_telegram_id)
                    created_slots += 1
                elif "Пробел в конце дня:" in message:
                    # Извлекаем время из сообщения
                    time_part = message.split("Пробел в конце дня: ")[1].split(" (")[0]
                    start_time_str, end_time_str = time_part.split(" - ")
                    
                    slot_data = {
                        "object_id": object_id,
                        "slot_date": gap["date"],
                        "start_time": start_time_str,
                        "end_time": end_time_str,
                        "hourly_rate": float(target_object.hourly_rate),
                        "max_employees": 1,
                        "is_active": True,
                        "notes": "Автоматически создан для заполнения пробела в конце дня"
                    }
                    
                    await timeslot_service.create_timeslot(slot_data, object_id, owner_telegram_id)
                    created_slots += 1
        
        return RedirectResponse(
            url=f"/owner/calendar/analysis?object_id={object_id}&days={days}&message=success&created={created_slots}",
            status_code=303
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error filling gaps: {e}")
        raise HTTPException(status_code=500, detail="Ошибка заполнения пробелов")


@router.get("/calendar/analysis/chart-data")
async def owner_analysis_chart_data(
    object_id: Optional[int] = Query(None),
    days: int = Query(30),
    request: Request = None,
    db: AsyncSession = Depends(get_db_session)
):
    """Получение данных для графика покрытия планирования."""
    try:
        # Получаем текущего пользователя
        current_user = await get_current_user(request)
        if current_user is None or isinstance(current_user, RedirectResponse):
            raise HTTPException(status_code=401, detail="Пользователь не авторизован")

        if isinstance(current_user, dict):
            user_roles = current_user.get("roles", [])
            if isinstance(user_roles, str):
                user_roles = [user_roles]
            if "owner" not in user_roles and "superadmin" not in user_roles:
                raise HTTPException(status_code=403, detail="Доступ запрещен")
        else:
            user_roles = getattr(current_user, "roles", [])
            if isinstance(user_roles, str):
                user_roles = [user_roles]
            if "owner" not in user_roles and "superadmin" not in user_roles:
                raise HTTPException(status_code=403, detail="Доступ запрещен")
        
        from apps.web.services.object_service import ObjectService, TimeSlotService
        
        object_service = ObjectService(db)
        timeslot_service = TimeSlotService(db)
        
        # Получаем объекты
        owner_telegram_id = current_user.get("telegram_id") or current_user.get("telegram_id") or current_user.get("id")
        objects = await object_service.get_objects_by_owner(owner_telegram_id)
        
        if object_id:
            objects = [obj for obj in objects if obj.id == object_id]
        
        if not objects:
            return {"error": "Нет объектов для анализа"}
        
        # Анализируем пробелы
        analysis_data = await _analyze_gaps(
            timeslot_service, 
            objects, 
            owner_telegram_id, 
            days
        )
        
        logger.info(f"Chart data analysis: {len(objects)} objects, {len(analysis_data.get('object_gaps', {}))} object_gaps")
        logger.info(f"Analysis data keys: {list(analysis_data.keys())}")
        
        # Отладочная информация
        for obj in objects:
            logger.info(f"Object {obj.id}: {obj.name}, opening: {obj.opening_time}, closing: {obj.closing_time}")
        
        if "object_gaps" not in analysis_data:
            logger.error("No object_gaps in analysis_data")
            return {"error": "Нет данных для анализа"}
        
        if not analysis_data["object_gaps"]:
            logger.error("object_gaps is empty")
            return {"error": "Нет объектов для анализа"}
        
        # Подготавливаем данные для графика
        chart_data = {
            "labels": [],  # Даты
            "datasets": []  # Данные по объектам
        }
        
        # Получаем все даты в периоде
        today = date.today()
        end_date = today + timedelta(days=days)
        current_date = today
        
        dates = []
        while current_date <= end_date:
            dates.append(current_date)
            current_date += timedelta(days=1)
        
        chart_data["labels"] = [d.strftime("%d.%m") for d in dates]
        
        # Создаем датасет для каждого объекта
        colors = ["#007bff", "#28a745", "#ffc107", "#dc3545", "#6f42c1", "#fd7e14", "#20c997", "#6c757d"]
        
        for i, (obj_id, obj_data) in enumerate(analysis_data["object_gaps"].items()):
            color = colors[i % len(colors)]
            
            # Создаем массив покрытия для каждого дня
            coverage_data = []
            gaps_by_date = {gap["date"]: gap for gap in obj_data["gaps"]}
            
            # Получаем все тайм-слоты для этого объекта
            from domain.entities.time_slot import TimeSlot
            from domain.entities.shift_schedule import ShiftSchedule
            
            timeslots_query = select(TimeSlot).where(
                TimeSlot.object_id == obj_id,
                TimeSlot.slot_date >= today,
                TimeSlot.slot_date <= end_date
            )
            timeslots_result = await db.execute(timeslots_query)
            timeslots = timeslots_result.scalars().all()
            timeslots_by_date = {}
            for ts in timeslots:
                if ts.slot_date not in timeslots_by_date:
                    timeslots_by_date[ts.slot_date] = []
                timeslots_by_date[ts.slot_date].append(ts)
            
            # Получаем запланированные смены для тайм-слотов
            timeslot_ids = [ts.id for ts in timeslots]
            scheduled_shifts_query = select(ShiftSchedule).where(
                ShiftSchedule.time_slot_id.in_(timeslot_ids),
                ShiftSchedule.status.in_(["planned", "confirmed"])
            )
            scheduled_result = await db.execute(scheduled_shifts_query)
            scheduled_shifts = scheduled_result.scalars().all()
            
            # Группируем смены по датам тайм-слотов
            scheduled_by_date = {}
            for shift in scheduled_shifts:
                # Находим тайм-слот для этой смены
                for ts in timeslots:
                    if ts.id == shift.time_slot_id:
                        if ts.slot_date not in scheduled_by_date:
                            scheduled_by_date[ts.slot_date] = []
                        scheduled_by_date[ts.slot_date].append(shift)
                        break
            
            for d in dates:
                if d in timeslots_by_date and d in scheduled_by_date:
                    # Есть тайм-слоты И есть запланированные смены - вычисляем процент покрытия
                    timeslots_for_day = timeslots_by_date[d]
                    scheduled_for_day = scheduled_by_date[d]
                    
                    # Получаем объект для расчета рабочего времени
                    obj = next((o for o in objects if o.id == obj_id), None)
                    if not obj:
                        coverage_data.append(0)
                        continue
                    
                    # Вычисляем общее рабочее время в день (в минутах)
                    opening_minutes = obj.opening_time.hour * 60 + obj.opening_time.minute
                    closing_minutes = obj.closing_time.hour * 60 + obj.closing_time.minute
                    total_work_minutes = closing_minutes - opening_minutes
                    
                    if total_work_minutes <= 0:
                        coverage_data.append(0)
                        continue
                    
                    # Вычисляем общее время запланированных смен (в минутах)
                    total_scheduled_minutes = 0
                    for shift in scheduled_for_day:
                        # Получаем тайм-слот для этой смены
                        timeslot = next((ts for ts in timeslots_for_day if ts.id == shift.time_slot_id), None)
                        if timeslot:
                            start_minutes = timeslot.start_time.hour * 60 + timeslot.start_time.minute
                            end_minutes = timeslot.end_time.hour * 60 + timeslot.end_time.minute
                            total_scheduled_minutes += (end_minutes - start_minutes)
                    
                    # Вычисляем процент покрытия
                    coverage_percent = min(100, (total_scheduled_minutes / total_work_minutes) * 100)
                    coverage_data.append(round(coverage_percent))
                    
                elif d in timeslots_by_date:
                    # Есть тайм-слоты, но нет запланированных смен - нет покрытия
                    coverage_data.append(0)
                else:
                    # Нет тайм-слотов - нет покрытия
                    coverage_data.append(0)
            
            chart_data["datasets"].append({
                "label": obj_data["object_name"],
                "data": coverage_data,
                "borderColor": color,
                "backgroundColor": color + "20",  # Прозрачность
                "fill": False,
                "tension": 0.1
            })
        
        return chart_data
        
    except Exception as e:
        logger.error(f"Error getting chart data: {e}")
        raise HTTPException(status_code=500, detail="Ошибка получения данных графика")


@router.get("/calendar/api/timeslots-status")
async def owner_calendar_api_timeslots_status(
    year: int = Query(...),
    month: int = Query(...),
    object_id: Optional[int] = Query(None),
    current_user: dict = Depends(get_current_user_dependency()),
    _: None = Depends(require_role(["owner", "superadmin"])),
    db: AsyncSession = Depends(get_db_session)
):
    """Получение статуса тайм-слотов для календаря"""
    try:
        logger.info(f"Getting timeslots status for {year}-{month}, object_id: {object_id}")
        
        # Получаем реальные тайм-слоты из базы
        from sqlalchemy import select, and_
        from sqlalchemy.orm import selectinload
        from domain.entities.time_slot import TimeSlot
        from domain.entities.object import Object
        from domain.entities.user import User
        
        # Получаем владельца из текущего пользователя (по telegram_id)
        if not current_user or not getattr(current_user, "telegram_id", None):
            return []
        owner_query = select(User).where(User.telegram_id == current_user.telegram_id)
        owner_result = await db.execute(owner_query)
        owner = owner_result.scalar_one_or_none()
        
        if not owner:
            return []
        
        # Получаем объекты владельца (используем ВНУТРЕНИЙ owner.id)
        objects_query = select(Object).where(Object.owner_id == owner.id)
        if object_id:
            objects_query = objects_query.where(Object.id == object_id)
        
        objects_result = await db.execute(objects_query)
        objects = objects_result.scalars().all()
        object_ids = [obj.id for obj in objects]
        
        if not object_ids:
            return []
        
        start_date = date(year, month, 1)
        last_day = calendar.monthrange(year, month)[1]
        end_date = date(year, month, last_day)
        
        calendar_service = CalendarFilterService(db)
        calendar_data = await calendar_service.get_calendar_data(
            user_telegram_id=owner.telegram_id,
            user_role='owner',
            date_range_start=start_date,
            date_range_end=end_date,
            object_filter=[object_id] if object_id else None
        )
        
        shifts_by_timeslot: Dict[int, Dict[str, list]] = {}
        for shift in calendar_data.shifts:
            if not shift.time_slot_id:
                continue
            entry = shifts_by_timeslot.setdefault(int(shift.time_slot_id), {"planned": [], "actual": []})
            shift_data = {
                "id": shift.id,
                "user_id": shift.user_id,
                "user_name": shift.user_name,
                "status": shift.status.value,
                "start_time": shift.start_time.isoformat() if shift.start_time else None,
                "end_time": shift.end_time.isoformat() if shift.end_time else None,
                "planned_start": shift.planned_start.isoformat() if shift.planned_start else None,
                "planned_end": shift.planned_end.isoformat() if shift.planned_end else None,
                "notes": shift.notes
            }
            if shift.shift_type == ShiftType.PLANNED:
                entry["planned"].append(shift_data)
            else:
                entry["actual"].append(shift_data)
        
        response_data = []
        for ts in calendar_data.timeslots:
            slot_capacity_minutes = (ts.occupied_minutes + ts.free_minutes)
            availability = f"{ts.current_employees}/{ts.max_employees}"
            slot_shifts = shifts_by_timeslot.get(ts.id, {"planned": [], "actual": []})
            response_data.append({
                "slot_id": ts.id,
                "object_id": ts.object_id,
                "object_name": ts.object_name,
                "date": ts.date.isoformat(),
                "start_time": ts.start_time.strftime("%H:%M"),
                "end_time": ts.end_time.strftime("%H:%M"),
                "hourly_rate": ts.hourly_rate,
                "status": ts.status.value,
                "status_label": ts.status_label,
                "current_employees": ts.current_employees,
                "available_slots": ts.available_slots,
                "occupancy_ratio": ts.occupancy_ratio,
                "occupied_minutes": ts.occupied_minutes,
                "free_minutes": ts.free_minutes,
                "capacity_minutes": slot_capacity_minutes,
                "availability": availability,
                "scheduled_shifts": slot_shifts["planned"],
                "actual_shifts": slot_shifts["actual"]
            })
        
        logger.info(f"Returning {len(response_data)} timeslots for calendar status")
        return response_data
        
    except Exception as e:
        logger.error(f"Error getting timeslots status: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки статуса тайм-слотов")


@router.get("/calendar/api/data")
async def owner_calendar_api_data(
    start_date: str = Query(..., description="Начальная дата в формате YYYY-MM-DD"),
    end_date: str = Query(..., description="Конечная дата в формате YYYY-MM-DD"),
    object_ids: Optional[str] = Query(None, description="ID объектов через запятую"),
    org_unit_id: Optional[int] = Query(None, description="ID подразделения для фильтрации (устаревший, используйте org_unit_ids)"),
    org_unit_ids: Optional[str] = Query(None, description="ID подразделений через запятую (включая потомков)"),
    current_user: dict = Depends(get_current_user_dependency()),
    db: AsyncSession = Depends(get_db_session)
):
    """
    Новый универсальный API для получения данных календаря владельца.
    Использует CalendarFilterService для правильной фильтрации смен.
    """
    try:
        # Генерируем ключ кэша
        import hashlib

        if current_user is None or isinstance(current_user, RedirectResponse):
            raise HTTPException(status_code=401, detail="Пользователь не авторизован")

        if isinstance(current_user, dict):
            user_id = current_user.get("telegram_id") or current_user.get("id")
        else:
            user_id = getattr(current_user, "telegram_id", None)

        if not user_id:
            raise HTTPException(status_code=401, detail="Пользователь не найден")
        cache_key_data = f"calendar_api_owner:{user_id}:{start_date}:{end_date}:{object_ids or 'all'}:{org_unit_ids or org_unit_id or 'all'}"
        cache_key = hashlib.md5(cache_key_data.encode()).hexdigest()
        
        # Проверяем кэш
        from core.cache.redis_cache import cache
        cached_response = await cache.get(f"api_response:{cache_key}", serialize="json")
        if cached_response:
            logger.info(f"Owner calendar API: cache HIT for {start_date} to {end_date}, org_unit_ids={org_unit_ids or org_unit_id}")
            return cached_response
        
        logger.info(f"Owner calendar API: cache MISS for {start_date} to {end_date}, org_unit_ids={org_unit_ids or org_unit_id}")
        
        # Парсим даты
        try:
            start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
            end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="Неверный формат даты. Используйте YYYY-MM-DD")
        
        # Парсим фильтр объектов
        object_filter = None
        
        # Если указаны подразделения (новый формат org_unit_ids или старый org_unit_id) - получаем объекты этих подразделений
        org_unit_ids_list = []
        if org_unit_ids:
            try:
                org_unit_ids_list = [int(unit_id.strip()) for unit_id in org_unit_ids.split(",") if unit_id.strip()]
            except ValueError:
                raise HTTPException(status_code=400, detail="Неверный формат ID подразделений")
        elif org_unit_id:
            # Старый формат - один ID
            org_unit_ids_list = [org_unit_id]
        
        if org_unit_ids_list and not object_ids:
            from apps.web.services.org_structure_service import OrgStructureService
            object_service = ObjectService(db)
            org_service = OrgStructureService(db)
            owner_telegram_id = user_id
            all_objects = await object_service.get_objects_by_owner(owner_telegram_id)
            
            # Получаем все потомки выбранных подразделений
            all_org_unit_ids = set(org_unit_ids_list)
            for unit_id in org_unit_ids_list:
                descendants = await org_service._get_all_descendants(unit_id)
                all_org_unit_ids.update([d.id for d in descendants])
            
            # Фильтруем объекты по подразделениям (включая потомков)
            filtered_by_org = [obj.id for obj in all_objects if obj.org_unit_id in all_org_unit_ids]
            if filtered_by_org:
                object_filter = filtered_by_org
                logger.info(f"Filtering by org_unit_ids={org_unit_ids_list} (including descendants): {len(filtered_by_org)} objects - {filtered_by_org}")
            else:
                logger.info(f"No objects found for org_unit_ids={org_unit_ids_list}")
                # Если нет объектов в подразделениях - вернем пустой результат
                object_filter = [-1]  # Несуществующий ID чтобы вернуть пустой результат
        elif object_ids:
            try:
                object_filter = [int(obj_id.strip()) for obj_id in object_ids.split(",") if obj_id.strip()]
            except ValueError:
                raise HTTPException(status_code=400, detail="Неверный формат ID объектов")
        
        # Получаем роль пользователя
        if isinstance(current_user, dict):
            user_role = current_user.get("role", "owner")
            user_telegram_id = current_user.get("telegram_id") or current_user.get("id")
        else:
            # current_user - это объект User
            user_role = getattr(current_user, "role", "owner")
            user_telegram_id = getattr(current_user, "telegram_id", None)
        
        if not user_telegram_id:
            raise HTTPException(status_code=401, detail="Пользователь не найден")
        
        # Получаем данные календаря через универсальный сервис
        calendar_service = CalendarFilterService(db)
        calendar_data = await calendar_service.get_calendar_data(
            user_telegram_id=user_telegram_id,
            user_role=user_role,
            date_range_start=start_date_obj,
            date_range_end=end_date_obj,
            object_filter=object_filter
        )
        
        # Логируем, сколько тайм-слотов пришло из CalendarFilterService
        logger.info(f"Owner calendar API: received {len(calendar_data.timeslots)} timeslots from CalendarFilterService")
        for ts in calendar_data.timeslots[:5]:  # Первые 5 для примера
            logger.info(f"  - Timeslot {ts.id}: date={ts.date}, status={ts.status}, start_time={ts.start_time}")
        
        # Преобразуем в формат, совместимый с существующим JavaScript
        # Для мобильного дневного вида показываем все тайм-слоты, включая свободные
        timeslots_data = []
        now = datetime.now()
        hidden_count = 0
        skipped_count = 0
        for ts in calendar_data.timeslots:
            # Пропускаем только тайм-слоты в прошлом со статусом HIDDEN
            # Для будущих дней показываем все тайм-слоты, включая свободные
            if ts.status == TimeslotStatus.HIDDEN:
                hidden_count += 1
                # Проверяем, не в прошлом ли тайм-слот
                slot_datetime = datetime.combine(ts.date, ts.start_time)
                if slot_datetime < now:
                    skipped_count += 1
                    continue  # Пропускаем только тайм-слоты в прошлом
                # Для будущих дней меняем статус на AVAILABLE, чтобы они отображались
                ts.status = TimeslotStatus.AVAILABLE
                ts.status_label = "Свободно"
            timeslots_data.append({
                "id": ts.id,
                "object_id": ts.object_id,
                "object_name": ts.object_name,
                "date": ts.date.isoformat(),
                "start_time": ts.start_time.strftime("%H:%M"),
                "end_time": ts.end_time.strftime("%H:%M"),
                "hourly_rate": ts.hourly_rate,
                "max_employees": ts.max_employees,
                "current_employees": ts.current_employees,
                "available_slots": ts.available_slots,
                "occupied_minutes": ts.occupied_minutes,
                "free_minutes": ts.free_minutes,
                "occupancy_ratio": ts.occupancy_ratio,
                "status": ts.status.value,
                "status_label": ts.status_label,
                "is_active": ts.is_active,
                "notes": ts.notes,
                "work_conditions": ts.work_conditions,
                "shift_tasks": ts.shift_tasks,
                "coordinates": ts.coordinates,
                "can_edit": ts.can_edit,
                "can_plan": ts.can_plan,
                "can_view": ts.can_view,
                "fully_occupied": getattr(ts, 'fully_occupied', False),
                "has_free_track": getattr(ts, 'has_free_track', True)
            })
        
        logger.info(f"Owner calendar API: processed {len(timeslots_data)} timeslots (hidden: {hidden_count}, skipped: {skipped_count})")
        
        shifts_data = []
        for s in calendar_data.shifts:
            # Получаем часовой пояс объекта
            object_timezone = s.timezone if hasattr(s, 'timezone') and s.timezone else 'Europe/Moscow'
            import pytz
            tz = pytz.timezone(object_timezone)
            
            # Конвертируем время в локальное время объекта
            def convert_to_local_time(utc_time):
                if utc_time:
                    # Если время уже имеет timezone info, конвертируем
                    if utc_time.tzinfo:
                        return utc_time.astimezone(tz).replace(tzinfo=None)
                    else:
                        # Если время naive, считаем его UTC и конвертируем
                        utc_aware = pytz.UTC.localize(utc_time)
                        return utc_aware.astimezone(tz).replace(tzinfo=None)
                return None
            
            shifts_data.append({
                "id": s.id,
                "user_id": s.user_id,
                "user_name": s.user_name,
                "object_id": s.object_id,
                "object_name": s.object_name,
                "time_slot_id": s.time_slot_id,
                "start_time": convert_to_local_time(s.start_time).isoformat() if s.start_time else None,
                "end_time": convert_to_local_time(s.end_time).isoformat() if s.end_time else None,
                "planned_start": convert_to_local_time(s.planned_start).isoformat() if s.planned_start else None,
                "planned_end": convert_to_local_time(s.planned_end).isoformat() if s.planned_end else None,
                "shift_type": s.shift_type.value,
                "status": s.status.value,
                "hourly_rate": s.hourly_rate,
                "total_hours": s.total_hours,
                "total_payment": s.total_payment,
                "notes": s.notes,
                "is_planned": s.is_planned,
                "schedule_id": s.schedule_id,
                "actual_shift_id": s.actual_shift_id,
                "start_coordinates": s.start_coordinates,
                "end_coordinates": s.end_coordinates,
                "can_edit": s.can_edit,
                "can_cancel": s.can_cancel,
                "can_view": s.can_view,
                "status_label": s.status_label
            })
        
        response_data = {
            "timeslots": timeslots_data,
            "shifts": shifts_data,
            "metadata": {
                "date_range_start": calendar_data.date_range_start.isoformat(),
                "date_range_end": calendar_data.date_range_end.isoformat(),
                "user_role": calendar_data.user_role,
                "total_timeslots": calendar_data.total_timeslots,
                "total_shifts": calendar_data.total_shifts,
                "planned_shifts": calendar_data.planned_shifts,
                "active_shifts": calendar_data.active_shifts,
                "completed_shifts": calendar_data.completed_shifts,
                "accessible_objects": calendar_data.accessible_objects
            }
        }
        
        # Сохраняем в кэш (TTL 2 минуты)
        await cache.set(f"api_response:{cache_key}", response_data, ttl=120, serialize="json")
        logger.info(f"Owner calendar API: response cached")
        
        return response_data
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting owner calendar data: {e}")
        raise HTTPException(status_code=500, detail="Ошибка получения данных календаря")


@router.get("/calendar/api/objects")
async def owner_calendar_api_objects(request: Request):
    """API: список объектов владельца (массив для drag&drop-панели)."""
    current_user = await get_current_user(request)
    if not current_user:
        raise HTTPException(status_code=403, detail="Доступ запрещен")
    
    # Проверяем роли (поддержка множественных ролей)
    user_roles = current_user.get("roles", [])
    if isinstance(user_roles, str):
        user_roles = [user_roles]
    if "owner" not in user_roles and "superadmin" not in user_roles:
        raise HTTPException(status_code=403, detail="Доступ запрещен")

    # Проверяем кэш
    user_id_key = current_user.get("telegram_id") or current_user.get("id")
    from core.cache.redis_cache import cache
    cache_key = f"api_objects:{user_id_key}"
    cached_data = await cache.get(cache_key, serialize="json")
    if cached_data:
        logger.info(f"Owner objects API: cache HIT for user {user_id_key}")
        return cached_data
    
    logger.info(f"Owner objects API: cache MISS for user {user_id_key}")

    try:
        async with get_async_session() as session:
            # Определяем владельца по telegram_id
            from sqlalchemy import select
            from domain.entities.user import User
            from domain.entities.object import Object

            if not current_user or not current_user.get("telegram_id"):
                return []

            owner_q = select(User).where(User.telegram_id == current_user["telegram_id"])
            owner = (await session.execute(owner_q)).scalar_one_or_none()
            if not owner:
                return []

            objects_q = select(Object).where(Object.owner_id == owner.id, Object.is_active == True).order_by(Object.created_at.desc())
            objects = (await session.execute(objects_q)).scalars().all()

            objects_data = [
                {
                    "id": obj.id,
                    "name": obj.name,
                    "hourly_rate": float(obj.hourly_rate),
                    "opening_time": obj.opening_time.strftime("%H:%M") if obj.opening_time else "09:00",
                    "closing_time": obj.closing_time.strftime("%H:%M") if obj.closing_time else "21:00",
                }
                for obj in objects
            ]
            
            # Сохраняем в кэш (TTL 2 минуты)
            await cache.set(cache_key, objects_data, ttl=120, serialize="json")
            logger.info(f"Owner objects API: cached {len(objects_data)} objects")
            
            return objects_data

    except Exception as e:
        logger.error(f"Error getting objects: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки объектов")


@router.get("/api/employees")
async def api_employees(request: Request):
    """API: список сотрудников для drag&drop панели."""
    current_user = await get_current_user(request)
    if not current_user:
        raise HTTPException(status_code=403, detail="Доступ запрещен")
    
    # Проверяем роли (поддержка множественных ролей)
    user_roles = current_user.get("roles", [])
    if isinstance(user_roles, str):
        user_roles = [user_roles]
    if "owner" not in user_roles and "superadmin" not in user_roles:
        raise HTTPException(status_code=403, detail="Доступ запрещен")

    # Проверяем кэш
    user_id_key = current_user.get("telegram_id") or current_user.get("id")
    from core.cache.redis_cache import cache
    cache_key = f"api_employees:{user_id_key}"
    cached_data = await cache.get(cache_key, serialize="json")
    if cached_data:
        logger.info(f"Owner employees API: cache HIT for user {user_id_key}")
        return cached_data
    
    logger.info(f"Owner employees API: cache MISS for user {user_id_key}")

    try:
        async with get_async_session() as session:
            from sqlalchemy import select
            from domain.entities.user import User
            from domain.entities.object import Object
            from domain.entities.shift import Shift

            # Получаем внутренний ID владельца
            user_id = await get_user_id_from_current_user(current_user, session)
            if not user_id:
                raise HTTPException(status_code=404, detail="Владелец не найден")

            # Получаем сотрудников, с которыми у владельца есть активные договоры
            from domain.entities.contract import Contract
            employees_q = select(User).distinct().join(Contract, User.id == Contract.employee_id).where(
                Contract.owner_id == user_id,  # Только договоры с текущим владельцем
                Contract.status == "active",
                Contract.is_active == True
            )
            all_employees = (await session.execute(employees_q)).scalars().all()
            
            # Получаем самого владельца для добавления в список
            owner_user = await session.execute(select(User).where(User.id == user_id))
            owner = owner_user.scalar_one_or_none()
            
            # Собираем всех сотрудников + владельца
            employees = []
            
            # Добавляем сотрудников с договорами
            for emp in all_employees:
                employee_roles = emp.roles if isinstance(emp.roles, list) else [emp.role]
                if "employee" in employee_roles:
                    employees.append(emp)
            
            # Добавляем самого владельца в начало списка
            if owner:
                employees.insert(0, owner)

            employees_data = [
                {
                    "id": emp.id,
                    "name": f"{emp.first_name or ''} {emp.last_name or ''}".strip() or emp.username,
                    "username": emp.username,
                    "role": emp.role,
                    "is_active": emp.is_active,
                    "telegram_id": emp.telegram_id,
                    "is_owner": emp.id == user_id  # Флаг для отличия владельца
                }
                for emp in employees
            ]
            
            # Сохраняем в кэш (TTL 2 минуты)
            await cache.set(cache_key, employees_data, ttl=120, serialize="json")
            logger.info(f"Owner employees API: cached {len(employees_data)} employees")
            
            return employees_data

    except Exception as e:
        logger.error(f"Error getting employees: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки сотрудников")
@router.get("/api/contracts/my-contracts")
async def api_my_contracts(request: Request):
    """API: получение договоров владельца."""
    current_user = await get_current_user(request)
    if not current_user:
        raise HTTPException(status_code=403, detail="Доступ запрещен")
    
    # Проверяем роли (поддержка множественных ролей)
    user_roles = current_user.get("roles", [])
    if isinstance(user_roles, str):
        user_roles = [user_roles]
    if "owner" not in user_roles and "superadmin" not in user_roles:
        raise HTTPException(status_code=403, detail="Доступ запрещен")
    
    try:
        async with get_async_session() as session:
            # Получаем внутренний ID пользователя
            telegram_id = current_user.get("telegram_id") or current_user.get("id")
            user_query = select(User).where(User.telegram_id == telegram_id)
            user_result = await session.execute(user_query)
            user_obj = user_result.scalar_one_or_none()
            
            if not user_obj:
                raise HTTPException(status_code=404, detail="Пользователь не найден")
            
            # Получаем договоры владельца
            contracts_query = select(Contract).where(Contract.owner_id == user_obj.id)
            contracts_result = await session.execute(contracts_query)
            contracts = contracts_result.scalars().all()
            
            # Формируем ответ
            contracts_data = []
            for contract in contracts:
                # Получаем информацию о сотруднике
                employee_query = select(User).where(User.id == contract.employee_id)
                employee_result = await session.execute(employee_query)
                employee = employee_result.scalar_one_or_none()
                
                contracts_data.append({
                    "id": contract.id,
                    "contract_number": contract.contract_number,
                    "title": contract.title,
                    "employee_id": contract.employee_id,
                    "employee_name": f"{employee.first_name} {employee.last_name}" if employee else "Неизвестно",
                    "status": contract.status,
                    "start_date": contract.start_date.isoformat() if contract.start_date else None,
                    "end_date": contract.end_date.isoformat() if contract.end_date else None
                })
            
            return JSONResponse(content={
                "success": True,
                "contracts": contracts_data,
                "count": len(contracts_data)
            })
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting contracts: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки договоров")


@router.get("/api/employees/for-object/{object_id}")
async def api_employees_for_object(object_id: int, request: Request):
    """API: получение списка сотрудников с доступом к конкретному объекту."""
    current_user = await get_current_user(request)
    if not current_user:
        raise HTTPException(status_code=403, detail="Доступ запрещен")
    
    # Проверяем роли (поддержка множественных ролей)
    user_roles = current_user.get("roles", [])
    if isinstance(user_roles, str):
        user_roles = [user_roles]
    if "owner" not in user_roles and "superadmin" not in user_roles:
        raise HTTPException(status_code=403, detail="Доступ запрещен")
    
    try:
        async with get_async_session() as session:
            from sqlalchemy import select
            from domain.entities.user import User
            from domain.entities.contract import Contract
            from domain.entities.object import Object
            import json
            
            # Получаем внутренний ID владельца
            user_id = await get_user_id_from_current_user(current_user, session)
            if not user_id:
                raise HTTPException(status_code=404, detail="Владелец не найден")
            
            # Проверяем, что объект принадлежит владельцу
            object_query = select(Object).where(Object.id == object_id, Object.owner_id == user_id)
            object_result = await session.execute(object_query)
            if not object_result.scalar_one_or_none():
                raise HTTPException(status_code=404, detail="Объект не найден")
            
            # Получаем сотрудников владельца с доступом к конкретному объекту
            from sqlalchemy.dialects.postgresql import JSONB
            from sqlalchemy import cast
            
            employees_with_access = []
            
            # Получаем сотрудников с договорами у ЭТОГО владельца, имеющих доступ к объекту (включая уволенных)
            employees_query = select(User).distinct().join(
                Contract, User.id == Contract.employee_id
            ).where(
                Contract.owner_id == user_id,  # Только сотрудники этого владельца
                # Включаем активные и уволенные договоры
                Contract.is_active == True,
                # Проверяем, что object_id есть в allowed_objects (JSONB массив)
                cast(Contract.allowed_objects, JSONB).op('@>')(cast([object_id], JSONB))
            )
            employees_result = await session.execute(employees_query)
            all_employees = employees_result.scalars().all()
            
            # Формируем список сотрудников (включая уволенных)
            # Получаем договоры для определения статуса
            contracts_query = select(Contract).where(
                Contract.employee_id.in_([emp.id for emp in all_employees]),
                Contract.owner_id == user_id,
                cast(Contract.allowed_objects, JSONB).op('@>')(cast([object_id], JSONB))
            )
            contracts_result = await session.execute(contracts_query)
            contracts_list = contracts_result.scalars().all()
            contracts_by_employee = {c.employee_id: c for c in contracts_list}
            
            for employee in all_employees:
                contract = contracts_by_employee.get(employee.id)
                # Определяем is_active на основе статуса договора
                is_active_contract = contract and contract.status == "active" if contract else False
                is_former = not is_active_contract
                
                employee_data = {
                    "id": int(employee.id),
                    "name": str(f"{employee.first_name or ''} {employee.last_name or ''}".strip() or employee.username or f"ID {employee.id}"),
                    "username": str(employee.username or ""),
                    "role": str(employee.role),
                    "is_active": is_active_contract,  # На основе статуса договора
                    "isFormer": is_former,  # Для совместимости с plan_shift.js
                    "telegram_id": int(employee.telegram_id) if employee.telegram_id else None
                }
                employees_with_access.append(employee_data)
            
            logger.info(f"Found {len(employees_with_access)} employees for object {object_id}, owner {user_id}")
            
            return employees_with_access
            
    except Exception as e:
        logger.error(f"Error fetching employees for object {object_id}: {e}")
        raise HTTPException(status_code=500, detail="Ошибка получения списка сотрудников для объекта")


@router.post("/calendar/api/quick-create-timeslot")
async def owner_calendar_quick_create_timeslot(
    request: Request,
    object_id: int = Form(...),
    slot_date: str = Form(...),
    start_time: str = Form(...),
    end_time: str = Form(...),
    hourly_rate: str = Form(...),
):
    """API: быстрое создание тайм-слота из drag&drop-панели."""
    current_user = await get_current_user(request)
    if current_user.get("role") != "owner":
        raise HTTPException(status_code=403, detail="Доступ запрещен")

    try:
        # Валидация
        try:
            slot_date_obj = datetime.strptime(slot_date, "%Y-%m-%d").date()
            start_time_obj = time.fromisoformat(start_time)
            end_time_obj = time.fromisoformat(end_time)
            hourly_rate_decimal = Decimal(hourly_rate) if hourly_rate else Decimal(0)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=f"Неверный формат данных: {str(e)}")

        if start_time_obj >= end_time_obj:
            raise HTTPException(status_code=400, detail="Время начала должно быть меньше времени окончания")
        if hourly_rate_decimal <= 0:
            raise HTTPException(status_code=400, detail="Ставка должна быть больше 0")

        async with get_async_session() as session:
            from apps.web.services.object_service import ObjectService, TimeSlotService

            timeslot_service = TimeSlotService(session)
            object_service = ObjectService(session)

            # Создание слота
            timeslot_data = {
                "slot_date": slot_date_obj,
                "start_time": start_time,
                "end_time": end_time,
                "hourly_rate": hourly_rate_decimal,
                "is_active": True,
            }

            telegram_id = int(current_user.get("telegram_id") or current_user.get("id"))

            new_slot = await timeslot_service.create_timeslot(
                timeslot_data,
                object_id,
                telegram_id,
            )

            # Очищаем кэш календаря безопасно (не роняем при ошибке Redis)
            try:
                from core.cache.redis_cache import cache
                await cache.clear_pattern("calendar_shifts:*")
                await cache.clear_pattern("api_response:*")
                if new_slot:
                    logger.info(f"Calendar cache cleared after creating timeslot {new_slot.id}")
                else:
                    logger.info("Calendar cache cleared after duplicate timeslot attempt")
            except Exception:
                logger.warning("Cache clear skipped (redis not available)")

            # Если дубликат (new_slot is None) — возвращаем идемпотентный успех
            if not new_slot:
                return {"success": True, "already_exists": True, "message": "Тайм-слот уже существует"}

            # Инфо об объекте для ответа
            obj = await object_service.get_object_by_id(object_id, telegram_id)

            return {
                "success": True,
                "timeslot": {
                    "id": new_slot.id,
                    "object_id": new_slot.object_id,
                    "object_name": obj.name if obj else "Неизвестный объект",
                    "date": new_slot.slot_date.strftime("%Y-%m-%d"),
                    "start_time": new_slot.start_time.strftime("%H:%M"),
                    "end_time": new_slot.end_time.strftime("%H:%M"),
                    "hourly_rate": float(new_slot.hourly_rate) if new_slot.hourly_rate else 0,
                    "is_active": new_slot.is_active,
                },
            }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating quick timeslot: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка создания тайм-слота: {str(e)}")


@router.post("/api/calendar/plan-shift")
@router.put("/api/calendar/plan-shift")
async def api_calendar_plan_shift(
    request: Request,
    timeslot_id: int = Body(...),
    employee_id: int = Body(...),
    start_time: Optional[str] = Body(None),
    end_time: Optional[str] = Body(None),
    schedule_id: Optional[int] = Body(None)
):
    """API: планирование или обновление смены для сотрудника в тайм-слот."""
    current_user = await get_current_user(request)
    if not current_user:
        raise HTTPException(status_code=403, detail="Доступ запрещен")
    
    # Проверяем роли (поддержка множественных ролей)
    user_roles = current_user.get("roles", [])
    if isinstance(user_roles, str):
        user_roles = [user_roles]
    if "owner" not in user_roles and "superadmin" not in user_roles:
        raise HTTPException(status_code=403, detail="Доступ запрещен")

    try:
        async with get_async_session() as session:
            from sqlalchemy import select
            from domain.entities.user import User
            from domain.entities.object import Object
            from domain.entities.time_slot import TimeSlot
            from domain.entities.shift_schedule import ShiftSchedule
            from domain.entities.contract import Contract
            from datetime import datetime
            import json

            # Получаем внутренний ID владельца
            user_id = await get_user_id_from_current_user(current_user, session)
            if not user_id:
                raise HTTPException(status_code=404, detail="Владелец не найден")
            
            logger.info(f"Planning shift for user_id: {user_id}, timeslot_id: {timeslot_id}, employee_id: {employee_id}")

            # Проверяем, что тайм-слот существует и принадлежит владельцу
            timeslot_query = select(TimeSlot).options(selectinload(TimeSlot.object)).join(Object).where(
                TimeSlot.id == timeslot_id,
                Object.owner_id == user_id
            )
            timeslot = (await session.execute(timeslot_query)).scalar_one_or_none()
            if not timeslot:
                raise HTTPException(status_code=404, detail="Тайм-слот не найден")

            # Проверяем, что сотрудник существует
            employee_query = select(User).where(User.id == employee_id)
            employee = (await session.execute(employee_query)).scalar_one_or_none()
            if not employee:
                raise HTTPException(status_code=404, detail="Сотрудник не найден")
            
            # Проверяем роли; владельцу разрешаем планирование независимо от employee-роли
            employee_roles = employee.roles if isinstance(employee.roles, list) else [employee.role]
            is_owner_assignee = employee_id == user_id or "owner" in employee_roles
            if "employee" not in employee_roles and not is_owner_assignee:
                raise HTTPException(status_code=400, detail="Пользователь не является сотрудником")

            # Проверяем, что у сотрудника есть договор с владельцем (кроме случая, когда сотрудник — это сам владелец)
            from shared.services.contract_validation_service import build_active_contract_filter
            from datetime import date as date_type
            
            contract = None
            if not is_owner_assignee:
                # Для планирования используем дату тайм-слота
                shift_date = timeslot.slot_date if timeslot.slot_date else date_type.today()
                
                contract_query = select(Contract).where(
                    and_(
                        Contract.employee_id == employee_id,
                        Contract.owner_id == user_id,
                        build_active_contract_filter(shift_date)
                    )
                )
                contracts = (await session.execute(contract_query)).scalars().all()
                if not contracts:
                    raise HTTPException(status_code=400, detail="У сотрудника нет активного договора с вами")
                contract = contracts[0]
                logger.info(f"Found contract {contract.id} for employee {employee_id}")
                logger.info(f"Contract allowed_objects type: {type(contract.allowed_objects)}")
                logger.info(f"Contract allowed_objects value: {contract.allowed_objects}")
            else:
                logger.info("Assigning owner to own timeslot without contract check")
            
            # Проверяем, что сотрудник имеет доступ к объекту тайм-слота (для владельца пропускаем проверку)
            allowed_objects = []
            if not is_owner_assignee and contract:
                allowed_objects = contract.allowed_objects if contract.allowed_objects else []
                logger.info(f"Contract {contract.id} allowed_objects: {allowed_objects}")
                logger.info(f"Timeslot object_id: {timeslot.object_id}")
                logger.info(f"Timeslot object_id type: {type(timeslot.object_id)}")
                logger.info(f"Timeslot object_id in allowed_objects: {timeslot.object_id in allowed_objects}")
                
                # Дополнительная проверка типов
                if allowed_objects:
                    logger.info(f"allowed_objects elements types: {[type(x) for x in allowed_objects]}")
                    logger.info(f"Comparison: {timeslot.object_id} in {allowed_objects} = {timeslot.object_id in allowed_objects}")
                
                if timeslot.object_id not in allowed_objects:
                    raise HTTPException(status_code=400, detail=f"У сотрудника нет доступа к объекту ID {timeslot.object_id}")

            # Создаем datetime объекты для planned_start и planned_end
            from datetime import datetime, time, date
            import pytz
            
            # Получаем временную зону объекта
            object_timezone = timeslot.object.timezone if timeslot.object and timeslot.object.timezone else 'Europe/Moscow'
            tz = pytz.timezone(object_timezone)

            slot_start_time = timeslot.start_time
            slot_end_time = timeslot.end_time

            if slot_start_time is None or slot_end_time is None:
                raise HTTPException(status_code=400, detail="У тайм-слота не указано время работы")

            if start_time:
                try:
                    custom_start_time = datetime.strptime(start_time, "%H:%M").time()
                except ValueError:
                    raise HTTPException(status_code=400, detail="Неверный формат времени начала (используйте ЧЧ:ММ)")
            else:
                custom_start_time = slot_start_time

            if end_time:
                try:
                    custom_end_time = datetime.strptime(end_time, "%H:%M").time()
                except ValueError:
                    raise HTTPException(status_code=400, detail="Неверный формат времени окончания (используйте ЧЧ:ММ)")
            else:
                custom_end_time = slot_end_time

            if custom_start_time >= custom_end_time:
                raise HTTPException(status_code=400, detail="Время окончания должно быть позже времени начала")

            if custom_start_time < slot_start_time or custom_end_time > slot_end_time:
                raise HTTPException(
                    status_code=400,
                    detail=f"Смена должна укладываться в границы тайм-слота: {slot_start_time.strftime('%H:%M')} - {slot_end_time.strftime('%H:%M')}"
                )
            
            # Создаем naive datetime в локальной временной зоне объекта
            slot_datetime_naive = datetime.combine(timeslot.slot_date, custom_start_time)
            end_datetime_naive = datetime.combine(timeslot.slot_date, custom_end_time)
            
            # Локализуем в временную зону объекта, затем конвертируем в UTC для сохранения
            slot_datetime = tz.localize(slot_datetime_naive).astimezone(pytz.UTC)
            end_datetime = tz.localize(end_datetime_naive).astimezone(pytz.UTC)
            
            # Режим редактирования: обновление существующей смены
            is_edit_mode = schedule_id is not None
            existing_schedule = None
            
            if is_edit_mode:
                # Получаем существующую смену для редактирования
                existing_schedule_query = select(ShiftSchedule).where(ShiftSchedule.id == schedule_id)
                existing_schedule = (await session.execute(existing_schedule_query)).scalar_one_or_none()
                if not existing_schedule:
                    raise HTTPException(status_code=404, detail="Запланированная смена не найдена")
                
                # Проверяем права доступа: смена должна принадлежать объекту владельца
                if existing_schedule.object_id != timeslot.object_id:
                    raise HTTPException(status_code=403, detail="Нет доступа к этой смене")
            
            # Проверяем, что сотрудник не занят в это время (исключая редактируемую смену)
            existing_schedule_query = select(ShiftSchedule).where(
                ShiftSchedule.user_id == employee_id,
                ShiftSchedule.planned_start < end_datetime,
                ShiftSchedule.planned_end > slot_datetime,
                ShiftSchedule.status.in_(["planned", "confirmed"])
            )
            if is_edit_mode:
                existing_schedule_query = existing_schedule_query.where(ShiftSchedule.id != schedule_id)
            existing_schedules = (await session.execute(existing_schedule_query)).scalars().all()
            if existing_schedules:
                raise HTTPException(status_code=400, detail="Сотрудник уже запланирован на это время")

            # Проверяем лимит по количеству сотрудников в тайм-слоте (исключая редактируемую смену)
            max_employees = timeslot.max_employees or 1
            current_slot_schedules_query = select(ShiftSchedule).where(
                ShiftSchedule.time_slot_id == timeslot_id,
                ShiftSchedule.status.in_(["planned", "confirmed"])
            )
            if is_edit_mode:
                current_slot_schedules_query = current_slot_schedules_query.where(ShiftSchedule.id != schedule_id)
            current_slot_schedules = (await session.execute(current_slot_schedules_query)).scalars().all()
            overlapping_schedules = [
                sched for sched in current_slot_schedules
                if not (sched.planned_end <= slot_datetime or sched.planned_start >= end_datetime)
            ]
            if len(overlapping_schedules) >= max_employees:
                raise HTTPException(status_code=400, detail="На выбранное время нет свободных мест в тайм-слоте")

            # Определяем ставку с учетом флага use_contract_rate
            effective_rate = None
            
            if not is_owner_assignee and contract:
                # Используем метод модели Contract для определения эффективной ставки
                timeslot_rate = float(timeslot.hourly_rate) if timeslot.hourly_rate and float(timeslot.hourly_rate) > 0 else None
                object_rate = float(timeslot.object.hourly_rate) if timeslot.object and timeslot.object.hourly_rate else None
                effective_rate = contract.get_effective_hourly_rate(
                    timeslot_rate=timeslot_rate,
                    object_rate=object_rate
                )
            else:
                # Для владельца или без договора: тайм-слот > объект
                if timeslot.hourly_rate and float(timeslot.hourly_rate) > 0:
                    effective_rate = timeslot.hourly_rate
                elif timeslot.object and timeslot.object.hourly_rate:
                    effective_rate = timeslot.object.hourly_rate
            
            actor_role = "owner" if "owner" in user_roles else "superadmin"
            history_service = ShiftHistoryService(session)
            
            if is_edit_mode:
                # Обновляем существующую смену
                old_status = existing_schedule.status
                old_employee_id = existing_schedule.user_id
                old_start = existing_schedule.planned_start
                old_end = existing_schedule.planned_end
                
                existing_schedule.user_id = int(employee_id)
                existing_schedule.planned_start = slot_datetime
                existing_schedule.planned_end = end_datetime
                if effective_rate is not None:
                    existing_schedule.hourly_rate = effective_rate
                
                await history_service.log_event(
                    operation="schedule_update",
                    source="web",
                    actor_id=user_id,
                    actor_role=actor_role,
                    schedule_id=existing_schedule.id,
                    old_status=old_status,
                    new_status=old_status,  # Статус не меняется при обновлении
                    payload={
                        "object_id": int(timeslot.object_id),
                        "time_slot_id": int(timeslot_id),
                        "old_employee_id": int(old_employee_id),
                        "new_employee_id": int(employee_id),
                        "old_planned_start": old_start.isoformat() if old_start else None,
                        "old_planned_end": old_end.isoformat() if old_end else None,
                        "new_planned_start": slot_datetime.isoformat(),
                        "new_planned_end": end_datetime.isoformat(),
                        "origin": "owner_calendar",
                    },
                )
                await session.commit()
                
                # Очищаем кэш календаря
                from core.cache.redis_cache import cache
                await cache.clear_pattern("calendar_shifts:*")
                await cache.clear_pattern("api_response:*")
                logger.info(f"Calendar cache cleared after updating shift {existing_schedule.id}")

                return {
                    "success": True,
                    "message": f"Смена обновлена для {employee.first_name or employee.username}",
                    "schedule_id": existing_schedule.id
                }
            else:
                # Создаем новую запланированную смену
                new_schedule = ShiftSchedule(
                    user_id=int(employee_id),
                    object_id=int(timeslot.object_id),
                    time_slot_id=int(timeslot_id),
                    planned_start=slot_datetime,
                    planned_end=end_datetime,
                    status="planned",
                    hourly_rate=effective_rate,
                    notes="Запланировано через drag&drop"
                )
                session.add(new_schedule)
                await session.flush()

                await history_service.log_event(
                    operation="schedule_plan",
                    source="web",
                    actor_id=user_id,
                    actor_role=actor_role,
                    schedule_id=new_schedule.id,
                    old_status=None,
                    new_status="planned",
                    payload={
                        "object_id": int(timeslot.object_id),
                        "time_slot_id": int(timeslot_id),
                        "employee_id": int(employee_id),
                        "planned_start": slot_datetime.isoformat(),
                        "planned_end": end_datetime.isoformat(),
                        "origin": "owner_calendar",
                    },
                )
                await session.commit()
                
                # Очищаем кэш календаря для немедленного отображения
                from core.cache.redis_cache import cache
                await cache.clear_pattern("calendar_shifts:*")
                await cache.clear_pattern("api_response:*")
                logger.info(f"Calendar cache cleared after planning shift {new_schedule.id}")

                return {
                    "success": True,
                    "message": f"Смена запланирована для {employee.first_name or employee.username}",
                    "schedule_id": new_schedule.id
                }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error planning shift: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка планирования смены: {str(e)}")


@router.delete("/api/calendar/plan-shift/{schedule_id}")
async def api_calendar_delete_shift(
    request: Request,
    schedule_id: int
):
    """API: удаление запланированной смены."""
    current_user = await get_current_user(request)
    if not current_user:
        raise HTTPException(status_code=403, detail="Доступ запрещен")
    
    # Проверяем роли (поддержка множественных ролей)
    user_roles = current_user.get("roles", [])
    if isinstance(user_roles, str):
        user_roles = [user_roles]
    if "owner" not in user_roles and "superadmin" not in user_roles:
        raise HTTPException(status_code=403, detail="Доступ запрещен")

    try:
        async with get_async_session() as session:
            from sqlalchemy import select
            from domain.entities.user import User
            from domain.entities.object import Object
            from domain.entities.shift_schedule import ShiftSchedule
            from shared.services.shift_history_service import ShiftHistoryService

            # Получаем внутренний ID владельца
            user_id = await get_user_id_from_current_user(current_user, session)
            if not user_id:
                raise HTTPException(status_code=404, detail="Владелец не найден")
            
            logger.info(f"Deleting shift schedule {schedule_id} by user_id: {user_id}")

            # Получаем запланированную смену
            schedule_query = select(ShiftSchedule).options(selectinload(ShiftSchedule.object)).where(
                ShiftSchedule.id == schedule_id
            )
            schedule = (await session.execute(schedule_query)).scalar_one_or_none()
            if not schedule:
                raise HTTPException(status_code=404, detail="Запланированная смена не найдена")
            
            # Проверяем, что смена принадлежит объекту владельца
            object_query = select(Object).where(
                Object.id == schedule.object_id,
                Object.owner_id == user_id
            )
            obj = (await session.execute(object_query)).scalar_one_or_none()
            if not obj:
                raise HTTPException(status_code=403, detail="Нет доступа к этой смене")
            
            # Проверяем, что смена может быть удалена (только planned или confirmed)
            if schedule.status not in ["planned", "confirmed"]:
                raise HTTPException(status_code=400, detail="Нельзя удалить смену со статусом " + schedule.status)
            
            # Логируем удаление в историю
            actor_role = "owner" if "owner" in user_roles else "superadmin"
            history_service = ShiftHistoryService(session)
            await history_service.log_event(
                operation="schedule_delete",
                source="web",
                actor_id=user_id,
                actor_role=actor_role,
                schedule_id=schedule.id,
                old_status=schedule.status,
                new_status="cancelled",
                payload={
                    "object_id": schedule.object_id,
                    "time_slot_id": schedule.time_slot_id,
                    "employee_id": schedule.user_id,
                    "planned_start": schedule.planned_start.isoformat() if schedule.planned_start else None,
                    "planned_end": schedule.planned_end.isoformat() if schedule.planned_end else None,
                    "origin": "owner_calendar",
                },
            )
            
            # Удаляем смену
            await session.delete(schedule)
            await session.commit()
            
            # Очищаем кэш календаря
            from core.cache.redis_cache import cache
            await cache.clear_pattern("calendar_shifts:*")
            await cache.clear_pattern("api_response:*")
            logger.info(f"Calendar cache cleared after deleting shift {schedule_id}")

            return {
                "success": True,
                "message": "Смена успешно удалена"
            }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting shift: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка удаления смены: {str(e)}")


@router.post("/api/calendar/check-availability")
async def check_employee_availability_owner(
    request: Request,
    timeslot_id: int = Body(...),
    employee_id: int = Body(...)
):
    """Проверка доступности сотрудника для планирования смены (владелец)"""
    current_user = await get_current_user(request)
    if not current_user:
        raise HTTPException(status_code=403, detail="Доступ запрещен")
    
    # Проверяем роли (поддержка множественных ролей)
    user_roles = current_user.get("roles", [])
    if isinstance(user_roles, str):
        user_roles = [user_roles]
    if "owner" not in user_roles and "superadmin" not in user_roles:
        raise HTTPException(status_code=403, detail="Доступ запрещен")

    try:
        async with get_async_session() as session:
            from sqlalchemy import select
            from domain.entities.user import User
            from domain.entities.object import Object
            from domain.entities.time_slot import TimeSlot
            from domain.entities.shift_schedule import ShiftSchedule
            from domain.entities.contract import Contract
            from domain.entities.shift import Shift
            from datetime import datetime
            import pytz

            # Получаем внутренний ID владельца
            user_id = await get_user_id_from_current_user(current_user, session)
            if not user_id:
                raise HTTPException(status_code=404, detail="Владелец не найден")
            
            # Проверяем, что тайм-слот существует и принадлежит владельцу
            timeslot_query = select(TimeSlot).options(selectinload(TimeSlot.object)).join(Object).where(
                TimeSlot.id == timeslot_id,
                Object.owner_id == user_id
            )
            timeslot = (await session.execute(timeslot_query)).scalar_one_or_none()
            if not timeslot:
                raise HTTPException(status_code=404, detail="Тайм-слот не найден")

            # Проверяем, что сотрудник существует
            employee_query = select(User).where(User.id == employee_id)
            employee = (await session.execute(employee_query)).scalar_one_or_none()
            if not employee:
                raise HTTPException(status_code=404, detail="Сотрудник не найден")

            # Проверяем, что у сотрудника есть договор с владельцем
            from shared.services.contract_validation_service import build_active_contract_filter
            from datetime import date as date_type
            
            # Для проверки доступности используем дату тайм-слота
            shift_date = timeslot.slot_date if timeslot.slot_date else date_type.today()
            
            contract_query = select(Contract).where(
                and_(
                    Contract.employee_id == employee_id,
                    Contract.owner_id == user_id,
                    build_active_contract_filter(shift_date)
                )
            )
            contract = (await session.execute(contract_query)).scalar_one_or_none()
            if not contract:
                return {
                    "available": False,
                    "message": "У сотрудника нет договора с владельцем"
                }

            # Проверяем, что у сотрудника есть доступ к объекту тайм-слота
            if contract.allowed_objects and timeslot.object_id not in contract.allowed_objects:
                return {
                    "available": False,
                    "message": "У сотрудника нет доступа к объекту"
                }

            # Получаем временную зону объекта для корректного сравнения времени
            object_timezone = timeslot.object.timezone if timeslot.object and timeslot.object.timezone else 'Europe/Moscow'
            tz = pytz.timezone(object_timezone)
            
            # Создаем naive datetime в локальной временной зоне объекта
            slot_datetime_naive = datetime.combine(timeslot.slot_date, timeslot.start_time)
            end_datetime_naive = datetime.combine(timeslot.slot_date, timeslot.end_time)
            
            # Локализуем в временную зону объекта, затем конвертируем в UTC для сравнения
            slot_datetime_utc = tz.localize(slot_datetime_naive).astimezone(pytz.UTC)
            end_datetime_utc = tz.localize(end_datetime_naive).astimezone(pytz.UTC)
            
            # Проверяем пересечение с активными сменами
            active_shifts_query = select(Shift).where(
                Shift.user_id == employee_id,
                Shift.status == "active",
                Shift.start_time < end_datetime_utc,
                Shift.end_time > slot_datetime_utc
            )
            active_shifts = (await session.execute(active_shifts_query)).scalars().all()
            
            if active_shifts:
                # Берем первую активную смену для детальной информации
                shift = active_shifts[0]
                local_start = shift.start_time.astimezone(tz).strftime('%H:%M')
                local_end = shift.end_time.astimezone(tz).strftime('%H:%M')
                
                # Получаем название объекта смены
                object_name = "Неизвестный объект"
                if hasattr(shift, 'object') and shift.object:
                    object_name = shift.object.name
                elif hasattr(shift, 'object_id'):
                    # Если объект не загружен, получаем его отдельно
                    from domain.entities.object import Object
                    obj_query = select(Object).where(Object.id == shift.object_id)
                    obj = (await session.execute(obj_query)).scalar_one_or_none()
                    if obj:
                        object_name = obj.name
                
                return {
                    "available": False,
                    "message": f"У сотрудника уже есть активная смена в это время",
                    "conflict_info": {
                        "object_name": object_name,
                        "start_time": local_start,
                        "end_time": local_end
                    }
                }
            
            # Проверяем пересечение с запланированными сменами
            planned_shifts_query = select(ShiftSchedule).where(
                ShiftSchedule.user_id == employee_id,
                ShiftSchedule.status == "planned",
                ShiftSchedule.planned_start < end_datetime_utc,
                ShiftSchedule.planned_end > slot_datetime_utc
            )
            planned_shifts = (await session.execute(planned_shifts_query)).scalars().all()
            
            if planned_shifts:
                # Берем первую запланированную смену для детальной информации
                shift = planned_shifts[0]
                local_start = shift.planned_start.astimezone(tz).strftime('%H:%M')
                local_end = shift.planned_end.astimezone(tz).strftime('%H:%M')
                
                # Получаем название объекта смены
                object_name = "Неизвестный объект"
                if hasattr(shift, 'object') and shift.object:
                    object_name = shift.object.name
                elif hasattr(shift, 'object_id'):
                    # Если объект не загружен, получаем его отдельно
                    from domain.entities.object import Object
                    obj_query = select(Object).where(Object.id == shift.object_id)
                    obj = (await session.execute(obj_query)).scalar_one_or_none()
                    if obj:
                        object_name = obj.name
                
                return {
                    "available": False,
                    "message": f"У сотрудника уже есть запланированная смена в это время",
                    "conflict_info": {
                        "object_name": object_name,
                        "start_time": local_start,
                        "end_time": local_end
                    }
                }
            
            return {
                "available": True,
                "message": "Сотрудник доступен"
            }
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error checking employee availability: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка проверки доступности: {str(e)}")


@router.get("/calendar/api/timeslot/{timeslot_id}")
async def owner_calendar_api_timeslot_detail(
    request: Request,
    timeslot_id: int
):
    """Детали конкретного тайм-слота"""
    current_user = await get_current_user(request)
    if not current_user or current_user.get("role") != "owner":
        raise HTTPException(status_code=403, detail="Доступ запрещен")
    
    try:
        async with get_async_session() as session:
            from sqlalchemy import select, and_, func
            from sqlalchemy.orm import selectinload
            from domain.entities.time_slot import TimeSlot
            from domain.entities.object import Object
            from domain.entities.user import User

            # Владелец по текущему пользователю
            if not current_user or not current_user.get("telegram_id"):
                raise HTTPException(status_code=403, detail="Нет доступа")
            owner_q = select(User).where(User.telegram_id == current_user["telegram_id"])
            owner = (await session.execute(owner_q)).scalar_one_or_none()
            if not owner:
                raise HTTPException(status_code=404, detail="Пользователь не найден")

            # Слот + проверка принадлежности через объект
            slot_q = select(TimeSlot).options(selectinload(TimeSlot.object)).where(TimeSlot.id == timeslot_id)
            slot = (await session.execute(slot_q)).scalar_one_or_none()
            if not slot:
                raise HTTPException(status_code=404, detail="Тайм-слот не найден")
            obj_q = select(Object).where(Object.id == slot.object_id, Object.owner_id == owner.id)
            if (await session.execute(obj_q)).scalar_one_or_none() is None:
                raise HTTPException(status_code=403, detail="Нет доступа к тайм-слоту")

            # Загружаем запланированные смены (ShiftSchedule)
            from domain.entities.shift_schedule import ShiftSchedule
            scheduled_query = select(ShiftSchedule).options(
                selectinload(ShiftSchedule.user)
            ).where(
                and_(
                    ShiftSchedule.time_slot_id == timeslot_id,
                    ShiftSchedule.status.in_(["planned", "confirmed"])
                )
            )
            scheduled_result = await session.execute(scheduled_query)
            scheduled_shifts = scheduled_result.scalars().all()
            
            # Загружаем фактические смены (Shift) по объекту и дате
            from domain.entities.shift import Shift
            from sqlalchemy import func
            actual_query = select(Shift).options(
                selectinload(Shift.user)
            ).where(
                and_(
                    Shift.object_id == slot.object_id,
                    func.date(Shift.start_time) == slot.slot_date,
                    Shift.status.in_(["active", "completed"])
                )
            )
            actual_result = await session.execute(actual_query)
            actual_shifts = actual_result.scalars().all()
            
            tz_name = slot.object.timezone if slot.object and slot.object.timezone else 'Europe/Moscow'

            # Формируем данные для запланированных смен
            scheduled_data = []
            for shift in scheduled_shifts:
                user_name = "Неизвестно"
                if shift.user:
                    first_name = shift.user.first_name or ""
                    last_name = shift.user.last_name or ""
                    user_name = f"{last_name} {first_name}".strip() or shift.user.username or f"ID {shift.user.id}"

                start_formatted = (
                    web_timezone_helper.format_datetime_with_timezone(shift.planned_start, tz_name, '%H:%M')
                    if shift.planned_start
                    else None
                )
                end_formatted = (
                    web_timezone_helper.format_datetime_with_timezone(shift.planned_end, tz_name, '%H:%M')
                    if shift.planned_end
                    else None
                )

                scheduled_data.append({
                    "id": shift.id,
                    "user_id": shift.user_id,
                    "user_name": user_name,
                    "status": shift.status,
                    "planned_hours": shift.planned_duration_hours,
                    "notes": shift.notes or "",
                    "planned_start": shift.planned_start.isoformat() if shift.planned_start else None,
                    "planned_end": shift.planned_end.isoformat() if shift.planned_end else None,
                    "start_time": start_formatted,
                    "end_time": end_formatted,
                    "hourly_rate": float(shift.hourly_rate) if shift.hourly_rate else None
                })
            
            # Формируем данные для фактических смен
            actual_data = []
            total_hours = 0
            total_payment = 0
            for shift in actual_shifts:
                hours = shift.total_hours or 0
                payment = shift.total_payment or (hours * (slot.hourly_rate or 0))
                total_hours += hours
                total_payment += payment
                
                actual_data.append({
                    "id": shift.id,
                    "user_name": f"{shift.user.first_name} {shift.user.last_name}".strip() if shift.user else "Неизвестно",
                    "status": shift.status,
                    "actual_hours": hours,
                    "payment": payment,
                    "notes": shift.notes or ""
                })

            return {
                "slot": {
                    "id": slot.id,
                    "object_id": slot.object_id,
                    "object_name": slot.object.name if slot.object else None,
                    "date": slot.slot_date.strftime("%Y-%m-%d"),
                    "start_time": slot.start_time.strftime("%H:%M"),
                    "end_time": slot.end_time.strftime("%H:%M"),
                    "hourly_rate": float(slot.hourly_rate) if slot.hourly_rate else None,
                    "max_employees": slot.max_employees or 1,
                    "is_active": slot.is_active,
                    "notes": slot.notes or "",
                },
                "scheduled": scheduled_data,
                "actual": actual_data,
                "summary": {
                    "total_hours": total_hours,
                    "total_payment": total_payment
                }
            }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting timeslot details: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки деталей тайм-слота")


def _create_calendar_grid(year: int, month: int, timeslots: List[Dict[str, Any]], shifts: List[Dict[str, Any]] = None) -> List[List[Dict[str, Any]]]:
    """Создает календарную сетку с тайм-слотами и сменами"""
    import calendar as py_calendar
    if shifts is None:
        shifts = []
        
    # Получаем первый день месяца и количество дней
    first_day = date(year, month, 1)
    last_day = date(year, month, py_calendar.monthrange(year, month)[1])
    
    # Находим понедельник для начала календаря
    today = date.today()
    if today.year == year and today.month == month:
        # Если смотрим текущий месяц - начинаем за 1 неделю до текущей
        current_monday = today - timedelta(days=today.weekday())
        first_monday = current_monday - timedelta(weeks=1)
    else:
        # Для других месяцев - начинаем с первого понедельника месяца
        first_monday = first_day - timedelta(days=first_day.weekday())
    
    # Создаем сетку 4x7 (4 недели, 7 дней) - 1 назад + текущая + 2 вперед
    calendar_grid = []
    current_date = first_monday
    
    for week in range(4):
        week_data = []
        for day in range(7):
            # Фильтруем смены: скрываем запланированные, если есть активные или завершенные для того же объекта
            day_shifts = []
            active_shifts_by_object = {}
            completed_shifts_by_object = {}
            
            # Сначала собираем все смены за день
            all_day_shifts = [
                shift for shift in shifts 
                if shift["date"] == current_date
            ]
            
            # Группируем активные и завершенные смены по объектам
            for shift in all_day_shifts:
                object_id = shift.get("object_id")
                if object_id:
                    if shift["status"] == "active":
                        active_shifts_by_object[object_id] = shift
                    elif shift["status"] == "completed":
                        completed_shifts_by_object[object_id] = shift
            
            # Показываем все смены - не фильтруем запланированные смены
            for shift in all_day_shifts:
                day_shifts.append(shift)
            
            # Фильтруем тайм-слоты: показываем только те, у которых нет связанных смен
            day_timeslots = []
            for slot in timeslots:
                if slot["date"] == current_date and slot.get("is_active", True):
                    # Проверяем, есть ли смены для этого тайм-слота
                    has_related_shift = False
                    for shift in day_shifts:
                        # Проверяем, что смена не отменена и относится к тому же объекту
                        if (shift["object_id"] == slot["object_id"] and 
                            shift["status"] not in ['cancelled']):
                            has_related_shift = True
                            break
                    
                    # Показываем тайм-слот всегда, но с правильным индикатором занятости
                    day_timeslots.append(slot)
                    
                    # Debug для тайм-слота 508
                    if slot["id"] == 508:
                        logger.info(f"Timeslot 508: date={slot['date']}, current_date={current_date}, is_active={slot.get('is_active', True)}, has_related_shift={has_related_shift}, max_employees={slot.get('max_employees', 1)}")
                        logger.info(f"Timeslot 508: day_shifts count={len(day_shifts)}")
                        for shift in day_shifts:
                            if shift.get("object_id") == slot["object_id"]:
                                logger.info(f"  Related shift: {shift['id']} status={shift['status']}")
            
            if day_timeslots:
                logger.info(f"Found {len(day_timeslots)} timeslots for {current_date}")
            if day_shifts:
                logger.info(f"Found {len(day_shifts)} shifts for {current_date}")
            
            week_data.append({
                "date": current_date,
                "day": current_date.day,
                "is_current_month": current_date.month == month,
                "is_other_month": current_date.month != month,
                "is_today": current_date == date.today(),
                "timeslots": day_timeslots,
                "timeslots_count": len(day_timeslots),
                "shifts": day_shifts,
                "shifts_count": len(day_shifts)
            })
            current_date += timedelta(days=1)
        
        calendar_grid.append(week_data)
    
    return calendar_grid


# Русские названия месяцев (И.п.)
RU_MONTHS = [
    "",
    "Январь",
    "Февраль",
    "Март",
    "Апрель",
    "Май",
    "Июнь",
    "Июль",
    "Август",
    "Сентябрь",
    "Октябрь",
    "Ноябрь",
    "Декабрь",
]
# ШАБЛОНЫ ДОГОВОРОВ
# ===============================

@router.get("/templates/contracts", response_class=HTMLResponse, name="owner_contract_templates")
async def owner_contract_templates(request: Request):
    """Список шаблонов договоров."""
    # Проверяем авторизацию и роль владельца
    current_user = await get_current_user(request)
    if not current_user:
        return RedirectResponse(url="/auth/login", status_code=status.HTTP_302_FOUND)
    
    user_role = current_user.get("role", "employee")
    if user_role != "owner":
        return RedirectResponse(url="/auth/login", status_code=status.HTTP_302_FOUND)
    
    try:
        from apps.web.services.contract_service import ContractService
        
        contract_service = ContractService()
        templates_list = await contract_service.get_contract_templates()
        
        # Получаем данные для переключения интерфейсов
        async with get_async_session() as session:
            user_id = await get_user_id_from_current_user(current_user, session)
            available_interfaces = await get_available_interfaces_for_user(user_id)
        
        return templates.TemplateResponse(
            "owner/templates/contracts/list.html",
            {
                "request": request,
                "templates": templates_list,
                "title": "Шаблоны договоров",
                "current_user": current_user,
                "available_interfaces": available_interfaces
            }
        )
        
    except Exception as e:
        logger.error(f"Error loading contract templates: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки шаблонов договоров")


@router.get("/templates/contracts/create", response_class=HTMLResponse)
async def owner_contract_templates_create_form(request: Request):
    """Форма создания шаблона договора для владельца."""
    # Проверяем авторизацию и роль владельца
    current_user = await get_current_user(request)
    user_role = current_user.get("role", "employee")
    if user_role != "owner":
        return RedirectResponse(url="/auth/login", status_code=status.HTTP_302_FOUND)
    
    # Получаем справочник всех тегов для подсказок
    async with get_async_session() as session:
        from apps.web.services.tag_service import TagService
        tag_service = TagService()
        all_tags = await tag_service.get_all_tags(session)
    
    return templates.TemplateResponse(
        "owner/templates/contracts/create.html",
        {
            "request": request,
            "current_user": current_user,
            "all_tags": all_tags  # Справочник всех тегов для подсказок
        }
    )
@router.post("/templates/contracts/create")
async def owner_create_contract_template(request: Request):
    """Создание шаблона договора владельцем."""
    # Проверяем авторизацию и роль владельца  
    current_user = await get_current_user(request)
    user_role = current_user.get("role", "employee")
    if user_role != "owner":
        return RedirectResponse(url="/auth/login", status_code=status.HTTP_302_FOUND)
    
    try:
        from fastapi import Form
        from apps.web.services.contract_service import ContractService
        
        form_data = await request.form()
        
        name = form_data.get("name", "")
        description = form_data.get("description", "") 
        content = form_data.get("content", "")
        version = form_data.get("version", "1.0")
        is_public = form_data.get("is_public") == "on"
        fields_schema = form_data.get("fields_schema")
        
        if not name or not content:
            raise HTTPException(status_code=400, detail="Название и содержимое обязательны")
            
        contract_service = ContractService()
        
        template_data = {
            "name": name,
            "description": description,
            "content": content,
            "version": version,
            "created_by": current_user["id"],  # telegram_id
            "is_public": is_public,
            "fields_schema": None
        }
        
        if fields_schema:
            try:
                import json
                template_data["fields_schema"] = json.loads(fields_schema)
            except Exception:
                template_data["fields_schema"] = None
        
        template = await contract_service.create_contract_template(template_data)
        
        if template:
            return RedirectResponse(url="/owner/templates/contracts", status_code=303)
        else:
            raise HTTPException(status_code=400, detail="Ошибка создания шаблона")
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating contract template: {e}")
        raise HTTPException(status_code=400, detail=f"Ошибка создания шаблона: {str(e)}")


@router.get("/templates/contracts/{template_id}", response_class=HTMLResponse)
async def owner_contract_template_detail(request: Request, template_id: int):
    """Просмотр шаблона договора владельцем."""
    current_user = await get_current_user(request)
    user_role = current_user.get("role", "employee")
    if user_role != "owner":
        return RedirectResponse(url="/auth/login", status_code=status.HTTP_302_FOUND)

    from apps.web.services.contract_service import ContractService

    contract_service = ContractService()
    template = await contract_service.get_contract_template(template_id)

    if not template:
        raise HTTPException(status_code=404, detail="Шаблон не найден")

    async with get_async_session() as session:
        user_id = await get_user_id_from_current_user(current_user, session)
        owner_context = await get_owner_context(user_id, session)

    context = {
        "request": request,
        "current_user": current_user,
        "template": template,
        "title": template.name,
        **owner_context,
    }

    return templates.TemplateResponse("owner/templates/contracts/detail.html", context)


@router.get("/templates/contracts/{template_id}/edit", response_class=HTMLResponse)
async def owner_contract_template_edit_form(request: Request, template_id: int):
    """Форма редактирования шаблона договора владельцем."""
    current_user = await get_current_user(request)
    user_role = current_user.get("role", "employee")
    if user_role != "owner":
        return RedirectResponse(url="/auth/login", status_code=status.HTTP_302_FOUND)

    from apps.web.services.contract_service import ContractService

    contract_service = ContractService()
    template = await contract_service.get_contract_template(template_id)

    if not template:
        raise HTTPException(status_code=404, detail="Шаблон не найден")

    async with get_async_session() as session:
        user_id = await get_user_id_from_current_user(current_user, session)
        owner_context = await get_owner_context(user_id, session)

    context = {
        "request": request,
        "current_user": current_user,
        "template": template,
        "title": "Редактирование шаблона",
        **owner_context,
    }

    return templates.TemplateResponse("owner/templates/contracts/edit.html", context)


@router.post("/templates/contracts/{template_id}/edit")
async def owner_update_contract_template(request: Request, template_id: int):
    """Обновление шаблона договора владельцем."""
    current_user = await get_current_user(request)
    user_role = current_user.get("role", "employee")
    if user_role != "owner":
        return RedirectResponse(url="/auth/login", status_code=status.HTTP_302_FOUND)

    form_data = await request.form()

    name = form_data.get("name", "").strip()
    description = form_data.get("description", "").strip()
    content = form_data.get("content", "").strip()
    version = form_data.get("version", "1.0").strip()
    is_public = form_data.get("is_public") == "on"
    fields_schema_raw = form_data.get("fields_schema")

    if not name or not content:
        raise HTTPException(status_code=400, detail="Название и содержимое обязательны")

    from apps.web.services.contract_service import ContractService
    import json

    contract_service = ContractService()

    template_data = {
        "name": name,
        "description": description,
        "content": content,
        "version": version,
        "is_public": is_public,
        "fields_schema": None,
    }

    if fields_schema_raw:
        try:
            template_data["fields_schema"] = json.loads(fields_schema_raw)
        except json.JSONDecodeError:
            template_data["fields_schema"] = None

    success = await contract_service.update_contract_template(template_id, template_data)

    if success:
        return RedirectResponse(url=f"/owner/templates/contracts/{template_id}", status_code=303)

    raise HTTPException(status_code=400, detail="Ошибка обновления шаблона")


# ===============================
# ТАЙМ-СЛОТЫ
# ===============================

# Дубликат списка тайм-слотов отключён. Используйте маршруты из apps/web/routes/owner_timeslots.py


@router.get("/timeslots/{timeslot_id}", response_class=HTMLResponse)
async def owner_timeslot_detail(
    request: Request,
    timeslot_id: int,
    current_user: dict = Depends(require_owner_or_superadmin),
    db: AsyncSession = Depends(get_db_session)
):
    """Детали тайм-слота владельца"""
    try:
        # Проверяем, что current_user - это словарь, а не RedirectResponse
        if isinstance(current_user, RedirectResponse):
            return current_user
            
        # Получаем внутренний ID пользователя
        telegram_id = current_user.get("telegram_id") or current_user.get("id")
        user_query = select(User).where(User.telegram_id == telegram_id)
        user_result = await db.execute(user_query)
        user_obj = user_result.scalar_one_or_none()
        user_id = user_obj.id if user_obj else None
        
        if not user_id:
            raise HTTPException(status_code=400, detail="Пользователь не найден")
        
        # Получаем тайм-слот
        from domain.entities.time_slot import TimeSlot
        from domain.entities.object import Object
        
        timeslot_query = select(TimeSlot).options(
            selectinload(TimeSlot.object),
            selectinload(TimeSlot.task_templates)
        ).where(TimeSlot.id == timeslot_id)
        
        timeslot_result = await db.execute(timeslot_query)
        timeslot = timeslot_result.scalar_one_or_none()
        
        if not timeslot:
            raise HTTPException(status_code=404, detail="Тайм-слот не найден")
        
        # Проверяем права доступа
        if timeslot.object.owner_id != user_id:
            raise HTTPException(status_code=403, detail="Нет доступа к этому тайм-слоту")
        
        # Получаем связанные смены
        from domain.entities.shift import Shift
        from domain.entities.shift_schedule import ShiftSchedule
        
        shifts_query = select(Shift).options(
            selectinload(Shift.user)
        ).where(Shift.time_slot_id == timeslot_id)
        
        shifts_result = await db.execute(shifts_query)
        shifts = shifts_result.scalars().all()
        
        # Получаем запланированные смены
        schedules_query = select(ShiftSchedule).options(
            selectinload(ShiftSchedule.user)
        ).where(ShiftSchedule.time_slot_id == timeslot_id)
        
        schedules_result = await db.execute(schedules_query)
        schedules = schedules_result.scalars().all()
        
        return templates.TemplateResponse("owner/timeslots/detail.html", {
            "request": request,
            "title": f"Детали тайм-слота {timeslot.start_time.strftime('%H:%M')} - {timeslot.end_time.strftime('%H:%M')}",
            "current_user": current_user,
            "timeslot": timeslot,
            "shifts": shifts,
            "schedules": schedules
        })
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error loading timeslot details: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки деталей тайм-слота")


# ДУБЛИКАТ: этот роут перенесён в owner_timeslots.py (специализированный роутер)
# Оставлен закомментированным для истории. См. apps/web/routes/owner_timeslots.py
"""
@router.get("/timeslots/{timeslot_id}/edit", response_class=HTMLResponse)
async def owner_timeslot_edit_form(
    request: Request,
    timeslot_id: int,
    current_user: dict = Depends(get_current_user_dependency()),
    _: None = Depends(require_role(["owner", "superadmin"]))
):
    # Форма редактирования тайм-слота владельца.
    try:
        # Получаем telegram_id из current_user
        if isinstance(current_user, dict):
            telegram_id = current_user.get("telegram_id") or current_user.get("id")
        else:
            telegram_id = current_user.telegram_id
        
        # Получение тайм-слота из базы данных
        async with get_async_session() as db:
            timeslot_service = TimeSlotService(db)
            object_service = ObjectService(db)
            
            # Получаем тайм-слот с проверкой владельца
            timeslot = await timeslot_service.get_timeslot_by_id(timeslot_id, telegram_id)
            if not timeslot:
                raise HTTPException(status_code=404, detail="Тайм-слот не найден")
            
            # Получаем объект
            obj = await object_service.get_object_by_id(timeslot.object_id, telegram_id)
            if not obj:
                raise HTTPException(status_code=404, detail="Объект не найден")
            
            timeslot_data = {
                "id": timeslot.id,
                "object_id": timeslot.object_id,
                "slot_date": timeslot.slot_date.strftime("%Y-%m-%d"),
                "start_time": timeslot.start_time.strftime("%H:%M"),
                "end_time": timeslot.end_time.strftime("%H:%M"),
                "hourly_rate": float(timeslot.hourly_rate) if timeslot.hourly_rate else float(obj.hourly_rate),
                "max_employees": timeslot.max_employees or 1,
                "is_active": timeslot.is_active
            }
            
            object_data = {
                "id": obj.id,
                "name": obj.name,
                "address": obj.address or "",
                "hourly_rate": float(obj.hourly_rate) if obj.hourly_rate else 0,
                "opening_time": obj.opening_time.strftime("%H:%M") if obj.opening_time else "00:00",
                "closing_time": obj.closing_time.strftime("%H:%M") if obj.closing_time else "23:59",
                "max_distance": obj.max_distance_meters or 0
            }
            
            # Получаем данные для переключения интерфейсов
            user_id = await get_user_id_from_current_user(current_user, db)
            available_interfaces = await get_available_interfaces_for_user(user_id)
            
            return templates.TemplateResponse("owner/timeslots/edit.html", {
                "request": request,
                "title": f"Редактирование тайм-слота: {object_data['name']}",
                "timeslot": timeslot_data,
                "object_id": timeslot.object_id,
                "object": object_data,
                "current_user": current_user,
                "available_interfaces": available_interfaces
            })
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error loading edit form for timeslot {timeslot_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Ошибка загрузки формы редактирования: {str(e)}")
"""

@router.post("/timeslots/{timeslot_id}/edit")
async def owner_timeslot_update(
    request: Request,
    timeslot_id: int,
    current_user: dict = Depends(get_current_user_dependency()),
    _: None = Depends(require_role(["owner", "superadmin"])),
    db: AsyncSession = Depends(get_db_session)
):
    """Обновление тайм-слота владельца."""
    try:
        # Получаем telegram_id из current_user
        if isinstance(current_user, dict):
            telegram_id = current_user.get("telegram_id") or current_user.get("id")
        else:
            telegram_id = current_user.telegram_id
            
        logger.info(f"Updating timeslot {timeslot_id}")
        
        # Получение данных формы
        form_data = await request.form()
        start_time = form_data.get("start_time", "")
        end_time = form_data.get("end_time", "")
        hourly_rate_str = form_data.get("hourly_rate", "0")
        is_active = "is_active" in form_data
        
        # Логирование для отладки
        logger.info(f"Form data: start_time={start_time}, end_time={end_time}, hourly_rate_str='{hourly_rate_str}', is_active={is_active}")
        
        # Валидация и преобразование данных
        try:
            # Очищаем строку от пробелов и проверяем на пустоту
            hourly_rate_str = hourly_rate_str.strip()
            if not hourly_rate_str:
                raise ValueError("Пустое значение ставки")
            hourly_rate = float(hourly_rate_str)
        except ValueError as e:
            logger.error(f"Error parsing hourly_rate '{hourly_rate_str}': {e}")
            raise HTTPException(status_code=400, detail=f"Неверный формат ставки: '{hourly_rate_str}'")
        
        if hourly_rate <= 0:
            raise HTTPException(status_code=400, detail="Ставка должна быть больше 0")
        
        # Валидация времени
        from datetime import time
        try:
            start = time.fromisoformat(start_time)
            end = time.fromisoformat(end_time)
            if start >= end:
                raise HTTPException(status_code=400, detail="Время начала должно быть меньше времени окончания")
        except ValueError:
            raise HTTPException(status_code=400, detail="Неверный формат времени")
        
        # Обработка задач тайм-слота
        task_texts = form_data.getlist("task_texts[]")
        task_amounts = form_data.getlist("task_amounts[]")
        task_mandatory_indices = [int(i) for i in form_data.getlist("task_mandatory[]")]
        task_media_indices = [int(i) for i in form_data.getlist("task_requires_media[]")]
        
        shift_tasks = []
        for idx, text in enumerate(task_texts):
            if text.strip():
                amount = float(task_amounts[idx]) if idx < len(task_amounts) and task_amounts[idx] else 0
                task = {
                    "text": text.strip(),
                    "is_mandatory": idx in task_mandatory_indices,
                    "requires_media": idx in task_media_indices,
                    "bonus_amount": amount if amount >= 0 else 0,
                    "deduction_amount": amount if amount < 0 else 0  # Сохраняем отрицательное!
                }
                shift_tasks.append(task)
        
        # Обновление тайм-слота в базе данных
        timeslot_service = TimeSlotService(db)
        timeslot_data = {
            "start_time": start_time,
            "end_time": end_time,
            "hourly_rate": hourly_rate,
            "max_employees": int(form_data.get("max_employees", 1)),
            "is_active": is_active,
            "penalize_late_start": "penalize_late_start" in form_data and form_data.get("penalize_late_start") not in ["false", ""],
            "shift_tasks": shift_tasks if shift_tasks else None
        }
        
        updated_timeslot = await timeslot_service.update_timeslot(timeslot_id, timeslot_data, telegram_id)
        if not updated_timeslot:
            raise HTTPException(status_code=404, detail="Тайм-слот не найден или нет доступа")
        
        logger.info(f"Timeslot {timeslot_id} updated successfully")
        
        return RedirectResponse(url=f"/owner/timeslots/object/{updated_timeslot.object_id}", status_code=status.HTTP_302_FOUND)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating timeslot: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка обновления тайм-слота: {str(e)}")


@router.post("/timeslots/{timeslot_id}/delete")
async def owner_timeslot_delete(
    timeslot_id: int,
    current_user: dict = Depends(get_current_user_dependency()),
    _: None = Depends(require_role(["owner", "superadmin"])),
    db: AsyncSession = Depends(get_db_session)
):
    """Удаление тайм-слота владельца."""
    try:
        # Получаем telegram_id из current_user
        if isinstance(current_user, dict):
            telegram_id = current_user.get("telegram_id") or current_user.get("id")
        else:
            telegram_id = current_user.telegram_id
            
        logger.info(f"Deleting timeslot {timeslot_id}")
        
        # Удаление тайм-слота из базы данных
        timeslot_service = TimeSlotService(db)
        
        # Получаем тайм-слот для получения object_id
        timeslot = await timeslot_service.get_timeslot_by_id(timeslot_id, telegram_id)
        if not timeslot:
            raise HTTPException(status_code=404, detail="Тайм-слот не найден")
        
        object_id = timeslot.object_id
        
        # Удаляем тайм-слот
        success = await timeslot_service.delete_timeslot(timeslot_id, telegram_id)
        if not success:
            raise HTTPException(status_code=404, detail="Тайм-слот не найден или нет доступа")
        
        logger.info(f"Timeslot {timeslot_id} deleted successfully")
        
        return RedirectResponse(url=f"/owner/timeslots/object/{object_id}", status_code=status.HTTP_302_FOUND)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting timeslot: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка удаления тайм-слота: {str(e)}")


## Дублирующий маршрут формы создания тайм-слота удалён. Используйте маршруты из owner_timeslots.py


@router.post("/timeslots/bulk-edit")
async def owner_timeslots_bulk_edit(
    request: Request,
    current_user: dict = Depends(get_current_user_dependency()),
    _: None = Depends(require_role(["owner", "superadmin"])),
    db: AsyncSession = Depends(get_db_session)
):
    """Массовое редактирование тайм-слотов владельца."""
    try:
        # Получаем telegram_id из current_user
        if isinstance(current_user, dict):
            telegram_id = current_user.get("telegram_id") or current_user.get("id")
        else:
            telegram_id = current_user.telegram_id
            
        logger.info(f"Bulk editing timeslots for user {telegram_id}")
        
        # Получение данных формы
        form_data = await request.form()
        object_id = int(form_data.get("object_id", 0))
        timeslot_ids_str = form_data.get("timeslot_ids", "")
        date_from = form_data.get("date_from", "")
        date_to = form_data.get("date_to", "")
        
        # Параметры для обновления
        start_time = form_data.get("start_time", "").strip()
        end_time = form_data.get("end_time", "").strip()
        hourly_rate_str = form_data.get("hourly_rate", "").strip()
        max_employees_str = form_data.get("max_employees", "").strip()
        is_active = "is_active" in form_data
        is_inactive = "is_inactive" in form_data
        
        # Валидация периода (если указан)
        date_from_obj = None
        date_to_obj = None
        
        if date_from and date_to:
            from datetime import date
            try:
                date_from_obj = date.fromisoformat(date_from)
                date_to_obj = date.fromisoformat(date_to)
                if date_from_obj > date_to_obj:
                    raise HTTPException(status_code=400, detail="Дата начала не может быть больше даты окончания")
            except ValueError:
                raise HTTPException(status_code=400, detail="Неверный формат даты")
        elif date_from or date_to:
            raise HTTPException(status_code=400, detail="Укажите обе даты периода или оставьте пустыми")
        
        # Подготовка данных для обновления
        update_data = {}
        
        from datetime import time

        start_time_obj = None
        if start_time:
            try:
                start_time_obj = time.fromisoformat(start_time)
                update_data["start_time"] = start_time
            except ValueError:
                raise HTTPException(status_code=400, detail="Неверный формат времени начала")

        end_time_obj = None
        if end_time:
            try:
                end_time_obj = time.fromisoformat(end_time)
                update_data["end_time"] = end_time
            except ValueError:
                raise HTTPException(status_code=400, detail="Неверный формат времени окончания")

        if start_time_obj and end_time_obj and start_time_obj >= end_time_obj:
            raise HTTPException(status_code=400, detail="Время начала должно быть меньше времени окончания")
        
        if hourly_rate_str:
            try:
                hourly_rate = float(hourly_rate_str)
                if hourly_rate <= 0:
                    raise HTTPException(status_code=400, detail="Ставка должна быть больше 0")
                update_data["hourly_rate"] = hourly_rate
            except ValueError:
                raise HTTPException(status_code=400, detail="Неверный формат ставки")
        
        if max_employees_str:
            try:
                max_employees = int(max_employees_str)
                if max_employees < 1:
                    raise HTTPException(status_code=400, detail="Лимит должен быть больше 0")
                update_data["max_employees"] = max_employees
            except ValueError:
                raise HTTPException(status_code=400, detail="Неверный формат лимита")
        
        if is_active and not is_inactive:
            update_data["is_active"] = True
        elif is_inactive and not is_active:
            update_data["is_active"] = False
        
        # Новые поля для Phase 4B/4C
        if "penalize_late_start" in form_data:
            update_data["penalize_late_start"] = form_data.get("penalize_late_start") not in ["false", ""]
        if "cancel_late_penalties" in form_data:
            update_data["penalize_late_start"] = False
        
        # Обработка задач из формы (JSONB)
        shift_tasks = []
        task_index = 0
        while f"task_description_{task_index}" in form_data:
            task_desc = form_data.get(f"task_description_{task_index}", "").strip()
            if task_desc:
                task = {"description": task_desc}
                
                amount_str = form_data.get(f"task_amount_{task_index}", "").strip()
                if amount_str:
                    try:
                        task["amount"] = float(amount_str)
                    except ValueError:
                        pass
                
                task["is_mandatory"] = f"task_mandatory_{task_index}" in form_data
                task["requires_media"] = f"task_media_{task_index}" in form_data
                
                shift_tasks.append(task)
            task_index += 1
        
        if shift_tasks:
            update_data["shift_tasks"] = shift_tasks
        
        if not update_data:
            raise HTTPException(status_code=400, detail="Не указано ни одного параметра для изменения")
        
        # Получаем тайм-слоты
        from domain.entities.time_slot import TimeSlot
        timeslot_service = TimeSlotService(db)
        
        # Если переданы конкретные ID, используем их
        if timeslot_ids_str:
            timeslot_ids = [int(id_str.strip()) for id_str in timeslot_ids_str.split(",") if id_str.strip()]
            
            # Если указан период, фильтруем по нему
            if date_from_obj and date_to_obj:
                timeslots_query = select(TimeSlot).where(
                    and_(
                        TimeSlot.id.in_(timeslot_ids),
                        TimeSlot.object_id == object_id,
                        TimeSlot.slot_date >= date_from_obj,
                        TimeSlot.slot_date <= date_to_obj
                    )
                )
            else:
                # Без периода - только по ID
                timeslots_query = select(TimeSlot).where(
                    and_(
                        TimeSlot.id.in_(timeslot_ids),
                        TimeSlot.object_id == object_id
                    )
                )
        else:
            # Если не переданы ID, но указан период - ищем по периоду
            if date_from_obj and date_to_obj:
                timeslots_query = select(TimeSlot).where(
                    and_(
                        TimeSlot.object_id == object_id,
                        TimeSlot.slot_date >= date_from_obj,
                        TimeSlot.slot_date <= date_to_obj
                    )
                )
            else:
                raise HTTPException(status_code=400, detail="Укажите тайм-слоты для изменения или период")
        
        timeslots_result = await db.execute(timeslots_query)
        timeslots = timeslots_result.scalars().all()
        
        if not timeslots:
            raise HTTPException(status_code=404, detail="Тайм-слоты в указанном периоде не найдены")
        
        # Обновляем каждый тайм-слот
        updated_count = 0
        for timeslot in timeslots:
            slot = await timeslot_service.get_timeslot_by_id(timeslot.id, telegram_id)
            if not slot:
                continue

            slot_data = {
                "slot_date": timeslot.slot_date,
                "start_time": update_data.get("start_time", timeslot.start_time.strftime("%H:%M")),
                "end_time": update_data.get("end_time", timeslot.end_time.strftime("%H:%M")),
                "hourly_rate": update_data.get("hourly_rate", timeslot.hourly_rate),
                "max_employees": update_data.get("max_employees", timeslot.max_employees),
                "is_active": update_data.get("is_active", timeslot.is_active),
                "notes": timeslot.notes or ""
            }

            try:
                await timeslot_service.update_timeslot(timeslot.id, slot_data, telegram_id)
                updated_count += 1
            except Exception as update_error:
                logger.warning(
                    "Bulk update skipped for timeslot %s: %s", timeslot.id, update_error
                )
        
        logger.info(f"Bulk updated {updated_count} timeslots for user {telegram_id}")
        
        return RedirectResponse(url=f"/owner/timeslots/object/{object_id}?success=bulk_updated&count={updated_count}", status_code=status.HTTP_302_FOUND)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error bulk editing timeslots: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка массового редактирования: {str(e)}")


## Дублирующий маршрут создания тайм-слота удалён. Используйте маршруты из owner_timeslots.py


@router.get("/shifts_legacy", response_class=HTMLResponse, name="owner_shifts_legacy")
async def owner_shifts_list_legacy(
    request: Request,
    status: Optional[str] = Query(None, description="Фильтр по статусу: active, planned, completed, cancelled"),
    date_from: Optional[str] = Query(None, description="Дата начала (YYYY-MM-DD)"),
    date_to: Optional[str] = Query(None, description="Дата окончания (YYYY-MM-DD)"),
    object_id: Optional[str] = Query(None, description="ID объекта"),
    sort: Optional[str] = Query(None, description="Поле для сортировки"),
    order: Optional[str] = Query("asc", description="Порядок сортировки: asc, desc"),
    page: int = Query(1, ge=1, description="Номер страницы"),
    per_page: int = Query(20, ge=1, le=100, description="Количество на странице"),
    current_user: dict = Depends(require_owner_or_superadmin),
    db: AsyncSession = Depends(get_db_session)
):
    """Список смен владельца."""
    try:
        # Проверяем, что current_user - это словарь, а не RedirectResponse
        if current_user is None or isinstance(current_user, RedirectResponse):
            raise HTTPException(status_code=401, detail="Пользователь не авторизован")
            
        # Получаем telegram_id из current_user
        telegram_id = current_user.get("telegram_id") or current_user.get("id")
        user_role = current_user.get("role")
        
        # Получаем внутренний ID пользователя из БД
        user_query = select(User).where(User.telegram_id == telegram_id)
        user_result = await db.execute(user_query)
        user_obj = user_result.scalar_one_or_none()
        user_id = user_obj.id if user_obj else None
        
        # Базовый запрос для смен
        shifts_query = select(Shift).options(
            selectinload(Shift.object),
            selectinload(Shift.user)
        )
        
        # Базовый запрос для запланированных смен
        schedules_query = select(ShiftSchedule).options(
            selectinload(ShiftSchedule.object),
            selectinload(ShiftSchedule.user)
        )
        
        # Фильтрация по владельцу
        if user_role != "superadmin":
            # Получаем объекты владельца
            owner_objects = select(Object.id).where(Object.owner_id == user_id)
            shifts_query = shifts_query.where(Shift.object_id.in_(owner_objects))
            schedules_query = schedules_query.where(ShiftSchedule.object_id.in_(owner_objects))
        
        # Получение объектов для фильтра
        objects_query = select(Object)
        if user_role != "superadmin":
            objects_query = objects_query.where(Object.owner_id == user_id)
        objects_result = await db.execute(objects_query)
        objects = objects_result.scalars().all()
        
        # Применение фильтров
        if object_id:
            shifts_query = shifts_query.where(Shift.object_id == int(object_id))
            schedules_query = schedules_query.where(ShiftSchedule.object_id == int(object_id))
        
        # Фильтрация по статусу
        if status:
            if status == "active":
                shifts_query = shifts_query.where(Shift.status == "active")
                schedules_query = schedules_query.where(False)  # Исключаем запланированные
            elif status == "planned":
                shifts_query = shifts_query.where(False)  # Исключаем обычные
                schedules_query = schedules_query.where(ShiftSchedule.status == "planned")
            elif status == "completed":
                shifts_query = shifts_query.where(Shift.status == "completed")
                schedules_query = schedules_query.where(False)  # Исключаем запланированные
            elif status == "cancelled":
                shifts_query = shifts_query.where(Shift.status == "cancelled")
                schedules_query = schedules_query.where(ShiftSchedule.status == "cancelled")
        
        # Фильтрация по датам
        if date_from:
            try:
                from_date = datetime.strptime(date_from, '%Y-%m-%d').date()
                shifts_query = shifts_query.where(Shift.start_time >= from_date)
                schedules_query = schedules_query.where(ShiftSchedule.planned_start >= from_date)
            except ValueError:
                pass
        
        if date_to:
            try:
                to_date = datetime.strptime(date_to, '%Y-%m-%d').date()
                shifts_query = shifts_query.where(Shift.start_time <= to_date)
                schedules_query = schedules_query.where(ShiftSchedule.planned_start <= to_date)
            except ValueError:
                pass
        
        # Получение данных
        shifts_result = await db.execute(shifts_query.order_by(desc(Shift.created_at)))
        shifts = shifts_result.scalars().all()
        
        schedules_result = await db.execute(schedules_query.order_by(desc(ShiftSchedule.created_at)))
        schedules = schedules_result.scalars().all()
        
        # Объединение и форматирование данных
        all_shifts = []
        
        # Добавляем обычные смены
        for shift in shifts:
            all_shifts.append({
                'id': shift.id,
                'type': 'shift',
                'object_id': shift.object_id,
                'object_name': shift.object.name if shift.object else 'Неизвестный объект',
                'user_id': shift.user_id,
                'user_name': f"{shift.user.first_name} {shift.user.last_name or ''}".strip() if shift.user else 'Неизвестный пользователь',
                'start_time': web_timezone_helper.format_datetime_with_timezone(shift.start_time, shift.object.timezone if shift.object else 'Europe/Moscow', '%Y-%m-%d %H:%M') if shift.start_time else '-',
                'end_time': web_timezone_helper.format_datetime_with_timezone(shift.end_time, shift.object.timezone if shift.object else 'Europe/Moscow', '%Y-%m-%d %H:%M') if shift.end_time else '-',
                'status': shift.status,
                'total_hours': shift.total_hours,
                'total_payment': shift.total_payment,
                'is_planned': shift.is_planned,
                'created_at': shift.created_at
            })
        
        # Добавляем запланированные смены
        for schedule in schedules:
            all_shifts.append({
                'id': schedule.id,
                'type': 'schedule',
                'object_id': schedule.object_id,
                'object_name': schedule.object.name if schedule.object else 'Неизвестный объект',
                'user_id': schedule.user_id,
                'user_name': f"{schedule.user.first_name} {schedule.user.last_name or ''}".strip() if schedule.user else 'Неизвестный пользователь',
                'start_time': web_timezone_helper.format_datetime_with_timezone(schedule.planned_start, schedule.object.timezone if schedule.object else 'Europe/Moscow', '%Y-%m-%d %H:%M') if schedule.planned_start else '-',
                'end_time': web_timezone_helper.format_datetime_with_timezone(schedule.planned_end, schedule.object.timezone if schedule.object else 'Europe/Moscow', '%Y-%m-%d %H:%M') if schedule.planned_end else '-',
                'status': schedule.status,
                'total_hours': None,
                'total_payment': None,
                'is_planned': True,
                'created_at': schedule.created_at
            })
        
        # Сортировка данных
        if sort:
            reverse = order == 'desc'
            if sort == "user_name":
                all_shifts.sort(key=lambda x: x['user_name'].lower(), reverse=reverse)
            elif sort == "object_name":
                all_shifts.sort(key=lambda x: x['object_name'].lower(), reverse=reverse)
            elif sort == "start_time":
                all_shifts.sort(key=lambda x: x['start_time'], reverse=reverse)
            elif sort == "status":
                all_shifts.sort(key=lambda x: x['status'], reverse=reverse)
            elif sort == "created_at":
                all_shifts.sort(key=lambda x: x['created_at'], reverse=reverse)
            else:
                # По умолчанию сортируем по дате создания
                all_shifts.sort(key=lambda x: x['created_at'], reverse=True)
        else:
            # По умолчанию сортируем по дате создания
            all_shifts.sort(key=lambda x: x['created_at'], reverse=True)
        
        # Пагинация
        total_shifts = len(all_shifts)
        start_idx = (page - 1) * per_page
        end_idx = start_idx + per_page
        paginated_shifts = all_shifts[start_idx:end_idx]
        
        # Статистика
        stats = {
            'total': total_shifts,
            'active': len([s for s in all_shifts if s['status'] == 'active']),
            'planned': len([s for s in all_shifts if s['type'] == 'schedule' and s['status'] == 'planned']),
            'completed': len([s for s in all_shifts if s['status'] == 'completed'])
        }
        
        # Получаем данные для переключения интерфейсов
        available_interfaces = await get_available_interfaces_for_user(user_id)
        
        return templates.TemplateResponse("owner/shifts/list.html", {
            "request": request,
            "current_user": current_user,
            "available_interfaces": available_interfaces,
            "shifts": paginated_shifts,
            "objects": objects,
            "stats": stats,
            "filters": {
                "status": status,
                "date_from": date_from,
                "date_to": date_to,
                "object_id": object_id
            },
            "sort": {
                "field": sort,
                "order": order
            },
            "pagination": {
                "page": page,
                "per_page": per_page,
                "total": total_shifts,
                "pages": (total_shifts + per_page - 1) // per_page
            }
        })
        
    except Exception as e:
        logger.error(f"Error loading shifts: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки смен")


@router.get("/shifts_legacy/{shift_id}", response_class=HTMLResponse)
async def owner_shift_detail_legacy(
    request: Request, 
    shift_id: str,  # Изменено на str для поддержки префикса schedule_
    shift_type: Optional[str] = Query("shift"),
    current_user: dict = Depends(require_owner_or_superadmin),
    db: AsyncSession = Depends(get_db_session)
):
    """Детали смены владельца"""
    try:
        # Проверяем, что current_user - это словарь, а не RedirectResponse
        if current_user is None or isinstance(current_user, RedirectResponse):
            raise HTTPException(status_code=401, detail="Пользователь не авторизован")
        
        # Определяем тип смены по ID
        if shift_id.startswith('schedule_'):
            actual_shift_id = int(shift_id.replace('schedule_', ''))
            actual_shift_type = "schedule"
        else:
            actual_shift_id = int(shift_id)
            actual_shift_type = shift_type or "shift"
            
        # Получаем роль пользователя
        user_role = current_user.get("role")
        
        # Получаем внутренний ID пользователя
        telegram_id = current_user.get("telegram_id") or current_user.get("id")
        user_query = select(User).where(User.telegram_id == telegram_id)
        user_result = await db.execute(user_query)
        user_obj = user_result.scalar_one_or_none()
        user_id = user_obj.id if user_obj else None
        
        if actual_shift_type == "schedule":
            # Запланированная смена
            query = select(ShiftSchedule).options(
                selectinload(ShiftSchedule.object),
                selectinload(ShiftSchedule.user)
            ).where(ShiftSchedule.id == actual_shift_id)
        else:
            # Реальная смена
            query = select(Shift).options(
                selectinload(Shift.object),
                selectinload(Shift.user)
            ).where(Shift.id == actual_shift_id)
        
        result = await db.execute(query)
        shift = result.scalar_one_or_none()
        
        if not shift:
            return templates.TemplateResponse("owner/shifts/not_found.html", {
                "request": request,
                "current_user": current_user
            })
        
        # Проверка прав доступа
        if user_role != "superadmin":
            if shift.object.owner_id != user_id:
                return templates.TemplateResponse("owner/shifts/access_denied.html", {
                    "request": request,
                    "current_user": current_user
                })
        
        # Получаем объект для передачи в шаблон
        object = shift.object
        
        # Получаем задачи смены (только для реальных смен)
        shift_tasks = []
        if actual_shift_type != "schedule":
            from apps.web.services.shift_task_service import ShiftTaskService
            task_service = ShiftTaskService(db)
            shift_tasks = await task_service.get_shift_tasks(actual_shift_id)
        
        # История операций
        history_service = ShiftHistoryService(db)
        history_entries = []
        schedule_id_for_history = None
        shift_id_for_history = None

        if actual_shift_type == "schedule":
            schedule_id_for_history = actual_shift_id
        else:
            shift_id_for_history = actual_shift_id
            schedule_id_for_history = getattr(shift, "schedule_id", None)

        if schedule_id_for_history:
            history_entries.extend(
                await history_service.fetch_history(schedule_id=schedule_id_for_history)
            )
        if shift_id_for_history:
            history_entries.extend(
                await history_service.fetch_history(shift_id=shift_id_for_history)
            )

        actor_ids = {entry.actor_id for entry in history_entries if entry.actor_id}
        actor_names: Dict[int, str] = {}
        if actor_ids:
            users_result = await db.execute(
                select(User.id, User.first_name, User.last_name).where(User.id.in_(actor_ids))
            )
            for row in users_result.all():
                full_name = " ".join(filter(None, [row.last_name, row.first_name])).strip()
                actor_names[row.id] = full_name or f"ID {row.id}"

        reason_titles: Dict[str, str] = {}
        owner_id = getattr(object, "owner_id", None)
        if owner_id:
            reasons = await CancellationPolicyService.get_owner_reasons(
                db,
                owner_id,
                only_visible=False,
                only_active=True,
            )
            reason_titles = {reason.code: reason.title for reason in reasons}

        timezone = getattr(object, "timezone", None) or "Europe/Moscow"
        history_items = build_shift_history_items(
            history_entries,
            timezone=timezone,
            actor_names=actor_names,
            reason_titles=reason_titles,
        )

        # Получаем данные для переключения интерфейсов
        available_interfaces = await get_available_interfaces_for_user(user_id)
        
        return templates.TemplateResponse("owner/shifts/detail.html", {
            "request": request,
            "current_user": current_user,
            "shift": shift,
            "shift_type": shift_type,
            "shift_tasks": shift_tasks,
            "available_interfaces": available_interfaces,
            "object": object,
            "web_timezone_helper": web_timezone_helper,
            "history_items": history_items,
        })
        
    except Exception as e:
        logger.error(f"Error loading shift detail: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки деталей смены")
# Phase 4A: shift-tasks роуты удалены, используйте /owner/payroll/adjustments для управления корректировками
# См. apps/web/routes/owner_payroll_adjustments.py




@router.post("/shifts_legacy/{shift_id}/cancel")
async def owner_cancel_shift_legacy(
    request: Request, 
    shift_id: str,  # Изменено на str для поддержки префикса schedule_
    shift_type: Optional[str] = Query("shift"),
    current_user: dict = Depends(require_owner_or_superadmin),
    db: AsyncSession = Depends(get_db_session)
):
    """Отмена смены владельца"""
    from fastapi.responses import JSONResponse
    from datetime import datetime
    
    try:
        # Определяем тип смены по ID
        if shift_id.startswith('schedule_'):
            actual_shift_id = int(shift_id.replace('schedule_', ''))
            actual_shift_type = "schedule"
        else:
            actual_shift_id = int(shift_id)
            actual_shift_type = shift_type or "shift"
        
        # Получаем роль пользователя
        user_role = current_user.get("role")
        
        # Получаем внутренний ID пользователя
        telegram_id = current_user.get("telegram_id") or current_user.get("id")
        user_query = select(User).where(User.telegram_id == telegram_id)
        user_result = await db.execute(user_query)
        user_obj = user_result.scalar_one_or_none()
        user_id = user_obj.id if user_obj else None
        
        cancellation_service = ShiftCancellationService(db)
        sync_service = ShiftStatusSyncService(db)
        history_service = ShiftHistoryService(db)

        cancelled_by_type = "owner"
        actor_role = user_role if user_role != "superadmin" else "superadmin"

        if actual_shift_type == "schedule":
            query = select(ShiftSchedule).options(
                selectinload(ShiftSchedule.object),
                selectinload(ShiftSchedule.actual_shifts),
            ).where(ShiftSchedule.id == actual_shift_id)
            result = await db.execute(query)
            schedule = result.scalar_one_or_none()

            if not schedule or schedule.status not in {"planned", "confirmed"}:
                return JSONResponse(
                    status_code=400,
                    content={"success": False, "message": "Смена не найдена или уже отменена"},
                )

            if user_role != "superadmin" and schedule.object.owner_id != user_id:
                return JSONResponse(
                    status_code=403,
                    content={"success": False, "message": "Доступ запрещен"},
                )

            cancel_result = await cancellation_service.cancel_shift(
                shift_schedule_id=schedule.id,
                cancelled_by_user_id=user_id,
                cancelled_by_type=cancelled_by_type,
                cancellation_reason="owner_decision",
                actor_role=actor_role,
                source="web",
                extra_payload={"origin": "owner_cancel_legacy"},
            )
            if not cancel_result.get("success"):
                return JSONResponse(
                    status_code=400,
                    content={"success": False, "message": cancel_result.get("message", "Не удалось отменить смену")},
                )

            await cache.clear_pattern("calendar_shifts:*")
            await cache.clear_pattern("api_response:*")
            return JSONResponse(
                status_code=200,
                content={"success": True, "message": cancel_result.get("message", "Смена отменена")},
            )

        shift_query = select(Shift).options(
            selectinload(Shift.object)
        ).where(Shift.id == actual_shift_id)
        shift_result = await db.execute(shift_query)
        shift = shift_result.scalar_one_or_none()

        if not shift or shift.status not in {"active", "planned"}:
            return JSONResponse(
                status_code=400,
                content={"success": False, "message": "Смена не найдена или уже завершена"},
            )

        if user_role != "superadmin" and shift.object.owner_id != user_id:
            return JSONResponse(
                status_code=403,
                content={"success": False, "message": "Доступ запрещен"},
            )

        if shift.schedule_id:
            cancel_result = await cancellation_service.cancel_shift(
                shift_schedule_id=shift.schedule_id,
                cancelled_by_user_id=user_id,
                cancelled_by_type=cancelled_by_type,
                cancellation_reason="owner_decision",
                actor_role=actor_role,
                source="web",
                extra_payload={"origin": "owner_cancel_legacy"},
            )
            if not cancel_result.get("success"):
                return JSONResponse(
                    status_code=400,
                    content={"success": False, "message": cancel_result.get("message", "Не удалось отменить смену")},
                )
        else:
            previous_status = shift.status
            shift.status = "cancelled"
            if not shift.end_time:
                shift.end_time = datetime.utcnow()
            shift.updated_at = datetime.utcnow()

            await history_service.log_event(
                operation="shift_cancel",
                source="web",
                actor_id=user_id,
                actor_role=actor_role,
                shift_id=shift.id,
                schedule_id=None,
                old_status=previous_status,
                new_status="cancelled",
                payload={
                    "reason_code": "owner_decision",
                    "origin": "owner_cancel_legacy",
                },
            )
            await sync_service.cancel_linked_shifts(
                None,
                actor_id=user_id,
                actor_role=actor_role,
                source="web",
                payload={"origin": "owner_cancel_legacy"},
            )
            await db.commit()

        await cache.clear_pattern("calendar_shifts:*")
        await cache.clear_pattern("api_response:*")
        return JSONResponse(
            status_code=200,
            content={"success": True, "message": "Смена отменена"},
        )
                
    except Exception as e:
        logger.error(f"Error canceling shift: {e}")
        return JSONResponse(
            status_code=500,
            content={"success": False, "message": "Ошибка отмены смены"}
        )


@router.get("/templates", response_class=HTMLResponse, name="owner_templates")
async def owner_templates(request: Request):
    """Шаблоны договоров владельца"""
    current_user = await get_current_user(request)
    user_role = current_user.get("role", "employee")
    if user_role != "owner":
        return RedirectResponse(url="/dashboard", status_code=status.HTTP_302_FOUND)
    
    # Перенаправляем на существующую страницу шаблонов
    return RedirectResponse(url="/templates", status_code=status.HTTP_302_FOUND)


@router.get("/reports", response_class=HTMLResponse, name="owner_reports")
async def owner_reports(
    request: Request,
    current_user: dict = Depends(require_owner_or_superadmin),
    db: AsyncSession = Depends(get_db_session)
):
    """Отчеты владельца"""
    try:
        # Получаем внутренний user_id (как в оригинале)
        user_id = await get_user_id_from_current_user(current_user, db)
        
        # Получаем объекты владельца (как в оригинале)
        objects_query = select(Object).where(Object.owner_id == user_id)
        objects_result = await db.execute(objects_query)
        objects = objects_result.scalars().all()
        
        # Получаем всех пользователей, которые работали на объектах владельца (как в оригинале)
        employees_query = select(User.id, User.telegram_id, User.username, User.first_name, User.last_name, User.phone, User.role, User.is_active, User.created_at, User.updated_at).distinct().join(Shift, User.id == Shift.user_id).where(
            Shift.object_id.in_([obj.id for obj in objects])
        )
        employees_result = await db.execute(employees_query)
        employees = employees_result.all()
        
        # Если нет сотрудников из смен, показываем всех пользователей кроме текущего владельца (как в оригинале)
        if not employees:
            all_employees_query = select(User.id, User.telegram_id, User.username, User.first_name, User.last_name, User.phone, User.role, User.is_active, User.created_at, User.updated_at).where(User.id != user_id)
            all_employees_result = await db.execute(all_employees_query)
            employees = all_employees_result.all()
        
        # Статистика за последний месяц (как в оригинале)
        month_ago = datetime.now() - timedelta(days=30)
        
        shifts_query = select(Shift).options(
            selectinload(Shift.object),
            selectinload(Shift.user)
        ).where(
            and_(
                Shift.object_id.in_([obj.id for obj in objects]),
                Shift.start_time >= month_ago
            )
        )
        shifts_result = await db.execute(shifts_query)
        recent_shifts = shifts_result.scalars().all()
        
        stats = {
            "total_shifts": len(recent_shifts),
            "total_hours": sum(s.total_hours or 0 for s in recent_shifts if s.total_hours),
            "total_payment": sum(s.total_payment or 0 for s in recent_shifts if s.total_payment),
            "active_objects": len(objects),
            "employees": len(employees)
        }
        
        # Получаем данные для переключения интерфейсов
        available_interfaces = await get_available_interfaces_for_user(user_id)
        
        return templates.TemplateResponse("owner/reports/index.html", {
            "request": request,
            "current_user": current_user,
            "objects": objects,
            "employees": employees,
            "available_interfaces": available_interfaces,
            "stats": stats
        })
        
    except Exception as e:
        logger.error(f"Error loading reports: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки отчетов")


@router.post("/reports/generate")
async def owner_generate_report(
    request: Request,
    report_type: str = Form(...),
    date_from: str = Form(...),
    date_to: str = Form(...),
    object_id: Optional[int] = Form(None),
    employee_id: Optional[int] = Form(None),
    format: str = Form("excel"),
    current_user: dict = Depends(require_owner_or_superadmin),
    db: AsyncSession = Depends(get_db_session)
):
    """Генерация отчета (как в оригинале)"""
    # Парсинг дат (как в оригинале)
    try:
        start_date = datetime.strptime(date_from, "%Y-%m-%d").date()
        end_date = datetime.strptime(date_to, "%Y-%m-%d").date()
    except ValueError:
        return {"error": "Неверный формат даты"}
    
    # Получаем внутренний user_id (как в оригинале)
    user_id = await get_user_id_from_current_user(current_user, db)
    
    # Получаем объекты владельца (как в оригинале)
    owner_objects = select(Object.id).where(Object.owner_id == user_id)
    objects_result = await db.execute(owner_objects)
    owner_object_ids = [obj[0] for obj in objects_result.all()]
    
    # Базовый запрос для смен (как в оригинале)
    shifts_query = select(Shift).options(
        selectinload(Shift.object),
        selectinload(Shift.user)
    ).where(
        and_(
            Shift.object_id.in_(owner_object_ids),
            Shift.start_time >= start_date,
            Shift.start_time <= end_date + timedelta(days=1)
        )
    )
    
    # Применение фильтров (как в оригинале)
    if object_id and object_id in owner_object_ids:
        shifts_query = shifts_query.where(Shift.object_id == object_id)
    
    if employee_id:
        shifts_query = shifts_query.where(Shift.user_id == employee_id)
    
    # Выполнение запроса (как в оригинале)
    shifts_result = await db.execute(shifts_query.order_by(desc(Shift.start_time)))
    shifts = shifts_result.scalars().all()
    
    # Генерация отчета в зависимости от типа (как в оригинале)
    if report_type == "shifts":
        return await _generate_shifts_report(shifts, format, start_date, end_date)
    elif report_type == "employees":
        return await _generate_employees_report(shifts, format, start_date, end_date)
    elif report_type == "objects":
        return await _generate_objects_report(shifts, format, start_date, end_date)
    else:
        return {"error": "Неизвестный тип отчета"}


@router.get("/reports/stats/period")
async def owner_reports_stats_period(
    request: Request,
    date_from: str = Query(...),
    date_to: str = Query(...),
    object_id: Optional[int] = Query(None),
    current_user: dict = Depends(require_owner_or_superadmin),
    db: AsyncSession = Depends(get_db_session)
):
    """Статистика за период (как в оригинале)"""
    try:
        start_date = datetime.strptime(date_from, "%Y-%m-%d").date()
        end_date = datetime.strptime(date_to, "%Y-%m-%d").date()
    except ValueError:
        return {"error": "Неверный формат даты"}
    
    # Получаем внутренний user_id (как в оригинале)
    user_id = await get_user_id_from_current_user(current_user, db)
    
    # Получаем объекты владельца (как в оригинале)
    owner_objects = select(Object.id).where(Object.owner_id == user_id)
    objects_result = await db.execute(owner_objects)
    owner_object_ids = [obj[0] for obj in objects_result.all()]
    
    # Запрос смен за период (как в оригинале)
    shifts_query = select(Shift).options(
        selectinload(Shift.object),
        selectinload(Shift.user)
    ).where(
        and_(
            Shift.object_id.in_(owner_object_ids),
            Shift.start_time >= start_date,
            Shift.start_time <= end_date + timedelta(days=1)
        )
    )
    
    if object_id and object_id in owner_object_ids:
        shifts_query = shifts_query.where(Shift.object_id == object_id)
    
    shifts_result = await db.execute(shifts_query)
    shifts = shifts_result.scalars().all()
    
    # Расчет статистики (как в оригинале)
    stats = {
        "period": {
            "from": start_date.strftime("%d.%m.%Y"),
            "to": end_date.strftime("%d.%m.%Y")
        },
        "total_shifts": len(shifts),
        "total_hours": sum(s.total_hours or 0 for s in shifts if s.total_hours),
        "total_payment": sum(s.total_payment or 0 for s in shifts if s.total_payment),
        "avg_hours_per_shift": 0,
        "avg_payment_per_shift": 0,
        "by_status": {},
        "by_object": {},
        "by_employee": {}
    }
    
    if stats["total_shifts"] > 0:
        stats["avg_hours_per_shift"] = stats["total_hours"] / stats["total_shifts"]
        stats["avg_payment_per_shift"] = stats["total_payment"] / stats["total_shifts"]
    
    # Статистика по статусам (как в оригинале)
    status_counts = {}
    for shift in shifts:
        status = shift.status
        status_counts[status] = status_counts.get(status, 0) + 1
    stats["by_status"] = status_counts
    
    # Статистика по объектам (как в оригинале)
    object_stats = {}
    for shift in shifts:
        obj_name = shift.object.name
        if obj_name not in object_stats:
            object_stats[obj_name] = {
                "shifts": 0,
                "hours": 0,
                "payment": 0
            }
        object_stats[obj_name]["shifts"] += 1
        object_stats[obj_name]["hours"] += shift.total_hours or 0
        object_stats[obj_name]["payment"] += shift.total_payment or 0
    stats["by_object"] = object_stats
    
    # Статистика по сотрудникам (как в оригинале)
    employee_stats = {}
    for shift in shifts:
        emp_name = f"{shift.user.first_name} {shift.user.last_name or ''}".strip()
        if emp_name not in employee_stats:
            employee_stats[emp_name] = {
                "shifts": 0,
                "hours": 0,
                "payment": 0
            }
        employee_stats[emp_name]["shifts"] += 1
        employee_stats[emp_name]["hours"] += shift.total_hours or 0
        employee_stats[emp_name]["payment"] += shift.total_payment or 0
    stats["by_employee"] = employee_stats
    
    return stats


async def _generate_shifts_report(shifts: List[Shift], format: str, start_date: date, end_date: date):
    """Генерация отчета по сменам (как в оригинале)"""
    data = []
    
    for shift in shifts:
        data.append({
            "ID": shift.id,
            "Сотрудник": f"{shift.user.first_name} {shift.user.last_name or ''}".strip(),
            "Объект": shift.object.name,
            "Дата начала": web_timezone_helper.format_datetime_with_timezone(shift.start_time, shift.object.timezone if shift.object else 'Europe/Moscow'),
            "Дата окончания": web_timezone_helper.format_datetime_with_timezone(shift.end_time, shift.object.timezone if shift.object else 'Europe/Moscow') if shift.end_time else "Не завершена",
            "Статус": shift.status,
            "Часов": shift.total_hours or 0,
            "Ставка": shift.hourly_rate or 0,
            "Сумма": shift.total_payment or 0,
            "Заметки": shift.notes or ""
        })
    
    if format == "excel":
        return await _create_excel_file(data, f"shifts_report_{start_date}_{end_date}")
    else:
        return {"data": data, "total": len(data)}


async def _generate_employees_report(shifts: List[Shift], format: str, start_date: date, end_date: date):
    """Генерация отчета по сотрудникам (как в оригинале)"""
    # Группировка по сотрудникам (как в оригинале)
    employees_data = {}
    
    for shift in shifts:
        employee_id = shift.user_id
        if employee_id not in employees_data:
            employees_data[employee_id] = {
                "employee": shift.user,
                "shifts": [],
                "total_hours": 0,
                "total_payment": 0
            }
        
        employees_data[employee_id]["shifts"].append(shift)
        employees_data[employee_id]["total_hours"] += shift.total_hours or 0
        employees_data[employee_id]["total_payment"] += shift.total_payment or 0
    
    data = []
    for emp_data in employees_data.values():
        data.append({
            "Сотрудник": f"{emp_data['employee'].first_name} {emp_data['employee'].last_name or ''}".strip(),
            "Количество смен": len(emp_data["shifts"]),
            "Общее время": emp_data["total_hours"],
            "Общая сумма": emp_data["total_payment"],
            "Средняя ставка": emp_data["total_payment"] / emp_data["total_hours"] if emp_data["total_hours"] > 0 else 0
        })
    
    if format == "excel":
        return await _create_excel_file(data, f"employees_report_{start_date}_{end_date}")
    else:
        return {"data": data, "total": len(data)}


async def _generate_objects_report(shifts: List[Shift], format: str, start_date: date, end_date: date):
    """Генерация отчета по объектам (как в оригинале)"""
    # Группировка по объектам (как в оригинале)
    objects_data = {}
    
    for shift in shifts:
        object_id = shift.object_id
        if object_id not in objects_data:
            objects_data[object_id] = {
                "object": shift.object,
                "shifts": [],
                "total_hours": 0,
                "total_payment": 0,
                "employees": set()
            }
        
        objects_data[object_id]["shifts"].append(shift)
        objects_data[object_id]["total_hours"] += shift.total_hours or 0
        objects_data[object_id]["total_payment"] += shift.total_payment or 0
        objects_data[object_id]["employees"].add(shift.user_id)
    
    data = []
    for obj_data in objects_data.values():
        data.append({
            "Объект": obj_data["object"].name,
            "Адрес": obj_data["object"].address or "",
            "Количество смен": len(obj_data["shifts"]),
            "Количество сотрудников": len(obj_data["employees"]),
            "Общее время": obj_data["total_hours"],
            "Общая сумма": obj_data["total_payment"],
            "Средняя ставка": obj_data["total_payment"] / obj_data["total_hours"] if obj_data["total_hours"] > 0 else 0
        })
    
    if format == "excel":
        return await _create_excel_file(data, f"objects_report_{start_date}_{end_date}")
    else:
        return {"data": data, "total": len(data)}


async def _create_excel_file(data: List[dict], filename: str):
    """Создание Excel файла (как в оригинале)"""
    if not data:
        return {"error": "Нет данных для отчета"}
    
    # Создание DataFrame (как в оригинале)
    
    df = pd.DataFrame(data)
    
    # Создание Excel файла (как в оригинале)
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет"
    
    # Добавление данных (как в оригинале)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # Стилизация (как в оригинале)
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Автоширина колонок (как в оригинале)
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохранение в память (как в оригинале)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"}
    )


@router.get("/profile", response_class=HTMLResponse, name="owner_profile")
async def owner_profile(
    request: Request,
    current_user: dict = Depends(require_owner_or_superadmin),
    db: AsyncSession = Depends(get_db_session)
):
    """Страница профиля владельца (как в оригинале)"""
    try:
        # Проверка типа current_user
        if current_user is None or isinstance(current_user, RedirectResponse):
            raise HTTPException(status_code=401, detail="Пользователь не авторизован")
        
        # Получаем внутренний user_id (как в оригинале)
        user_id = await get_user_id_from_current_user(current_user, db)
        
        # Получаем существующий профиль (как в оригинале)
        from apps.web.services.tag_service import TagService
        tag_service = TagService()
        profile = await tag_service.get_owner_profile(db, user_id)
        
        # Получаем все доступные теги (как в оригинале)
        all_tags = await tag_service.get_all_tags(db)
        
        # Группируем теги по категориям (для HTML) (как в оригинале)
        tags_by_category = {}
        # Группируем теги по категориям (для JSON) (как в оригинале)
        tags_by_category_json = {}
        
        for tag in all_tags:
            if tag.category not in tags_by_category:
                tags_by_category[tag.category] = []
                tags_by_category_json[tag.category] = []
            tags_by_category[tag.category].append(tag)
            tags_by_category_json[tag.category].append(tag.to_dict())
        
        # Получаем данные для переключения интерфейсов
        available_interfaces = await get_available_interfaces_for_user(user_id)
        
        # Получаем enabled_features для меню
        from shared.services.system_features_service import SystemFeaturesService
        features_service = SystemFeaturesService()
        enabled_features = await features_service.get_enabled_features(db, user_id)
        
        # Получаем информацию о текущем тарифе
        from apps.web.services.limits_service import LimitsService
        limits_service = LimitsService(db)
        limits_summary = await limits_service.get_user_limits_summary(user_id)
        
        # Извлекаем данные о тарифе для отображения
        subscription_info = None
        if limits_summary.get('has_subscription'):
            # Получаем информацию о подписке из БД для получения notes
            from domain.entities.user_subscription import UserSubscription, SubscriptionStatus
            subscription_result = await db.execute(
                select(UserSubscription).where(
                    UserSubscription.user_id == user_id,
                    UserSubscription.status == SubscriptionStatus.ACTIVE
                ).limit(1)
            )
            subscription = subscription_result.scalar_one_or_none()
            
            subscription_info = {
                'tariff_name': limits_summary['subscription']['tariff_name'],
                'expires_at': limits_summary['subscription'].get('expires_at'),
                'status': limits_summary['subscription'].get('status', 'active'),
                'notes': subscription.notes if subscription else None
            }
        
        return templates.TemplateResponse("owner/profile/index.html", {
            "request": request,
            "current_user": current_user,
            "available_interfaces": available_interfaces,
            "enabled_features": enabled_features,
            "profile": profile,
            "subscription_info": subscription_info,
            "tags_by_category": tags_by_category,
            "tags_by_category_json": tags_by_category_json,
            "legal_types": [
                {"value": "individual", "label": "Физическое лицо (ИП)"},
                {"value": "legal", "label": "Юридическое лицо (ООО)"}
            ]
        })
        
    except Exception as e:
        logger.error(f"Error loading profile: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки профиля")


@router.post("/profile/save")
async def owner_profile_save(
    request: Request,
    current_user: dict = Depends(require_owner_or_superadmin),
    db: AsyncSession = Depends(get_db_session)
):
    """Сохранение профиля владельца (как в оригинале)"""
    try:
        # Получаем данные формы (как в оригинале)
        form_data = await request.form()
        
        # Получаем внутренний user_id (как в оригинале)
        user_id = await get_user_id_from_current_user(current_user, db)
        
        # Извлекаем основные данные профиля (как в оригинале)
        profile_name = form_data.get("profile_name", "Мой профиль")
        legal_type = form_data.get("legal_type", "individual")
        is_public = form_data.get("is_public") == "on"
        
        # Извлекаем выбранные теги (как в оригинале)
        selected_tags = []
        for key, value in form_data.items():
            if key.startswith("tag_") and value == "on":
                tag_key = key[4:]  # убираем префикс "tag_"
                selected_tags.append(tag_key)
        
        # Извлекаем значения тегов (как в оригинале)
        profile_data = {}
        for key, value in form_data.items():
            if key.startswith("value_") and value.strip():
                tag_key = key[6:]  # убираем префикс "value_"
                if tag_key in selected_tags:  # только для выбранных тегов
                    profile_data[tag_key] = value.strip()
        
        # Извлекаем новые поля профиля
        about_company = form_data.get("about_company", "")
        values = form_data.get("values", "")
        contact_phone = form_data.get("contact_phone", "")
        
        # Парсим JSON поля
        import json
        photos = json.loads(form_data.get("photos", "[]"))
        contact_messengers = json.loads(form_data.get("contact_messengers", "[]"))
        
        # Сохраняем профиль (как в оригинале)
        from apps.web.services.tag_service import TagService
        tag_service = TagService()
        profile = await tag_service.create_or_update_owner_profile(
            db,
            user_id,
            {
                "profile_name": profile_name,
                "profile_data": profile_data,
                "active_tags": selected_tags,
                "is_public": is_public,
                "about_company": about_company,
                "values": values,
                "photos": photos,
                "contact_phone": contact_phone,
                "contact_messengers": contact_messengers
            },
            legal_type
        )
        
        logger.info(f"Profile saved for user {user_id}: {len(selected_tags)} tags, {len(profile_data)} values")
        
        return RedirectResponse(url="/owner/profile?success=1", status_code=303)
        
    except Exception as e:
        logger.error(f"Error saving profile: {e}")
        raise HTTPException(status_code=500, detail="Ошибка сохранения профиля")


@router.post("/profile/api/autosave")
async def owner_profile_autosave(
    request: Request,
    current_user: dict = Depends(require_owner_or_superadmin),
    db: AsyncSession = Depends(get_db_session)
):
    """Автосохранение отдельных полей профиля (JSON)."""
    try:
        data = await request.json()
        user_id = await get_user_id_from_current_user(current_user, db)
        from apps.web.services.tag_service import TagService
        tag_service = TagService()
        profile = await tag_service.update_owner_profile_fields(db, user_id, data)
        return {"success": True, "profile_id": profile.id}
    except Exception as e:
        logger.error(f"Error autosaving profile: {e}")
        return {"success": False, "error": str(e)}


@router.post("/profile/api/photo")
async def owner_upload_photo(
    request: Request,
    current_user: dict = Depends(require_owner_or_superadmin),
    db: AsyncSession = Depends(get_db_session)
):
    """Загрузка фото для профиля владельца."""
    try:
        user_id = await get_user_id_from_current_user(current_user, db)
        if not user_id:
            raise HTTPException(status_code=401, detail="Пользователь не найден")
        
        # Получаем файл из формы
        form = await request.form()
        photo_file = form.get("photo")
        
        if not photo_file:
            raise HTTPException(status_code=400, detail="Файл не предоставлен")
        
        # Проверяем тип файла
        if not hasattr(photo_file, 'content_type') or not photo_file.content_type.startswith('image/'):
            raise HTTPException(status_code=400, detail="Файл должен быть изображением")
        
        # Проверяем размер (максимум 5MB)
        file_content = await photo_file.read()
        if len(file_content) > 5 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="Размер файла не должен превышать 5MB")
        
        # Загружаем файл в хранилище
        from shared.services.media_storage import get_media_storage_client
        import uuid
        storage_client = get_media_storage_client()
        
        # Определяем имя файла
        file_name = photo_file.filename or f"photo_{user_id}_{uuid.uuid4().hex[:8]}.jpg"
        folder = f"profiles/{user_id}/photos"
        
        # Убеждаемся, что folder не содержит двойных слешей
        folder = folder.replace('//', '/')
        
        # Загружаем
        media_file = await storage_client.upload(
            file_content=file_content,
            file_name=file_name,
            content_type=photo_file.content_type,
            folder=folder,
            metadata={"user_id": user_id, "type": "profile_photo"}
        )
        
        # Получаем URL
        photo_url = await storage_client.get_url(media_file.key, expires_in=31536000)  # 1 год
        
        logger.info(f"Profile photo uploaded: user_id={user_id}, photo_url={photo_url}, storage_key={media_file.key}")
        
        return JSONResponse(content={
            "success": True,
            "photo_url": photo_url,
            "message": "Фото успешно загружено"
        })
        
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Error uploading photo: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка загрузки фото: {str(e)}")


@router.get("/profile/tags/{category}")
async def owner_profile_tags_category(
    request: Request,
    category: str,
    current_user: dict = Depends(require_owner_or_superadmin),
    db: AsyncSession = Depends(get_db_session)
):
    """API для получения тегов по категории (как в оригинале)"""
    try:
        from apps.web.services.tag_service import TagService
        tag_service = TagService()
        tags = await tag_service.get_tags_by_category(db, category)
        
        return {
            "tags": [tag.to_dict() for tag in tags]
        }
        
    except Exception as e:
        logger.error(f"Error loading tags for category {category}: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки тегов")


@router.get("/profile/preview")
async def owner_profile_preview(
    request: Request,
    current_user: dict = Depends(require_owner_or_superadmin),
    db: AsyncSession = Depends(get_db_session)
):
    """Предпросмотр данных профиля для использования в договорах (как в оригинале)"""
    try:
        # Получаем внутренний user_id (как в оригинале)
        user_id = await get_user_id_from_current_user(current_user, db)
        
        # Получаем профиль (как в оригинале)
        from apps.web.services.tag_service import TagService
        tag_service = TagService()
        profile = await tag_service.get_owner_profile(db, user_id)
        
        if not profile:
            return {"error": "Профиль не найден"}
        
        # Получаем все теги для подстановки в договоры (как в оригинале)
        tags_for_templates = profile.get_tags_for_templates()
        
        return {
            "profile": profile.to_dict(),
            "tags_for_templates": tags_for_templates,
            "completion": profile.get_completion_percentage()
        }
        
    except Exception as e:
        logger.error(f"Error loading profile preview: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки предпросмотра профиля")


@router.get("/settings", response_class=HTMLResponse, name="owner_settings")
async def owner_settings(request: Request):
    """Настройки владельца"""
    current_user = await get_current_user(request)
    user_role = current_user.get("role", "employee")
    if user_role != "owner":
        return RedirectResponse(url="/dashboard", status_code=status.HTTP_302_FOUND)
    
    # Получаем данные для переключения интерфейсов
    async with get_async_session() as session:
        user_id = await get_user_id_from_current_user(current_user, session)
        available_interfaces = await get_available_interfaces_for_user(user_id)
    
    return templates.TemplateResponse("owner/settings.html", {
        "request": request,
        "current_user": current_user,
        "title": "Настройки владельца",
        "message": "Настройки в разработке",
        "available_interfaces": available_interfaces
    })


# ===== ШАБЛОНЫ ПЛАНИРОВАНИЯ =====

@router.get("/templates/planning", response_class=HTMLResponse, name="owner_planning_templates_list")
async def owner_planning_templates_list(request: Request):
    """Список шаблонов планирования."""
    current_user = await get_current_user(request)
    if current_user.get("role") != "owner":
        return RedirectResponse(url="/auth/login", status_code=status.HTTP_302_FOUND)
    
    try:
        async with get_async_session() as session:
            from apps.web.services.template_service import TemplateService
            template_service = TemplateService(session)
            templates_list = await template_service.get_templates_by_owner(current_user["id"])
        
        # Получаем данные для переключения интерфейсов
        async with get_async_session() as session:
            user_id = await get_user_id_from_current_user(current_user, session)
            available_interfaces = await get_available_interfaces_for_user(user_id)
        
        return templates.TemplateResponse(
            "owner/templates/planning/list.html",
            {
                "request": request,
                "templates": templates_list,
                "template_type": "planning",
                "title": "Шаблоны планирования",
                "current_user": current_user,
                "available_interfaces": available_interfaces
            }
        )
    except Exception as e:
        logger.error(f"Error loading planning templates: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки шаблонов планирования")


@router.get("/templates/planning/create", response_class=HTMLResponse, name="owner_planning_template_create")
async def owner_planning_template_create(request: Request):
    """Форма создания шаблона планирования."""
    current_user = await get_current_user(request)
    if current_user.get("role") != "owner":
        return RedirectResponse(url="/auth/login", status_code=status.HTTP_302_FOUND)
    
    # Получаем данные для переключения интерфейсов
    async with get_async_session() as session:
        user_id = await get_user_id_from_current_user(current_user, session)
        available_interfaces = await get_available_interfaces_for_user(user_id)
    
    return templates.TemplateResponse(
        "owner/templates/planning/create.html",
        {
            "request": request,
            "template_type": "planning",
            "title": "Создание шаблона планирования",
            "current_user": current_user,
            "available_interfaces": available_interfaces
        }
    )


@router.post("/templates/planning/create", name="owner_planning_template_create_post")
async def owner_planning_template_create_post(
    request: Request,
    name: str = Form(...),
    description: str = Form(""),
    start_time: str = Form(""),
    end_time: str = Form(""),
    hourly_rate: int = Form(0),
    repeat_type: str = Form("none"),
    repeat_days: str = Form(""),
    is_public: bool = Form(False)
):
    """Создание шаблона планирования."""
    current_user = await get_current_user(request)
    if current_user.get("role") != "owner":
        return RedirectResponse(url="/auth/login", status_code=status.HTTP_302_FOUND)
    
    try:
        async with get_async_session() as session:
            from apps.web.services.template_service import TemplateService
            template_service = TemplateService(session)
            
            template_data = {
                "name": name,
                "description": description,
                "start_time": start_time,
                "end_time": end_time,
                "hourly_rate": hourly_rate,
                "repeat_type": repeat_type,
                "repeat_days": repeat_days,
                "is_public": is_public
            }
            
            template = await template_service.create_template(template_data, current_user["id"])
            
            if template:
                return RedirectResponse(url="/owner/templates/planning", status_code=303)
            else:
                raise HTTPException(status_code=400, detail="Ошибка создания шаблона планирования")
                
    except Exception as e:
        logger.error(f"Error creating planning template: {e}")
        raise HTTPException(status_code=400, detail=f"Ошибка создания шаблона планирования: {str(e)}")


@router.get("/templates/planning/{template_id}", response_class=HTMLResponse, name="owner_planning_template_detail")
async def owner_planning_template_detail(request: Request, template_id: int):
    """Детали шаблона планирования."""
    current_user = await get_current_user(request)
    if current_user.get("role") != "owner":
        return RedirectResponse(url="/auth/login", status_code=status.HTTP_302_FOUND)
    
    try:
        async with get_async_session() as session:
            from apps.web.services.template_service import TemplateService
            template_service = TemplateService(session)
            template = await template_service.get_template_by_id(template_id, current_user["id"])
        
        if not template:
            raise HTTPException(status_code=404, detail="Шаблон планирования не найден")
        
        # Получаем данные для переключения интерфейсов
        async with get_async_session() as session:
            user_id = await get_user_id_from_current_user(current_user, session)
            available_interfaces = await get_available_interfaces_for_user(user_id)
        
        return templates.TemplateResponse(
            "owner/templates/planning/detail.html",
            {
                "request": request,
                "template": template,
                "template_type": "planning",
                "title": "Детали шаблона планирования",
                "current_user": current_user,
                "available_interfaces": available_interfaces
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error loading planning template detail: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки шаблона планирования")


@router.get("/templates/planning/{template_id}/edit", response_class=HTMLResponse, name="owner_planning_template_edit")
async def owner_planning_template_edit(request: Request, template_id: int):
    """Форма редактирования шаблона планирования."""
    current_user = await get_current_user(request)
    if current_user.get("role") != "owner":
        return RedirectResponse(url="/auth/login", status_code=status.HTTP_302_FOUND)
    
    try:
        async with get_async_session() as session:
            from apps.web.services.template_service import TemplateService
            template_service = TemplateService(session)
            template = await template_service.get_template_by_id(template_id, current_user["id"])
        
        if not template:
            raise HTTPException(status_code=404, detail="Шаблон планирования не найден")
        
        # Получаем данные для переключения интерфейсов
        async with get_async_session() as session:
            user_id = await get_user_id_from_current_user(current_user, session)
            available_interfaces = await get_available_interfaces_for_user(user_id)
        
        return templates.TemplateResponse(
            "owner/templates/planning/edit.html",
            {
                "request": request,
                "template": template,
                "template_type": "planning",
                "title": "Редактирование шаблона планирования",
                "current_user": current_user,
                "available_interfaces": available_interfaces
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error loading planning template edit: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки формы редактирования")


@router.post("/templates/planning/{template_id}/edit", name="owner_planning_template_edit_post")
async def owner_planning_template_edit_post(
    request: Request,
    template_id: int,
    name: str = Form(...),
    description: str = Form(""),
    start_time: str = Form(""),
    end_time: str = Form(""),
    hourly_rate: int = Form(0),
    repeat_type: str = Form("none"),
    repeat_days: str = Form(""),
    is_public: bool = Form(False)
):
    """Обновление шаблона планирования."""
    current_user = await get_current_user(request)
    if current_user.get("role") != "owner":
        return RedirectResponse(url="/auth/login", status_code=status.HTTP_302_FOUND)
    
    try:
        async with get_async_session() as session:
            from apps.web.services.template_service import TemplateService
            template_service = TemplateService(session)
            
            template_data = {
                "name": name,
                "description": description,
                "start_time": start_time,
                "end_time": end_time,
                "hourly_rate": hourly_rate,
                "repeat_type": repeat_type,
                "repeat_days": repeat_days,
                "is_public": is_public
            }
            
            success = await template_service.update_template(template_id, template_data, current_user["id"])
            
            if success:
                return RedirectResponse(url=f"/owner/templates/planning/{template_id}", status_code=303)
            else:
                raise HTTPException(status_code=400, detail="Ошибка обновления шаблона планирования")
                
    except Exception as e:
        logger.error(f"Error updating planning template: {e}")
        raise HTTPException(status_code=400, detail=f"Ошибка обновления шаблона планирования: {str(e)}")


@router.post("/templates/planning/{template_id}/delete", name="owner_planning_template_delete")
async def owner_planning_template_delete(request: Request, template_id: int):
    """Удаление шаблона планирования."""
    current_user = await get_current_user(request)
    if current_user.get("role") != "owner":
        return RedirectResponse(url="/auth/login", status_code=status.HTTP_302_FOUND)
    
    try:
        async with get_async_session() as session:
            from apps.web.services.template_service import TemplateService
            template_service = TemplateService(session)
            success = await template_service.delete_template(template_id, current_user["id"])
        
        if success:
            return RedirectResponse(url="/owner/templates/planning", status_code=303)
        else:
            raise HTTPException(status_code=400, detail="Ошибка удаления шаблона планирования")
            
    except Exception as e:
        logger.error(f"Error deleting planning template: {e}")
        raise HTTPException(status_code=400, detail=f"Ошибка удаления шаблона планирования: {str(e)}")


# ===============================
# СОТРУДНИКИ
# ===============================

@router.get("/employees", response_class=HTMLResponse)
async def owner_employees_list(
    request: Request,
    view_mode: str = Query("list", description="Режим отображения: cards или list"),
    sort_by: str = Query("employee", description="Сортировка: employee | telegram_id | status"),
    sort_order: str = Query("asc", description="Порядок сортировки: asc | desc"),
    q_employee: str = Query("", description="Фильтр по Фамилия Имя"),
    q_telegram: str = Query("", description="Фильтр по Telegram ID"),
    q_status: str = Query("", description="Фильтр по статусу: active|former"),
    show_former: bool = Query(False, description="Показать бывших сотрудников"),
    current_user: dict = Depends(require_owner_or_superadmin),
    db: AsyncSession = Depends(get_db_session)
):
    """Список сотрудников владельца."""
    try:
        from apps.web.services.contract_service import ContractService
        
        # Проверяем, что current_user - это словарь, а не RedirectResponse
        if current_user is None or isinstance(current_user, RedirectResponse):
            raise HTTPException(status_code=401, detail="Пользователь не авторизован")
        
        # Получаем реальных сотрудников из базы данных
        contract_service = ContractService()
        # Используем telegram_id для поиска пользователя в БД
        user_id = current_user["id"]  # Это telegram_id из токена
        
        if show_former:
            employees = await contract_service.get_all_contract_employees_by_telegram_id(user_id)
        else:
            employees = await contract_service.get_contract_employees_by_telegram_id(user_id)

        # Фильтрация
        q_emp = (q_employee or "").strip().lower()
        q_tel = (q_telegram or "").strip()
        q_sts = (q_status or "").strip().lower()

        def is_active_employee(emp: dict) -> bool:
            contracts = emp.get("contracts") or []
            for c in contracts:
                if c.get("status") == "active" and c.get("is_active") is True:
                    return True
            return False

        def name_matches(emp: dict) -> bool:
            if not q_emp:
                return True
            last_name = (emp.get('last_name') or '').lower()
            first_name = (emp.get('first_name') or '').lower()
            full1 = f"{last_name} {first_name}".strip()
            full2 = f"{first_name} {last_name}".strip()
            return q_emp in full1 or q_emp in full2

        def telegram_matches(emp: dict) -> bool:
            if not q_tel:
                return True
            return q_tel in str(emp.get('telegram_id') or '')

        def status_matches(emp: dict) -> bool:
            if not q_sts:
                return True
            if q_sts == 'active':
                return is_active_employee(emp)
            if q_sts == 'former':
                return not is_active_employee(emp)
            return True

        employees = [e for e in employees if name_matches(e) and telegram_matches(e) and status_matches(e)]

        # Нормализация параметров сортировки
        sort_by_norm = (sort_by or "employee").strip().lower()
        allowed_sort_by = {"employee", "telegram_id", "status"}
        if sort_by_norm not in allowed_sort_by:
            sort_by_norm = "employee"

        sort_order_norm = (sort_order or "asc").strip().lower()
        if sort_order_norm not in {"asc", "desc"}:
            sort_order_norm = "asc"

        # Вспомогательные вычисления
        # Ключи сортировки
        def key_employee(emp: dict):
            return (
                (emp.get("last_name") or "").strip().lower(),
                (emp.get("first_name") or "").strip().lower(),
            )

        def key_telegram(emp: dict):
            try:
                return int(emp.get("telegram_id") or 0)
            except Exception:
                return 0

        def key_status(emp: dict):
            # Активные первыми при asc
            return 0 if is_active_employee(emp) else 1

        key_func = key_employee
        if sort_by_norm == "telegram_id":
            key_func = key_telegram
        elif sort_by_norm == "status":
            key_func = key_status

        try:
            employees = sorted(employees, key=key_func, reverse=(sort_order_norm == "desc"))
        except Exception:
            pass
        
        # Получаем данные для переключения интерфейсов
        available_interfaces = await get_available_interfaces_for_user(user_id)
        
        return templates.TemplateResponse(
            "owner/employees/list.html",
            {
                "request": request,
                "employees": employees,
                "title": "Управление сотрудниками",
                "current_user": current_user,
                "view_mode": view_mode,
                "sort_by": sort_by_norm,
                "sort_order": sort_order_norm,
                "q_employee": q_emp,
                "q_telegram": q_tel,
                "q_status": q_sts,
                "show_former": show_former,
                "available_interfaces": available_interfaces
            }
        )
    except Exception as e:
        logger.error(f"Error loading employees list: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка загрузки списка сотрудников: {str(e)}")


@router.get("/employees/create", response_class=HTMLResponse)
async def owner_employees_create_form(
    request: Request,
    employee_telegram_id: int = Query(None, description="Telegram ID сотрудника для предзаполнения"),
    current_user: dict = Depends(require_owner_or_superadmin),
    db: AsyncSession = Depends(get_db_session)
):
    """Форма создания договора с сотрудником."""
    try:
        from apps.web.services.contract_service import ContractService
        from apps.web.services.tag_service import TagService
        
        # Получаем доступных сотрудников и объекты
        contract_service = ContractService()
        # Используем telegram_id для поиска пользователя в БД
        user_id = current_user["id"]  # Это telegram_id из токена
        available_employees = await contract_service.get_available_employees(user_id)
        objects = await contract_service.get_owner_objects(user_id)
        
        # Получаем внутренний ID пользователя для шаблонов
        user_query = select(User).where(User.telegram_id == user_id)
        user_result = await db.execute(user_query)
        user_obj = user_result.scalar_one_or_none()
        internal_user_id = user_obj.id if user_obj else None
        
        # Получаем профиль владельца для тегов
        tag_service = TagService()
        owner_profile = await tag_service.get_owner_profile(db, internal_user_id)
        
        # Получаем шаблоны с учетом владельца и публичных
        templates_list = await contract_service.get_contract_templates_for_user(internal_user_id)
        
        # Текущая дата для шаблона (формат YYYY-MM-DD)
        from datetime import date
        current_date = date.today().strftime("%Y-%m-%d")
        
        # Подготавливаем шаблоны в JSON формате для JavaScript
        templates_json = []
        for template in templates_list:
            templates_json.append({
                "id": template.id,
                "name": template.name,
                "content": template.content,
                "version": template.version,
                "fields_schema": template.fields_schema or []
            })
        
        # Получаем теги владельца для подстановки
        owner_tags = {}
        if owner_profile:
            owner_tags = owner_profile.get_tags_for_templates()
            # Добавляем системные теги
            from datetime import datetime
            owner_tags.update({
                'current_date': datetime.now().strftime('%d.%m.%Y'),
                'current_time': datetime.now().strftime('%H:%M'),
                'current_year': str(datetime.now().year)
            })
        
        # Получаем данные для переключения интерфейсов
        available_interfaces = await get_available_interfaces_for_user(internal_user_id)
        
        # Получаем графики выплат (системные + кастомные владельца)
        from domain.entities.payment_schedule import PaymentSchedule
        schedules_query = select(PaymentSchedule).where(
            PaymentSchedule.is_active == True
        ).where(
            (PaymentSchedule.owner_id == None) |  # Системные
            (PaymentSchedule.owner_id == internal_user_id)  # Кастомные владельца
        ).order_by(PaymentSchedule.is_custom.asc(), PaymentSchedule.id.asc())
        schedules_result = await db.execute(schedules_query)
        payment_schedules = schedules_result.scalars().all()
        
        return templates.TemplateResponse(
            "owner/employees/create.html",
            {
                "request": request,
                "title": "Создание договора",
                "current_user": current_user,
                "available_employees": available_employees,
                "objects": objects,
                "templates": templates_list,
                "templates_json": templates_json,
                "available_interfaces": available_interfaces,
                "owner_tags": owner_tags,
                "current_date": current_date,
                "employee_telegram_id": employee_telegram_id,
                "payment_schedules": payment_schedules  # Графики выплат
            }
        )
    except Exception as e:
        logger.error(f"Error loading create contract form: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка загрузки формы: {str(e)}")


@router.get("/employees/{employee_id}", response_class=HTMLResponse)
async def owner_employee_detail(
    employee_id: int,
    request: Request,
    current_user: dict = Depends(require_owner_or_superadmin),
    db: AsyncSession = Depends(get_db_session)
):
    """Детальная информация о сотруднике."""
    try:
        from apps.web.services.contract_service import ContractService
        
        # Получаем внутренний user_id владельца
        user_id = await get_user_id_from_current_user(current_user, db)
        if not user_id:
            raise HTTPException(status_code=401, detail="Пользователь не найден")
        
        # Получаем информацию о сотруднике через сервис (employee_id - это внутренний user_id)
        contract_service = ContractService()
        employee_info = await contract_service.get_employee_by_id(employee_id, user_id)
        
        if not employee_info:
            raise HTTPException(status_code=404, detail="У вас нет договоров с этим сотрудником")
        
        # Получаем данные для переключения интерфейсов
        available_interfaces = await get_available_interfaces_for_user(user_id)
        
        return templates.TemplateResponse(
            "owner/employees/detail.html",
            {
                "request": request,
                "title": f"Сотрудник {employee_info['first_name']} {employee_info['last_name'] or ''}",
                "current_user": current_user,
                "employee": employee_info,  # Полная информация из сервиса
                "employee_contracts": employee_info["contracts"],  # Договоры из сервиса
                "employee_accessible_objects": employee_info["accessible_objects"],  # Объекты из сервиса
                "available_interfaces": available_interfaces
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error loading employee detail: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка загрузки информации о сотруднике: {str(e)}")


@router.get("/employees/{employee_id}/edit", response_class=HTMLResponse)
async def owner_employee_edit_form(
    employee_id: int,
    request: Request,
    current_user: dict = Depends(require_owner_or_superadmin),
    db: AsyncSession = Depends(get_db_session)
):
    """Форма редактирования сотрудника."""
    try:
        if current_user is None or isinstance(current_user, RedirectResponse):
            raise HTTPException(status_code=401, detail="Пользователь не авторизован")
        
        # Получаем внутренний user_id владельца
        user_id = await get_user_id_from_current_user(current_user, db)
        if not user_id:
            raise HTTPException(status_code=401, detail="Пользователь не найден")
        
        # Получаем информацию о сотруднике
        from apps.web.services.contract_service import ContractService
        contract_service = ContractService()
        # employee_id - это внутренний user_id
        employee_info = await contract_service.get_employee_by_id(employee_id, user_id)
        
        if not employee_info:
            raise HTTPException(status_code=404, detail="Сотрудник не найден или у вас нет прав на его редактирование")
        
        # Получаем данные для переключения интерфейсов
        available_interfaces = await get_available_interfaces_for_user(user_id)
        
        # Получаем объект User из БД для редактирования
        from sqlalchemy import select
        from domain.entities.user import User
        employee_user_query = select(User).where(User.id == employee_id)
        result = await db.execute(employee_user_query)
        employee_user = result.scalar_one_or_none()
        
        if not employee_user:
            raise HTTPException(status_code=404, detail="Пользователь не найден")
        
        return templates.TemplateResponse("owner/employees/edit.html", {
            "request": request,
            "current_user": current_user,
            "employee": employee_user,  # Объект User из БД
            "available_interfaces": available_interfaces
        })
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in owner employee edit form: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки формы")


@router.post("/employees/{employee_id}/edit")
async def owner_employee_edit(
    employee_id: int,
    request: Request,
    current_user: dict = Depends(require_owner_or_superadmin)
):
    """Обработка формы редактирования сотрудника."""
    try:
        if current_user is None or isinstance(current_user, RedirectResponse):
            raise HTTPException(status_code=401, detail="Пользователь не авторизован")
        
        async with get_async_session() as db:
            user_id = await get_user_id_from_current_user(current_user, db)
            if not user_id:
                raise HTTPException(status_code=401, detail="Пользователь не найден")
            
            form_data = await request.form()
            
            # Получаем сотрудника (employee_id - это внутренний user_id)
            from sqlalchemy import select
            from domain.entities.user import User
            
            employee_query = select(User).where(User.id == employee_id)
            result = await db.execute(employee_query)
            employee = result.scalar_one_or_none()
            
            if not employee:
                raise HTTPException(status_code=404, detail="Сотрудник не найден")
            
            # Обновляем данные
            employee.first_name = form_data.get("first_name", "").strip()
            employee.last_name = form_data.get("last_name", "").strip()
            employee.username = form_data.get("username", "").strip()
            employee.phone = form_data.get("phone", "").strip()
            employee.email = form_data.get("email", "").strip() or None
            
            # Обработка даты рождения
            birth_date_str = form_data.get("birth_date", "").strip()
            if birth_date_str:
                try:
                    employee.birth_date = datetime.strptime(birth_date_str, "%Y-%m-%d")
                except ValueError:
                    pass
            else:
                employee.birth_date = None
            
            await db.commit()
            await db.refresh(employee)
            
            logger.info(f"Updated employee {employee_id} by owner {user_id}")
            
            return RedirectResponse(url=f"/owner/employees/{employee_id}", status_code=302)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating employee: {e}")
        raise HTTPException(status_code=500, detail="Ошибка обновления сотрудника")


@router.post("/employees/create")
async def owner_employees_create_contract(
    request: Request,
    employee_telegram_id: int = Form(...),
    title: str = Form(...),
    content: str = Form(""),
    hourly_rate: Optional[float] = Form(None),
    use_contract_rate: bool = Form(False),
    payment_system_id: Optional[int] = Form(1),  # По умолчанию simple_hourly
    use_contract_payment_system: bool = Form(False),
    start_date: str = Form(...),
    end_date: Optional[str] = Form(None),
    template_id: Optional[int] = Form(None),
    allowed_objects: List[int] = Form(default=[]),
    is_manager: bool = Form(False),
    manager_permissions: List[str] = Form(default=[]),
    first_name: Optional[str] = Form(None),
    last_name: Optional[str] = Form(None),
    phone: Optional[str] = Form(None),
    email: Optional[str] = Form(None),
    birth_date: Optional[str] = Form(None),
    inherit_payment_schedule: bool = Form(True),
    payment_schedule_id: Optional[int] = Form(None),
    current_user: dict = Depends(require_owner_or_superadmin),
    db: AsyncSession = Depends(get_db_session)
):
    """Создание договора с сотрудником."""
    try:
        from apps.web.services.contract_service import ContractService
        
        contract_service = ContractService()
        
        # Парсим даты
        start_date_obj = datetime.strptime(start_date, "%Y-%m-%d")
        end_date_obj = datetime.strptime(end_date, "%Y-%m-%d") if end_date else None
        
        # Обновляем профиль сотрудника, если указаны поля
        from sqlalchemy import select
        from domain.entities.user import User, UserRole
        
        employee_query = select(User).where(User.telegram_id == employee_telegram_id)
        result = await db.execute(employee_query)
        employee_user = result.scalar_one_or_none()
        
        first_name_value = (first_name or "").strip()
        last_name_value = (last_name or "").strip()
        phone_value = (phone or "").strip()
        email_value = (email or "").strip()
        birth_date_value = (birth_date or "").strip() if birth_date is not None else ""
        
        if not employee_user:
            # Создаем нового пользователя, если его нет в базе
            if not first_name_value:
                first_name_value = f"Сотрудник {employee_telegram_id}"
            employee_user = User(
                telegram_id=employee_telegram_id,
                username=None,
                first_name=first_name_value,
                last_name=last_name_value or None,
                phone=phone_value or None,
                email=email_value or None,
                role=UserRole.EMPLOYEE.value,
                roles=[UserRole.EMPLOYEE.value],
                is_active=True
            )
            if birth_date_value:
                try:
                    employee_user.birth_date = datetime.strptime(birth_date_value, "%Y-%m-%d")
                except ValueError:
                    employee_user.birth_date = None
            db.add(employee_user)
            await db.commit()
            await db.refresh(employee_user)
        else:
            updated = False
            if first_name_value:
                employee_user.first_name = first_name_value
                updated = True
            if last_name is not None:
                employee_user.last_name = last_name_value or None
                updated = True
            if phone is not None:
                employee_user.phone = phone_value or None
                updated = True
            if email is not None:
                employee_user.email = email_value or None
                updated = True
            if birth_date is not None:
                if birth_date_value:
                    try:
                        employee_user.birth_date = datetime.strptime(birth_date_value, "%Y-%m-%d")
                    except ValueError:
                        employee_user.birth_date = None
                else:
                    employee_user.birth_date = None
                updated = True
            if updated:
                await db.commit()
        
        # Получаем данные формы для динамических полей
        form_data = await request.form()
        dynamic_values = {}
        
        # Извлекаем значения динамических полей
        for key, value in form_data.items():
            if key.startswith("field_"):
                field_key = key[6:]  # Убираем префикс "field_"
                dynamic_values[field_key] = value
        
        # Обрабатываем права управляющего
        manager_permissions_dict = {}
        if is_manager and manager_permissions:
            for permission in manager_permissions:
                manager_permissions_dict[permission] = True
        
        # Создаем договор
        contract_data = {
            "employee_telegram_id": employee_telegram_id,
            "title": title,
            "content": content if content else None,
            "hourly_rate": hourly_rate,
            "use_contract_rate": use_contract_rate,
            "payment_system_id": payment_system_id,
            "use_contract_payment_system": use_contract_payment_system,
            "start_date": start_date_obj,
            "end_date": end_date_obj,
            "template_id": template_id,
            "allowed_objects": allowed_objects,
            "is_manager": is_manager,
            "manager_permissions": manager_permissions_dict if manager_permissions_dict else None,
            "values": dynamic_values if dynamic_values else None,
            "inherit_payment_schedule": inherit_payment_schedule,
            "payment_schedule_id": payment_schedule_id
        }
        
        contract = await contract_service.create_contract(current_user["id"], contract_data)
        
        if contract:
            return RedirectResponse(url="/owner/employees", status_code=303)
        else:
            raise HTTPException(status_code=400, detail="Ошибка создания договора")
            
    except Exception as e:
        logger.error(f"Error creating contract: {e}")
        raise HTTPException(status_code=400, detail=f"Ошибка создания договора: {str(e)}")


@router.get("/employees/contract/{contract_id}", response_class=HTMLResponse)
async def owner_contract_detail(
    contract_id: int,
    request: Request,
    current_user: dict = Depends(require_owner_or_superadmin),
    db: AsyncSession = Depends(get_db_session)
):
    """Детали договора."""
    try:
        from apps.web.services.contract_service import ContractService
        
        contract_service = ContractService()
        contract = await contract_service.get_contract_by_telegram_id(contract_id, current_user["id"])
        
        if not contract:
            raise HTTPException(status_code=404, detail="Договор не найден")
        
        # Получаем данные для переключения интерфейсов
        available_interfaces = await get_available_interfaces_for_user(current_user["id"])
        
        # Получаем историю изменений договора
        from shared.services.contract_history_service import ContractHistoryService
        from sqlalchemy import select
        from domain.entities.contract import Contract
        
        # Получаем внутренний ID договора
        contract_query = select(Contract).where(
            Contract.id == contract_id
        )
        contract_result = await db.execute(contract_query)
        contract_entity = contract_result.scalar_one_or_none()
        
        contract_history = []
        if contract_entity:
            history_service = ContractHistoryService(db)
            contract_history = await history_service.get_contract_history(
                contract_id=contract_entity.id,
                limit=100
            )
        
        return templates.TemplateResponse(
            "owner/employees/contract_detail.html",
            {
                "request": request,
                "contract": contract,
                "contract_history": contract_history,
                "title": f"Договор {contract.get('title', 'Без названия')}",
                "current_user": current_user,
                "available_interfaces": available_interfaces
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error loading contract detail: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка загрузки информации о договоре: {str(e)}")


@router.get("/contracts/{contract_id}/history", response_class=JSONResponse)
async def owner_contract_history_api(
    contract_id: int,
    field_name: Optional[str] = Query(None),
    limit: int = Query(100, ge=1, le=1000),
    offset: int = Query(0, ge=0),
    current_user: dict = Depends(require_owner_or_superadmin),
    db: AsyncSession = Depends(get_db_session)
):
    """API для получения истории изменений договора."""
    try:
        from apps.web.services.contract_service import ContractService
        from shared.services.contract_history_service import ContractHistoryService
        from sqlalchemy import select
        from domain.entities.contract import Contract
        
        # Проверяем права доступа
        contract_service = ContractService()
        contract = await contract_service.get_contract_by_telegram_id(contract_id, current_user["id"])
        
        if not contract:
            raise HTTPException(status_code=404, detail="Договор не найден")
        
        # Получаем внутренний ID договора
        contract_query = select(Contract).where(Contract.id == contract_id)
        contract_result = await db.execute(contract_query)
        contract_entity = contract_result.scalar_one_or_none()
        
        if not contract_entity:
            raise HTTPException(status_code=404, detail="Договор не найден")
        
        # Получаем историю
        history_service = ContractHistoryService(db)
        history = await history_service.get_contract_history(
            contract_id=contract_entity.id,
            field_name=field_name,
            limit=limit,
            offset=offset
        )
        
        # Сериализуем историю
        history_data = []
        for entry in history:
            history_data.append({
                "id": entry.id,
                "contract_id": entry.contract_id,
                "changed_at": entry.changed_at.isoformat() if entry.changed_at else None,
                "changed_by": entry.changed_by,
                "changed_by_user": {
                    "id": entry.changed_by_user.id,
                    "first_name": entry.changed_by_user.first_name,
                    "last_name": entry.changed_by_user.last_name
                } if entry.changed_by_user else None,
                "change_type": entry.change_type,
                "field_name": entry.field_name,
                "old_value": entry.old_value,
                "new_value": entry.new_value,
                "change_reason": entry.change_reason,
                "effective_from": entry.effective_from.isoformat() if entry.effective_from else None,
            })
        
        return {
            "contract_id": contract_entity.id,
            "history": history_data,
            "total": len(history_data),
            "limit": limit,
            "offset": offset
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error loading contract history: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка загрузки истории договора: {str(e)}")


@router.get("/contracts/{contract_id}/snapshot", response_class=JSONResponse)
async def owner_contract_snapshot_api(
    contract_id: int,
    date: str = Query(..., description="Дата в формате YYYY-MM-DD"),
    current_user: dict = Depends(require_owner_or_superadmin),
    db: AsyncSession = Depends(get_db_session)
):
    """API для получения снимка договора на конкретную дату."""
    try:
        from apps.web.services.contract_service import ContractService
        from shared.services.contract_history_service import ContractHistoryService
        from sqlalchemy import select
        from domain.entities.contract import Contract
        from datetime import date as date_type
        
        # Проверяем права доступа
        contract_service = ContractService()
        contract = await contract_service.get_contract_by_telegram_id(contract_id, current_user["id"])
        
        if not contract:
            raise HTTPException(status_code=404, detail="Договор не найден")
        
        # Получаем внутренний ID договора
        contract_query = select(Contract).where(Contract.id == contract_id)
        contract_result = await db.execute(contract_query)
        contract_entity = contract_result.scalar_one_or_none()
        
        if not contract_entity:
            raise HTTPException(status_code=404, detail="Договор не найден")
        
        # Парсим дату
        try:
            target_date = date_type.fromisoformat(date)
        except ValueError:
            raise HTTPException(status_code=400, detail="Неверный формат даты. Используйте YYYY-MM-DD")
        
        # Получаем снимок
        history_service = ContractHistoryService(db)
        snapshot = await history_service.get_contract_snapshot(
            contract_id=contract_entity.id,
            target_date=target_date
        )
        
        return {
            "contract_id": contract_entity.id,
            "target_date": date,
            "snapshot": snapshot
        }
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"Error loading contract snapshot: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка загрузки снимка договора: {str(e)}")


@router.get("/employees/contract/{contract_id}/edit", response_class=HTMLResponse)
async def owner_contract_edit_form(
    contract_id: int,
    request: Request,
    current_user: dict = Depends(require_owner_or_superadmin),
    db: AsyncSession = Depends(get_db_session)
):
    """Форма редактирования договора."""
    try:
        from apps.web.services.contract_service import ContractService
        from apps.web.services.tag_service import TagService
        
        contract_service = ContractService()
        contract = await contract_service.get_contract_by_telegram_id(contract_id, current_user["id"])
        
        if not contract:
            raise HTTPException(status_code=404, detail="Договор не найден")
        
        # Получаем доступные объекты
        objects = await contract_service.get_owner_objects(current_user["id"])
        
        # Получаем внутренний ID пользователя для шаблонов
        user_query = select(User).where(User.telegram_id == current_user["id"])
        user_result = await db.execute(user_query)
        user_obj = user_result.scalar_one_or_none()
        internal_user_id = user_obj.id if user_obj else None
        
        # Получаем шаблоны
        templates_list = await contract_service.get_contract_templates_for_user(internal_user_id)
        
        # Получаем профиль владельца для тегов
        tag_service = TagService()
        owner_profile = await tag_service.get_owner_profile(db, internal_user_id)
        
        # Подготавливаем шаблоны в JSON формате
        templates_json = []
        for template in templates_list:
            templates_json.append({
                "id": template.id,
                "name": template.name,
                "content": template.content,
                "version": template.version,
                "fields_schema": template.fields_schema or []
            })
        
        # Получаем теги владельца
        owner_tags = {}
        if owner_profile:
            owner_tags = owner_profile.get_tags_for_templates()
            from datetime import datetime
            owner_tags.update({
                'current_date': datetime.now().strftime('%d.%m.%Y'),
                'current_time': datetime.now().strftime('%H:%M'),
                'current_year': str(datetime.now().year)
            })
        
        # Получаем данные для переключения интерфейсов
        available_interfaces = await get_available_interfaces_for_user(internal_user_id)
        
        # Получаем графики выплат (системные + кастомные владельца)
        from domain.entities.payment_schedule import PaymentSchedule
        schedules_query = select(PaymentSchedule).where(
            PaymentSchedule.is_active == True
        ).where(
            (PaymentSchedule.owner_id == None) |  # Системные
            (PaymentSchedule.owner_id == internal_user_id)  # Кастомные владельца
        ).order_by(PaymentSchedule.is_custom.asc(), PaymentSchedule.id.asc())
        schedules_result = await db.execute(schedules_query)
        payment_schedules = schedules_result.scalars().all()
        
        return templates.TemplateResponse(
            "owner/employees/edit_contract.html",
            {
                "request": request,
                "contract": contract,
                "objects": objects,
                "templates": templates_list,
                "templates_json": templates_json,
                "available_interfaces": available_interfaces,
                "owner_tags": owner_tags,
                "title": f"Редактирование договора {contract.get('title', 'Без названия')}",
                "current_user": current_user,
                "payment_schedules": payment_schedules  # Графики выплат
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error loading contract edit form: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка загрузки формы: {str(e)}")


@router.post("/employees/contract/{contract_id}/edit")
async def owner_contract_edit(
    contract_id: int,
    request: Request,
    title: str = Form(...),
    content: str = Form(""),
    hourly_rate: Optional[float] = Form(None),
    use_contract_rate: bool = Form(False),
    payment_system_id: Optional[int] = Form(1),
    use_contract_payment_system: bool = Form(False),
    start_date: str = Form(...),
    end_date: Optional[str] = Form(None),
    template_id: Optional[int] = Form(None),
    allowed_objects: List[int] = Form(default=[]),
    is_manager: bool = Form(False),
    manager_permissions: List[str] = Form(default=[]),
    inherit_payment_schedule: bool = Form(True),
    payment_schedule_id: Optional[int] = Form(None),
    current_user: dict = Depends(require_owner_or_superadmin),
    db: AsyncSession = Depends(get_db_session)
):
    """Редактирование договора."""
    try:
        from apps.web.services.contract_service import ContractService
        
        contract_service = ContractService()
        
        # Парсим даты
        start_date_obj = datetime.strptime(start_date, "%Y-%m-%d")
        end_date_obj = datetime.strptime(end_date, "%Y-%m-%d") if end_date else None
        
        # Получаем данные формы для динамических полей
        form_data = await request.form()
        dynamic_values = {}
        
        # Извлекаем значения динамических полей
        for key, value in form_data.items():
            if key.startswith("field_"):
                field_key = key[6:]  # Убираем префикс "field_"
                dynamic_values[field_key] = value
        
        # Обрабатываем права управляющего
        manager_permissions_dict = {}
        if is_manager and manager_permissions:
            for permission in manager_permissions:
                manager_permissions_dict[permission] = True
        
        # Обновляем договор
        contract_data = {
            "title": title,
            "content": content if content else None,
            "hourly_rate": hourly_rate,
            "use_contract_rate": use_contract_rate,
            "payment_system_id": payment_system_id,
            "use_contract_payment_system": use_contract_payment_system,
            "start_date": start_date_obj,
            "end_date": end_date_obj,
            "template_id": template_id,
            "allowed_objects": allowed_objects,
            "is_manager": is_manager,
            "manager_permissions": manager_permissions_dict if manager_permissions_dict else None,
            "values": dynamic_values if dynamic_values else None,
            "inherit_payment_schedule": inherit_payment_schedule,
            "payment_schedule_id": payment_schedule_id
        }
        
        # Получаем внутренний ID пользователя
        user_id = await get_user_id_from_current_user(current_user, db)
        if not user_id:
            raise HTTPException(status_code=400, detail="Пользователь не найден")
        
        success = await contract_service.update_contract(contract_id, user_id, contract_data)
        
        if success:
            return RedirectResponse(url=f"/owner/employees/contract/{contract_id}", status_code=303)
        else:
            raise HTTPException(status_code=400, detail="Ошибка обновления договора")
            
    except Exception as e:
        logger.error(f"Error updating contract: {e}")
        raise HTTPException(status_code=400, detail=f"Ошибка обновления договора: {str(e)}")


@router.post("/employees/contract/{contract_id}/activate")
async def owner_contract_activate(
    contract_id: int,
    request: Request,
    current_user: dict = Depends(require_owner_or_superadmin),
    db: AsyncSession = Depends(get_db_session)
):
    """Активация договора."""
    try:
        from apps.web.services.contract_service import ContractService
        
        contract_service = ContractService()
        # Получаем внутренний ID пользователя
        user_id = await get_user_id_from_current_user(current_user, db)
        if not user_id:
            raise HTTPException(status_code=400, detail="Пользователь не найден")
        
        success = await contract_service.activate_contract(contract_id, user_id)
        
        if success:
            return RedirectResponse(url=f"/owner/employees/contract/{contract_id}", status_code=303)
        else:
            raise HTTPException(status_code=400, detail="Ошибка активации договора")
            
    except Exception as e:
        logger.error(f"Error activating contract: {e}")
        raise HTTPException(status_code=400, detail=f"Ошибка активации договора: {str(e)}")


@router.post("/employees/contract/{contract_id}/terminate")
async def owner_contract_terminate(
    contract_id: int,
    request: Request,
    current_user: dict = Depends(require_owner_or_superadmin),
    db: AsyncSession = Depends(get_db_session)
):
    """Расторжение договора."""
    try:
        # Получаем reason из запроса (может быть JSON или Form)
        content_type = request.headers.get("content-type", "")
        if "application/json" in content_type:
            # JSON запрос от JavaScript
            data = await request.json()
            reason = data.get("reason", "Расторжение по кнопке")
        else:
            # Form запрос от HTML формы
            form_data = await request.form()
            reason = form_data.get("reason", "Расторжение по кнопке")
        
        logger.info(f"=== ROUTE: Starting contract termination ===")
        logger.info(f"Route parameters: contract_id={contract_id}, reason='{reason}'")
        logger.info(f"Current user: {current_user}")
        
        from apps.web.services.contract_service import ContractService
        
        # Получаем внутренний ID пользователя
        user_id = await get_user_id_from_current_user(current_user, db)
        if not user_id:
            raise HTTPException(status_code=401, detail="Пользователь не найден")
        
        contract_service = ContractService()
        success = await contract_service.terminate_contract(contract_id, user_id, reason)
        
        if success:
            logger.info(f"Contract terminated successfully")
            # Проверяем тип запроса для правильного ответа
            if "application/json" in content_type:
                # JSON ответ для JavaScript
                return {"success": True, "message": "Договор расторгнут"}
            else:
                # Redirect для HTML формы
                return RedirectResponse(url="/owner/employees", status_code=303)
        else:
            logger.error(f"Contract termination returned False")
            if "application/json" in content_type:
                return {"success": False, "message": "Ошибка расторжения договора"}
            else:
                raise HTTPException(status_code=400, detail="Ошибка расторжения договора")
            
    except Exception as e:
        logger.error(f"=== ROUTE: Contract termination failed ===")
        logger.error(f"Error type: {type(e).__name__}")
        logger.error(f"Error message: {str(e)}")
        logger.error(f"Full traceback:")
        
        # Проверяем тип запроса для правильного ответа
        content_type = request.headers.get("content-type", "")
        if "application/json" in content_type:
            return {"success": False, "message": f"Ошибка расторжения договора: {str(e)}"}
        else:
            raise HTTPException(status_code=400, detail=f"Ошибка расторжения договора: {str(e)}")


@router.get("/employees/contract/{contract_id}/pdf")
async def owner_contract_pdf(
    contract_id: int,
    request: Request,
    current_user: dict = Depends(require_owner_or_superadmin),
    db: AsyncSession = Depends(get_db_session)
):
    """Генерация PDF договора."""
    try:
        from apps.web.services.contract_service import ContractService
        from apps.web.services.pdf_service import PDFService
        
        contract_service = ContractService()
        # Получаем внутренний ID пользователя
        user_id = await get_user_id_from_current_user(current_user, db)
        if not user_id:
            raise HTTPException(status_code=400, detail="Пользователь не найден")
        
        contract = await contract_service.get_contract_by_id_and_owner_telegram_id(contract_id, current_user.get("telegram_id") or current_user.get("telegram_id") or current_user.get("id"))
        
        if not contract:
            raise HTTPException(status_code=404, detail="Договор не найден")
        
        pdf_service = PDFService()
        pdf_content = await pdf_service.generate_contract_pdf(contract)
        
        return Response(
            content=pdf_content,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=contract_{contract_id}.pdf"}
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating contract PDF: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка генерации PDF: {str(e)}")

@router.get("/applications", response_class=HTMLResponse)
async def owner_applications(
    request: Request,
    current_user: dict = Depends(require_owner_or_superadmin),
    db: AsyncSession = Depends(get_db_session)
):
    """Страница заявок владельца"""
    try:
        if current_user is None or isinstance(current_user, RedirectResponse):
            raise HTTPException(status_code=401, detail="Пользователь не авторизован")
            
        user_id = await get_user_id_from_current_user(current_user, db)
        if not user_id:
            raise HTTPException(status_code=401, detail="Пользователь не найден")
        
        # Получаем заявки по объектам владельца
        applications_query = select(Application).join(Object).where(
            Object.owner_id == user_id
        ).options(
            selectinload(Application.applicant),
            selectinload(Application.object)
        ).order_by(desc(Application.created_at))
        
        applications_result = await db.execute(applications_query)
        applications = applications_result.scalars().all()
        
        # Получаем общий контекст владельца
        owner_context = await get_owner_context(user_id, db)
        
        return templates.TemplateResponse("owner/applications.html", {
            "request": request,
            "current_user": current_user,
            "applications": applications,
            "show_actions": True,
            "current_role": "owner",
            **owner_context
        })
        
    except Exception as e:
        logger.error(f"Ошибка загрузки заявок: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка загрузки заявок: {e}")

@router.post("/api/applications/approve")
async def approve_application(
    request: Request,
    current_user: dict = Depends(require_owner_or_superadmin),
    db: AsyncSession = Depends(get_db_session)
):
    """Одобрение заявки с назначением собеседования"""
    try:
        if current_user is None or isinstance(current_user, RedirectResponse):
            raise HTTPException(status_code=401, detail="Пользователь не авторизован")
            
        user_id = await get_user_id_from_current_user(current_user, db)
        if not user_id:
            raise HTTPException(status_code=401, detail="Пользователь не найден")
        
        form_data = await request.form()
        application_id = form_data.get("application_id")
        interview_datetime = form_data.get("interview_datetime")
        interview_type = form_data.get("interview_type")
        interview_notes = form_data.get("interview_notes", "").strip()
        
        if not application_id or not interview_datetime or not interview_type:
            raise HTTPException(status_code=400, detail="Не все поля заполнены")
        
        # Получаем заявку
        application_query = select(Application).join(Object).where(
            and_(
                Application.id == int(application_id),
                Object.owner_id == user_id
            )
        )
        application_result = await db.execute(application_query)
        application = application_result.scalar_one_or_none()
        
        if not application:
            raise HTTPException(status_code=404, detail="Заявка не найдена")
        
        # Обновляем заявку
        application.status = ApplicationStatus.INTERVIEW
        application.interview_scheduled_at = datetime.fromisoformat(interview_datetime.replace('T', ' '))
        application.interview_type = interview_type
        application.interview_result = interview_notes
        
        await db.commit()
        
        # Отправляем уведомления  
        try:
            from core.database.session import get_sync_session
            from shared.services.notification_service import NotificationService
            from core.config.settings import settings
            from domain.entities.user import User
            
            # Получаем синхронную сессию для NotificationService
            session_factory = get_sync_session
            with session_factory() as session:
                notification_service = NotificationService(
                    session=session,
                    telegram_token=settings.telegram_bot_token
                )
                
                # Получаем информацию о владельце для имени в уведомлении
                owner_query = select(User).where(User.id == user_id)
                owner_result = session.execute(owner_query)
                owner_user = owner_result.scalar_one_or_none()
                
                owner_name = "Владелец"
                if owner_user:
                    if owner_user.first_name or owner_user.last_name:
                        parts = []
                        if owner_user.first_name:
                            parts.append(owner_user.first_name.strip())
                        if owner_user.last_name:
                            parts.append(owner_user.last_name.strip())
                        owner_name = " ".join(parts) if parts else owner_user.username
                    elif owner_user.username:
                        owner_name = owner_user.username
                
                # Уведомляем соискателя
                notification_payload = {
                    "application_id": application.id,
                    "object_name": application.object.name if application.object else "Объект",
                    "object_address": application.object.address if application.object else "—",
                    "employee_position": application.object.employee_position if application.object and hasattr(application.object, 'employee_position') else "Должность не указана",
                    "scheduled_at": application.interview_scheduled_at.isoformat(),
                    "interview_type": interview_type,
                    "owner_name": owner_name
                }
                
                notification_service.create(
                    [application.applicant_id],
                    "interview_assigned",
                    notification_payload,
                    send_telegram=True
                )
                session.commit()
        except Exception as notification_error:
            logger.error(f"Ошибка отправки уведомлений: {notification_error}")
        
        logger.info(f"Заявка {application_id} одобрена, собеседование назначено на {interview_datetime}")
        
        return {
            "message": "Заявка одобрена и собеседование назначено",
            "interview_datetime": interview_datetime,
            "interview_type": interview_type
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка одобрения заявки: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка одобрения заявки: {e}")

@router.post("/api/applications/reject")
async def reject_application(
    application_id: int = Form(...),
    reject_reason: str = Form(""),
    current_user: dict = Depends(require_owner_or_superadmin),
    db: AsyncSession = Depends(get_db_session)
):
    """Отклонение заявки owner"""
    logger.info(f"🔥 [OWNER REJECT] === START owner.py reject_application function ===")
    logger.info(f"🔥 [OWNER REJECT] application_id={application_id}, reject_reason={reject_reason}")
    try:
        logger.info(f"=== REJECT APPLICATION STARTED ===")
        
        if current_user is None or isinstance(current_user, RedirectResponse):
            raise HTTPException(status_code=401, detail="Пользователь не авторизован")

        if isinstance(current_user, dict):
            user_id = current_user.get("telegram_id") or current_user.get("id")
        else:
            user_id = getattr(current_user, "telegram_id", None)

        if not user_id:
            raise HTTPException(status_code=401, detail="Пользователь не найден")
        
        logger.info(f"User ID resolved: {user_id}")
        
        logger.info(f"Form data received: application_id={application_id}, reject_reason={reject_reason}")
        
        if not application_id:
            raise HTTPException(status_code=400, detail="ID заявки не указан")
        
        # Получаем заявку
        application_query = select(Application).join(Object).where(
            and_(
                Application.id == int(application_id),
                Object.owner_id == user_id
            )
        )
        application_result = await db.execute(application_query)
        application = application_result.scalar_one_or_none()
        
        if not application:
            raise HTTPException(status_code=404, detail="Заявка не найдена")
        
        # Обновляем заявку
        application.status = ApplicationStatus.REJECTED
        application.interview_result = reject_reason
        
        await db.commit()
        
        logger.info(f"---> ПЕРЕД_ОТПРАВКОЙ_УВЕДОМЛЕНИЙ: Заявка {application_id} отклонена. Переходим в блок notification.")
        
        # Отправляем уведомления
        try:
            from core.database.session import get_sync_session
            from shared.services.notification_service import NotificationService
            logger.info(f"---> НАЧАЛО отправки уведомления о отклонении заявки {application_id} для соискателя {application.applicant_id}, причина: {reject_reason}")
            from core.config.settings import settings
            from domain.entities.user import User
            
            # Получаем синхронную сессию для NotificationService
            session_factory = get_sync_session
            with session_factory() as session:
                notification_service = NotificationService(
                    session=session,
                    telegram_token=settings.telegram_bot_token
                )
                
                # Получаем информацию о владельце для имени в уведомлении
                owner_query = select(User).where(User.id == user_id)
                owner_result = session.execute(owner_query)
                owner_user = owner_result.scalar_one_or_none()
                
                owner_name = "Владелец"
                if owner_user:
                    if owner_user.first_name or owner_user.last_name:
                        parts = []
                        if owner_user.first_name:
                            parts.append(owner_user.first_name.strip())
                        if owner_user.last_name:
                            parts.append(owner_user.last_name.strip())
                        owner_name = " ".join(parts) if parts else owner_user.username
                    elif owner_user.username:
                        owner_name = owner_user.username
                
                # Уведомляем соискателя об отклонении
                notification_payload = {
                    "application_id": application.id,
                    "object_name": application.object.name if hasattr(application, 'object') and application.object else "Объект",
                    "object_address": application.object.address if hasattr(application, 'object') and application.object else "—",
                    "employee_position": application.object.employee_position if hasattr(application, 'object') and application.object and hasattr(application.object, 'employee_position') else "Должность не указана",
                    "reason": reject_reason,
                    "owner_name": owner_name
                }
                
                logger.info(f"---> ВЫЗЫВАЕМ notification_service.create для пользователя {application.applicant_id}")
                
                notification_service.create(
                    [application.applicant_id],
                    "application_rejected",
                    notification_payload,
                    send_telegram=True
                )
                session.commit()
                logger.info(f"---> УВЕДОМЛЕНИЕ отправилось успешно для пользователя {application.applicant_id}")
        except Exception as notification_error:
            logger.error(f"Ошибка отправки уведомлений об отклонении: {notification_error}")
        
        logger.info(f"Заявка {application_id} отклонена")
        
        return {
            "message": "Заявка отклонена",
            "reason": reject_reason
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка отклонения заявки: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка отклонения заявки: {e}")


@router.get("/api/applications/count")
async def owner_applications_count(
    current_user: dict = Depends(require_owner_or_superadmin),
    db: AsyncSession = Depends(get_db_session)
) -> dict[str, int]:
    """Количество новых заявок для владельца."""
    if current_user is None or isinstance(current_user, RedirectResponse):
        raise HTTPException(status_code=401, detail="Пользователь не авторизован")

    user_id = await get_user_id_from_current_user(current_user, db)
    if not user_id:
        raise HTTPException(status_code=401, detail="Пользователь не найден")
    from apps.web.utils.applications_utils import get_new_applications_count
    count = await get_new_applications_count(user_id, db, "owner")
    return {"count": count}

@router.get("/api/applications/{application_id}")
async def owner_application_details_api(
    application_id: int,
    current_user: dict = Depends(require_owner_or_superadmin),
    db: AsyncSession = Depends(get_db_session)
):
    if current_user is None or isinstance(current_user, RedirectResponse):
        raise HTTPException(status_code=401, detail="Необходима авторизация")

    user_id = await get_user_id_from_current_user(current_user, db)
    if not user_id:
        raise HTTPException(status_code=401, detail="Пользователь не найден")

    query = select(Application, Object, User).select_from(Application).join(
        Object, Application.object_id == Object.id
    ).join(
        User, Application.applicant_id == User.id
    ).where(
        and_(Application.id == application_id, Object.owner_id == user_id)
    )

    result = await db.execute(query)
    row = result.first()
    if not row:
        raise HTTPException(status_code=404, detail="Заявка не найдена или нет доступа")

    application, obj, applicant = row
    return {
        "id": application.id,
        "object_name": obj.name,
        "object_address": obj.address,
        "status": application.status.value,
        "message": application.message,
        "created_at": application.created_at.isoformat() if application.created_at else None,
        "interview_scheduled_at": application.interview_scheduled_at.isoformat() if application.interview_scheduled_at else None,
        "interview_type": application.interview_type,
        "applicant": {
            "full_name": applicant.full_name,
            "first_name": applicant.first_name,
            "last_name": applicant.last_name,
            "username": applicant.username,
            "email": applicant.email,
            "phone": applicant.phone,
            "skills": applicant.skills,
            "about": applicant.about,
            "work_experience": applicant.work_experience,
            "preferred_schedule": applicant.preferred_schedule,
            "education": applicant.education,
        }
    }

@router.post("/api/applications/finalize-contract")
async def owner_finalize_contract(
    application_id: int = Form(...),
    current_user: dict = Depends(require_owner_or_superadmin),
    db: AsyncSession = Depends(get_db_session)
):
    try:
        if current_user is None or isinstance(current_user, RedirectResponse):
            raise HTTPException(status_code=401, detail="Пользователь не авторизован")

        user_id = await get_user_id_from_current_user(current_user, db)
        if not user_id:
            raise HTTPException(status_code=401, detail="Пользователь не найден")

        query = select(Application).select_from(Application).join(
            Object, Application.object_id == Object.id
        ).where(and_(Application.id == application_id, Object.owner_id == user_id))

        result = await db.execute(query)
        application = result.scalar_one_or_none()
        if not application:
            raise HTTPException(status_code=404, detail="Заявка не найдена или нет доступа")

        application.status = ApplicationStatus.APPROVED
        await db.commit()

        logger.info(f"Заявка {application_id} переведена в статус APPROVED владельцем {user_id}")
        return {"message": "Заявка одобрена", "status": application.status.value}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка одобрения заявки владельцем: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка одобрения заявки: {e}")


@router.get("/tariff/change", response_class=HTMLResponse, name="owner_change_tariff")
async def owner_change_tariff_page(
    request: Request,
    current_user: dict = Depends(require_owner_or_superadmin)
):
    """Страница смены тарифа для владельца."""
    try:
        # Получаем user_id из current_user
        if isinstance(current_user, dict):
            telegram_id = current_user.get("telegram_id")
            if not telegram_id:
                raise HTTPException(status_code=400, detail="Не удалось определить пользователя")
            
            async with get_async_session() as session:
                from sqlalchemy import select
                from domain.entities.user import User
                
                user_query = select(User).where(User.telegram_id == telegram_id)
                user_result = await session.execute(user_query)
                user_obj = user_result.scalar_one_or_none()
                user_id = user_obj.id if user_obj else None
        else:
            user_id = current_user.id
            
        if not user_id:
            raise HTTPException(status_code=400, detail="Не удалось определить пользователя")
        
        async with get_async_session() as session:
            from apps.web.services.tariff_service import TariffService
            from apps.web.services.limits_service import LimitsService
            
            # Получаем доступные тарифы
            tariff_service = TariffService(session)
            tariff_plans = await tariff_service.get_all_tariff_plans(active_only=True)
            
            # Получаем текущую подписку
            limits_service = LimitsService(session)
            limits_summary = await limits_service.get_user_limits_summary(user_id)
            
            # Загружаем системные функции для отображения названий
            from shared.services.system_features_service import SystemFeaturesService
            features_service = SystemFeaturesService()
            all_features = await features_service.get_all_features(session)
            feature_names_map = {f.key: f.name for f in all_features}
        
        return templates.TemplateResponse("owner/change_tariff.html", {
            "request": request,
            "current_user": current_user,
            "title": "Смена тарифа",
            "tariff_plans": tariff_plans,
            "limits_summary": limits_summary,
            "user_id": user_id,
            "feature_names": feature_names_map
        })
        
    except Exception as e:
        logger.error(f"Error loading change tariff page: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка загрузки страницы: {str(e)}")


@router.post("/tariff/change", response_class=JSONResponse, name="owner_change_tariff_post")
async def owner_change_tariff_post(
    request: Request,
    current_user: dict = Depends(require_owner_or_superadmin),
    db: AsyncSession = Depends(get_db_session)
):
    """Смена тарифа для владельца."""
    try:
        # Получаем данные из запроса
        data = await request.json()
        tariff_plan_id = data.get("tariff_plan_id")
        
        if not tariff_plan_id:
            raise HTTPException(status_code=400, detail="Не указан тариф")
        
        # Получаем user_id из current_user
        user_id = await get_user_id_from_current_user(current_user, db)
        if not user_id:
            raise HTTPException(status_code=400, detail="Не удалось определить пользователя")
        
        from domain.entities.tariff_plan import TariffPlan
        from domain.entities.user_subscription import UserSubscription, SubscriptionStatus
        from apps.web.services.billing_service import BillingService
        from apps.web.services.limits_service import LimitsService
        from core.utils.url_helper import URLHelper
        from datetime import datetime, timedelta, timezone
        
        # Получаем тарифный план
        tariff_result = await db.execute(
            select(TariffPlan).where(TariffPlan.id == tariff_plan_id)
        )
        tariff_plan = tariff_result.scalar_one_or_none()
        
        if not tariff_plan:
            raise HTTPException(status_code=404, detail="Тарифный план не найден")
        
        # Проверяем возможность понижения тарифа
        limits_service = LimitsService(db)
        downgrade_allowed, downgrade_message, downgrade_details = await limits_service.check_tariff_downgrade_allowed(
            user_id=user_id,
            new_tariff_plan_id=tariff_plan_id
        )
        
        if not downgrade_allowed:
            logger.info(
                f"Tariff downgrade blocked",
                user_id=user_id,
                tariff_plan_id=tariff_plan_id,
                message=downgrade_message,
                details=downgrade_details
            )
            return {
                "success": False,
                "error": downgrade_message,
                "details": downgrade_details,
                "requires_action": True
            }
        
        # Получаем текущую активную подписку для определения даты начала нового тарифа
        current_subscription = await limits_service._get_active_subscription(user_id)
        
        # Деактивируем старую подписку перед созданием новой
        if current_subscription:
            from apps.web.services.tariff_service import TariffService
            tariff_service = TariffService(db)
            await tariff_service.deactivate_user_subscriptions(user_id)
            logger.info(
                f"Deactivated old subscription before creating new one",
                user_id=user_id,
                old_subscription_id=current_subscription.id
            )
            # Обновляем current_subscription для правильного вычисления даты начала
            await db.refresh(current_subscription)
        
        # Определяем дату начала нового тарифа
        # Если есть активная подписка - новый тариф начнется после её окончания
        # Если нет - новый тариф начнется сейчас
        if current_subscription and current_subscription.expires_at and current_subscription.expires_at > datetime.now(timezone.utc):
            new_tariff_starts_at = current_subscription.expires_at
            logger.info(
                f"Scheduling new tariff to start after current expires",
                user_id=user_id,
                current_subscription_id=current_subscription.id,
                current_expires_at=current_subscription.expires_at,
                new_tariff_starts_at=new_tariff_starts_at
            )
        else:
            new_tariff_starts_at = datetime.now(timezone.utc)
        
        # Если тариф платный, создаем платеж через YooKassa
        if tariff_plan.price and float(tariff_plan.price) > 0:
            # Вычисляем expires_at нового тарифа: начинает действовать с new_tariff_starts_at + период тарифа
            new_tariff_expires_at = None
            if tariff_plan.billing_period == "month":
                new_tariff_expires_at = new_tariff_starts_at + timedelta(days=30)
            elif tariff_plan.billing_period == "year":
                new_tariff_expires_at = new_tariff_starts_at + timedelta(days=365)
            
            # Формируем notes с информацией о дате начала
            notes_parts = ["Смена тарифа владельцем (ожидает оплаты)"]
            if current_subscription and new_tariff_starts_at > datetime.now(timezone.utc):
                notes_parts.append(f"Начнет действовать с {new_tariff_starts_at.strftime('%d.%m.%Y')}")
            
            # Создаем подписку со статусом SUSPENDED (будет активирована после оплаты и начнет действовать с scheduled даты)
            new_subscription = UserSubscription(
                user_id=user_id,
                tariff_plan_id=tariff_plan_id,
                status=SubscriptionStatus.SUSPENDED,  # Будет активирована после оплаты
                started_at=new_tariff_starts_at,  # Начнет действовать с этой даты
                expires_at=new_tariff_expires_at,  # Окончится через период тарифа от даты начала
                auto_renewal=True,
                payment_method="yookassa",
                notes=" | ".join(notes_parts)
            )
            
            db.add(new_subscription)
            await db.commit()
            await db.refresh(new_subscription)
            
            # Создаем транзакцию и платеж через YooKassa
            billing_service = BillingService(db)
            amount = await billing_service.compute_subscription_amount(
                user_id, new_subscription, tariff_plan
            )
            return_url = await URLHelper.build_url("/owner/subscription/payment_success")
            
            logger.info(
                f"Building payment return URL",
                user_id=user_id,
                tariff_plan_id=tariff_plan_id,
                return_url=return_url
            )
            
            try:
                transaction, payment_url = await billing_service.create_payment_transaction(
                    user_id=user_id,
                    subscription_id=new_subscription.id,
                    amount=amount,
                    currency=tariff_plan.currency or "RUB",
                    description=f"Оплата подписки на тариф '{tariff_plan.name}'",
                    return_url=return_url
                )
                
                logger.info(
                    f"Created payment for tariff change",
                    user_id=user_id,
                    tariff_plan_id=tariff_plan_id,
                    transaction_id=transaction.id,
                    payment_url=payment_url
                )
                
                # Формируем сообщение для пользователя
                message = "Тариф выбран. Перейдите к оплате."
                if current_subscription and new_tariff_starts_at > datetime.now(timezone.utc):
                    message += f"\n\nНовый тариф начнет действовать с {new_tariff_starts_at.strftime('%d.%m.%Y')} (после окончания текущего тарифа)."
                
                return {
                    "success": True,
                    "requires_payment": True,
                    "payment_url": payment_url,
                    "message": message
                }
                
            except Exception as e:
                logger.error(
                    f"Error creating payment for tariff change: {e}",
                    error=str(e),
                    user_id=user_id,
                    tariff_plan_id=tariff_plan_id
                )
                # Откатываем создание подписки
                await db.rollback()
                raise HTTPException(status_code=500, detail=f"Ошибка создания платежа: {str(e)}")
        
        else:
            # Бесплатный тариф - создаем подписку напрямую
            # Вычисляем expires_at нового тарифа: начинает действовать с new_tariff_starts_at + период тарифа
            new_tariff_expires_at = None
            if tariff_plan.billing_period == "month":
                new_tariff_expires_at = new_tariff_starts_at + timedelta(days=30)
            elif tariff_plan.billing_period == "year":
                new_tariff_expires_at = new_tariff_starts_at + timedelta(days=365)
            
            # Формируем notes с информацией о дате начала
            notes_parts = ["Смена тарифа владельцем"]
            if current_subscription and new_tariff_starts_at > datetime.now(timezone.utc):
                notes_parts.append(f"Начнет действовать с {new_tariff_starts_at.strftime('%d.%m.%Y')}")
            
            # Если новый тариф начнется в будущем - создаем подписку со статусом SUSPENDED
            # Иначе - сразу ACTIVE
            subscription_status = SubscriptionStatus.SUSPENDED if new_tariff_starts_at > datetime.now(timezone.utc) else SubscriptionStatus.ACTIVE
            
            new_subscription = UserSubscription(
                user_id=user_id,
                tariff_plan_id=tariff_plan_id,
                status=subscription_status,
                started_at=new_tariff_starts_at,  # Начнет действовать с этой даты
                expires_at=new_tariff_expires_at,  # Окончится через период тарифа от даты начала
                auto_renewal=True,
                payment_method="manual",
                notes=" | ".join(notes_parts)
            )
            
            db.add(new_subscription)
            await db.commit()
            
            # Формируем сообщение для пользователя
            message = "Тариф изменен."
            if current_subscription and new_tariff_starts_at > datetime.now(timezone.utc):
                message += f"\n\nНовый тариф начнет действовать с {new_tariff_starts_at.strftime('%d.%m.%Y')} (после окончания текущего тарифа)."
            
            logger.info(
                f"Tariff changed successfully",
                user_id=user_id,
                tariff_plan_id=tariff_plan_id,
                new_tariff_starts_at=new_tariff_starts_at,
                subscription_status=subscription_status.value
            )
            
            # Обновляем enabled_features в профиле владельца
            from domain.entities.owner_profile import OwnerProfile
            from sqlalchemy.orm.attributes import flag_modified
            
            profile_result = await db.execute(
                select(OwnerProfile).where(OwnerProfile.user_id == user_id)
            )
            owner_profile = profile_result.scalar_one_or_none()
            
            if owner_profile:
                new_tariff_features = tariff_plan.features or []
                current_enabled = list(owner_profile.enabled_features) if owner_profile.enabled_features else []
                filtered_enabled = [f for f in current_enabled if f in new_tariff_features]
                
                owner_profile.enabled_features = filtered_enabled
                flag_modified(owner_profile, 'enabled_features')
                
                await db.commit()
                
                # Инвалидируем кэш Redis
                from core.cache.redis_cache import cache
                telegram_id = current_user.get("telegram_id") if isinstance(current_user, dict) else None
                if telegram_id:
                    cache_key = f"enabled_features:{telegram_id}"
                    await cache.delete(cache_key)
                
                logger.info(
                    f"Tariff changed for user {user_id} to tariff {tariff_plan_id}. "
                    f"Enabled features updated: {current_enabled} -> {filtered_enabled}"
                )
            else:
                logger.info(f"Tariff changed for user {user_id} to tariff {tariff_plan_id}")
            
            return {
                "success": True,
                "requires_payment": False,
                "message": "Тариф успешно изменен"
            }
        
    except HTTPException as e:
        return JSONResponse(
            status_code=e.status_code,
            content={
                "success": False,
                "error": e.detail,
                "message": str(e.detail)
            }
        )
    except Exception as e:
        logger.exception(f"Error changing tariff: {e}")
        return JSONResponse(
            status_code=500,
            content={
                "success": False,
                "error": str(e),
                "message": f"Ошибка смены тарифа: {str(e)}"
            }
        )


# Роут /payment-systems удален - функционал перенесен в /org-structure