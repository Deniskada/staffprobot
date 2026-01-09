"""Роуты для работы с начислениями и выплатами."""

from datetime import date, timedelta, datetime
from typing import Optional, Dict, Any, Tuple, List
from decimal import Decimal
from urllib.parse import urlencode, quote
import io

from fastapi import APIRouter, Depends, Request, Form, HTTPException, status, Query
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, Response
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, or_, func
from sqlalchemy.orm import selectinload
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from apps.web.jinja import templates
from apps.web.middleware.auth_middleware import get_current_user
from apps.web.dependencies import require_role
from apps.web.middleware.role_middleware import require_employee_or_applicant
from core.database.session import get_db_session
from core.logging.logger import logger
from apps.web.services.payroll_service import PayrollService
from apps.web.services.payroll_statement_exporter import build_statement_workbook
from domain.entities.user import User
from domain.entities.payroll_entry import PayrollEntry
from domain.entities.contract import Contract
from domain.entities.payroll_adjustment import PayrollAdjustment
from domain.entities.payment_schedule import PaymentSchedule
from domain.entities.object import Object
from domain.entities.org_structure import OrgStructureUnit
from sqlalchemy.dialects.postgresql import JSONB
from sqlalchemy import cast
from shared.services.payroll_adjustment_service import PayrollAdjustmentService
from shared.services.payment_schedule_service import get_payment_period_for_date
from shared.services.payroll_generation_service import PayrollGenerationService
from shared.services.payroll_statement_service import PayrollStatementService
from shared.services.contract_helper import get_inherited_payment_schedule_id

router = APIRouter()


async def get_user_id_from_current_user(current_user, session: AsyncSession) -> Optional[int]:
    """Получает внутренний ID пользователя из current_user."""
    # current_user может быть dict (из JWT) или User объект (из dependencies)
    if isinstance(current_user, dict):
        telegram_id = current_user.get("telegram_id") or current_user.get("id")
        user_query = select(User).where(User.telegram_id == telegram_id)
        user_result = await session.execute(user_query)
        user_obj = user_result.scalar_one_or_none()
        return user_obj.id if user_obj else None
    elif isinstance(current_user, User):
        return current_user.id
    else:
        return None


async def _get_payment_schedule_id_for_object(obj: Object, session: AsyncSession) -> Optional[int]:
    """
    Получить ID графика выплат для объекта с учетом наследования от подразделения.
    
    Args:
        obj: Объект
        session: Сессия БД
        
    Returns:
        ID графика выплат или None
    """
    # ПРИОРИТЕТ 1: Если у объекта есть прямая привязка к графику - вернуть её
    if obj.payment_schedule_id:
        return obj.payment_schedule_id
    
    # ПРИОРИТЕТ 2+: Если нет графика у объекта - идем к подразделению
    if not obj.org_unit_id:
        return None
    
    # ПРИОРИТЕТ 2+: Получить график от подразделения (с учетом наследования по цепочке)
    # Реализуем логику вручную, чтобы избежать lazy loading
    current_unit_id = obj.org_unit_id
    
    while current_unit_id:
        result = await session.execute(
            select(OrgStructureUnit).where(OrgStructureUnit.id == current_unit_id)
        )
        unit = result.scalar_one_or_none()
        
        if not unit:
            break
        
        # Если у подразделения есть явный график - вернуть его
        if unit.payment_schedule_id:
            return unit.payment_schedule_id
        
        # Иначе идем к родителю
        current_unit_id = unit.parent_id
    
    return None


async def calculate_payment_date_for_entry(entry: PayrollEntry, session: AsyncSession) -> Optional[date]:
    """
    Вычисляет дату выплаты для начисления на основе графика выплат.
    
    Args:
        entry: Начисление
        session: Сессия БД
        
    Returns:
        Дата выплаты или None, если график не найден
    """
    try:
        # Загрузить объект и договор
        if not entry.object_id:
            return None
        
        obj_query = select(Object).where(Object.id == entry.object_id)
        obj_result = await session.execute(obj_query)
        obj = obj_result.scalar_one_or_none()
        
        if not obj:
            return None
        
        # Найти график выплат
        schedule_id = None
        
        # Проверить договор
        if entry.contract_id:
            contract_query = select(Contract).where(Contract.id == entry.contract_id)
            contract_result = await session.execute(contract_query)
            contract = contract_result.scalar_one_or_none()
            
            if contract:
                if not contract.inherit_payment_schedule and contract.payment_schedule_id:
                    schedule_id = contract.payment_schedule_id
                else:
                    # Наследуем от объекта
                    schedule_id = await _get_payment_schedule_id_for_object(obj, session)
        else:
            # Нет договора - используем график объекта
            schedule_id = await _get_payment_schedule_id_for_object(obj, session)
        
        if not schedule_id:
            return None
        
        # Загрузить график
        schedule_query = select(PaymentSchedule).where(PaymentSchedule.id == schedule_id)
        schedule_result = await session.execute(schedule_query)
        schedule = schedule_result.scalar_one_or_none()
        
        if not schedule or not schedule.payment_period:
            return None
        
        # Вычислить дату выплаты обратно из периода
        period_config = schedule.payment_period
        
        if schedule.frequency == "weekly":
            end_offset = period_config.get("end_offset", -16)
            payment_date = entry.period_end - timedelta(days=end_offset)
            return payment_date
        
        elif schedule.frequency == "biweekly":
            end_offset = period_config.get("end_offset", -14)
            payment_date = entry.period_end - timedelta(days=end_offset)
            return payment_date
        
        elif schedule.frequency == "monthly":
            payments = period_config.get("payments", [])
            if payments:
                # Для месячных графиков сложнее - нужно найти payment, который соответствует периоду
                # Используем period_end для определения
                for payment in payments:
                    start_offset = payment.get("start_offset", 0)
                    end_offset = payment.get("end_offset", 0)
                    is_end_of_month = payment.get("is_end_of_month", False)
                    
                    if is_end_of_month:
                        # Для второй выплаты period_end - это последний день предыдущего месяца
                        # payment_date должен быть днём месяца из next_payment_date или payment_day
                        if payment.get("is_start_of_month", False):
                            payment_day = schedule.payment_day
                        else:
                            next_payment_str = payment.get("next_payment_date")
                            if next_payment_str:
                                try:
                                    next_payment = date.fromisoformat(next_payment_str)
                                    payment_day = next_payment.day
                                except (ValueError, TypeError):
                                    continue
                            else:
                                continue
                        
                        # Проверяем, соответствует ли период
                        # Для is_end_of_month period_end - это последний день предыдущего месяца
                        # period_start - это день после окончания первой выплаты
                        # Проверяем по period_end
                        if entry.period_end.month == (entry.period_start.month - 1 if entry.period_start.month > 1 else 12):
                            # Это вторая выплата
                            # payment_date - это день месяца payment_day в месяце period_end
                            payment_date = date(entry.period_end.year, entry.period_end.month, payment_day)
                            return payment_date
                    else:
                        # Обычная выплата
                        calculated_payment_date = entry.period_end - timedelta(days=end_offset)
                        if calculated_payment_date.month == entry.period_end.month:
                            return calculated_payment_date
                
                # Fallback: используем period_end
                return entry.period_end
            else:
                # Старый формат
                end_offset = period_config.get("end_offset", -30)
                payment_date = entry.period_end - timedelta(days=end_offset)
                return payment_date
        
        elif schedule.frequency == "daily":
            end_offset = period_config.get("end_offset", -1)
            payment_date = entry.period_end - timedelta(days=end_offset)
            return payment_date
        
        return None
        
    except Exception as e:
        logger.warning(f"Error calculating payment date for entry {entry.id}: {e}")
        return None


async def _fetch_owner_payroll_report_data(
    owner_id: int,
    period_start_date: date,
    period_end_date: date,
    db: AsyncSession
) -> Tuple[list, Optional[Dict[str, Any]], Optional[Dict[str, Any]]]:
    """Возвращает данные для HTML и Excel отчётов владельца."""
    from sqlalchemy.orm import selectinload

    entries_query = select(PayrollEntry).options(
        selectinload(PayrollEntry.employee),
        selectinload(PayrollEntry.object_),
        selectinload(PayrollEntry.contract)
    ).join(
        Contract, Contract.id == PayrollEntry.contract_id
    ).where(
        Contract.owner_id == owner_id,
        PayrollEntry.period_start >= period_start_date,
        PayrollEntry.period_end <= period_end_date
    ).order_by(PayrollEntry.object_id, PayrollEntry.employee_id)

    entries_result = await db.execute(entries_query)
    entries = list(entries_result.scalars().all())

    if not entries:
        return [], None, None

    # Группировка по сотрудникам: считаем на скольких объектах работал каждый
    employee_objects_count = {}
    for entry in entries:
        employee_objects_count.setdefault(entry.employee_id, set()).add(entry.object_id)

    single_object_employees = {emp_id for emp_id, objs in employee_objects_count.items() if len(objs) == 1}
    multi_object_employees = {emp_id for emp_id, objs in employee_objects_count.items() if len(objs) > 1}

    objects_data: Dict[int, Dict[str, Any]] = {}
    multi_object_rows = []

    for entry in entries:
        shifts_count = len(entry.calculation_details.get("shifts", [])) if entry.calculation_details else 0

        contract = entry.contract
        if contract.status == 'active':
            status = "Работает"
        elif contract.status == 'terminated':
            status = "Уволен"
        else:
            status = contract.status.capitalize()

        row_data = {
            "employee_id": entry.employee_id,
            "last_name": entry.employee.last_name or "",
            "first_name": entry.employee.first_name or "",
            "status": status,
            "shifts_count": shifts_count,
            "hours": float(entry.hours_worked or 0),
            "rate": float(entry.hourly_rate or 0),
            "bonus": float(entry.total_bonuses or 0),
            "penalty": float(entry.total_deductions or 0),
            "total": float(entry.net_amount or 0)
        }

        if entry.employee_id in multi_object_employees:
            row_data["object_name"] = entry.object_.name if entry.object_ else "—"
            multi_object_rows.append(row_data)
        else:
            object_id = entry.object_id
            if object_id not in objects_data:
                objects_data[object_id] = {
                    "object_name": entry.object_.name if entry.object_ else f"Объект #{object_id}",
                    "rows": [],
                    "subtotal": 0.0
                }

            objects_data[object_id]["rows"].append(row_data)
            objects_data[object_id]["subtotal"] += row_data["total"]

    objects_list = sorted(objects_data.values(), key=lambda x: x["object_name"])
    grand_total = sum(obj["subtotal"] for obj in objects_list) + sum(row["total"] for row in multi_object_rows)

    report_data = {
        "objects": objects_list,
        "multi_object_employees": sorted(multi_object_rows, key=lambda x: (x["last_name"], x["first_name"])),
        "grand_total": grand_total
    }

    from domain.entities.employee_payment import EmployeePayment

    employee_payment_data: Dict[int, Dict[str, Any]] = {}
    for entry in entries:
        emp_id = entry.employee_id
        emp_data = employee_payment_data.setdefault(emp_id, {
            "employee_id": emp_id,
            "last_name": entry.employee.last_name or "",
            "first_name": entry.employee.first_name or "",
            "gross_total": 0.0,
            "net_total": 0.0,
            "paid": 0.0,
            "payment_methods": set()
        })

        emp_data["gross_total"] += float(entry.gross_amount or 0)
        emp_data["net_total"] += float(entry.net_amount or 0)

    payments_query = select(EmployeePayment).join(
        PayrollEntry, EmployeePayment.payroll_entry_id == PayrollEntry.id
    ).join(
        Contract, Contract.id == PayrollEntry.contract_id
    ).where(
        Contract.owner_id == owner_id,
        PayrollEntry.period_start >= period_start_date,
        PayrollEntry.period_end <= period_end_date
    )

    payments_result = await db.execute(payments_query)
    payments = list(payments_result.scalars().all())

    for payment in payments:
        emp_id = payment.employee_id
        if emp_id in employee_payment_data:
            employee_payment_data[emp_id]["paid"] += float(payment.amount or 0)
            if payment.payment_method:
                employee_payment_data[emp_id]["payment_methods"].add(payment.payment_method)

    employees_list = []
    total_gross = 0.0
    total_net = 0.0
    total_paid = 0.0
    total_remainder = 0.0

    for emp_data in employee_payment_data.values():
        remainder = emp_data["net_total"] - emp_data["paid"]
        emp_data["remainder"] = remainder
        emp_data["payment_methods"] = ", ".join(sorted(emp_data["payment_methods"])) if emp_data["payment_methods"] else ""

        employees_list.append(emp_data)
        total_gross += emp_data["gross_total"]
        total_net += emp_data["net_total"]
        total_paid += emp_data["paid"]
        total_remainder += remainder

    employees_list.sort(key=lambda x: (x["last_name"], x["first_name"]))

    payments_data = {
        "employees": employees_list,
        "total_gross": total_gross,
        "total_net": total_net,
        "total_paid": total_paid,
        "total_remainder": total_remainder
    }

    return entries, report_data, payments_data


@router.get("/payroll", response_class=HTMLResponse, name="owner_payroll_list")
async def owner_payroll_list(
    request: Request,
    current_user: dict = Depends(require_role(["owner", "superadmin"])),
    db: AsyncSession = Depends(get_db_session),
    object_id: Optional[str] = Query(None),
    payment_date: Optional[str] = Query(None),
    employee_id: Optional[str] = Query(None),
    q_employee: Optional[str] = Query(None),
    sort: Optional[str] = Query(None),
    order: Optional[str] = Query("asc"),
    page: int = Query(1, ge=1),
    per_page: int = Query(25, ge=1, le=100),
    view: str = Query("summary"),
    show_inactive: bool = Query(True)  # Всегда включено по умолчанию
):
    """Список всех начислений."""
    try:
        payroll_service = PayrollService(db)
        
        # Получить внутренний ID владельца
        owner_id = await get_user_id_from_current_user(current_user, db)
        if not owner_id:
            raise HTTPException(status_code=403, detail="Пользователь не найден")
        
        # Фильтр по дате выплаты
        payment_date_obj = None
        if payment_date:
            try:
                payment_date_obj = date.fromisoformat(payment_date)
            except Exception:
                raise HTTPException(status_code=400, detail="Неверный формат даты выплаты. Используйте YYYY-MM-DD")

        # Получить объекты владельца
        objects_query = select(Object).where(Object.owner_id == owner_id, Object.is_active == True).order_by(Object.name)
        objects_result = await db.execute(objects_query)
        objects = objects_result.scalars().all()

        # Нормализация object_id: пустое значение → None
        object_id_int = None
        if object_id is not None and str(object_id).strip() != "":
            try:
                object_id_int = int(object_id)
            except ValueError:
                raise HTTPException(status_code=400, detail="object_id должен быть числом")
        
        # Нормализация employee_id: пустое значение → None
        employee_id_int = None
        if employee_id is not None and str(employee_id).strip() != "":
            try:
                employee_id_int = int(employee_id)
            except ValueError:
                raise HTTPException(status_code=400, detail="employee_id должен быть числом")

        # Получить ВСЕХ сотрудников владельца (включая уволенных)
        # для возможности просмотра и создания начислений задним числом
        all_emps_query = select(User).join(Contract, Contract.employee_id == User.id).where(
            Contract.owner_id == owner_id
        ).distinct()
        all_emps_result = await db.execute(all_emps_query)
        employees_all = all_emps_result.scalars().all()
        
        # Получить все начисления владельца за последние 2 года (широкий период для фильтрации по дате выплаты)
        # Если payment_date указана, будем фильтровать по ней после вычисления дат выплат
        wide_period_start = date.today() - timedelta(days=730)  # 2 года назад
        wide_period_end = date.today() + timedelta(days=365)  # 1 год вперед
        
        # Получить начисления для всех сотрудников (включая неактивных)
        # Используем прямой запрос с selectinload для загрузки contract и employee
        from sqlalchemy.orm import selectinload
        entries_query = select(PayrollEntry).options(
            selectinload(PayrollEntry.contract),
            selectinload(PayrollEntry.employee),
            selectinload(PayrollEntry.object_)
        ).join(
            Contract, PayrollEntry.contract_id == Contract.id
        ).where(
            Contract.owner_id == owner_id,
            PayrollEntry.period_start >= wide_period_start,
            PayrollEntry.period_end <= wide_period_end
        )
        entries_result = await db.execute(entries_query)
        entries = list(entries_result.scalars().all())
        
        # Вычислить даты выплат для всех начислений и отфильтровать по payment_date
        entries_filtered = []
        entries_payment_dates_map = {}  # Кэш для дат выплат
        for e in entries:
            entry_payment_date = await calculate_payment_date_for_entry(e, db)
            entries_payment_dates_map[e.id] = entry_payment_date
            # Если payment_date указана, фильтруем только начисления с этой датой выплаты
            if payment_date_obj:
                if entry_payment_date and entry_payment_date == payment_date_obj:
                    entries_filtered.append(e)
            else:
                # Если payment_date не указана, включаем все начисления
                entries_filtered.append(e)
        
        # Использовать отфильтрованные начисления
        entries = entries_filtered
        
        # Фильтрация по объекту (если указан)
        if object_id_int:
            entries = [e for e in entries if e.object_id == object_id_int]
        
        # Фильтрация по сотруднику (если указан)
        if employee_id_int:
            entries = [e for e in entries if e.employee_id == employee_id_int]
        
        # Фильтрация по текстовому поиску сотрудника (клиентский фильтр, но применяем и на сервере для пагинации)
        if q_employee and q_employee.strip():
            q_employee_lower = q_employee.strip().lower()
            entries = [
                e for e in entries
                if e.employee and (
                    (e.employee.last_name or "").lower().startswith(q_employee_lower) or
                    (e.employee.first_name or "").lower().startswith(q_employee_lower) or
                    f"{(e.employee.last_name or '')} {(e.employee.first_name or '')}".strip().lower().startswith(q_employee_lower)
                )
            ]
        
        # Сортировка
        sort_key = None
        if sort == "id":
            sort_key = lambda e: e.id
        elif sort == "employee":
            sort_key = lambda e: (
                (e.employee.last_name or "").lower() if e.employee else "",
                (e.employee.first_name or "").lower() if e.employee else ""
            )
        elif sort == "period":
            sort_key = lambda e: e.period_end
        else:
            # По умолчанию - по дате периода (последние сначала)
            sort_key = lambda e: e.period_end
        
        reverse_order = order == "desc"
        entries.sort(key=sort_key, reverse=reverse_order)
        
        # Пагинация
        total = len(entries)
        pages = (total + per_page - 1) // per_page if total > 0 else 0
        start_idx = (page - 1) * per_page
        end_idx = start_idx + per_page
        paginated_entries = entries[start_idx:end_idx]
        
        # Загрузить выплаты для всех начислений
        from domain.entities.employee_payment import EmployeePayment
        entry_ids = [e.id for e in paginated_entries]
        payments_map = {}
        if entry_ids:
            payments_query = select(EmployeePayment).where(
                EmployeePayment.payroll_entry_id.in_(entry_ids)
            )
            payments_result = await db.execute(payments_query)
            payments_list = payments_result.scalars().all()
            for payment in payments_list:
                if payment.payroll_entry_id not in payments_map:
                    payments_map[payment.payroll_entry_id] = []
                payments_map[payment.payroll_entry_id].append(payment)
        
        # Вычислить даты выплат для каждого начисления (для пагинированных)
        entries_with_payment_dates = []
        for e in paginated_entries:
            # Использовать кэшированную дату выплаты
            entry_payment_date = entries_payment_dates_map.get(e.id)
            if entry_payment_date is None:
                entry_payment_date = await calculate_payment_date_for_entry(e, db)
                entries_payment_dates_map[e.id] = entry_payment_date
            
            has_payments = e.id in payments_map and len(payments_map[e.id]) > 0
            # Найти выплату со статусом pending (если есть)
            pending_payment = None
            payments_added = 0.0  # Сумма всех добавленных выплат
            payments_completed = 0.0  # Сумма подтвержденных выплат
            if e.id in payments_map:
                for payment in payments_map[e.id]:
                    if payment.status == 'pending':
                        pending_payment = payment
                    # Суммируем все выплаты (добавленные)
                    payments_added += float(payment.amount or 0)
                    # Суммируем только подтвержденные выплаты
                    if payment.status == 'completed':
                        payments_completed += float(payment.amount or 0)
            entries_with_payment_dates.append({
                "entry": e,
                "payment_date": entry_payment_date,
                "has_payments": has_payments,
                "pending_payment": pending_payment,  # Выплата со статусом pending (если есть)
                "payments_added": payments_added,  # Сумма всех добавленных выплат
                "payments_completed": payments_completed  # Сумма подтвержденных выплат
            })
        
        # Подготовка фильтров для шаблона
        filters = {
            "object_id": object_id_int,
            "payment_date": payment_date or "",
            "employee_id": employee_id_int,
            "q_employee": q_employee or ""
        }
        
        # Подготовка сортировки для шаблона (список начислений)
        sort_info = {
            "field": sort if view == "entries" and sort else "period",
            "order": order or "desc"
        }

        # Подготовка данных для вкладки "Сводка"
        employee_objs_map: Dict[int, User] = {emp.id: emp for emp in employees_all}
        summary_map: Dict[int, Dict[str, Any]] = {}

        for emp in employees_all:
            summary_map[emp.id] = {
                "employee_id": emp.id,
                "employee": emp,
                "entries_count": 0,
                "total_amount": 0.0,
                "latest_entry": None,
                "latest_entry_id": None,
                "is_active": False,
            }

        from shared.services.contract_validation_service import is_contract_terminated_for_payroll
        
        contracts_query = select(Contract).where(Contract.owner_id == owner_id)
        contracts_result = await db.execute(contracts_query)
        contracts = contracts_result.scalars().all()
        
        active_employee_ids: set[int] = set()
        for contract in contracts:
            emp_id = int(contract.employee_id)
            if emp_id not in summary_map:
                summary_map[emp_id] = {
                    "employee_id": emp_id,
                    "employee": None,
                    "entries_count": 0,
                    "total_amount": 0.0,
                    "latest_entry": None,
                    "latest_entry_id": None,
                    "is_active": False,
                }
            # Контракт активен для работы, если не уволен для расчётного листа
            if not is_contract_terminated_for_payroll(contract):
                active_employee_ids.add(emp_id)

        for entry_obj in entries:
            emp_id = entry_obj.employee_id
            if emp_id is None:
                continue
            emp_id = int(emp_id)
            if emp_id not in summary_map:
                summary_map[emp_id] = {
                    "employee_id": emp_id,
                    "employee": entry_obj.employee if entry_obj.employee else None,
                    "entries_count": 0,
                    "total_amount": 0.0,
                    "latest_entry": None,
                    "latest_entry_id": None,
                    "is_active": False,
                    "hourly_rate": None,
                    "contract_name": None,
                }
            row = summary_map[emp_id]
            row["entries_count"] += 1
            row["total_amount"] += float(entry_obj.net_amount or 0)
            # Получаем ставку и название договора из начисления
            if entry_obj.contract:
                row["contract_name"] = entry_obj.contract.contract_number or f"Договор #{entry_obj.contract.id}"
            if entry_obj.hourly_rate:
                row["hourly_rate"] = float(entry_obj.hourly_rate)
            if row["latest_entry"] is None or (entry_obj.id and entry_obj.id > (row["latest_entry_id"] or 0)):
                row["latest_entry"] = entry_obj
                row["latest_entry_id"] = entry_obj.id
                if entry_obj.employee:
                    row["employee"] = entry_obj.employee

        missing_employee_ids = [emp_id for emp_id, row in summary_map.items() if row["employee"] is None]
        if missing_employee_ids:
            users_query = select(User).where(User.id.in_(missing_employee_ids))
            users_result = await db.execute(users_query)
            for user in users_result.scalars().all():
                if user.id in summary_map:
                    summary_map[user.id]["employee"] = user

        for emp_id, row in summary_map.items():
            row["is_active"] = emp_id in active_employee_ids
            row["total_amount"] = float(row["total_amount"])

        summary_rows = list(summary_map.values())
        if not show_inactive:
            summary_rows = [row for row in summary_rows if row["is_active"]]

        # Сортировка сводки
        allowed_summary_fields = {"latest_period", "employee", "entries_count", "total_amount", "rate"}
        summary_sort_field = (sort or "latest_period").lower() if view == "summary" else "latest_period"
        if summary_sort_field not in allowed_summary_fields:
            summary_sort_field = "latest_period"
        summary_sort_order = order.lower() if (view == "summary" and sort) else "desc"
        if summary_sort_order not in {"asc", "desc"}:
            summary_sort_order = "desc"

        def summary_sort_key(row: Dict[str, Any]):
            if summary_sort_field == "employee":
                employee = row.get("employee")
                last_name = (employee.last_name if employee else "") or ""
                first_name = (employee.first_name if employee else "") or ""
                return (last_name.lower(), first_name.lower(), row["employee_id"])
            if summary_sort_field == "entries_count":
                return (row["entries_count"], row["employee_id"])
            if summary_sort_field == "total_amount":
                return (row["total_amount"], row["employee_id"])
            if summary_sort_field == "rate":
                return (row.get("hourly_rate") or 0.0, row["employee_id"])
            latest_entry = row.get("latest_entry")
            if latest_entry:
                return (latest_entry.period_end or date.min, latest_entry.id or 0)
            return (date.min, row["employee_id"])

        summary_rows.sort(key=summary_sort_key, reverse=(summary_sort_order == "desc"))

        summary_total = len(summary_rows)
        summary_pages = (summary_total + per_page - 1) // per_page if summary_total else 0
        summary_page = page if view == "summary" else 1
        if summary_pages and summary_page > summary_pages:
            summary_page = summary_pages
        if summary_page < 1:
            summary_page = 1 if summary_pages else 0
        summary_start_idx = (summary_page - 1) * per_page if summary_page else 0
        summary_end_idx = summary_start_idx + per_page if summary_page else 0
        summary_paginated_rows = summary_rows[summary_start_idx:summary_end_idx] if summary_rows else []

        summary_filters = {
            "object_id": object_id_int,
            "payment_date": payment_date or "",
            "employee_id": employee_id_int,
            "show_inactive": show_inactive,
        }
        summary_sort_info = {"field": summary_sort_field, "order": summary_sort_order}
        summary_pagination = {
            "page": summary_page,
            "per_page": per_page,
            "total": summary_total,
            "pages": summary_pages,
        }

        base_nav_params = {}
        if object_id_int is not None:
            base_nav_params["object_id"] = object_id_int
        if payment_date:
            base_nav_params["payment_date"] = payment_date
        if employee_id_int is not None:
            base_nav_params["employee_id"] = employee_id_int

        summary_nav_params = dict(base_nav_params)
        summary_nav_params["view"] = "summary"
        summary_nav_params["per_page"] = per_page
        summary_nav_params["page"] = 1
        if show_inactive:
            summary_nav_params["show_inactive"] = "1"
        summary_nav_params["sort"] = summary_sort_field
        summary_nav_params["order"] = summary_sort_order
        summary_nav_query = urlencode(summary_nav_params)

        entries_nav_params = dict(base_nav_params)
        entries_nav_params["view"] = "entries"
        entries_nav_params["per_page"] = per_page
        entries_nav_params["page"] = page
        if show_inactive:
            entries_nav_params["show_inactive"] = "1"
        if q_employee:
            entries_nav_params["q_employee"] = q_employee
        if view == "entries" and sort:
            entries_nav_params["sort"] = sort
            entries_nav_params["order"] = order
        entries_nav_query = urlencode(entries_nav_params)

        return templates.TemplateResponse(
            "owner/payroll/list.html",
            {
                "request": request,
                "current_user": current_user,
                "view": view,
                "entries": entries_with_payment_dates,
                "employees_all": employees_all,  # Все сотрудники для фильтра
                "objects": objects,
                "filters": filters,
                "sort": sort_info,
                "pagination": {
                    "page": page,
                    "per_page": per_page,
                    "total": total,
                    "pages": pages
                },
                "payment_date": payment_date or "",
                "employee_id": employee_id_int,
                "summary_rows": summary_paginated_rows,
                "summary_filters": summary_filters,
                "summary_sort": summary_sort_info,
                "summary_pagination": summary_pagination,
                "summary_nav_query": summary_nav_query,
                "entries_nav_query": entries_nav_query,
                "summary_show_inactive": show_inactive
            }
        )
        
    except Exception as e:
        logger.error(f"Error loading payroll list: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка загрузки начислений: {str(e)}")


@router.get("/payroll/report", response_class=HTMLResponse, name="owner_payroll_report")
async def owner_payroll_report(
    request: Request,
    _: None = Depends(require_role(["owner", "superadmin"])),
    db: AsyncSession = Depends(get_db_session),
    period_start: str = None,
    period_end: str = None
):
    """Отчёт по начислениям с группировкой по объектам."""
    try:
        # Получить текущего пользователя из request
        current_user = await get_current_user(request)
        if not current_user:
            raise HTTPException(status_code=401, detail="Требуется авторизация")
        
        # Получить внутренний ID владельца
        owner_id = await get_user_id_from_current_user(current_user, db)
        if not owner_id:
            raise HTTPException(status_code=403, detail="Пользователь не найден")
        
        # Парсинг дат
        if not period_start or not period_end:
            raise HTTPException(status_code=400, detail="Укажите период")
        
        period_start_date = date.fromisoformat(period_start)
        period_end_date = date.fromisoformat(period_end)
        
        # Получить все начисления за период
        entries, report_data, payments_data = await _fetch_owner_payroll_report_data(
            owner_id=owner_id,
            period_start_date=period_start_date,
            period_end_date=period_end_date,
            db=db
        )
        
        if not entries:
            return templates.TemplateResponse(
                "owner/payroll/report.html",
                {
                    "request": request,
                    "current_user": current_user,
                    "period_start": period_start,
                    "period_end": period_end,
                    "report_data": None,
                    "error": "Нет начислений за выбранный период"
                }
            )
        return templates.TemplateResponse(
            "owner/payroll/report.html",
            {
                "request": request,
                "current_user": current_user,
                "period_start": period_start,
                "period_end": period_end,
                "report_data": report_data,
                "payments_data": payments_data
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Error generating payroll report: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка формирования отчёта: {str(e)}")


@router.get("/payroll/report/export", name="owner_payroll_report_export")
async def owner_payroll_report_export(
    request: Request,
    _: None = Depends(require_role(["owner", "superadmin"])),
    db: AsyncSession = Depends(get_db_session),
    period_start: str = None,
    period_end: str = None
):
    """Экспорт отчёта по начислениям и выплатам в Excel."""
    try:
        current_user = await get_current_user(request)
        if not current_user:
            raise HTTPException(status_code=401, detail="Требуется авторизация")

        owner_id = await get_user_id_from_current_user(current_user, db)
        if not owner_id:
            raise HTTPException(status_code=403, detail="Пользователь не найден")

        if not period_start or not period_end:
            raise HTTPException(status_code=400, detail="Укажите период")

        period_start_date = date.fromisoformat(period_start)
        period_end_date = date.fromisoformat(period_end)

        entries, report_data, payments_data = await _fetch_owner_payroll_report_data(
            owner_id=owner_id,
            period_start_date=period_start_date,
            period_end_date=period_end_date,
            db=db
        )

        if not entries or not report_data:
            raise HTTPException(status_code=404, detail="Нет данных за выбранный период")

        wb = Workbook()
        ws_accruals = wb.active
        ws_accruals.title = "Отчет по начислениям"

        accruals_headers = [
            "Объект",
            "Сотрудник",
            "Статус",
            "Смен",
            "Часов",
            "Ставка",
            "Доплачено",
            "Удержано",
            "Сумма"
        ]
        ws_accruals.append(accruals_headers)

        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="E5E5E5", end_color="E5E5E5", fill_type="solid")

        for cell in ws_accruals[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for obj in report_data["objects"]:
            ws_accruals.append([obj["object_name"], "", "", "", "", "", "", "", ""])
            for row in obj["rows"]:
                ws_accruals.append([
                    "",
                    f"{row['last_name']} {row['first_name']}".strip(),
                    row["status"],
                    row["shifts_count"],
                    row["hours"],
                    row["rate"],
                    row["bonus"],
                    row["penalty"],
                    row["total"],
                ])
            ws_accruals.append(["", "", "", "", "", "", "", "Итого по объекту", obj["subtotal"]])
            ws_accruals.append([])

        if report_data["multi_object_employees"]:
            ws_accruals.append(["Сотрудники на нескольких объектах", "", "", "", "", "", "", "", ""])
            for row in report_data["multi_object_employees"]:
                ws_accruals.append([
                    row["object_name"],
                    f"{row['last_name']} {row['first_name']}".strip(),
                    row["status"],
                    row["shifts_count"],
                    row["hours"],
                    row["rate"],
                    row["bonus"],
                    row["penalty"],
                    row["total"],
                ])
            ws_accruals.append([])

        ws_accruals.append(["", "", "", "", "", "", "", "ОБЩИЙ ИТОГ", report_data["grand_total"]])

        for column in ws_accruals.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws_accruals.column_dimensions[column_letter].width = min(max_length + 2, 60)

        ws_payments = wb.create_sheet(title="Отчет по выплатам")
        payments_headers = ["Сотрудник", "К выплате", "Выплачено", "Способы оплаты", "Остаток"]
        ws_payments.append(payments_headers)

        for cell in ws_payments[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        if payments_data and payments_data["employees"]:
            for row in payments_data["employees"]:
                ws_payments.append([
                    f"{row['last_name']} {row['first_name']}".strip(),
                    row["net_total"],
                    row["paid"],
                    row["payment_methods"],
                    row["remainder"]
                ])

            ws_payments.append([
                "ИТОГО:",
                payments_data["total_net"],
                payments_data["total_paid"],
                "",
                payments_data["total_remainder"]
            ])
        else:
            ws_payments.append(["Нет данных", "", "", "", ""])

        for column in ws_payments.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws_payments.column_dimensions[column_letter].width = min(max_length + 2, 60)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"payroll_report_{period_start}_{period_end}.xlsx"
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Error exporting payroll report: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка экспорта отчёта: {str(e)}")


@router.get("/payroll/{entry_id}", response_class=HTMLResponse, name="owner_payroll_detail")
async def owner_payroll_detail(
    request: Request,
    entry_id: int,
    current_user: dict = Depends(require_role(["owner", "superadmin"])),
    db: AsyncSession = Depends(get_db_session)
):
    """Детальная страница начисления."""
    try:
        payroll_service = PayrollService(db)
        entry = await payroll_service.get_payroll_entry_by_id(entry_id)
        
        if not entry:
            raise HTTPException(status_code=404, detail="Начисление не найдено")
        
        # Получить внутренний ID владельца
        owner_id = await get_user_id_from_current_user(current_user, db)
        if not owner_id:
            raise HTTPException(status_code=403, detail="Пользователь не найден")
        
        # Проверить доступ: владелец должен быть owner этого сотрудника
        query = select(Contract).where(
            Contract.employee_id == entry.employee_id,
            Contract.owner_id == owner_id
        )
        result = await db.execute(query)
        owner_contracts = result.scalars().all()
        
        if not owner_contracts:
            raise HTTPException(status_code=403, detail="Доступ запрещен")
        
        # Получить связанные adjustments (для старого шаблона - как deductions, bonuses)
        from domain.entities.payroll_adjustment import PayrollAdjustment
        from sqlalchemy.orm import selectinload
        adjustments_query = select(PayrollAdjustment).where(
            PayrollAdjustment.payroll_entry_id == entry_id
        ).options(
            selectinload(PayrollAdjustment.creator),
            selectinload(PayrollAdjustment.updater)
        ).order_by(PayrollAdjustment.created_at)
        adjustments_result = await db.execute(adjustments_query)
        adjustments = adjustments_result.scalars().all()
        
        # Нормализовать timestamp в edit_history (конвертировать строки в datetime)
        # И загрузить информацию о пользователях
        from datetime import timezone
        user_ids_to_load = set()
        for adj in adjustments:
            if adj.edit_history:
                for change in adj.edit_history:
                    if isinstance(change.get('timestamp'), str):
                        try:
                            dt = datetime.fromisoformat(change['timestamp'])
                            # Если datetime naive, делаем его UTC aware
                            if dt.tzinfo is None:
                                dt = dt.replace(tzinfo=timezone.utc)
                            change['timestamp'] = dt
                        except (ValueError, AttributeError):
                            pass  # Оставляем как есть если не удалось распарсить
                    
                    # Собираем ID пользователей для загрузки
                    if change.get('user_id'):
                        user_ids_to_load.add(change['user_id'])
        
        # Загрузить пользователей
        users_map = {}
        if user_ids_to_load:
            users_query = select(User).where(User.id.in_(list(user_ids_to_load)))
            users_result = await db.execute(users_query)
            users_list = users_result.scalars().all()
            users_map = {u.id: f"{u.last_name} {u.first_name}" for u in users_list}
        
        # Добавить user_name в edit_history
        for adj in adjustments:
            if adj.edit_history:
                for change in adj.edit_history:
                    uid = change.get('user_id')
                    if uid:
                        change['user_name'] = users_map.get(uid, f"ID: {uid}")
        
        # Разделить на deductions и bonuses для совместимости со старым шаблоном
        deductions = [adj for adj in adjustments if float(adj.amount) < 0]
        bonuses = [adj for adj in adjustments if float(adj.amount) > 0 and adj.adjustment_type != 'shift_base']
        
        # Получить выплаты (payments)
        from domain.entities.employee_payment import EmployeePayment
        payments_query = select(EmployeePayment).where(
            EmployeePayment.payroll_entry_id == entry_id
        ).order_by(EmployeePayment.payment_date)
        payments_result = await db.execute(payments_query)
        payments = payments_result.scalars().all()
        has_payments = len(payments) > 0

        # Список начислений сотрудника (для навигации, аналог менеджера)
        employee_entries = await payroll_service.get_payroll_entries_by_employee(
            employee_id=entry.employee_id,
            limit=500,
            owner_id=owner_id
        )
        employee_entries.sort(key=lambda e: (e.period_end, e.id), reverse=True)
        
        # Вычислить даты выплат для каждого начисления
        employee_entries_with_payment_dates = []
        for e in employee_entries:
            payment_date = await calculate_payment_date_for_entry(e, db)
            employee_entries_with_payment_dates.append({
                "entry": e,
                "payment_date": payment_date
            })

        # Определить активную запись (подсветка)
        origin = request.query_params.get("origin")
        selected_entry_id = request.query_params.get("selected_entry_id")
        action = request.query_params.get("action")  # Параметр для открытия модального окна
        try:
            selected_entry_id_int = int(selected_entry_id) if selected_entry_id else None
        except ValueError:
            selected_entry_id_int = None

        if origin == "entries" and selected_entry_id_int:
            active_entry_id = selected_entry_id_int
        elif origin == "summary" and employee_entries:
            active_entry_id = employee_entries[0].id
        else:
            active_entry_id = entry.id
        
        return templates.TemplateResponse(
            "owner/payroll/detail.html",
            {
                "request": request,
                "current_user": current_user,
                "entry": entry,
                "deductions": deductions,
                "bonuses": bonuses,
                "payments": payments,
                "employee_entries": employee_entries_with_payment_dates,
                "active_entry_id": active_entry_id,
                "has_payments": has_payments,
                "origin": origin,
                "action": action,  # Передаем action в шаблон
                "today": date.today()
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error loading payroll detail: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка загрузки начисления: {str(e)}")


@router.post("/payroll/{entry_id}/add-deduction", name="owner_payroll_add_deduction")
async def owner_payroll_add_deduction(
    request: Request,
    entry_id: int,
    current_user: dict = Depends(require_role(["owner", "superadmin"])),
    db: AsyncSession = Depends(get_db_session),
    deduction_type: str = Form(...),
    amount: float = Form(...),
    description: str = Form(...),
    adjustment_date: Optional[str] = Form(None),
    return_url: Optional[str] = Form(None)
):
    """Добавить удержание к начислению."""
    try:
        payroll_service = PayrollService(db)
        adjustment_service = PayrollAdjustmentService(db)
        
        # Получить внутренний ID владельца
        owner_id = await get_user_id_from_current_user(current_user, db)
        if not owner_id:
            raise HTTPException(status_code=403, detail="Пользователь не найден")
        
        # Проверить доступ
        entry = await payroll_service.get_payroll_entry_by_id(entry_id)
        if not entry:
            raise HTTPException(status_code=404, detail="Начисление не найдено")
        
        query = select(Contract).where(
            Contract.employee_id == entry.employee_id,
            Contract.owner_id == owner_id
        )
        result = await db.execute(query)
        owner_contracts = result.scalars().all()
        
        if not owner_contracts:
            raise HTTPException(status_code=403, detail="Доступ запрещен")
        
        # Парсинг даты
        adjustment_date_obj = None
        if adjustment_date:
            try:
                adjustment_date_obj = date.fromisoformat(adjustment_date)
            except ValueError:
                pass  # Используем текущую дату
        
        # Добавить удержание через PayrollAdjustmentService
        adjustment = await adjustment_service.create_manual_adjustment(
            employee_id=entry.employee_id,
            amount=-abs(Decimal(str(amount))),  # Отрицательное для удержания
            adjustment_type='manual_deduction',
            description=description,
            created_by=owner_id,
            object_id=entry.object_id,
            adjustment_date=adjustment_date_obj,
            details={'deduction_type': deduction_type}  # Сохраняем тип удержания
        )
        
        # Привязать к payroll_entry
        adjustment.payroll_entry_id = entry_id
        adjustment.is_applied = True
        
        # Пересчитать суммы в entry (все в Decimal)
        entry.total_deductions = Decimal(str(entry.total_deductions)) + abs(Decimal(str(amount)))
        entry.net_amount = Decimal(str(entry.gross_amount)) + Decimal(str(entry.total_bonuses)) - Decimal(str(entry.total_deductions))
        
        await db.commit()
        
        logger.info(
            f"Deduction added by owner",
            entry_id=entry_id,
            amount=amount,
            type=deduction_type
        )
        
        # Редирект на сохраненный URL или на страницу деталей
        redirect_url = return_url if return_url else f"/owner/payroll/{entry_id}"
        return RedirectResponse(
            url=redirect_url,
            status_code=status.HTTP_302_FOUND
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error adding deduction: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка добавления удержания: {str(e)}")


@router.post("/payroll/{entry_id}/add-bonus", name="owner_payroll_add_bonus")
async def owner_payroll_add_bonus(
    request: Request,
    entry_id: int,
    current_user: dict = Depends(require_role(["owner", "superadmin"])),
    db: AsyncSession = Depends(get_db_session),
    bonus_type: str = Form(...),
    amount: float = Form(...),
    description: str = Form(...),
    adjustment_date: Optional[str] = Form(None),
    return_url: Optional[str] = Form(None)
):
    """Добавить доплату к начислению."""
    try:
        
        payroll_service = PayrollService(db)
        adjustment_service = PayrollAdjustmentService(db)
        
        # Получить внутренний ID владельца
        owner_id = await get_user_id_from_current_user(current_user, db)
        if not owner_id:
            raise HTTPException(status_code=403, detail="Пользователь не найден")
        
        # Проверить доступ
        entry = await payroll_service.get_payroll_entry_by_id(entry_id)
        if not entry:
            raise HTTPException(status_code=404, detail="Начисление не найдено")
        
        query = select(Contract).where(
            Contract.employee_id == entry.employee_id,
            Contract.owner_id == owner_id
        )
        result = await db.execute(query)
        owner_contracts = result.scalars().all()
        
        if not owner_contracts:
            raise HTTPException(status_code=403, detail="Доступ запрещен")
        
        # Парсинг даты
        adjustment_date_obj = None
        if adjustment_date:
            try:
                adjustment_date_obj = date.fromisoformat(adjustment_date)
            except ValueError:
                pass  # Используем текущую дату
        
        # Добавить доплату через PayrollAdjustmentService
        adjustment = await adjustment_service.create_manual_adjustment(
            employee_id=entry.employee_id,
            amount=abs(Decimal(str(amount))),  # Положительное для бонуса
            adjustment_type='manual_bonus',
            description=description,
            created_by=owner_id,
            object_id=entry.object_id,
            adjustment_date=adjustment_date_obj,
            details={'bonus_type': bonus_type}  # Сохраняем тип доплаты
        )
        
        # Привязать к payroll_entry
        adjustment.payroll_entry_id = entry_id
        adjustment.is_applied = True
        
        # Пересчитать суммы в entry (все в Decimal)
        entry.total_bonuses = Decimal(str(entry.total_bonuses)) + abs(Decimal(str(amount)))
        entry.net_amount = Decimal(str(entry.gross_amount)) + Decimal(str(entry.total_bonuses)) - Decimal(str(entry.total_deductions))
        
        await db.commit()
        
        logger.info(
            f"Bonus added by owner",
            entry_id=entry_id,
            amount=amount,
            type=bonus_type
        )
        
        # Редирект на сохраненный URL или на страницу деталей
        redirect_url = return_url if return_url else f"/owner/payroll/{entry_id}"
        return RedirectResponse(
            url=redirect_url,
            status_code=status.HTTP_302_FOUND
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error adding bonus: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка добавления доплаты: {str(e)}")


@router.post("/payroll/{entry_id}/payments/{payment_id}/complete", name="owner_payment_complete")
async def complete_payment(
    request: Request,
    entry_id: int,
    payment_id: int,
    current_user: dict = Depends(require_role(["owner", "superadmin"])),
    db: AsyncSession = Depends(get_db_session),
    confirmation_code: Optional[str] = Form(None),
    return_url: Optional[str] = Form(None)
):
    """Подтвердить выплату."""
    try:
        payroll_service = PayrollService(db)
        
        # Получить внутренний ID владельца
        owner_id = await get_user_id_from_current_user(current_user, db)
        if not owner_id:
            raise HTTPException(status_code=403, detail="Пользователь не найден")
        
        # Проверить доступ к entry
        entry = await payroll_service.get_payroll_entry_by_id(entry_id)
        if not entry:
            raise HTTPException(status_code=404, detail="Начисление не найдено")
        
        query = select(Contract).where(
            Contract.employee_id == entry.employee_id,
            Contract.owner_id == owner_id
        )
        result = await db.execute(query)
        owner_contracts = result.scalars().all()
        
        if not owner_contracts:
            raise HTTPException(status_code=403, detail="Доступ запрещен")
        
        # Подтвердить выплату
        await payroll_service.mark_payment_completed(
            payment_id=payment_id,
            confirmation_code=confirmation_code
        )
        
        logger.info(
            f"Payment completed by owner",
            payment_id=payment_id,
            entry_id=entry_id,
            confirmation_code=confirmation_code
        )
        
        # Редирект на сохраненный URL или на страницу деталей
        redirect_url = return_url if return_url else f"/owner/payroll/{entry_id}"
        return RedirectResponse(
            url=redirect_url,
            status_code=status.HTTP_302_FOUND
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error completing payment: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка подтверждения выплаты: {str(e)}")


@router.post("/payroll/{entry_id}/create-payment", name="owner_payroll_create_payment")
async def owner_payroll_create_payment(
    request: Request,
    entry_id: int,
    current_user: dict = Depends(require_role(["owner", "superadmin"])),
    db: AsyncSession = Depends(get_db_session),
    amount: float = Form(...),
    payment_date: str = Form(...),
    payment_method: str = Form(...),
    notes: Optional[str] = Form(None),
    return_url: Optional[str] = Form(None)
):
    """Создать выплату для начисления."""
    try:
        payroll_service = PayrollService(db)
        
        # Получить внутренний ID владельца
        owner_id = await get_user_id_from_current_user(current_user, db)
        if not owner_id:
            raise HTTPException(status_code=403, detail="Пользователь не найден")
        
        # Проверить доступ
        entry = await payroll_service.get_payroll_entry_by_id(entry_id)
        if not entry:
            raise HTTPException(status_code=404, detail="Начисление не найдено")
        
        query = select(Contract).where(
            Contract.employee_id == entry.employee_id,
            Contract.owner_id == owner_id
        )
        result = await db.execute(query)
        owner_contracts = result.scalars().all()
        
        if not owner_contracts:
            raise HTTPException(status_code=403, detail="Доступ запрещен")
        
        # Создать выплату
        await payroll_service.create_payment(
            payroll_entry_id=entry_id,
            amount=Decimal(str(amount)),
            payment_date=date.fromisoformat(payment_date),
            payment_method=payment_method,
            created_by_id=owner_id,
            notes=notes
        )
        
        logger.info(
            f"Payment created by owner",
            entry_id=entry_id,
            amount=amount,
            method=payment_method
        )
        
        # Редирект на сохраненный URL или на страницу деталей
        redirect_url = return_url if return_url else f"/owner/payroll/{entry_id}"
        return RedirectResponse(
            url=redirect_url,
            status_code=status.HTTP_302_FOUND
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating payment: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка создания выплаты: {str(e)}")


@router.post("/payroll/manual-recalculate", response_class=JSONResponse, name="owner_payroll_manual_recalculate")
async def owner_payroll_manual_recalculate(
    request: Request,
    target_date: str = Form(...),
    current_user: dict = Depends(require_role(["owner", "superadmin"])),
    db: AsyncSession = Depends(get_db_session)
):
    """Ручной запуск пересчёта выплат на указанную дату (идемпотентно)."""
    try:
        # Парсинг даты
        try:
            target_date_obj = date.fromisoformat(target_date)
        except ValueError:
            return JSONResponse(
                status_code=400,
                content={"success": False, "error": "Неверный формат даты"}
            )
        
        # Получить внутренний ID владельца
        owner_id = await get_user_id_from_current_user(current_user, db)
        if not owner_id:
            return JSONResponse(
                status_code=403,
                content={"success": False, "error": "Пользователь не найден"}
            )
        
        logger.info(
            f"Manual payroll recalculation started",
            owner_id=owner_id,
            target_date=target_date_obj.isoformat()
        )

        generation_service = PayrollGenerationService(db)
        
        # Импорт функции для расчёта периода

        # Найти все активные payment_schedules (системные + кастомные владельца)
        # Логика такая же как в автоматическом пересчете - берем ВСЕ активные графики,
        # а объекты фильтруем по owner_id
        schedules_query = select(PaymentSchedule).where(
            PaymentSchedule.is_active == True
        )
        schedules_result = await db.execute(schedules_query)
        schedules = schedules_result.scalars().all()
        
        logger.info(f"Found {len(schedules)} active schedules total")
        
        total_entries_created = 0
        total_entries_updated = 0
        total_adjustments_applied = 0
        errors = []
        
        for schedule in schedules:
            try:
                # Проверяем, является ли target_date днём выплаты
                payment_period = await get_payment_period_for_date(schedule, target_date_obj)
                
                if not payment_period:
                    logger.debug(
                        f"Skip schedule {schedule.id}: target date is not a payment day",
                        schedule_id=schedule.id,
                        target_date=target_date_obj.isoformat()
                    )
                    continue
                
                period_start = payment_period['period_start']
                period_end = payment_period['period_end']
                
                logger.info(
                    f"Manual recalculation for schedule {schedule.id}: {schedule.name}",
                    period_start=period_start.isoformat(),
                    period_end=period_end.isoformat()
                )
                
                # Найти объекты через подразделения с учетом наследования (как в автоматической задаче)
                # Функция для рекурсивного поиска всех потомков подразделения
                async def get_all_descendant_unit_ids(unit_id: int) -> List[int]:
                    """Рекурсивно найти все дочерние подразделения"""
                    result = [unit_id]
                    children_query = select(OrgStructureUnit.id).where(
                        OrgStructureUnit.parent_id == unit_id,
                        OrgStructureUnit.is_active == True
                    )
                    children_result = await db.execute(children_query)
                    children_ids = [row[0] for row in children_result.all()]
                    
                    for child_id in children_ids:
                        result.extend(await get_all_descendant_unit_ids(child_id))
                    
                    return result
                
                # Найти подразделения, использующие этот график (с учетом наследования)
                units_with_schedule = []
                
                # Найти все подразделения владельца
                all_units_query = select(OrgStructureUnit).options(
                    selectinload(OrgStructureUnit.parent)
                ).where(
                    OrgStructureUnit.owner_id == owner_id,
                    OrgStructureUnit.is_active == True
                )
                
                all_units_result = await db.execute(all_units_query)
                all_units = all_units_result.scalars().all()
                
                # Проверить каждое подразделение
                for unit in all_units:
                    # Проверить, использует ли подразделение этот график (с учетом наследования)
                    inherited_schedule_id = unit.get_inherited_payment_schedule_id()
                    if inherited_schedule_id == schedule.id:
                        # Добавить само подразделение и все его дочерние подразделения
                        unit_ids = await get_all_descendant_unit_ids(unit.id)
                        units_with_schedule.extend(unit_ids)
                
                # Убрать дубликаты
                units_with_schedule = list(set(units_with_schedule))
                
                # Найти объекты:
                # - с прямой привязкой к графику ИЛИ
                # - принадлежащие подразделению с этим графиком (с учетом дочерних)
                if units_with_schedule:
                    objects_query = select(Object).where(
                        Object.is_active == True,
                        Object.owner_id == owner_id,  # Дополнительная проверка на владельца
                        or_(
                            Object.payment_schedule_id == schedule.id,
                            Object.org_unit_id.in_(units_with_schedule)
                        )
                    )
                else:
                    # Если нет подразделений с этим графиком, искать только объекты с прямой привязкой
                    objects_query = select(Object).where(
                        Object.is_active == True,
                        Object.owner_id == owner_id,
                        Object.payment_schedule_id == schedule.id
                    )
                
                objects_result = await db.execute(objects_query)
                objects = objects_result.scalars().all()
                
                logger.info(
                    f"Found {len(objects)} objects for schedule {schedule.id} (owner {owner_id})",
                    units_with_schedule=len(units_with_schedule)
                )
                
                for obj in objects:
                    try:
                        # Найти контракты:
                        # - активные
                        # - terminated + settlement_policy='schedule'
                        # - terminated + settlement_policy='termination_date' И конец платёжного периода <= termination_date
                        contracts_query = select(Contract).where(
                            and_(
                                Contract.allowed_objects.isnot(None),
                                cast(Contract.allowed_objects, JSONB).op('@>')(cast([obj.id], JSONB)),
                                or_(
                                    Contract.status == 'active',
                                    and_(
                                        Contract.status == 'terminated',
                                        Contract.settlement_policy == 'schedule'
                                    ),
                                    and_(
                                        Contract.status == 'terminated',
                                        Contract.settlement_policy == 'termination_date',
                                        Contract.termination_date.isnot(None),
                                        func.date(period_end) <= Contract.termination_date
                                    )
                                )
                            )
                        )
                        contracts_result = await db.execute(contracts_query)
                        contracts = contracts_result.scalars().all()
                        
                        logger.debug(f"Found {len(contracts)} contracts for object {obj.id}")
                        
                        for contract in contracts:
                            try:
                                # Проверить график выплат контракта
                                effective_payment_schedule_id = None
                                
                                if contract.inherit_payment_schedule:
                                    # Наследуем от подразделения
                                    effective_payment_schedule_id = await get_inherited_payment_schedule_id(contract, db)
                                    logger.debug(
                                        f"Contract {contract.id} inherits payment schedule from org unit",
                                        inherited_schedule_id=effective_payment_schedule_id
                                    )
                                else:
                                    # Используем явно указанный график (ПРИОРИТЕТ!)
                                    effective_payment_schedule_id = contract.payment_schedule_id
                                    logger.debug(
                                        f"Contract {contract.id} has explicit payment schedule",
                                        schedule_id=effective_payment_schedule_id
                                    )
                                
                                # Проверяем, совпадает ли график контракта с текущим графиком
                                if effective_payment_schedule_id and effective_payment_schedule_id != schedule.id:
                                    logger.debug(
                                        f"Skip contract {contract.id}: different payment schedule",
                                        contract_schedule=effective_payment_schedule_id,
                                        current_schedule=schedule.id
                                    )
                                    continue
                                
                                # Если у контракта нет графика (ни явного, ни наследуемого) - пропускаем
                                if not effective_payment_schedule_id:
                                    logger.debug(
                                        f"Skip contract {contract.id}: no payment schedule"
                                    )
                                    continue
                                
                                result = await generation_service.process_contract_period(
                                    contract=contract,
                                    obj=obj,
                                    period_start=period_start,
                                    period_end=period_end,
                                    calculation_date=target_date_obj,
                                    created_by_id=owner_id,
                                    source="manual_recalculate",
                                )
                                total_entries_created += result.created_entries
                                total_entries_updated += result.updated_entries
                                total_adjustments_applied += result.applied_adjustments
                            except Exception as e:
                                error_msg = f"Error processing contract {contract.id}: {e}"
                                logger.error(error_msg)
                                errors.append(error_msg)
                                continue
                    
                    except Exception as e:
                        error_msg = f"Error processing object {obj.id}: {e}"
                        logger.error(error_msg)
                        errors.append(error_msg)
                        continue
            
            except Exception as e:
                error_msg = f"Error processing schedule {schedule.id}: {e}"
                logger.error(error_msg)
                errors.append(error_msg)
                continue
        
        # ===== ОБРАБОТКА ИНДИВИДУАЛЬНЫХ ГРАФИКОВ СОТРУДНИКОВ =====
        # Обрабатываем графики еще раз для индивидуальных договоров
        for schedule in schedules:
            try:
                # Проверяем, является ли target_date днём выплаты
                payment_period = await get_payment_period_for_date(schedule, target_date_obj)
                
                if not payment_period:
                    continue
                
                period_start = payment_period['period_start']
                period_end = payment_period['period_end']
                
                # Найти все contracts с payment_schedule_id=schedule.id и inherit_payment_schedule=false
                # принадлежащие владельцу
                individual_contracts_query = select(Contract).where(
                    and_(
                        Contract.owner_id == owner_id,
                        Contract.payment_schedule_id == schedule.id,
                        Contract.inherit_payment_schedule == False,
                        or_(
                            Contract.status == 'active',
                            and_(
                                Contract.status == 'terminated',
                                Contract.settlement_policy == 'schedule'
                            )
                        )
                    )
                )
                individual_contracts_result = await db.execute(individual_contracts_query)
                individual_contracts = individual_contracts_result.scalars().all()
                
                logger.info(f"Found {len(individual_contracts)} contracts with individual schedule {schedule.id} (owner {owner_id})")
                
                for contract in individual_contracts:
                    try:
                        # Используем первый объект из allowed_objects
                        if not contract.allowed_objects or len(contract.allowed_objects) == 0:
                            logger.warning(f"Contract {contract.id} has no allowed_objects, skipping")
                            continue
                        
                        first_object_id = contract.allowed_objects[0]
                        
                        # Загрузить объект
                        obj_result = await db.execute(
                            select(Object).where(Object.id == first_object_id)
                        )
                        obj = obj_result.scalar_one_or_none()
                        
                        if not obj:
                            logger.warning(f"Object {first_object_id} not found for contract {contract.id}")
                            continue
                        
                        # Проверить, что объект принадлежит владельцу
                        if obj.owner_id != owner_id:
                            logger.debug(f"Skip contract {contract.id}: object {obj.id} belongs to different owner")
                            continue
                        
                        result = await generation_service.process_contract_period(
                            contract=contract,
                            obj=obj,
                            period_start=period_start,
                            period_end=period_end,
                            calculation_date=target_date_obj,
                            created_by_id=owner_id,
                            source="manual_recalculate",
                        )
                        total_entries_created += result.created_entries
                        total_entries_updated += result.updated_entries
                        total_adjustments_applied += result.applied_adjustments
                        
                    except Exception as e:
                        error_msg = f"Error processing individual contract {contract.id}: {e}"
                        logger.error(error_msg)
                        errors.append(error_msg)
                        continue
                        
            except Exception as e:
                error_msg = f"Error processing individual schedules for schedule {schedule.id}: {e}"
                logger.error(error_msg)
                errors.append(error_msg)
                continue
        
        # Commit всех изменений
        await db.commit()
        
        logger.info(
            f"Manual payroll recalculation completed",
            owner_id=owner_id,
            target_date=target_date_obj.isoformat(),
            entries_created=total_entries_created,
            entries_updated=total_entries_updated,
            adjustments_applied=total_adjustments_applied,
            errors_count=len(errors)
        )
        
        return JSONResponse(content={
            "success": True,
            "target_date": target_date_obj.isoformat(),
            "entries_created": total_entries_created,
            "entries_updated": total_entries_updated,
            "adjustments_applied": total_adjustments_applied,
            "errors": errors
        })
        
    except Exception as e:
        logger.error(f"Error in manual payroll recalculation: {e}")
        await db.rollback()
        return JSONResponse(
            status_code=500,
            content={"success": False, "error": str(e)}
        )


@router.get(
    "/payroll/statement/{employee_id}",
    response_class=HTMLResponse,
    name="owner_payroll_statement",
)
async def owner_payroll_statement(
    request: Request,
    employee_id: int,
    current_user: dict = Depends(require_role(["owner", "superadmin"])),
    db: AsyncSession = Depends(get_db_session),
):
    """Страница расчётного листа сотрудника."""
    statement_service = PayrollStatementService(db)
    try:
        owner_id = await get_user_id_from_current_user(current_user, db)
        if not owner_id:
            raise HTTPException(status_code=403, detail="Пользователь не найден")

        statement = await statement_service.generate_statement(
            employee_id=employee_id,
            requested_by_id=owner_id,
            requested_role="owner",
            owner_id=owner_id,
        )
        await db.commit()

        return templates.TemplateResponse(
            "owner/payroll/statement.html",
            {
                "request": request,
                "current_user": current_user,
                "statement": statement,
            },
        )
    except HTTPException:
        await db.rollback()
        raise
    except Exception as e:
        await db.rollback()
        logger.exception(f"Error generating payroll statement: {e}")
        raise HTTPException(status_code=500, detail="Не удалось сформировать расчётный лист")


@router.get(
    "/payroll/statement/{employee_id}/export",
    name="owner_payroll_statement_export",
)
async def owner_payroll_statement_export(
    request: Request,
    employee_id: int,
    current_user: dict = Depends(require_role(["owner", "superadmin"])),
    db: AsyncSession = Depends(get_db_session),
):
    """Экспорт расчётного листа в Excel."""
    statement_service = PayrollStatementService(db)
    try:
        owner_id = await get_user_id_from_current_user(current_user, db)
        if not owner_id:
            raise HTTPException(status_code=403, detail="Пользователь не найден")

        statement = await statement_service.generate_statement(
            employee_id=employee_id,
            requested_by_id=owner_id,
            requested_role="owner",
            owner_id=owner_id,
            log_result=False,
        )
        await db.commit()

        content = build_statement_workbook(statement)
        employee = statement["employee"]
        # Используем только ASCII для имени файла, чтобы избежать проблем с кодировкой
        safe_name = (employee.last_name or "employee").encode("ascii", "ignore").decode("ascii") or "employee"
        filename = f"payroll_statement_{safe_name}_{employee_id}.xlsx"
        # Кодируем имя файла для Content-Disposition заголовка
        encoded_filename = quote(filename, safe="")
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{encoded_filename}"'},
        )
    except HTTPException:
        await db.rollback()
        raise
    except Exception as e:
        await db.rollback()
        logger.exception(f"Error exporting payroll statement: {e}")
        raise HTTPException(status_code=500, detail="Не удалось выгрузить отчёт")


# ==================== EMPLOYEE ROUTES ====================


@router.get("/employee/payroll", response_class=HTMLResponse, name="employee_payroll_list")
async def employee_payroll_list(
    request: Request,
    current_user: dict = Depends(require_employee_or_applicant),
    db: AsyncSession = Depends(get_db_session),
    period_start: Optional[str] = None,
    period_end: Optional[str] = None
):
    """Список начислений для сотрудника."""
    try:
        payroll_service = PayrollService(db)
        
        # Получить внутренний ID сотрудника
        employee_id = await get_user_id_from_current_user(current_user, db)
        if not employee_id:
            raise HTTPException(status_code=403, detail="Пользователь не найден")
        
        # Фильтры по дате
        if not period_start:
            period_start = (date.today() - timedelta(days=90)).isoformat()
        if not period_end:
            period_end = date.today().isoformat()
        
        # Получить начисления текущего сотрудника
        entries = await payroll_service.get_payroll_entries_by_employee(
            employee_id=employee_id,
            period_start=date.fromisoformat(period_start),
            period_end=date.fromisoformat(period_end)
        )
        
        # Рассчитать сводку
        summary = await payroll_service.get_employee_payroll_summary(
            employee_id=employee_id,
            period_start=date.fromisoformat(period_start),
            period_end=date.fromisoformat(period_end)
        )
        
        return templates.TemplateResponse(
            "employee/payroll/list.html",
            {
                "request": request,
                "current_user": current_user,
                "entries": entries,
                "summary": summary,
                "period_start": period_start,
                "period_end": period_end
            }
        )
        
    except Exception as e:
        logger.error(f"Error loading employee payroll list: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка загрузки начислений: {str(e)}")


@router.get("/employee/payroll/{entry_id}", response_class=HTMLResponse, name="employee_payroll_detail")
async def employee_payroll_detail(
    request: Request,
    entry_id: int,
    current_user: dict = Depends(require_employee_or_applicant),
    db: AsyncSession = Depends(get_db_session)
):
    """Детальная страница начисления для сотрудника."""
    try:
        payroll_service = PayrollService(db)
        
        # Получить внутренний ID сотрудника
        employee_id = await get_user_id_from_current_user(current_user, db)
        if not employee_id:
            raise HTTPException(status_code=403, detail="Пользователь не найден")
        
        entry = await payroll_service.get_payroll_entry_by_id(entry_id)
        
        if not entry:
            raise HTTPException(status_code=404, detail="Начисление не найдено")
        
        # Проверить доступ: сотрудник может видеть только свои начисления
        if entry.employee_id != employee_id:
            raise HTTPException(status_code=403, detail="Доступ запрещен")
        
        return templates.TemplateResponse(
            "employee/payroll/detail.html",
            {
                "request": request,
                "current_user": current_user,
                "entry": entry
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error loading employee payroll detail: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка загрузки начисления: {str(e)}")

