"""Роуты владельца для управления инцидентами - использует shared IncidentService."""

from __future__ import annotations

from fastapi import APIRouter, Depends, Request, Form, Query
from fastapi.responses import RedirectResponse, Response
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, or_, func, desc, asc
from sqlalchemy.orm import selectinload
import io
from datetime import datetime
from typing import List, Dict, Any, Optional

from apps.web.jinja import templates
from apps.web.dependencies import get_current_user_dependency, require_role
from core.database.session import get_db_session
from core.config.settings import settings
from domain.entities.user import User
from domain.entities.object import Object
from shared.services.incident_service import IncidentService
from shared.services.incident_category_service import IncidentCategoryService
from shared.services.employee_selector_service import EmployeeSelectorService
from fastapi import HTTPException, Query
from fastapi.responses import JSONResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm


router = APIRouter()


async def _get_db_user_id(current_user, session: AsyncSession) -> int | None:
    """Локальный helper: получить внутренний user_id по current_user (JWT словарь или ORM User)."""
    from sqlalchemy import select as sql_select
    from domain.entities.user import User as UserEntity
    if isinstance(current_user, dict):
        telegram_id = current_user.get("telegram_id") or current_user.get("id")
        res = await session.execute(sql_select(UserEntity).where(UserEntity.telegram_id == telegram_id))
        u = res.scalar_one_or_none()
        return u.id if u else None
    return current_user.id


async def _get_owner_template_context(owner_id: int, session: AsyncSession):
    """Получить контекст для owner шаблонов."""
    from apps.web.routes.owner import get_owner_context
    return await get_owner_context(owner_id, session)


@router.get("/owner/incidents")
async def owner_incidents_list(
    request: Request,
    incident_type: str = Query("deduction"),
    status: str = Query(None),
    statuses: str = Query(None),  # Множественный фильтр через запятую: "new,in_review"
    category: str = Query(None),
    object_id: int = Query(None),
    employee_id: int = Query(None),
    sort_by: str = Query("created_at"),
    sort_order: str = Query("desc"),
    page: int = Query(1, ge=1),
    per_page: int = Query(25, ge=10, le=200),
    current_user: User = Depends(get_current_user_dependency()),
    _: User = Depends(require_role(["owner", "superadmin"])),
    session: AsyncSession = Depends(get_db_session)
):
    """Список тикетов с пагинацией, сортировкой и множественными фильтрами."""
    if not settings.enable_incidents:
        raise HTTPException(
            status_code=404,
            detail="Incidents отключены. Включите enable_incidents в настройках."
        )
    
    from domain.entities.incident import Incident
    
    owner_id = await _get_db_user_id(current_user, session)
    
    # Базовый запрос — фильтр по типу тикета
    query = select(Incident).where(
        Incident.owner_id == owner_id,
        Incident.incident_type == incident_type
    ).options(
        selectinload(Incident.object),
        selectinload(Incident.employee),
        selectinload(Incident.shift_schedule),
        selectinload(Incident.creator)
    )
    
    # Множественный фильтр по статусам (приоритет над одиночным фильтром)
    if statuses:
        status_list = [s.strip() for s in statuses.split(",") if s.strip()]
        if status_list:
            query = query.where(Incident.status.in_(status_list))
    elif status:
        # Одиночный фильтр для обратной совместимости
        query = query.where(Incident.status == status)
    
    # Фильтр по категории
    if category:
        query = query.where(Incident.category == category)
    
    # Фильтр по объекту
    if object_id:
        query = query.where(Incident.object_id == object_id)
    
    # Фильтр по сотруднику
    if employee_id:
        query = query.where(Incident.employee_id == employee_id)
    
    # Сортировка
    sort_column_map = {
        "created_at": Incident.created_at,
        "custom_date": Incident.custom_date,
        "category": Incident.category,
        "severity": Incident.severity,
        "status": Incident.status,
        "damage_amount": Incident.damage_amount,
        "number": Incident.custom_number
    }
    sort_column = sort_column_map.get(sort_by, Incident.created_at)
    if sort_order == "asc":
        query = query.order_by(asc(sort_column))
    else:
        query = query.order_by(desc(sort_column))
    
    # Подсчет общего количества для пагинации (применяем те же фильтры)
    count_query = select(func.count(Incident.id)).where(
        Incident.owner_id == owner_id,
        Incident.incident_type == incident_type
    )
    
    # Применяем те же фильтры к подсчету
    if statuses:
        status_list = [s.strip() for s in statuses.split(",") if s.strip()]
        if status_list:
            count_query = count_query.where(Incident.status.in_(status_list))
    elif status:
        count_query = count_query.where(Incident.status == status)
    
    if category:
        count_query = count_query.where(Incident.category == category)
    if object_id:
        count_query = count_query.where(Incident.object_id == object_id)
    if employee_id:
        count_query = count_query.where(Incident.employee_id == employee_id)
    
    count_result = await session.execute(count_query)
    total_count = count_result.scalar_one() or 0
    
    # Пагинация
    offset = (page - 1) * per_page
    query = query.limit(per_page).offset(offset)
    
    result = await session.execute(query)
    incidents = result.scalars().all()
    
    # Вычисляем параметры пагинации
    total_pages = (total_count + per_page - 1) // per_page if total_count > 0 else 1

    # Категории владельца (по типу тикета)
    categories = await IncidentCategoryService(session).list_categories(owner_id, incident_type=incident_type)

    # Объекты владельца
    from domain.entities.object import Object
    obj_res = await session.execute(select(Object).where(Object.owner_id == owner_id, Object.is_active == True).order_by(Object.name))
    objects = obj_res.scalars().all()

    # Сотрудники владельца по активным договорам (не менеджеры)
    from domain.entities.contract import Contract
    emp_ids_res = await session.execute(
        select(Contract.employee_id)
        .where(
            Contract.owner_id == owner_id,
            Contract.is_active == True
        )
        .distinct()
    )
    employee_ids = [row[0] for row in emp_ids_res.all()]
    from domain.entities.user import User as UserEntity
    employees = []
    if employee_ids:
        emp_res = await session.execute(
            select(UserEntity)
            .where(UserEntity.id.in_(employee_ids))
            .order_by(UserEntity.last_name, UserEntity.first_name)
        )
        employees = emp_res.scalars().all()

    # Счётчики по типам для табов
    for_tabs_query = select(
        Incident.incident_type, func.count(Incident.id)
    ).where(Incident.owner_id == owner_id).group_by(Incident.incident_type)
    tabs_result = await session.execute(for_tabs_query)
    type_counts = dict(tabs_result.all())
    
    # Формируем параметры для сохранения фильтров в URL
    filter_params = {}
    filter_params["incident_type"] = incident_type
    if statuses:
        filter_params["statuses"] = statuses
    elif status:
        filter_params["status"] = status
    if category:
        filter_params["category"] = category
    if object_id:
        filter_params["object_id"] = object_id
    if employee_id:
        filter_params["employee_id"] = employee_id
    if sort_by != "created_at":
        filter_params["sort_by"] = sort_by
    if sort_order != "desc":
        filter_params["sort_order"] = sort_order
    if per_page != 25:
        filter_params["per_page"] = per_page
    
    # Получаем контекст для owner базового шаблона
    owner_context = await _get_owner_template_context(owner_id, session)
    
    return templates.TemplateResponse(
        "owner/incidents/list.html",
        {
            "request": request,
            "current_user": current_user,
            "available_interfaces": owner_context.get("available_interfaces", []),
            "new_applications_count": owner_context.get("new_applications_count", 0),
            "incidents": incidents,
            "incident_type": incident_type,
            "type_counts": type_counts,
            "status_filter": status,
            "statuses_filter": statuses,
            "category_filter": category,
            "object_id_filter": object_id,
            "employee_id_filter": employee_id,
            "sort_by": sort_by,
            "sort_order": sort_order,
            "page": page,
            "per_page": per_page,
            "total_count": total_count,
            "total_pages": total_pages,
            "categories": categories,
            "objects": objects,
            "employees": employees,
            "filter_params": filter_params
        }
    )


@router.get("/owner/incidents/create")
async def owner_incident_create_page(
    request: Request,
    incident_type: str = Query("deduction"),
    current_user: User = Depends(get_current_user_dependency()),
    _: User = Depends(require_role(["owner", "superadmin"])),
    session: AsyncSession = Depends(get_db_session)
):
    owner_id = await _get_db_user_id(current_user, session)
    categories = await IncidentCategoryService(session).list_categories(owner_id, incident_type=incident_type)
    from domain.entities.object import Object
    obj_res = await session.execute(select(Object).where(Object.owner_id == owner_id, Object.is_active == True).order_by(Object.name))
    objects = obj_res.scalars().all()

    # Для типа "request" — справочник товаров
    products = []
    if incident_type == "request":
        from shared.services.product_service import ProductService
        products = await ProductService(session).list_products(owner_id)

    owner_context = await _get_owner_template_context(owner_id, session)
    return templates.TemplateResponse(
        "owner/incidents/create.html",
        {
            "request": request,
            "current_user": current_user,
            "available_interfaces": owner_context.get("available_interfaces", []),
            "new_applications_count": owner_context.get("new_applications_count", 0),
            "incident_type": incident_type,
            "categories": categories,
            "objects": objects,
            "products": products,
        }
    )


@router.post("/owner/incidents/create")
async def owner_incidents_create(
    request: Request,
    incident_type: str = Form("deduction"),
    category: str = Form(...),
    severity: str = Form("medium"),
    object_id: int = Form(None),
    shift_schedule_id: int = Form(None),
    employee_id: int = Form(None),
    notes: str = Form(None),
    custom_number: str = Form(None),
    custom_date: str = Form(None),
    damage_amount: str = Form(None),
    current_user: User = Depends(get_current_user_dependency()),
    _: User = Depends(require_role(["owner", "superadmin"])),
    session: AsyncSession = Depends(get_db_session)
):
    """Создать тикет вручную."""
    incident_service = IncidentService(session)
    from datetime import datetime
    from decimal import Decimal
    parsed_date = None
    if custom_date:
        try:
            parsed_date = datetime.strptime(custom_date, "%Y-%m-%d").date()
        except Exception:
            parsed_date = None
    parsed_damage = None
    if damage_amount:
        try:
            parsed_damage = Decimal(damage_amount)
        except Exception:
            parsed_damage = None

    owner_db_id = await _get_db_user_id(current_user, session)
    incident = await incident_service.create_incident(
        owner_id=owner_db_id,
        incident_type=incident_type,
        category=category,
        severity=severity,
        object_id=object_id,
        shift_schedule_id=shift_schedule_id,
        employee_id=employee_id,
        notes=notes,
        created_by=current_user.id,
        custom_number=custom_number,
        custom_date=parsed_date,
        damage_amount=parsed_damage
    )

    # Для обращений — добавить позиции товаров из формы
    if incident_type == "request":
        form_data = await request.form()
        await _save_incident_items_from_form(session, incident.id, form_data, owner_db_id)
    
    return RedirectResponse(url=f"/owner/incidents?incident_type={incident_type}", status_code=303)


async def _save_incident_items_from_form(
    session: AsyncSession, incident_id: int, form_data, added_by: int
):
    """Извлечь позиции товаров из формы и сохранить."""
    from shared.services.incident_item_service import IncidentItemService
    item_service = IncidentItemService(session)

    idx = 0
    while True:
        product_id_key = f"items[{idx}][product_id]"
        if product_id_key not in form_data:
            break
        try:
            product_id = int(form_data[product_id_key]) if form_data[product_id_key] else None
            product_name = form_data.get(f"items[{idx}][product_name]", "")
            quantity = float(form_data.get(f"items[{idx}][quantity]", 1))
            price = float(form_data.get(f"items[{idx}][price]", 0))
            if product_name and quantity > 0:
                await item_service.add_item(
                    incident_id=incident_id,
                    product_id=product_id,
                    product_name=product_name,
                    quantity=quantity,
                    price=price,
                    added_by=added_by,
                )
        except (ValueError, TypeError):
            pass
        idx += 1


@router.post("/owner/incidents/{incident_id}/status")
async def owner_incidents_update_status(
    request: Request,
    incident_id: int,
    new_status: str = Form(...),
    notes: str = Form(None),
    return_url: str = Query(None),
    current_user: User = Depends(get_current_user_dependency()),
    _: User = Depends(require_role(["owner", "superadmin"])),
    session: AsyncSession = Depends(get_db_session)
):
    """Изменить статус инцидента."""
    from domain.entities.incident import Incident
    
    owner_db_id = await _get_db_user_id(current_user, session)
    incident = await session.get(Incident, incident_id)
    if incident and incident.owner_id == owner_db_id:
        incident_service = IncidentService(session)
        changed_by_id = owner_db_id
        await incident_service.update_incident_status(
            incident_id=incident_id,
            new_status=new_status,
            notes=notes,
            changed_by=changed_by_id
        )
    
    # Возвращаемся на указанную страницу или на список с фильтром
    if return_url:
        return RedirectResponse(url=return_url, status_code=303)
    
    status_param = f"?status={new_status}" if new_status else ""
    return RedirectResponse(url=f"/owner/incidents{status_param}", status_code=303)


@router.get("/owner/incidents/{incident_id}/edit")
async def owner_incident_edit_page(
    request: Request,
    incident_id: int,
    return_url: str = Query(None),
    current_user: User = Depends(get_current_user_dependency()),
    _: User = Depends(require_role(["owner", "superadmin"])),
    session: AsyncSession = Depends(get_db_session)
):
    from domain.entities.incident import Incident
    from sqlalchemy import select
    from sqlalchemy.orm import selectinload
    # Eager load basic relations
    result = await session.execute(
        select(Incident)
        .options(
            selectinload(Incident.object),
            selectinload(Incident.employee)
        )
        .where(Incident.id == incident_id)
    )
    incident = result.scalar_one_or_none()
    owner_db_id = await _get_db_user_id(current_user, session)
    if not incident or incident.owner_id != owner_db_id:
        raise HTTPException(status_code=404, detail="Инцидент не найден")
    # История изменений
    from domain.entities.incident_history import IncidentHistory
    hist_res = await session.execute(
        select(IncidentHistory)
        .options(selectinload(IncidentHistory.changer))
        .where(IncidentHistory.incident_id == incident.id)
        .order_by(IncidentHistory.changed_at.desc())
    )
    history = hist_res.scalars().all()
    categories = await IncidentCategoryService(session).list_categories(owner_db_id)
    # Списки объектов и сотрудников владельца
    from domain.entities.object import Object
    obj_res = await session.execute(select(Object).where(Object.owner_id == owner_db_id, Object.is_active == True).order_by(Object.name))
    objects = obj_res.scalars().all()
    employee_groups = {"active": [], "former": []}
    if incident.object_id:
        selector = EmployeeSelectorService(session)
        employee_groups = await selector.get_employees_for_owner(owner_db_id, object_id=incident.object_id)
    adjustments = await IncidentService(session).get_adjustments_by_incident(incident.id)

    # Позиции и товары (для обращений)
    incident_items = []
    products = []
    if incident.incident_type == "request":
        from shared.services.incident_item_service import IncidentItemService
        from shared.services.product_service import ProductService
        incident_items = await IncidentItemService(session).list_items(incident.id)
        products = await ProductService(session).list_products(owner_db_id)
    
    # Формируем URL возврата к списку
    back_url = return_url or f"/owner/incidents?incident_type={incident.incident_type}"
    
    # Получаем контекст для owner базового шаблона
    owner_context = await _get_owner_template_context(owner_db_id, session)
    
    return templates.TemplateResponse(
        "owner/incidents/edit.html",
        {
            "request": request,
            "current_user": current_user,
            "available_interfaces": owner_context.get("available_interfaces", []),
            "new_applications_count": owner_context.get("new_applications_count", 0),
            "incident": incident,
            "categories": categories,
            "history": history,
            "objects": objects,
            "employee_groups": employee_groups,
            "adjustments": adjustments,
            "incident_items": incident_items,
            "products": products,
            "back_url": back_url,
        }
    )


@router.get("/owner/incidents/api/employees", response_class=JSONResponse)
async def owner_incident_employees(
    object_id: int = Query(..., ge=1),
    current_user: User = Depends(get_current_user_dependency()),
    _: User = Depends(require_role(["owner", "superadmin"])),
    session: AsyncSession = Depends(get_db_session),
):
    """Вернуть сотрудников (активных и уволенных) для выбранного объекта."""
    owner_id = await _get_db_user_id(current_user, session)
    if not owner_id:
        raise HTTPException(status_code=403, detail="Пользователь не найден")

    selector = EmployeeSelectorService(session)
    obj = await session.get(Object, object_id)
    if not obj or obj.owner_id != owner_id:
        raise HTTPException(status_code=404, detail="Объект не найден")

    grouped = await selector.get_employees_for_owner(owner_id, object_id=object_id)
    return JSONResponse(grouped)


@router.get("/owner/incidents/{incident_id}/items", response_class=JSONResponse)
async def owner_incident_items_list(
    incident_id: int,
    current_user: User = Depends(get_current_user_dependency()),
    _: User = Depends(require_role(["owner", "superadmin"])),
    session: AsyncSession = Depends(get_db_session),
):
    """Получить позиции тикета (JSON)."""
    from shared.services.incident_item_service import IncidentItemService
    from domain.entities.incident import Incident
    owner_db_id = await _get_db_user_id(current_user, session)
    incident = await session.get(Incident, incident_id)
    if not incident or incident.owner_id != owner_db_id:
        raise HTTPException(status_code=404, detail="Тикет не найден")
    items = await IncidentItemService(session).list_items(incident_id)
    return JSONResponse([{
        "id": i.id,
        "product_id": i.product_id,
        "product_name": i.product_name,
        "quantity": float(i.quantity),
        "price": float(i.price),
        "total": i.total,
    } for i in items])


@router.post("/owner/incidents/{incident_id}/items", response_class=JSONResponse)
async def owner_incident_item_add(
    request: Request,
    incident_id: int,
    current_user: User = Depends(get_current_user_dependency()),
    _: User = Depends(require_role(["owner", "superadmin"])),
    session: AsyncSession = Depends(get_db_session),
):
    """Добавить позицию к тикету (JSON)."""
    from shared.services.incident_item_service import IncidentItemService
    from domain.entities.incident import Incident
    owner_db_id = await _get_db_user_id(current_user, session)
    incident = await session.get(Incident, incident_id)
    if not incident or incident.owner_id != owner_db_id:
        raise HTTPException(status_code=404, detail="Тикет не найден")
    body = await request.json()
    item_svc = IncidentItemService(session)
    item = await item_svc.add_item(
        incident_id=incident_id,
        product_id=body.get("product_id"),
        product_name=body.get("product_name", ""),
        quantity=float(body.get("quantity", 1)),
        price=float(body.get("price", 0)),
        added_by=owner_db_id,
    )
    total = await item_svc.calculate_total(incident_id)
    return JSONResponse({
        "id": item.id, "product_name": item.product_name,
        "quantity": float(item.quantity), "price": float(item.price),
        "total": item.total, "incident_total": float(total),
    })


@router.put("/owner/incidents/{incident_id}/items/{item_id}", response_class=JSONResponse)
async def owner_incident_item_update(
    request: Request,
    incident_id: int,
    item_id: int,
    current_user: User = Depends(get_current_user_dependency()),
    _: User = Depends(require_role(["owner", "superadmin"])),
    session: AsyncSession = Depends(get_db_session),
):
    """Обновить позицию тикета (JSON)."""
    from shared.services.incident_item_service import IncidentItemService
    from domain.entities.incident import Incident
    owner_db_id = await _get_db_user_id(current_user, session)
    incident = await session.get(Incident, incident_id)
    if not incident or incident.owner_id != owner_db_id:
        raise HTTPException(status_code=404, detail="Тикет не найден")
    body = await request.json()
    item_svc = IncidentItemService(session)
    item = await item_svc.update_item(
        item_id=item_id,
        modified_by=owner_db_id,
        quantity=body.get("quantity"),
        price=body.get("price"),
        product_name=body.get("product_name"),
    )
    if not item:
        raise HTTPException(status_code=404, detail="Позиция не найдена")
    total = await item_svc.calculate_total(incident_id)
    return JSONResponse({
        "id": item.id, "product_name": item.product_name,
        "quantity": float(item.quantity), "price": float(item.price),
        "total": item.total, "incident_total": float(total),
    })


@router.delete("/owner/incidents/{incident_id}/items/{item_id}", response_class=JSONResponse)
async def owner_incident_item_delete(
    incident_id: int,
    item_id: int,
    current_user: User = Depends(get_current_user_dependency()),
    _: User = Depends(require_role(["owner", "superadmin"])),
    session: AsyncSession = Depends(get_db_session),
):
    """Удалить позицию тикета (JSON)."""
    from shared.services.incident_item_service import IncidentItemService
    from domain.entities.incident import Incident
    owner_db_id = await _get_db_user_id(current_user, session)
    incident = await session.get(Incident, incident_id)
    if not incident or incident.owner_id != owner_db_id:
        raise HTTPException(status_code=404, detail="Тикет не найден")
    item_svc = IncidentItemService(session)
    ok = await item_svc.remove_item(item_id, removed_by=owner_db_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Позиция не найдена")
    total = await item_svc.calculate_total(incident_id)
    return JSONResponse({"ok": True, "incident_total": float(total)})


@router.post("/owner/incidents/{incident_id}/edit")
async def owner_incident_edit_save(
    request: Request,
    incident_id: int,
    category: str = Form(None),
    severity: str = Form(None),
    object_id: int = Form(None),
    shift_schedule_id: int = Form(None),
    employee_id: int = Form(None),
    notes: str = Form(None),
    custom_number: str = Form(None),
    custom_date: str = Form(None),
    damage_amount: str = Form(None),
    status: str = Form(None),
    compensate_purchase: str = Form(None),
    return_url: str = Form(None),
    current_user: User = Depends(get_current_user_dependency()),
    _: User = Depends(require_role(["owner", "superadmin"])),
    session: AsyncSession = Depends(get_db_session)
):
    from decimal import Decimal
    from datetime import datetime
    from domain.entities.incident import Incident
    from sqlalchemy import select
    from fastapi import HTTPException
    
    # Проверяем статус инцидента перед редактированием
    result = await session.execute(select(Incident).where(Incident.id == incident_id))
    incident = result.scalar_one_or_none()
    owner_db_id = await _get_db_user_id(current_user, session)
    if not incident or incident.owner_id != owner_db_id:
        raise HTTPException(status_code=404, detail="Инцидент не найден")
    
    # Проверка статуса: решенные, отклоненные и отмененные инциденты нельзя редактировать
    if incident.status in ['resolved', 'rejected', 'cancelled']:
        raise HTTPException(status_code=400, detail=f"Нельзя редактировать инцидент со статусом '{incident.status}'")
    
    service = IncidentService(session)
    changed_by_id = owner_db_id
    
    # Если статус изменился — применяем через IncidentService для истории
    if status and status != incident.status:
        if status not in ["new", "in_review", "resolved", "rejected", "cancelled"]:
            raise HTTPException(status_code=400, detail="Некорректный статус")
        await service.update_incident_status(
            incident_id=incident_id,
            new_status=status,
            notes=None,
            changed_by=changed_by_id
        )
        await session.commit()
        # После изменения статуса нужно обновить инцидент из БД
        await session.refresh(incident)
    
    data = {
        k: v for k, v in {
            "category": category,
            "severity": severity,
            "object_id": object_id,
            "shift_schedule_id": shift_schedule_id,
            "employee_id": employee_id,
            "notes": notes,
            "custom_number": custom_number
        }.items() if v is not None and v != ""
    }
    if custom_date:
        try:
            data["custom_date"] = datetime.strptime(custom_date, "%Y-%m-%d").date()
        except Exception:
            pass
    if damage_amount:
        try:
            data["damage_amount"] = Decimal(damage_amount)
        except Exception:
            pass

    # Обработка compensate_purchase (checkbox — приходит "on" или отсутствует)
    if incident.incident_type == "request":
        data["compensate_purchase"] = compensate_purchase == "on"
    
    try:
        await service.update_incident(incident_id, data, changed_by_id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    
    # Возвращаемся на указанную страницу или на список
    if return_url:
        return RedirectResponse(url=return_url, status_code=303)
    
    return RedirectResponse(url=f"/owner/incidents?incident_type={incident.incident_type}", status_code=303)


@router.get("/owner/incidents/categories")
async def owner_incident_categories(
    request: Request,
    incident_type: str = Query("deduction"),
    current_user: User = Depends(get_current_user_dependency()),
    _: User = Depends(require_role(["owner", "superadmin"])),
    session: AsyncSession = Depends(get_db_session)
):
    owner_id = await _get_db_user_id(current_user, session)
    cat_svc = IncidentCategoryService(session)
    deduction_cats = await cat_svc.list_categories(owner_id, incident_type="deduction")
    request_cats = await cat_svc.list_categories(owner_id, incident_type="request")
    owner_context = await _get_owner_template_context(owner_id, session)
    return templates.TemplateResponse(
        "owner/incidents/categories.html",
        {
            "request": request,
            "current_user": current_user,
            "available_interfaces": owner_context.get("available_interfaces", []),
            "new_applications_count": owner_context.get("new_applications_count", 0),
            "deduction_categories": deduction_cats,
            "request_categories": request_cats,
            "incident_type": incident_type,
        }
    )


@router.post("/owner/incidents/categories")
async def owner_incident_categories_save(
    request: Request,
    name: str = Form(...),
    incident_type: str = Form("deduction"),
    category_id: int = Form(None),
    action: str = Form("save"),
    current_user: User = Depends(get_current_user_dependency()),
    _: User = Depends(require_role(["owner", "superadmin"])),
    session: AsyncSession = Depends(get_db_session)
):
    svc = IncidentCategoryService(session)
    if action == "deactivate" and category_id:
        await svc.deactivate(category_id)
    elif action == "activate" and category_id:
        await svc.activate(category_id)
    else:
        await svc.create_or_update(current_user.id, name=name, category_id=category_id, incident_type=incident_type)
    return RedirectResponse(url=f"/owner/incidents/categories?incident_type={incident_type}", status_code=303)


@router.get("/owner/incidents/reports")
async def owner_incidents_reports_page(
    request: Request,
    report_type: str = Query("deductions"),
    date_from: str = Query(None),
    date_to: str = Query(None),
    object_id: int = Query(None),
    employee_id: int = Query(None),
    current_user: User = Depends(get_current_user_dependency()),
    _: User = Depends(require_role(["owner", "superadmin"])),
    session: AsyncSession = Depends(get_db_session)
):
    from domain.entities.incident import Incident
    from domain.entities.incident_item import IncidentItem
    from decimal import Decimal

    owner_id = await _get_db_user_id(current_user, session)
    owner_context = await _get_owner_template_context(owner_id, session)

    # Объекты и сотрудники для фильтров
    obj_res = await session.execute(
        select(Object).where(Object.owner_id == owner_id, Object.is_active == True).order_by(Object.name)
    )
    objects = obj_res.scalars().all()

    emp_grouped = await EmployeeSelectorService(session).get_employees_for_owner(owner_id)
    employees = emp_grouped.get("active", []) + emp_grouped.get("former", [])

    # Разбор дат
    from datetime import date as date_type, timedelta
    d_from = None
    d_to = None
    try:
        if date_from:
            d_from = datetime.strptime(date_from, "%Y-%m-%d")
        if date_to:
            d_to = datetime.strptime(date_to, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
    except ValueError:
        pass

    report_data = []
    summary = {}

    if report_type == "requests":
        # Отчёт по заявкам (обращения)
        q = select(Incident).where(
            Incident.owner_id == owner_id,
            Incident.incident_type == "request",
        ).options(
            selectinload(Incident.object),
            selectinload(Incident.employee),
        ).order_by(Incident.created_at.desc())

        if d_from:
            q = q.where(Incident.created_at >= d_from)
        if d_to:
            q = q.where(Incident.created_at <= d_to)
        if object_id:
            q = q.where(Incident.object_id == object_id)
        if employee_id:
            q = q.where(Incident.employee_id == employee_id)

        res = await session.execute(q)
        incidents = res.scalars().all()

        total_all = Decimal("0")
        total_compensated = Decimal("0")
        total_not_compensated = Decimal("0")
        count_compensated = 0
        count_not_compensated = 0

        for inc in incidents:
            # Получить позиции
            items_res = await session.execute(
                select(IncidentItem).where(IncidentItem.incident_id == inc.id)
            )
            items = items_res.scalars().all()
            inc_total = sum((it.quantity or 0) * (it.price or 0) for it in items)

            emp_name = ""
            if inc.employee:
                emp_name = f"{inc.employee.first_name or ''} {inc.employee.last_name or ''}".strip()

            row = {
                "id": inc.id,
                "date": inc.created_at,
                "employee": emp_name,
                "object": inc.object.name if inc.object else "—",
                "category": inc.category,
                "status": inc.status,
                "compensated": inc.compensate_purchase,
                "items_count": len(items),
                "total": float(inc_total),
            }
            report_data.append(row)
            total_all += Decimal(str(inc_total))
            if inc.compensate_purchase:
                total_compensated += Decimal(str(inc_total))
                count_compensated += 1
            else:
                total_not_compensated += Decimal(str(inc_total))
                count_not_compensated += 1

        summary = {
            "total_count": len(incidents),
            "total_all": float(total_all),
            "count_compensated": count_compensated,
            "total_compensated": float(total_compensated),
            "count_not_compensated": count_not_compensated,
            "total_not_compensated": float(total_not_compensated),
        }

    return templates.TemplateResponse(
        "owner/incidents/reports.html",
        {
            "request": request,
            "current_user": current_user,
            "available_interfaces": owner_context.get("available_interfaces", []),
            "new_applications_count": owner_context.get("new_applications_count", 0),
            "report_type": report_type,
            "date_from": date_from or "",
            "date_to": date_to or "",
            "object_id": object_id,
            "employee_id": employee_id,
            "objects": objects,
            "employees": employees,
            "report_data": report_data,
            "summary": summary,
        }
    )


@router.post("/owner/incidents/{incident_id}/cancel")
async def owner_incidents_cancel(
    request: Request,
    incident_id: int,
    cancellation_reason: str = Form(...),
    notes: str = Form(None),
    return_url: str = Form(None),
    current_user: User = Depends(get_current_user_dependency()),
    _: User = Depends(require_role(["owner", "superadmin"])),
    session: AsyncSession = Depends(get_db_session)
):
    """Отменить инцидент с указанием причины."""
    from domain.entities.incident import Incident
    from fastapi import HTTPException
    
    owner_db_id = await _get_db_user_id(current_user, session)
    incident = await session.get(Incident, incident_id)
    if not incident or incident.owner_id != owner_db_id:
        raise HTTPException(status_code=404, detail="Инцидент не найден")
    
    incident_service = IncidentService(session)
    cancelled_by_id = owner_db_id
    
    try:
        await incident_service.cancel_incident(
            incident_id=incident_id,
            cancellation_reason=cancellation_reason,
            cancelled_by=cancelled_by_id,
            notes=notes
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        from core.logging.logger import logger
        logger.error("Error cancelling incident", incident_id=incident_id, error=str(e), exc_info=True)
        raise HTTPException(status_code=500, detail=f"Ошибка при отмене инцидента: {str(e)}")
    
    # Возвращаемся на указанную страницу или на список
    if return_url:
        return RedirectResponse(url=return_url, status_code=303)
    
    return RedirectResponse(url="/owner/incidents?statuses=cancelled", status_code=303)


@router.post("/owner/incidents/{incident_id}/apply-adjustments")
async def owner_incidents_apply_adjustments(
    request: Request,
    incident_id: int,
    return_url: str = Query(None),
    current_user: User = Depends(get_current_user_dependency()),
    _: User = Depends(require_role(["owner", "superadmin"])),
    session: AsyncSession = Depends(get_db_session)
):
    """Применить предложенные корректировки."""
    from domain.entities.incident import Incident
    
    owner_db_id = await _get_db_user_id(current_user, session)
    incident = await session.get(Incident, incident_id)
    if incident and incident.owner_id == owner_db_id:
        incident_service = IncidentService(session)
        await incident_service.apply_suggested_adjustments(incident_id)
    
    # Возвращаемся на указанную страницу или на список
    if return_url:
        return RedirectResponse(url=return_url, status_code=303)
    
    return RedirectResponse(url="/owner/incidents", status_code=303)


async def _get_filtered_incidents_for_export(
    session: AsyncSession,
    owner_id: int,
    status: Optional[str] = None,
    statuses: Optional[str] = None,
    category: Optional[str] = None,
    object_id: Optional[int] = None,
    employee_id: Optional[int] = None,
    sort_by: str = "created_at",
    sort_order: str = "desc"
) -> List:
    """Получить отфильтрованные инциденты для экспорта (без пагинации)."""
    from domain.entities.incident import Incident
    
    query = select(Incident).where(Incident.owner_id == owner_id).options(
        selectinload(Incident.object),
        selectinload(Incident.employee),
        selectinload(Incident.shift_schedule),
        selectinload(Incident.creator)
    )
    
    # Множественный фильтр по статусам
    if statuses:
        status_list = [s.strip() for s in statuses.split(",") if s.strip()]
        if status_list:
            query = query.where(Incident.status.in_(status_list))
    elif status:
        query = query.where(Incident.status == status)
    
    # Фильтр по категории
    if category:
        query = query.where(Incident.category == category)
    
    # Фильтр по объекту
    if object_id:
        query = query.where(Incident.object_id == object_id)
    
    # Фильтр по сотруднику
    if employee_id:
        query = query.where(Incident.employee_id == employee_id)
    
    # Сортировка
    sort_column_map = {
        "created_at": Incident.created_at,
        "custom_date": Incident.custom_date,
        "category": Incident.category,
        "severity": Incident.severity,
        "status": Incident.status,
        "damage_amount": Incident.damage_amount,
        "number": Incident.custom_number
    }
    sort_column = sort_column_map.get(sort_by, Incident.created_at)
    if sort_order == "asc":
        query = query.order_by(asc(sort_column))
    else:
        query = query.order_by(desc(sort_column))
    
    result = await session.execute(query)
    return result.scalars().all()


def _get_category_display_export(category: str) -> str:
    """Получить отображаемое название категории для экспорта."""
    category_map = {
        "late_arrival": "Опоздание",
        "task_non_completion": "Невыполнение задачи",
        "damage": "Повреждение",
        "violation": "Нарушение"
    }
    return category_map.get(category, category)


def _get_severity_display_export(severity: Optional[str]) -> str:
    """Получить отображаемое название критичности для экспорта."""
    if not severity:
        return "Средняя"
    severity_map = {
        "low": "Низкая",
        "medium": "Средняя",
        "high": "Высокая",
        "critical": "Критично"
    }
    return severity_map.get(severity, severity)


def _get_status_display_export(status: str) -> str:
    """Получить отображаемое название статуса для экспорта."""
    status_map = {
        "new": "Новый",
        "in_review": "На рассмотрении",
        "resolved": "Решён",
        "rejected": "Отклонён",
        "cancelled": "Отменён"
    }
    return status_map.get(status, status)


@router.get("/owner/incidents/export/excel")
async def owner_incidents_export_excel(
    request: Request,
    status: str = Query(None),
    statuses: str = Query(None),
    category: str = Query(None),
    object_id: int = Query(None),
    employee_id: int = Query(None),
    sort_by: str = Query("created_at"),
    sort_order: str = Query("desc"),
    current_user: User = Depends(get_current_user_dependency()),
    _: User = Depends(require_role(["owner", "superadmin"])),
    session: AsyncSession = Depends(get_db_session)
):
    """Экспорт инцидентов в Excel."""
    if not settings.enable_incidents:
        raise HTTPException(status_code=404, detail="Incidents отключены")
    
    owner_id = await _get_db_user_id(current_user, session)
    incidents = await _get_filtered_incidents_for_export(
        session, owner_id, status, statuses, category, object_id, employee_id, sort_by, sort_order
    )
    
    # Подготовка данных для Excel
    data = []
    for inc in incidents:
        data.append({
            "Номер": inc.custom_number or f"#{inc.id}",
            "Дата": inc.custom_date.strftime('%d.%m.%Y') if inc.custom_date else inc.created_at.strftime('%d.%m.%Y'),
            "Категория": _get_category_display_export(inc.category),
            "Критичность": _get_severity_display_export(inc.severity),
            "Статус": _get_status_display_export(inc.status),
            "Объект": inc.object.name if inc.object else "—",
            "Сотрудник": f"{inc.employee.last_name} {inc.employee.first_name}".strip() if inc.employee else "—",
            "Ущерб, ₽": f"{inc.damage_amount:,.2f}".replace(',', ' ') if inc.damage_amount else "—",
            "Создал": inc.creator.first_name if inc.creator else "—",
            "Создан": inc.created_at.strftime('%d.%m.%Y %H:%M'),
            "Примечания": (inc.notes or "")[:100] + "..." if inc.notes and len(inc.notes) > 100 else (inc.notes or "—")
        })
    
    if not data:
        raise HTTPException(status_code=404, detail="Нет данных для экспорта")
    
    # Создание Excel файла
    df = pd.DataFrame(data)
    wb = Workbook()
    ws = wb.active
    ws.title = "Инциденты"
    
    # Добавление данных
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # Стилизация
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Автоширина колонок
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохранение в память
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"incidents_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"}
    )


@router.get("/owner/incidents/export/pdf")
async def owner_incidents_export_pdf(
    request: Request,
    status: str = Query(None),
    statuses: str = Query(None),
    category: str = Query(None),
    object_id: int = Query(None),
    employee_id: int = Query(None),
    sort_by: str = Query("created_at"),
    sort_order: str = Query("desc"),
    current_user: User = Depends(get_current_user_dependency()),
    _: User = Depends(require_role(["owner", "superadmin"])),
    session: AsyncSession = Depends(get_db_session)
):
    """Экспорт инцидентов в PDF."""
    if not settings.enable_incidents:
        raise HTTPException(status_code=404, detail="Incidents отключены")
    
    owner_id = await _get_db_user_id(current_user, session)
    incidents = await _get_filtered_incidents_for_export(
        session, owner_id, status, statuses, category, object_id, employee_id, sort_by, sort_order
    )
    
    if not incidents:
        raise HTTPException(status_code=404, detail="Нет данных для экспорта")
    
    # Создание PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    
    # Определяем доступный шрифт
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        import os
        font_path = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
        if os.path.exists(font_path):
            pdfmetrics.registerFont(TTFont('DejaVuSans', font_path))
            font_name = 'DejaVuSans'
        else:
            font_name = 'Helvetica'
    except:
        font_name = 'Helvetica'
    
    # Стили
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=20,
        alignment=1,  # Центрирование
        fontName=font_name
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=6,
        fontName=font_name
    )
    
    story = []
    
    # Заголовок
    story.append(Paragraph("Отчет по инцидентам", title_style))
    story.append(Spacer(1, 12))
    
    # Информация о фильтрах
    filter_info = []
    if statuses:
        filter_info.append(f"Статусы: {statuses}")
    elif status:
        filter_info.append(f"Статус: {status}")
    if category:
        filter_info.append(f"Категория: {category}")
    if object_id:
        obj = await session.get(Object, object_id)
        if obj:
            filter_info.append(f"Объект: {obj.name}")
    if employee_id:
        from domain.entities.user import User as UserEntity
        emp = await session.get(UserEntity, employee_id)
        if emp:
            filter_info.append(f"Сотрудник: {emp.last_name} {emp.first_name}")
    
    if filter_info:
        story.append(Paragraph("Фильтры: " + ", ".join(filter_info), normal_style))
        story.append(Spacer(1, 12))
    
    story.append(Paragraph(f"Всего инцидентов: {len(incidents)}", normal_style))
    story.append(Spacer(1, 12))
    
    # Таблица инцидентов
    table_data = [["Номер", "Дата", "Категория", "Критичность", "Статус", "Объект", "Сотрудник", "Ущерб"]]
    
    for inc in incidents:
        table_data.append([
            inc.custom_number or f"#{inc.id}",
            (inc.custom_date.strftime('%d.%m.%Y') if inc.custom_date else inc.created_at.strftime('%d.%m.%Y'))[:10],
            _get_category_display_export(inc.category)[:15],
            _get_severity_display_export(inc.severity)[:10],
            _get_status_display_export(inc.status)[:15],
            (inc.object.name if inc.object else "—")[:20],
            (f"{inc.employee.last_name} {inc.employee.first_name}".strip() if inc.employee else "—")[:20],
            f"{inc.damage_amount:,.2f}".replace(',', ' ') if inc.damage_amount else "—"
        ])
    
    table = Table(table_data, colWidths=[2*cm, 2*cm, 2.5*cm, 2*cm, 2.5*cm, 3*cm, 3*cm, 2*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), font_name),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), font_name),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    story.append(table)
    
    # Генерация PDF
    doc.build(story)
    buffer.seek(0)
    
    filename = f"incidents_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}.pdf"}
    )
