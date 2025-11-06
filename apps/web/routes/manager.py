"""Роуты для интерфейса управляющего."""

from typing import List, Dict, Any, Optional
from datetime import date, datetime, time, timedelta
from fastapi import APIRouter, Request, Depends, HTTPException, Query, Form, status
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from sqlalchemy import select, and_, desc, func
from core.database.session import get_async_session, get_db_session
from shared.services.role_service import RoleService
from apps.bot.services.user_service import UserService
from shared.services.manager_permission_service import ManagerPermissionService
from apps.web.utils.timezone_utils import web_timezone_helper
from shared.services.role_based_login_service import RoleBasedLoginService
from apps.web.middleware.role_middleware import require_manager_or_owner
from apps.web.dependencies import get_current_user_dependency
from shared.services.calendar_filter_service import CalendarFilterService
from domain.entities.user import User
from domain.entities.object import Object
from domain.entities.shift import Shift
from domain.entities.shift_schedule import ShiftSchedule
from domain.entities.time_slot import TimeSlot
from domain.entities.application import Application, ApplicationStatus
from core.logging.logger import logger
from apps.web.utils.applications_utils import get_new_applications_count
from domain.entities.manager_object_permission import ManagerObjectPermission
from domain.entities.contract import Contract
from urllib.parse import quote

router = APIRouter(prefix="/manager", tags=["manager"])
from apps.web.jinja import templates
@router.get("/incidents", response_class=HTMLResponse)
async def manager_incidents_index(
    request: Request,
    status: Optional[str] = Query(None),
    object_id: Optional[str] = Query(None),
    current_user: dict = Depends(require_manager_or_owner)
):
    if isinstance(current_user, RedirectResponse):
        return current_user
    async with get_async_session() as db:
        user_id = await get_user_id_from_current_user(current_user, db)
        perm = ManagerPermissionService(db)
        accessible_objects = await perm.get_user_accessible_objects(user_id)
        accessible_ids = [o.id for o in accessible_objects]
        from sqlalchemy.orm import selectinload
        from domain.entities.incident import Incident
        conditions = [Incident.object_id.in_(accessible_ids)]
        if status:
            conditions.append(Incident.status == status)
        if object_id:
            try:
                oid = int(object_id)
                if oid in accessible_ids:
                    conditions.append(Incident.object_id == oid)
                else:
                    # Нет доступа — пустой список
                    return templates.TemplateResponse("manager/incidents/index.html", {"request": request, "current_user": current_user, "incidents": [], "objects": accessible_objects, "status_filter": status, "selected_object_id": None})
            except ValueError:
                pass
        query = select(Incident).where(and_(*conditions)).options(
            selectinload(Incident.object),
            selectinload(Incident.employee),
            selectinload(Incident.shift_schedule),
            selectinload(Incident.creator)
        ).order_by(Incident.created_at.desc()).limit(200)
        result = await db.execute(query)
        incidents = result.scalars().all()
        return templates.TemplateResponse("manager/incidents/index.html", {"request": request, "current_user": current_user, "incidents": incidents, "objects": accessible_objects, "status_filter": status, "selected_object_id": object_id})


@router.get("/incidents/{incident_id}", response_class=HTMLResponse)
async def manager_incident_detail(
    request: Request,
    incident_id: int,
    current_user: dict = Depends(require_manager_or_owner)
):
    # Требование: сразу открывать форму редактирования
    if isinstance(current_user, RedirectResponse):
        return current_user
    return RedirectResponse(url=f"/manager/incidents/{incident_id}/edit", status_code=302)


@router.get("/incidents/create", response_class=HTMLResponse)
async def manager_incident_create_form(
    request: Request,
    current_user: dict = Depends(require_manager_or_owner)
):
    if isinstance(current_user, RedirectResponse):
        return current_user
    async with get_async_session() as db:
        user_id = await get_user_id_from_current_user(current_user, db)
        perm = ManagerPermissionService(db)
        accessible_objects = await perm.get_user_accessible_objects(user_id)
        return templates.TemplateResponse("manager/incidents/create.html", {"request": request, "objects": accessible_objects, "current_user": current_user})


@router.post("/incidents/create")
async def manager_incident_create(
    request: Request,
    current_user: dict = Depends(require_manager_or_owner)
):
    if isinstance(current_user, RedirectResponse):
        return current_user
    form = await request.form()
    category = (form.get("category") or "").strip()
    severity = (form.get("severity") or "medium").strip()
    notes = (form.get("notes") or "").strip()
    object_id = form.get("object_id")
    employee_id = form.get("employee_id")
    try:
        object_id_int = int(object_id) if object_id else None
    except ValueError:
        object_id_int = None
    try:
        employee_id_int = int(employee_id) if employee_id else None
    except ValueError:
        employee_id_int = None
    async with get_async_session() as db:
        user_id = await get_user_id_from_current_user(current_user, db)
        perm = ManagerPermissionService(db)
        accessible_objects = await perm.get_user_accessible_objects(user_id)
        accessible_ids = [o.id for o in accessible_objects]
        if not object_id_int or object_id_int not in accessible_ids:
            raise HTTPException(status_code=400, detail="Некорректный объект")
        from domain.entities.incident import Incident
        from sqlalchemy import insert
        values = {
            "category": category or None,
            "severity": severity or "medium",
            "status": "new",
            "object_id": object_id_int,
            "employee_id": employee_id_int,
            "notes": notes or None,
        }
        await db.execute(insert(Incident).values(**values))
        await db.commit()
    return RedirectResponse(url="/manager/incidents", status_code=303)


@router.get("/incidents/{incident_id}/edit", response_class=HTMLResponse)
async def manager_incident_edit_form(
    request: Request,
    incident_id: int,
    current_user: dict = Depends(require_manager_or_owner)
):
    if isinstance(current_user, RedirectResponse):
        return current_user
    async with get_async_session() as db:
        from domain.entities.incident import Incident
        from sqlalchemy import select
        from sqlalchemy.orm import selectinload
        user_id = await get_user_id_from_current_user(current_user, db)
        perm = ManagerPermissionService(db)
        accessible_objects = await perm.get_user_accessible_objects(user_id)
        accessible_ids = [o.id for o in accessible_objects]
        res = await db.execute(select(Incident).where(Incident.id == incident_id).options(selectinload(Incident.object)))
        incident = res.scalar_one_or_none()
        if not incident:
            raise HTTPException(status_code=404, detail="Not Found")
        if incident.object_id not in accessible_ids:
            raise HTTPException(status_code=403, detail="Access denied")
        # Список сотрудников по владельцам доступных объектов (активные контракты)
        from sqlalchemy import select
        from domain.entities.contract import Contract
        from domain.entities.user import User as UserEntity
        owner_ids = sorted({getattr(o, 'owner_id', None) for o in accessible_objects if getattr(o, 'owner_id', None) is not None})
        emp_ids = set()
        if owner_ids:
            emp_rows = await db.execute(
                select(Contract.employee_id).where(Contract.owner_id.in_(owner_ids), Contract.is_active == True).distinct()
            )
            emp_ids = {row[0] for row in emp_rows.all()}
        employees = []
        if emp_ids:
            emp_res = await db.execute(select(UserEntity).where(UserEntity.id.in_(list(emp_ids))).order_by(UserEntity.last_name, UserEntity.first_name))
            employees = emp_res.scalars().all()
        return templates.TemplateResponse("manager/incidents/edit.html", {"request": request, "incident": incident, "objects": accessible_objects, "employees": employees})


@router.post("/incidents/{incident_id}/edit")
async def manager_incident_edit(
    request: Request,
    incident_id: int,
    current_user: dict = Depends(require_manager_or_owner)
):
    if isinstance(current_user, RedirectResponse):
        return current_user
    form = await request.form()
    category = (form.get("category") or "").strip()
    severity = (form.get("severity") or "medium").strip()
    notes = (form.get("notes") or "").strip()
    damage_amount_str = (form.get("damage_amount") or "").strip()
    object_id = form.get("object_id")
    employee_id = form.get("employee_id")
    try:
        object_id_int = int(object_id) if object_id else None
    except ValueError:
        object_id_int = None
    try:
        employee_id_int = int(employee_id) if employee_id else None
    except ValueError:
        employee_id_int = None
    from decimal import Decimal, InvalidOperation
    damage_amount = None
    if damage_amount_str:
        try:
            damage_amount = Decimal(damage_amount_str)
        except InvalidOperation:
            damage_amount = None
    async with get_async_session() as db:
        user_id = await get_user_id_from_current_user(current_user, db)
        perm = ManagerPermissionService(db)
        accessible_objects = await perm.get_user_accessible_objects(user_id)
        accessible_ids = [o.id for o in accessible_objects]
        from domain.entities.incident import Incident
        from sqlalchemy import select, update
        res = await db.execute(select(Incident).where(Incident.id == incident_id))
        incident = res.scalar_one_or_none()
        if not incident:
            raise HTTPException(status_code=404, detail="Not Found")
        if object_id_int and object_id_int not in accessible_ids:
            raise HTTPException(status_code=403, detail="Access denied")
        upd = update(Incident).where(Incident.id == incident_id).values(
            category=category or None,
            severity=severity or "medium",
            notes=notes or None,
            object_id=object_id_int or incident.object_id,
            employee_id=employee_id_int if employee_id_int is not None else incident.employee_id,
            damage_amount=damage_amount if damage_amount is not None else incident.damage_amount
        )
        await db.execute(upd)
        await db.commit()
    return RedirectResponse(url=f"/manager/incidents/{incident_id}", status_code=303)


@router.post("/incidents/{incident_id}/status")
async def manager_incident_change_status(
    incident_id: int,
    status: str = Form(...),
    request: Request = None,
    current_user: dict = Depends(require_manager_or_owner)
):
    if isinstance(current_user, RedirectResponse):
        return current_user
    async with get_async_session() as db:
        from domain.entities.incident import Incident
        from sqlalchemy import select, update
        user_id = await get_user_id_from_current_user(current_user, db)
        perm = ManagerPermissionService(db)
        accessible_objects = await perm.get_user_accessible_objects(user_id)
        accessible_ids = [o.id for o in accessible_objects]
        res = await db.execute(select(Incident).where(Incident.id == incident_id))
        inc = res.scalar_one_or_none()
        if not inc:
            raise HTTPException(status_code=404, detail="Not Found")
        if inc.object_id not in accessible_ids:
            raise HTTPException(status_code=403, detail="Access denied")
        if status not in ["new", "in_review", "resolved", "rejected"]:
            raise HTTPException(status_code=400, detail="Некорректный статус")
        await db.execute(update(Incident).where(Incident.id == incident_id).values(status=status))
        await db.commit()
    return RedirectResponse(url=f"/manager/incidents/{incident_id}", status_code=303)


async def get_user_id_from_current_user(current_user, session: AsyncSession) -> Optional[int]:
    """Получает внутренний ID пользователя из current_user."""
    logger.info(f"get_user_id_from_current_user: current_user type = {type(current_user)}")
    logger.info(f"get_user_id_from_current_user: current_user = {current_user}")
    
    if isinstance(current_user, dict):
        from sqlalchemy import select
        # Пробуем получить telegram_id из разных полей
        telegram_id = current_user.get("telegram_id") or current_user.get("id")
        logger.info(f"get_user_id_from_current_user: telegram_id = {telegram_id}")
        
        if not telegram_id:
            logger.error("telegram_id is None or empty")
            return None
            
        user_query = select(User).where(User.telegram_id == telegram_id)
        user_result = await session.execute(user_query)
        user_obj = user_result.scalar_one_or_none()
        logger.info(f"get_user_id_from_current_user: user_obj = {user_obj}")
        return user_obj.id if user_obj else None
    else:
        logger.info(f"get_user_id_from_current_user: returning current_user.id = {current_user.id}")
        return current_user.id


@router.get("/", response_class=HTMLResponse)
async def manager_index(request: Request):
    """Главная страница управляющего - перенаправляет на дашборд."""
    return RedirectResponse(url="/manager/dashboard", status_code=302)


@router.get("/dashboard", response_class=HTMLResponse)
async def manager_dashboard(
    request: Request,
    current_user: dict = Depends(require_manager_or_owner)
):
    """Дашборд управляющего."""
    try:
        # Проверяем, что current_user не является RedirectResponse
        if isinstance(current_user, RedirectResponse):
            return current_user
        
        async with get_async_session() as db:
            user_id = await get_user_id_from_current_user(current_user, db)
            if not user_id:
                raise HTTPException(status_code=401, detail="Пользователь не найден")
            
            # Получаем сервисы
            role_service = RoleService(db)
            permission_service = ManagerPermissionService(db)
            login_service = RoleBasedLoginService(db)
            
            # Получаем доступные объекты
            accessible_objects = await permission_service.get_user_accessible_objects(user_id)
            
            # Получаем права на каждый объект
            object_permissions = {}
            for obj in accessible_objects:
                # Находим договор управляющего для этого объекта
                manager_contracts = await permission_service.get_manager_contracts_for_user(user_id)
                for contract in manager_contracts:
                    permission = await permission_service.get_permission(contract.id, obj.id)
                    if permission:
                        object_permissions[obj.id] = permission.get_permissions_dict()
                        break
            
            # Получаем статистику
            accessible_objects_count = len(accessible_objects)
            
            # Получаем смены по доступным объектам
            object_ids = [obj.id for obj in accessible_objects]
            active_shifts = []
            scheduled_shifts_today = []
            
            if object_ids:
                from sqlalchemy import select, and_, func
                from datetime import datetime, date, timezone
                
                # Активные смены
                active_shifts_query = select(Shift).where(
                    and_(
                        Shift.object_id.in_(object_ids),
                        Shift.status == "active"
                    )
                ).options(
                    selectinload(Shift.user),
                    selectinload(Shift.object)
                )
                result = await db.execute(active_shifts_query)
                active_shifts = result.scalars().all()
                
                # Запланированные смены на сегодня (из ShiftSchedule)
                today = date.today()
                
                scheduled_shifts_today_query = select(ShiftSchedule).where(
                    and_(
                        ShiftSchedule.object_id.in_(object_ids),
                        ShiftSchedule.status.in_(["planned", "confirmed"]),
                        func.date(ShiftSchedule.planned_start) == today
                    )
                )
                result = await db.execute(scheduled_shifts_today_query)
                scheduled_shifts_today = result.scalars().all()
            
            # Получаем сотрудников с активными договорами у владельцев, с которыми есть активный контракт у управляющего
            employees_user_ids = set()
            
            # Получаем договоры управляющего, чтобы узнать владельцев
            manager_contracts = await permission_service.get_manager_contracts_for_user(user_id)
            owner_ids = [contract.owner_id for contract in manager_contracts]
            
            if owner_ids:
                from domain.entities.contract import Contract
                # Получаем сотрудников, которые работают по активным договорам с этими владельцами
                employees_query = select(User.id).join(
                    Contract, User.id == Contract.employee_id
                ).where(
                    and_(
                        Contract.owner_id.in_(owner_ids),  # Договоры с владельцами управляющего
                        Contract.is_active == True,
                        Contract.status == "active",
                        User.id != user_id  # Исключаем самого управляющего
                    )
                ).distinct()
                
                employees_result = await db.execute(employees_query)
                employees_user_ids.update([row[0] for row in employees_result.all() if row[0]])
            
            # Получаем последние смены
            recent_shifts = active_shifts
            recent_shifts = sorted(recent_shifts, key=lambda x: x.start_time if x.start_time else datetime.min, reverse=True)[:10]
            
            # Названия прав
            permission_names = {
                "can_view": "Просмотр",
                "can_edit": "Редактирование",
                "can_delete": "Удаление",
                "can_manage_employees": "Управление сотрудниками",
                "can_view_finances": "Просмотр финансов",
                "can_edit_rates": "Редактирование ставок",
                "can_edit_schedule": "Редактирование расписания"
            }
            
            # Получаем открытые объекты
            from domain.entities.object_opening import ObjectOpening
            open_objects_query = select(ObjectOpening).where(
                and_(
                    ObjectOpening.object_id.in_(object_ids) if object_ids else False,
                    ObjectOpening.closed_at.is_(None)
                )
            ).options(
                selectinload(ObjectOpening.object),
                selectinload(ObjectOpening.opener)
            ).order_by(ObjectOpening.opened_at.desc())
            result = await db.execute(open_objects_query)
            open_objects = result.scalars().all()
            
            # Получаем данные для переключения интерфейсов
            manager_context = await get_manager_context(user_id, db)

            return templates.TemplateResponse("manager/dashboard.html", {
                "request": request,
                "current_user": current_user,
                "accessible_objects": accessible_objects,
                "accessible_objects_count": accessible_objects_count,
                "open_objects": open_objects,
                "active_shifts_count": len(active_shifts),
                "scheduled_shifts_count_today": len(scheduled_shifts_today),
                "employees_count": len(employees_user_ids),
                "recent_shifts": recent_shifts,
                "object_permissions": object_permissions,
                "permission_names": permission_names,
                **manager_context
            })
        
    except Exception as e:
        logger.error(f"Error in manager dashboard: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки дашборда")


@router.get("/objects", response_class=HTMLResponse)
async def manager_objects(
    request: Request,
    current_user: dict = Depends(require_manager_or_owner),
    q_name: Optional[str] = Query(None, description="Поиск по названию"),
    q_address: Optional[str] = Query(None, description="Поиск по адресу"),
    sort: Optional[str] = Query(None, description="Поле сортировки: name, address, hourly_rate, opening_time, closing_time, is_active, created_at"),
    order: Optional[str] = Query("asc", description="Порядок сортировки: asc, desc"),
    page: int = Query(1, ge=1, description="Номер страницы"),
    per_page: int = Query(25, ge=1, le=100, description="Количество на странице")
):
    """Список объектов управляющего с фильтрацией, сортировкой и пагинацией."""
    try:
        # Проверяем, что current_user не является RedirectResponse
        if isinstance(current_user, RedirectResponse):
            return current_user
        
        async with get_async_session() as db:
            user_id = await get_user_id_from_current_user(current_user, db)
            if not user_id:
                raise HTTPException(status_code=401, detail="Пользователь не найден")
            
            permission_service = ManagerPermissionService(db)
            login_service = RoleBasedLoginService(db)
            
            # Получаем доступные объекты
            accessible_objects = await permission_service.get_user_accessible_objects(user_id)
            
            # Обрабатываем объекты для отображения
            processed_objects = []
            for obj in accessible_objects:
                processed_obj = {
                    "id": obj.id,
                    "name": obj.name,
                    "address": obj.address or "",
                    "coordinates": obj.coordinates or "",
                    "hourly_rate": float(obj.hourly_rate),
                    "opening_time": obj.opening_time.strftime("%H:%M") if obj.opening_time else "",
                    "closing_time": obj.closing_time.strftime("%H:%M") if obj.closing_time else "",
                    "max_distance_meters": obj.max_distance_meters or 500,
                    "available_for_applicants": obj.available_for_applicants,
                    "is_active": obj.is_active,
                    "work_conditions": obj.work_conditions or "",
                    "employee_position": obj.employee_position or "",
                    "shift_tasks": obj.shift_tasks or [],
                    "work_days_mask": obj.work_days_mask,
                    "schedule_repeat_weeks": obj.schedule_repeat_weeks,
                    "timezone": obj.timezone,
                    "created_at": obj.created_at
                }
                processed_objects.append(processed_obj)
            
            # Фильтрация по названию и адресу (contains, case-insensitive)
            if q_name:
                qn = q_name.strip().lower()
                processed_objects = [o for o in processed_objects if qn in (o["name"] or "").lower()]
            if q_address:
                qa = q_address.strip().lower()
                processed_objects = [o for o in processed_objects if qa in (o["address"] or "").lower()]
            
            # Сортировка
            if sort:
                key_map = {
                    "name": lambda o: (o["name"] or "").lower(),
                    "address": lambda o: (o["address"] or "").lower(),
                    "hourly_rate": lambda o: o["hourly_rate"],
                    "opening_time": lambda o: o["opening_time"],
                    "closing_time": lambda o: o["closing_time"],
                    "is_active": lambda o: o["is_active"],
                    "created_at": lambda o: o["created_at"] or datetime.min,
                }
                key_func = key_map.get(sort)
                if key_func:
                    reverse = (order == "desc")
                    processed_objects = sorted(processed_objects, key=key_func, reverse=reverse)
            else:
                # По умолчанию сортировка по названию
                processed_objects = sorted(processed_objects, key=lambda o: (o["name"] or "").lower())
            
            # Пагинация - подсчет ПОСЛЕ всех фильтров
            total = len(processed_objects)
            pages = (total + per_page - 1) // per_page if total > 0 else 0
            start = (page - 1) * per_page
            end = start + per_page
            paginated_objects = processed_objects[start:end]
            
            # Получаем права на каждый объект
            object_permissions = {}
            for obj in accessible_objects:
                manager_contracts = await permission_service.get_manager_contracts_for_user(user_id)
                for contract in manager_contracts:
                    permission = await permission_service.get_permission(contract.id, obj.id)
                    if permission:
                        object_permissions[obj.id] = permission.get_permissions_dict()
                        break
            
            # Названия прав
            permission_names = {
                "can_view": "Просмотр",
                "can_edit": "Редактирование",
                "can_delete": "Удаление",
                "can_manage_employees": "Управление сотрудниками",
                "can_view_finances": "Просмотр финансов",
                "can_edit_rates": "Редактирование ставок",
                "can_edit_schedule": "Редактирование расписания"
            }
            
            manager_context = await get_manager_context(user_id, db)
            
            return templates.TemplateResponse("manager/objects.html", {
                "request": request,
                "current_user": current_user,
                "accessible_objects": paginated_objects,
                "object_permissions": object_permissions,
                "permission_names": permission_names,
                "filters": {
                    "q_name": q_name,
                    "q_address": q_address
                },
                "sort": {
                    "field": sort,
                    "order": order
                },
                "pagination": {
                    "page": page,
                    "per_page": per_page,
                    "total": total,
                    "pages": pages
                },
                **manager_context
            })
        
    except Exception as e:
        logger.error(f"Error in manager objects: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки объектов")


@router.get("/objects/{object_id}", response_class=HTMLResponse)
async def manager_object_detail(
    object_id: int,
    request: Request,
    current_user: dict = Depends(require_manager_or_owner)
):
    """Детальная информация об объекте."""
    try:
        # Проверяем, что current_user не является RedirectResponse
        if isinstance(current_user, RedirectResponse):
            return current_user
        
        async with get_async_session() as db:
            user_id = await get_user_id_from_current_user(current_user, db)
            if not user_id:
                raise HTTPException(status_code=401, detail="Пользователь не найден")
            
            permission_service = ManagerPermissionService(db)
            
            # Проверяем доступ к объекту
            accessible_objects = await permission_service.get_user_accessible_objects(user_id)
            if not any(obj.id == object_id for obj in accessible_objects):
                raise HTTPException(status_code=403, detail="Нет доступа к объекту")
            
            # Получаем объект
            from sqlalchemy import select
            object_query = select(Object).where(Object.id == object_id)
            result = await db.execute(object_query)
            obj = result.scalar_one_or_none()
            
            if not obj:
                raise HTTPException(status_code=404, detail="Объект не найден")
            
            # Получаем права на объект
            manager_contracts = await permission_service.get_manager_contracts_for_user(user_id)
            object_permission = None
            for contract in manager_contracts:
                permission = await permission_service.get_permission(contract.id, object_id)
                if permission:
                    object_permission = permission
                    break
            
            manager_context = await get_manager_context(user_id, db)
            
            return templates.TemplateResponse("manager/objects/detail.html", {
                "request": request,
                "current_user": current_user,
                "object": obj,
                "object_permission": object_permission,
                **manager_context
            })
        
    except Exception as e:
        logger.error(f"Error in manager object detail: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки объекта")


@router.get("/objects/{object_id}/edit", response_class=HTMLResponse)
async def manager_object_edit(
    request: Request,
    object_id: int,
    current_user: dict = Depends(require_manager_or_owner)
):
    """Форма редактирования объекта управляющим."""
    try:
        # Проверяем, что current_user не является RedirectResponse
        if isinstance(current_user, RedirectResponse):
            return current_user
        
        async with get_async_session() as db:
            user_id = await get_user_id_from_current_user(current_user, db)
            if not user_id:
                raise HTTPException(status_code=401, detail="Пользователь не найден")
            
            # Получаем сервисы
            permission_service = ManagerPermissionService(db)
            
            # Проверяем доступ к объекту
            accessible_objects = await permission_service.get_user_accessible_objects(user_id)
            obj = next((o for o in accessible_objects if o.id == object_id), None)
            
            if not obj:
                raise HTTPException(status_code=404, detail="Объект не найден или нет доступа")
            
            # Получаем права на объект
            manager_contracts = await permission_service.get_manager_contracts_for_user(user_id)
            object_permission = None
            for contract in manager_contracts:
                permission = await permission_service.get_permission(contract.id, object_id)
                if permission:
                    object_permission = permission
                    break
            
            # Проверяем права на редактирование
            if not object_permission or not object_permission.can_edit:
                raise HTTPException(status_code=403, detail="Нет прав на редактирование объекта")
            
            # Преобразуем в формат для шаблона
            object_data = {
                "id": obj.id,
                "name": obj.name,
                "address": obj.address or "",
                "coordinates": obj.coordinates or "",
                "hourly_rate": obj.hourly_rate,
                "opening_time": obj.opening_time.strftime("%H:%M") if obj.opening_time else "",
                "closing_time": obj.closing_time.strftime("%H:%M") if obj.closing_time else "",
                "max_distance": obj.max_distance_meters or 500,
                "work_conditions": obj.work_conditions or "",
                "employee_position": obj.employee_position or "",
                "shift_tasks": obj.shift_tasks or [],
                "available_for_applicants": obj.available_for_applicants,
                "is_active": obj.is_active,
                "work_days_mask": obj.work_days_mask,
                "schedule_repeat_weeks": obj.schedule_repeat_weeks
            }

            manager_context = await get_manager_context(user_id, db)

            return templates.TemplateResponse("manager/objects/edit.html", {
                "request": request,
                "title": f"Редактирование: {object_data['name']}",
                "object": object_data,
                "current_user": current_user,
                "object_permission": object_permission,
                **manager_context
            })
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in manager object edit: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки формы редактирования")


@router.post("/objects/{object_id}/edit")
async def manager_object_edit_post(
    request: Request,
    object_id: int,
    current_user: dict = Depends(require_manager_or_owner)
):
    """Обновление объекта управляющим."""
    try:
        # Проверяем, что current_user не является RedirectResponse
        if isinstance(current_user, RedirectResponse):
            return current_user
        
        async with get_async_session() as db:
            user_id = await get_user_id_from_current_user(current_user, db)
            if not user_id:
                raise HTTPException(status_code=401, detail="Пользователь не найден")
            
            # Получаем сервисы
            permission_service = ManagerPermissionService(db)
            
            # Проверяем доступ к объекту
            accessible_objects = await permission_service.get_user_accessible_objects(user_id)
            obj = next((o for o in accessible_objects if o.id == object_id), None)
            
            if not obj:
                raise HTTPException(status_code=404, detail="Объект не найден или нет доступа")
            
            # Получаем права на объект
            manager_contracts = await permission_service.get_manager_contracts_for_user(user_id)
            object_permission = None
            for contract in manager_contracts:
                permission = await permission_service.get_permission(contract.id, object_id)
                if permission:
                    object_permission = permission
                    break
            
            # Проверяем права на редактирование
            if not object_permission or not object_permission.can_edit:
                raise HTTPException(status_code=403, detail="Нет прав на редактирование объекта")
            
            # Получение данных формы
            form_data = await request.form()
            
            name = form_data.get("name", "").strip()
            address = form_data.get("address", "").strip()
            hourly_rate_str = form_data.get("hourly_rate", "").strip()
            opening_time = form_data.get("opening_time", "").strip()
            closing_time = form_data.get("closing_time", "").strip()
            timezone = form_data.get("timezone", "Europe/Moscow").strip()
            max_distance_str = form_data.get("max_distance", "500").strip()
            latitude_str = form_data.get("latitude", "").strip()
            longitude_str = form_data.get("longitude", "").strip()
            available_for_applicants = form_data.get("available_for_applicants") == "true"
            is_active = form_data.get("is_active") == "true"
            work_conditions = form_data.get("work_conditions", "").strip()
            employee_position = form_data.get("employee_position", "").strip()
            
            # Парсинг задач с новой структурой
            task_texts = form_data.getlist("task_texts[]")
            task_deductions = form_data.getlist("task_deductions[]")
            task_mandatory = form_data.getlist("task_mandatory[]")
            
            logger.info(f"Task parsing (manager edit) - texts: {task_texts}, deductions: {task_deductions}, mandatory: {task_mandatory}")
            
            shift_tasks = []
            for idx, text in enumerate(task_texts):
                if text.strip():
                    is_mandatory = str(idx) in task_mandatory
                    logger.info(f"Task {idx}: text='{text}', is_mandatory={is_mandatory}")
                    shift_tasks.append({
                        "text": text.strip(),
                        "is_mandatory": is_mandatory,
                        "deduction_amount": float(task_deductions[idx]) if idx < len(task_deductions) else 100.0
                    })
            
            # Получение дней недели (битовая маска)
            work_days_mask_str = form_data.get("work_days_mask", "0").strip()
            try:
                work_days_mask = int(work_days_mask_str)
            except ValueError:
                work_days_mask = 0
            
            schedule_repeat_weeks_str = form_data.get("schedule_repeat_weeks", "1").strip()
            
            logger.info(f"Updating object {object_id} for manager user {user_id}")
            
            # Валидация обязательных полей
            if not name:
                raise HTTPException(status_code=400, detail="Название объекта обязательно")
            if not address:
                raise HTTPException(status_code=400, detail="Адрес объекта обязателен")
            
            # Валидация и преобразование числовых полей
            if not hourly_rate_str:
                raise HTTPException(status_code=400, detail="Часовая ставка обязательна")
            
            try:
                # Поддержка запятой как десятичного разделителя ("500,00")
                normalized_rate = hourly_rate_str.replace(",", ".")
                hourly_rate = int(float(normalized_rate))
                if hourly_rate <= 0:
                    raise ValueError("Ставка должна быть больше 0")
            except ValueError:
                raise HTTPException(status_code=400, detail="Неверный формат ставки")
            
            try:
                max_distance = int(max_distance_str) if max_distance_str else 500
            except ValueError:
                raise HTTPException(status_code=400, detail="Неверный формат максимального расстояния")
            
            try:
                schedule_repeat_weeks = int(schedule_repeat_weeks_str) if schedule_repeat_weeks_str else 1
            except ValueError:
                raise HTTPException(status_code=400, detail="Неверный формат повторения")
            
            if hourly_rate <= 0:
                raise HTTPException(status_code=400, detail="Ставка должна быть больше 0")
            
            if max_distance <= 0:
                raise HTTPException(status_code=400, detail="Максимальное расстояние должно быть больше 0")
            
            if schedule_repeat_weeks <= 0:
                raise HTTPException(status_code=400, detail="Повторение должно быть больше 0")
            
            # Обработка координат
            coordinates = None
            if latitude_str and longitude_str:
                try:
                    lat = float(latitude_str)
                    lon = float(longitude_str)
                    coordinates = f"{lat},{lon}"
                except ValueError:
                    raise HTTPException(status_code=400, detail="Неверный формат координат")
            
            # Обработка времени
            from datetime import datetime, time
            opening_time_obj = None
            closing_time_obj = None
            
            if opening_time:
                try:
                    opening_time_obj = datetime.strptime(opening_time, "%H:%M").time()
                except ValueError:
                    raise HTTPException(status_code=400, detail="Неверный формат времени открытия")
            
            if closing_time:
                try:
                    closing_time_obj = datetime.strptime(closing_time, "%H:%M").time()
                except ValueError:
                    raise HTTPException(status_code=400, detail="Неверный формат времени закрытия")
            
            # Обновление объекта
            from apps.web.services.object_service import ObjectService
            object_service = ObjectService(db)
            
            # Подготавливаем данные для обновления
            update_data = {
                "name": name,
                "address": address,
                "hourly_rate": hourly_rate,
                "opening_time": opening_time_obj,
                "closing_time": closing_time_obj,
                "timezone": timezone,
                "max_distance_meters": max_distance,
                "coordinates": coordinates,
                "available_for_applicants": available_for_applicants,
                "is_active": is_active,
                "work_days_mask": work_days_mask,
                "schedule_repeat_weeks": schedule_repeat_weeks,
                "work_conditions": work_conditions,
                "employee_position": employee_position,
                "shift_tasks": shift_tasks
            }
            
            # Обновляем объект (используем метод для управляющих)
            updated_object = await object_service.update_object_by_manager(object_id, update_data)
            
            if not updated_object:
                raise HTTPException(status_code=500, detail="Ошибка обновления объекта")
            
            logger.info(f"Object {object_id} updated successfully by manager {user_id}")
            
            # Перенаправляем на страницу объекта
            return RedirectResponse(url=f"/manager/objects/{object_id}", status_code=302)
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating object {object_id}: {e}")
        raise HTTPException(status_code=500, detail="Ошибка обновления объекта")


@router.get("/employees", response_class=HTMLResponse)
async def manager_employees(
    request: Request,
    current_user: dict = Depends(require_manager_or_owner),
    q_name: Optional[str] = Query(None, description="Поиск по ФИО"),
    sort: Optional[str] = Query(None, description="Поле сортировки: name, phone, created_at"),
    order: Optional[str] = Query("asc", description="Порядок сортировки: asc, desc"),
    page: int = Query(1, ge=1, description="Номер страницы"),
    per_page: int = Query(25, ge=1, le=100, description="Количество на странице")
):
    """Список сотрудников управляющего с фильтрацией, сортировкой и пагинацией."""
    try:
        logger.info("Starting manager_employees function")
        
        # Проверяем, что current_user не является RedirectResponse
        if isinstance(current_user, RedirectResponse):
            logger.info("Current user is RedirectResponse, redirecting")
            return current_user
        
        async with get_async_session() as db:
            logger.info("Got database session")
            
            user_id = await get_user_id_from_current_user(current_user, db)
            logger.info(f"User ID: {user_id}")
            if not user_id:
                raise HTTPException(status_code=401, detail="Пользователь не найден")
            
            # Получаем сервисы
            logger.info("Creating services")
            permission_service = ManagerPermissionService(db)
            login_service = RoleBasedLoginService(db)
            
            # Получаем доступные объекты управляющего
            logger.info("Getting accessible objects")
            accessible_objects = await permission_service.get_user_accessible_objects(user_id)
            logger.info(f"Accessible objects count: {len(accessible_objects)}")
            logger.info(f"Accessible objects: {[obj.id for obj in accessible_objects]}")
            
            object_ids = [obj.id for obj in accessible_objects]
            
            # Получаем сотрудников, работающих на доступных объектах
            employees = []
            if object_ids:
                logger.info(f"Getting employees for object IDs: {object_ids}")
                from sqlalchemy import select, distinct
                from domain.entities.contract import Contract
                from domain.entities.user import User
                
                # Получаем всех сотрудников, работающих на доступных объектах
                from sqlalchemy import func, or_, text, cast, String, JSON, any_, exists
                
                # Используем простой SQL с text() для проверки пересечения массивов
                # Исключаем самого управляющего из списка сотрудников
                employees_query = select(User).join(
                    Contract, User.id == Contract.employee_id
                ).where(
                    Contract.is_active == True,
                    User.id != user_id,  # Исключаем самого управляющего
                    text("EXISTS (SELECT 1 FROM json_array_elements(contracts.allowed_objects) AS elem WHERE elem::text::int = ANY(ARRAY[{}]))".format(','.join(map(str, object_ids))))
                ).distinct()
                
                logger.info(f"Executing query: {employees_query}")
                logger.info(f"Query parameters: object_ids={object_ids}")
                
                result = await db.execute(employees_query)
                employees = result.scalars().all()
                logger.info(f"Found {len(employees)} employees")
                
                # Логируем найденных сотрудников
                for emp in employees:
                    logger.info(f"Employee: {emp.id} - {emp.first_name} {emp.last_name}")
            else:
                logger.info("No accessible objects, returning empty employee list")
            
            # Преобразуем в список словарей для фильтрации и сортировки
            employees_list = []
            for emp in employees:
                employees_list.append({
                    "id": emp.id,
                    "telegram_id": emp.telegram_id,
                    "first_name": emp.first_name or "",
                    "last_name": emp.last_name or "",
                    "username": emp.username or "",
                    "phone": emp.phone or "",
                    "role": emp.role or "",
                    "roles": emp.roles or [],
                    "is_active": getattr(emp, 'is_active', True),
                    "created_at": emp.created_at
                })
            
            # Фильтрация по ФИО (contains, case-insensitive)
            if q_name:
                qn = q_name.strip().lower()
                employees_list = [
                    e for e in employees_list
                    if qn in (f"{(e['last_name'] or '')} {(e['first_name'] or '')}".strip().lower())
                    or qn in (f"{(e['first_name'] or '')} {(e['last_name'] or '')}".strip().lower())
                    or qn in (e['username'] or '').lower()
                ]
            
            # Сортировка
            if sort:
                key_map = {
                    "name": lambda e: (e.get("last_name") or "", e.get("first_name") or ""),
                    "phone": lambda e: e.get("phone") or "",
                    "created_at": lambda e: e.get("created_at") if e.get("created_at") else datetime.min,
                }
                key_func = key_map.get(sort)
                if key_func:
                    reverse = (order == "desc")
                    employees_list = sorted(employees_list, key=key_func, reverse=reverse)
            else:
                # По умолчанию сортировка по Фамилии, затем Имени
                employees_list = sorted(employees_list, key=lambda e: ((e.get("last_name") or "").lower(), (e.get("first_name") or "").lower()))
            
            # Пагинация - подсчет ПОСЛЕ всех фильтров
            total = len(employees_list)
            pages = (total + per_page - 1) // per_page if total > 0 else 0
            start = (page - 1) * per_page
            end = start + per_page
            paginated_employees = employees_list[start:end]
            
            # Получаем данные для переключения интерфейсов
            logger.info("Getting available interfaces")
            available_interfaces = await login_service.get_available_interfaces(user_id)
            logger.info(f"Available interfaces: {available_interfaces}")
            
            logger.info("Rendering template")
            manager_context = await get_manager_context(user_id, db)
            return templates.TemplateResponse("manager/employees.html", {
                "request": request,
                "current_user": current_user,
                "employees": paginated_employees,
                "filters": {
                    "q_name": q_name
                },
                "sort": {
                    "field": sort,
                    "order": order
                },
                "pagination": {
                    "page": page,
                    "per_page": per_page,
                    "total": total,
                    "pages": pages
                },
                **manager_context
            })
        
    except Exception as e:
        logger.error(f"Error in manager employees: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки сотрудников")


@router.get("/employees/add", response_class=HTMLResponse)
async def manager_employee_add_form(
    request: Request,
    current_user: dict = Depends(require_manager_or_owner)
):
    """Форма добавления нового сотрудника."""
    try:
        if isinstance(current_user, RedirectResponse):
            return current_user
        
        async with get_async_session() as db:
            user_id = await get_user_id_from_current_user(current_user, db)
            if not user_id:
                raise HTTPException(status_code=401, detail="Пользователь не найден")
            
            # Получаем шаблоны договоров
            from sqlalchemy import select
            from domain.entities.contract import ContractTemplate
            
            templates_query = select(ContractTemplate).where(ContractTemplate.is_active == True)
            result = await db.execute(templates_query)
            contract_templates = result.scalars().all()
            
            # Получаем доступные объекты
            permission_service = ManagerPermissionService(db)
            accessible_objects = await permission_service.get_user_accessible_objects(user_id)
            
            manager_context = await get_manager_context(user_id, db)
            
            return templates.TemplateResponse("manager/employees/add.html", {
                "request": request,
                "current_user": current_user,
                "contract_templates": contract_templates,
                "accessible_objects": accessible_objects,
                **manager_context
            })
        
    except Exception as e:
        logger.error(f"Error in manager employee add form: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки формы")


@router.post("/employees/add")
async def manager_employee_add(
    request: Request,
    current_user: dict = Depends(require_manager_or_owner)
):
    """Создание нового сотрудника."""
    try:
        if isinstance(current_user, RedirectResponse):
            return current_user
        
        async with get_async_session() as db:
            user_id = await get_user_id_from_current_user(current_user, db)
            if not user_id:
                raise HTTPException(status_code=401, detail="Пользователь не найден")
            
            form_data = await request.form()
            
            # Получаем данные формы
            telegram_id = int(form_data.get("telegram_id"))
            first_name = form_data.get("first_name", "").strip()
            last_name = form_data.get("last_name", "").strip()
            username = form_data.get("username", "").strip()
            phone = form_data.get("phone", "").strip()
            email = form_data.get("email", "").strip()
            birth_date_str = form_data.get("birth_date", "").strip()
            role = form_data.get("role", "employee")
            is_active = form_data.get("is_active") == "true"
            
            # Дополнительные поля профиля (только для сотрудников)
            work_experience = form_data.get("work_experience", "").strip() if role == "employee" else None
            education = form_data.get("education", "").strip() if role == "employee" else None
            skills = form_data.get("skills", "").strip() if role == "employee" else None
            about = form_data.get("about", "").strip() if role == "employee" else None
            preferred_schedule = form_data.get("preferred_schedule", "").strip() if role == "employee" else None
            min_salary_str = form_data.get("min_salary", "").strip() if role == "employee" else None
            availability_notes = form_data.get("availability_notes", "").strip() if role == "employee" else None
            
            # Данные договора
            contract_template_id = form_data.get("contract_template")
            contract_objects = form_data.getlist("contract_objects")
            hourly_rate_str = form_data.get("hourly_rate", "").strip()
            hourly_rate = int(hourly_rate_str) if hourly_rate_str else None
            start_date_str = form_data.get("start_date")
            end_date_str = form_data.get("end_date")
            
            # Валидация
            if not first_name:
                raise HTTPException(status_code=400, detail="Имя обязательно")
            
            if not hourly_rate:
                raise HTTPException(status_code=400, detail="Часовая ставка обязательна")
            
            # Проверяем, существует ли уже пользователь с таким telegram_id
            from sqlalchemy import select
            from domain.entities.user import User
            
            existing_user_query = select(User).where(User.telegram_id == telegram_id)
            existing_user_result = await db.execute(existing_user_query)
            existing_user = existing_user_result.scalar_one_or_none()
            
            if existing_user:
                # Пользователь уже существует - обновляем его данные и создаем договор
                logger.info(f"User {telegram_id} already exists, updating data and creating contract")
                
                # Импорты для обновления пользователя
                from shared.services.role_service import RoleService
                from apps.web.services.contract_service import ContractService
                from datetime import datetime, date
                
                # Обновляем данные пользователя
                existing_user.first_name = first_name
                existing_user.last_name = last_name
                existing_user.username = username
                existing_user.phone = phone
                existing_user.email = email if email else existing_user.email
                
                # Обработка даты рождения
                if birth_date_str:
                    try:
                        from datetime import datetime
                        existing_user.birth_date = datetime.strptime(birth_date_str, "%Y-%m-%d")
                    except ValueError:
                        pass
                
                # Обновляем профиль сотрудника
                if role == "employee":
                    existing_user.work_experience = work_experience if work_experience else existing_user.work_experience
                    existing_user.education = education if education else existing_user.education
                    existing_user.skills = skills if skills else existing_user.skills
                    existing_user.about = about if about else existing_user.about
                    existing_user.preferred_schedule = preferred_schedule if preferred_schedule else existing_user.preferred_schedule
                    existing_user.availability_notes = availability_notes if availability_notes else existing_user.availability_notes
                    
                    # Обработка минимальной зарплаты
                    if min_salary_str and min_salary_str.isdigit():
                        existing_user.min_salary = int(min_salary_str)
                
                existing_user.is_active = is_active
                
                # Обновляем роли - добавляем employee если его нет
                if hasattr(existing_user, 'roles') and existing_user.roles:
                    if "employee" not in existing_user.roles:
                        existing_user.roles.append("employee")
                else:
                    existing_user.roles = ["applicant", "employee"]
                
                await db.commit()
                user = existing_user
                
            else:
                # Пользователь не существует - создаем нового
                logger.info(f"Creating new user {telegram_id}")
                
                # Создаем пользователя
                from shared.services.role_service import RoleService
                from apps.web.services.contract_service import ContractService
                from datetime import datetime, date
            
                # Обработка даты рождения
                birth_date = None
                if birth_date_str:
                    try:
                        birth_date = datetime.strptime(birth_date_str, "%Y-%m-%d")
                    except ValueError:
                        birth_date = None
                
                # Обработка минимальной зарплаты
                min_salary = None
                if min_salary_str and min_salary_str.isdigit():
                    min_salary = int(min_salary_str)
                
                user = User(
                    telegram_id=telegram_id,
                    first_name=first_name,
                    last_name=last_name,
                    username=username,
                    phone=phone,
                    email=email if email else None,
                    birth_date=birth_date,
                    work_experience=work_experience if work_experience else None,
                    education=education if education else None,
                    skills=skills if skills else None,
                    about=about if about else None,
                    preferred_schedule=preferred_schedule if preferred_schedule else None,
                    min_salary=min_salary,
                    availability_notes=availability_notes if availability_notes else None,
                    role=role,
                    roles=[role],
                    is_active=is_active
                )
                
                db.add(user)
                await db.commit()
                await db.refresh(user)
                
                # Добавляем роль
                role_service = RoleService(db)
                from domain.entities.user import UserRole
                await role_service.add_role(user.id, UserRole(role))
                
                logger.info(f"Created new user {user.id}")
            
            # Создаем договор (общий блок для существующих и новых пользователей)
            if contract_objects:
                # Получаем доступные объекты для проверки владельца
                permission_service = ManagerPermissionService(db)
                accessible_objects = await permission_service.get_user_accessible_objects(user_id)
                
                if not accessible_objects:
                    raise HTTPException(status_code=403, detail="Нет доступных объектов")
                
                # Преобразуем contract_objects в список целых чисел
                selected_object_ids = [int(obj_id) for obj_id in contract_objects]
                
                # Находим выбранные объекты среди доступных
                accessible_obj_dict = {obj.id: obj for obj in accessible_objects}
                selected_objects = [accessible_obj_dict[obj_id] for obj_id in selected_object_ids if obj_id in accessible_obj_dict]
                
                if not selected_objects:
                    raise HTTPException(status_code=403, detail="Выбранные объекты недоступны")
                
                # Проверяем, что все выбранные объекты принадлежат одному владельцу
                owner_ids = set(obj.owner_id for obj in selected_objects)
                if len(owner_ids) > 1:
                    raise HTTPException(status_code=400, detail="Все объекты должны принадлежать одному владельцу")
                
                # Получаем owner_id из выбранных объектов
                owner_id = selected_objects[0].owner_id
                
                # Парсим даты
                start_date = None
                end_date = None
                if start_date_str:
                    start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
                if end_date_str:
                    end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
            
                # Создаем договор напрямую
                from domain.entities.contract import Contract
                
                # Генерируем номер договора
                contract_service = ContractService()
                contract_number = await contract_service._generate_contract_number(owner_id)
                
                # Генерируем заголовок договора
                title = f"Трудовой Договор с {first_name} {last_name}".strip()
                
                contract = Contract(
                    contract_number=contract_number,
                    employee_id=user.id,
                    owner_id=owner_id,  # Используем владельца объекта, а не управляющего
                    template_id=int(contract_template_id) if contract_template_id and contract_template_id.strip() else None,
                    title=title,
                    hourly_rate=hourly_rate,
                    start_date=start_date,
                    end_date=end_date,
                    allowed_objects=[int(obj_id) for obj_id in contract_objects if obj_id and obj_id.strip()],
                    is_active=True,
                    status="active"
                )
                
                db.add(contract)
                await db.commit()
                await db.refresh(contract)
                
                # Обновляем роль пользователя на employee
                contract_service = ContractService()
                await contract_service._update_employee_role(db, user.id)
                
                # Создаем права управляющего на объекты если роль manager
                if role == "manager":
                    permission_service = ManagerPermissionService(db)
                    for obj_id in contract_objects:
                        await permission_service.create_permission(
                            contract_id=contract.id,
                            object_id=int(obj_id),
                            permissions={
                                "can_view": True,
                                "can_edit": True,
                                "can_delete": False,
                                "can_manage_employees": True,
                                "can_view_finances": True,
                                "can_edit_rates": True,
                                "can_edit_schedule": True
                            }
                        )
                
                logger.info(f"Created contract {contract.id} for user {user.id} by manager {user_id}")
            
            return RedirectResponse(url=f"/manager/employees/{user.id}", status_code=302)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating employee: {e}")
        raise HTTPException(status_code=500, detail="Ошибка создания сотрудника")


@router.get("/employees/{employee_id}", response_class=HTMLResponse)
async def manager_employee_detail(
    employee_id: int,
    request: Request,
    current_user: dict = Depends(require_manager_or_owner),
    db: AsyncSession = Depends(get_db_session)
):
    """Детальная информация о сотруднике."""
    try:
        if isinstance(current_user, RedirectResponse):
            return current_user
        
        # Получаем внутренний user_id управляющего
        user_id = await get_user_id_from_current_user(current_user, db)
        if not user_id:
            raise HTTPException(status_code=401, detail="Пользователь не найден")
        
        # Запрещаем управляющему просматривать свою страницу как сотрудника
        if employee_id == user_id:
            raise HTTPException(status_code=403, detail="Вы не можете управлять своими договорами через этот интерфейс")
        
        # Получаем информацию о сотруднике через сервис
        from apps.web.services.contract_service import ContractService
        contract_service = ContractService()
        employee_info = await contract_service.get_employee_by_id(employee_id, current_user["id"])
        
        if not employee_info:
            raise HTTPException(status_code=404, detail="У вас нет договоров с этим сотрудником")
        
        # Получаем права управляющего для каждого договора
        from shared.services.manager_permission_service import ManagerPermissionService
        from domain.entities.manager_object_permission import ManagerObjectPermission
        permission_service = ManagerPermissionService(db)
        
        # Получаем договоры управляющего (где is_manager=True)
        manager_contracts_query = select(Contract).where(
            Contract.employee_id == user_id,
            Contract.is_manager == True,
            Contract.is_active == True
        )
        manager_contracts_result = await db.execute(manager_contracts_query)
        manager_contracts = manager_contracts_result.scalars().all()
        
        # Получаем все права управляющего на объекты
        manager_contract_ids = [mc.id for mc in manager_contracts]
        
        # Для каждого договора сотрудника проверяем права управляющего
        for contract in employee_info["contracts"]:
            contract_objects = contract["allowed_objects"]  # Объекты из договора сотрудника
            
            # Проверяем, есть ли у управляющего право can_manage_employees на эти объекты
            if manager_contract_ids and contract_objects:
                permissions_query = select(ManagerObjectPermission).where(
                    ManagerObjectPermission.contract_id.in_(manager_contract_ids),
                    ManagerObjectPermission.object_id.in_(contract_objects),
                    ManagerObjectPermission.can_manage_employees == True
                )
                permissions_result = await db.execute(permissions_query)
                permissions = permissions_result.scalars().all()
                
                contract["can_manage_employees"] = len(permissions) > 0
            else:
                contract["can_manage_employees"] = False
        
        # Получаем данные для переключения интерфейсов
        from shared.services.role_based_login_service import RoleBasedLoginService
        login_service = RoleBasedLoginService(db)
        available_interfaces = await login_service.get_available_interfaces(user_id)
        
        return templates.TemplateResponse("manager/employees/detail.html", {
            "request": request,
            "current_user": current_user,
            "employee": employee_info,  # Полная информация из сервиса
            "contracts": employee_info["contracts"],  # Все договоры
            "accessible_objects": employee_info["accessible_objects"],  # Доступные объекты
            "available_interfaces": available_interfaces
        })
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in manager employee detail: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки сотрудника")


@router.get("/employees/{employee_id}/edit", response_class=HTMLResponse)
async def manager_employee_edit_form(
    employee_id: int,
    request: Request,
    current_user: dict = Depends(require_manager_or_owner),
    db: AsyncSession = Depends(get_db_session)
):
    """Форма редактирования сотрудника."""
    try:
        if isinstance(current_user, RedirectResponse):
            return current_user
        
        # Получаем внутренний user_id управляющего
        user_id = await get_user_id_from_current_user(current_user, db)
        if not user_id:
            raise HTTPException(status_code=401, detail="Пользователь не найден")
        
        # Получаем информацию о сотруднике через сервис
        from apps.web.services.contract_service import ContractService
        contract_service = ContractService()
        employee_info = await contract_service.get_employee_by_id(employee_id, current_user["id"])
        
        if not employee_info:
            raise HTTPException(status_code=404, detail="Сотрудник не найден или у вас нет прав на его редактирование")
        
        # Получаем данные для переключения интерфейсов
        from shared.services.role_based_login_service import RoleBasedLoginService
        login_service = RoleBasedLoginService(db)
        available_interfaces = await login_service.get_available_interfaces(user_id)
        
        return templates.TemplateResponse("manager/employees/edit.html", {
            "request": request,
            "current_user": current_user,
            "employee": employee_info,  # Только информация о сотруднике, без договоров
            "available_interfaces": available_interfaces
        })
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in manager employee edit form: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки формы")


@router.post("/employees/{employee_id}/edit")
async def manager_employee_edit(
    employee_id: int,
    request: Request,
    current_user: dict = Depends(require_manager_or_owner)
):
    """Обновление сотрудника."""
    try:
        if isinstance(current_user, RedirectResponse):
            return current_user
        
        async with get_async_session() as db:
            user_id = await get_user_id_from_current_user(current_user, db)
            if not user_id:
                raise HTTPException(status_code=401, detail="Пользователь не найден")
            
            form_data = await request.form()
            
            # Получаем сотрудника
            from sqlalchemy import select
            from domain.entities.user import User
            
            employee_query = select(User).where(User.id == employee_id)
            result = await db.execute(employee_query)
            employee = result.scalar_one_or_none()
            
            if not employee:
                raise HTTPException(status_code=404, detail="Сотрудник не найден")
            
            # Обновляем данные
            employee.first_name = form_data.get("first_name", "").strip()
            employee.last_name = form_data.get("last_name", "").strip()
            employee.username = form_data.get("username", "").strip()
            employee.phone = form_data.get("phone", "").strip()
            employee.email = form_data.get("email", "").strip() or None
            employee.is_active = form_data.get("is_active") == "true"
            
            # Обработка даты рождения
            birth_date_str = form_data.get("birth_date", "").strip()
            if birth_date_str:
                try:
                    employee.birth_date = datetime.strptime(birth_date_str, "%Y-%m-%d")
                except ValueError:
                    pass
            else:
                employee.birth_date = None
            
            # Дополнительные поля профиля (только для сотрудников)
            if employee.role == "employee" or (hasattr(employee, 'roles') and employee.roles and "employee" in employee.roles):
                employee.work_experience = form_data.get("work_experience", "").strip() or None
                employee.education = form_data.get("education", "").strip() or None
                employee.skills = form_data.get("skills", "").strip() or None
                employee.about = form_data.get("about", "").strip() or None
                employee.preferred_schedule = form_data.get("preferred_schedule", "").strip() or None
                employee.availability_notes = form_data.get("availability_notes", "").strip() or None
                
                # Обработка минимальной зарплаты
                min_salary_str = form_data.get("min_salary", "").strip()
                if min_salary_str and min_salary_str.isdigit():
                    employee.min_salary = int(min_salary_str)
                else:
                    employee.min_salary = None
            
            await db.commit()
            await db.refresh(employee)
            
            logger.info(f"Updated employee {employee_id} by manager {user_id}")
            
            return RedirectResponse(url=f"/manager/employees/{employee_id}", status_code=302)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating employee: {e}")
        raise HTTPException(status_code=500, detail="Ошибка обновления сотрудника")


@router.get("/employees/{employee_id}/shifts", response_class=HTMLResponse)
async def manager_employee_shifts(
    employee_id: int,
    request: Request,
    current_user: dict = Depends(require_manager_or_owner)
):
    """Смены сотрудника."""
    try:
        if isinstance(current_user, RedirectResponse):
            return current_user
        
        async with get_async_session() as db:
            user_id = await get_user_id_from_current_user(current_user, db)
            if not user_id:
                raise HTTPException(status_code=401, detail="Пользователь не найден")
            
            # Получаем сотрудника
            from sqlalchemy import select
            from domain.entities.user import User
            from domain.entities.shift import Shift
            from domain.entities.object import Object
            
            employee_query = select(User).where(User.id == employee_id)
            result = await db.execute(employee_query)
            employee = result.scalar_one_or_none()
            
            if not employee:
                raise HTTPException(status_code=404, detail="Сотрудник не найден")
            
            # Получаем смены
            shifts_query = select(Shift).where(
                Shift.user_id == employee_id
            ).order_by(Shift.start_time.desc()).options(
                selectinload(Shift.object)
            )
            result = await db.execute(shifts_query)
            shifts = result.scalars().all()
            
            # Получаем доступные объекты для фильтра
            permission_service = ManagerPermissionService(db)
            accessible_objects = await permission_service.get_user_accessible_objects(user_id)
            
            # Подсчитываем статистику
            total_hours = sum(shift.duration_hours or 0 for shift in shifts)
            total_earnings = sum(shift.total_payment or 0 for shift in shifts)
            
            login_service = RoleBasedLoginService(db)
            available_interfaces = await login_service.get_available_interfaces(user_id)
            
            return templates.TemplateResponse("manager/employees/shifts.html", {
                "request": request,
                "current_user": current_user,
                "employee": employee,
                "shifts": shifts,
                "objects": accessible_objects,
                "total_hours": total_hours,
                "total_earnings": total_earnings,
                "available_interfaces": available_interfaces
            })
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in manager employee shifts: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки смен")


@router.post("/employees/{employee_id}/terminate")
async def manager_employee_terminate(
    employee_id: int,
    request: Request,
    current_user: dict = Depends(require_manager_or_owner)
):
    """Расторжение договора с сотрудником."""
    try:
        if isinstance(current_user, RedirectResponse):
            return current_user
        
        async with get_async_session() as db:
            user_id = await get_user_id_from_current_user(current_user, db)
            if not user_id:
                raise HTTPException(status_code=401, detail="Пользователь не найден")
            
            form_data = await request.form()
            reason = form_data.get("reason", "").strip()
            
            if not reason:
                raise HTTPException(status_code=400, detail="Причина расторжения обязательна")
            
            # Запрещаем управляющему расторгать свой собственный договор
            if employee_id == user_id:
                raise HTTPException(status_code=403, detail="Вы не можете расторгать свой собственный договор через этот интерфейс")
            
            # Получаем активный договор
            from sqlalchemy import select
            from domain.entities.contract import Contract
            from apps.web.services.contract_service import ContractService
            
            contract_query = select(Contract).where(
                Contract.employee_id == employee_id,
                Contract.is_active == True
            )
            result = await db.execute(contract_query)
            contract = result.scalar_one_or_none()
            
            if not contract:
                raise HTTPException(status_code=404, detail="Активный договор не найден")
            
            # Расторгаем договор
            contract.status = "terminated"
            contract.is_active = False
            contract.terminated_at = datetime.now()
            contract.termination_reason = reason
            
            await db.commit()
            await db.refresh(contract)
            
            # Проверяем, есть ли у сотрудника другие активные договоры
            from domain.entities.user import User
            remaining_contracts_query = select(Contract).where(
                Contract.employee_id == employee_id,
                Contract.is_active == True,
                Contract.status == 'active'
            )
            remaining_result = await db.execute(remaining_contracts_query)
            remaining_contracts = remaining_result.scalars().all()
            
            # Если это был последний активный договор - делаем сотрудника неактивным
            if not remaining_contracts:
                user_query = select(User).where(User.id == employee_id)
                user_result = await db.execute(user_query)
                user = user_result.scalar_one_or_none()
                if user:
                    user.is_active = False
                    await db.commit()
                    logger.info(f"User {employee_id} marked as inactive (no active contracts)")
            
            logger.info(f"Terminated contract {contract.id} for employee {employee_id} by manager {user_id}")
            
            return RedirectResponse(url=f"/manager/employees/{employee_id}", status_code=302)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error terminating employee contract: {e}")
        raise HTTPException(status_code=500, detail="Ошибка расторжения договора")


@router.get("/employees/contract/{contract_id}/edit", response_class=HTMLResponse)
async def manager_contract_edit_form(
    contract_id: int,
    request: Request,
    current_user: dict = Depends(require_manager_or_owner),
    db: AsyncSession = Depends(get_db_session)
):
    """Форма редактирования договора управляющим."""
    try:
        if isinstance(current_user, RedirectResponse):
            return current_user
        
        # Получаем внутренний user_id управляющего
        user_id = await get_user_id_from_current_user(current_user, db)
        if not user_id:
            raise HTTPException(status_code=401, detail="Пользователь не найден")
        
        # Получаем договор через сервис
        from apps.web.services.contract_service import ContractService
        contract_service = ContractService()
        contract_info = await contract_service.get_contract_by_id_for_manager(contract_id, current_user["id"])
        
        if not contract_info:
            raise HTTPException(status_code=404, detail="Договор не найден или у вас нет прав на его редактирование")
        
        # Запрещаем управляющему редактировать свой собственный договор
        if contract_info['employee']['id'] == user_id:
            raise HTTPException(status_code=403, detail="Вы не можете редактировать свой собственный договор через этот интерфейс")
        
        # Получаем доступные объекты для управляющего
        from shared.services.manager_permission_service import ManagerPermissionService
        permission_service = ManagerPermissionService(db)
        accessible_objects = await permission_service.get_user_accessible_objects(user_id)
        
        # Преобразуем объекты в список словарей для шаблона
        accessible_objects_list = []
        for obj in accessible_objects:
            accessible_objects_list.append({
                'id': obj.id,
                'name': obj.name,
                'address': obj.address
            })
        
        # Получаем данные для переключения интерфейсов
        from shared.services.role_based_login_service import RoleBasedLoginService
        login_service = RoleBasedLoginService(db)
        available_interfaces = await login_service.get_available_interfaces(user_id)
        
        return templates.TemplateResponse("manager/contracts/edit.html", {
            "request": request,
            "current_user": current_user,
            "contract": contract_info,
            "accessible_objects": accessible_objects_list,
            "available_interfaces": available_interfaces
        })
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in manager contract edit form: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки формы редактирования договора")


@router.post("/employees/{employee_id}/contract/{contract_id}/edit")
async def manager_contract_edit(
    employee_id: int,
    contract_id: int,
    request: Request,
    current_user: dict = Depends(require_manager_or_owner)
):
    """Редактирование договора сотрудника."""
    try:
        if isinstance(current_user, RedirectResponse):
            return current_user
        
        # Получаем внутренний user_id управляющего
        async with get_async_session() as db:
            user_id = await get_user_id_from_current_user(current_user, db)
            if not user_id:
                raise HTTPException(status_code=401, detail="Пользователь не найден")
            
            # Запрещаем управляющему редактировать свой собственный договор
            if employee_id == user_id:
                raise HTTPException(status_code=403, detail="Вы не можете редактировать свой собственный договор через этот интерфейс")
        
        form_data = await request.form()
        
        # Получаем данные формы
        title = form_data.get("title", "").strip()
        hourly_rate_str = form_data.get("hourly_rate", "").strip()
        hourly_rate = float(hourly_rate_str) if hourly_rate_str else None
        start_date_str = form_data.get("start_date")
        end_date_str = form_data.get("end_date")
        status = form_data.get("status", "active")
        content = form_data.get("content", "").strip()
        
        # Обработка объектов
        allowed_objects = []
        for key, value in form_data.items():
            if key == "allowed_objects" and value:
                allowed_objects.append(int(value))
        
        # Обработка прав управляющего
        is_manager = form_data.get("is_manager") == "true"
        manager_permissions = {}
        for key, value in form_data.items():
            if key == "manager_permissions" and value:
                manager_permissions[value] = True
        
        # Валидация
        if not title:
            raise HTTPException(status_code=400, detail="Название договора обязательно")
        
        if not content:
            raise HTTPException(status_code=400, detail="Содержание договора обязательно")
        
        # Обработка дат
        start_date = None
        end_date = None
        if start_date_str:
            try:
                from datetime import datetime
                start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
            except ValueError:
                pass
        
        if end_date_str:
            try:
                from datetime import datetime
                end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
            except ValueError:
                pass
        
        # Обновляем договор через сервис
        from apps.web.services.contract_service import ContractService
        contract_service = ContractService()
        
        update_data = {
            "title": title,
            "hourly_rate": hourly_rate,
            "start_date": start_date,
            "end_date": end_date,
            "status": status,
            "content": content,
            "allowed_objects": allowed_objects,
            "is_manager": is_manager,
            "manager_permissions": manager_permissions
        }
        
        success = await contract_service.update_contract_for_manager(
            contract_id, 
            current_user["id"], 
            update_data
        )
        
        if not success:
            raise HTTPException(status_code=404, detail="Договор не найден или у вас нет прав на его редактирование")
        
        logger.info(f"Updated contract {contract_id} for employee {employee_id} by manager {user_id}")
        
        return RedirectResponse(url=f"/manager/employees/{employee_id}", status_code=302)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating contract: {e}")
        raise HTTPException(status_code=500, detail="Ошибка обновления договора")


@router.get("/calendar", response_class=HTMLResponse)
async def manager_calendar(
    request: Request,
    year: int = Query(None),
    month: int = Query(None),
    object_id: int = Query(None),
    current_user: dict = Depends(require_manager_or_owner),
):
    """Календарь управляющего."""
    try:
        from datetime import date, timedelta
        
        logger.info("Starting manager_calendar function")
        
        # Проверяем, что current_user не является RedirectResponse
        if isinstance(current_user, RedirectResponse):
            logger.info("Current user is RedirectResponse, redirecting")
            return current_user
        
        async with get_async_session() as db:
            # Получаем доступные объекты
            from shared.services.object_access_service import ObjectAccessService
            object_service = ObjectAccessService(db)
            accessible_objects = await object_service.get_accessible_objects(
                user_telegram_id=current_user.get("telegram_id") or current_user.get("id"),
                user_role=current_user.get("role")
            )
            logger.info(f"Found {len(accessible_objects)} accessible objects")
        
        # Получаем доступные интерфейсы
        available_interfaces = []
        if current_user.get("role") == "owner":
            available_interfaces.append({
                'name': 'owner',
                'title': 'Владелец',
                'description': 'Полное управление',
                'icon': '👑',
                'url': '/owner/dashboard',
                'priority': 1
            })
        if current_user.get("role") == "manager" or "manager" in current_user.get("roles", []):
            available_interfaces.append({
                'name': 'manager',
                'title': 'Управляющий',
                'description': 'Управление по правам',
                'icon': '👨‍💼',
                'url': '/manager/dashboard',
                'priority': 4
            })
        if current_user.get("role") == "employee" or "employee" in current_user.get("roles", []):
            available_interfaces.append({
                'name': 'employee',
                'title': 'Сотрудник',
                'description': 'Работа на объектах',
                'icon': '👷',
                'url': '/employee/',
                'priority': 5
            })
        
        # Формируем заголовок календаря
        current_year = year or date.today().year
        current_month = month or date.today().month
        month_names = [
            "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
            "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"
        ]
        calendar_title = f"{month_names[current_month - 1]} {current_year}"
        
        logger.info(f"Available interfaces: {available_interfaces}")
        logger.info("Rendering template")
        
        return templates.TemplateResponse("manager/calendar.html", {
            "request": request,
            "current_user": current_user,
            "available_interfaces": available_interfaces,
            "objects": accessible_objects,
            "selected_object_id": object_id,
            "calendar_title": calendar_title,
            "show_today_button": True,
            "object_id": object_id,
            "year": current_year,
            "month": current_month,
            "date_range": {
                "start": (date.today().replace(day=1) - timedelta(days=30)).isoformat(),
                "end": (date.today().replace(day=1) + timedelta(days=60)).isoformat()
            }
        })
        
    except Exception as e:
        logger.error(f"Error in manager calendar: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки календаря")


@router.get("/calendar/api/data")
async def manager_calendar_api_data(
    start_date: str = Query(..., description="Начальная дата в формате YYYY-MM-DD"),
    end_date: str = Query(..., description="Конечная дата в формате YYYY-MM-DD"),
    object_ids: Optional[str] = Query(None, description="ID объектов для фильтрации через запятую"),
    current_user: dict = Depends(require_manager_or_owner),
    db: AsyncSession = Depends(get_db_session)
):
    """
    Новый универсальный API для получения данных календаря управляющего.
    Использует CalendarFilterService для правильной фильтрации смен.
    """
    try:
        # Генерируем ключ кэша для всего response
        import hashlib
        user_id = current_user.get("id") if isinstance(current_user, dict) else current_user.telegram_id
        cache_key_data = f"calendar_api:{user_id}:{start_date}:{end_date}:{object_ids or 'all'}"
        cache_key = hashlib.md5(cache_key_data.encode()).hexdigest()
        
        # Проверяем кэш (TTL 2 минуты для API)
        from core.cache.redis_cache import cache
        cached_response = await cache.get(f"api_response:{cache_key}", serialize="json")
        if cached_response:
            logger.info(f"Manager calendar API: cache HIT for {start_date} to {end_date}")
            return cached_response
        
        logger.info(f"Manager calendar API: cache MISS, fetching data for {start_date} to {end_date}, object_ids={object_ids}")
        
        # Парсим даты
        try:
            start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
            end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="Неверный формат даты. Используйте YYYY-MM-DD")
        
        # Парсим object_ids
        object_filter_list = None
        if object_ids:
            try:
                object_filter_list = [int(obj_id.strip()) for obj_id in object_ids.split(',') if obj_id.strip()]
            except ValueError:
                raise HTTPException(status_code=400, detail="Неверный формат object_ids. Используйте числа через запятую")
        
        # Получаем роль пользователя
        if isinstance(current_user, dict):
            user_role = current_user.get("role", "manager")
            user_telegram_id = current_user.get("id")
        elif hasattr(current_user, 'role'):
            # current_user - это объект User
            user_role = current_user.role
            user_telegram_id = current_user.telegram_id
        else:
            # current_user - это RedirectResponse (не аутентифицирован)
            raise HTTPException(status_code=401, detail="Требуется аутентификация")
        
        # Если пользователь - владелец, используем роль owner для CalendarFilterService
        if user_role == "owner":
            user_role = "owner"
        else:
            user_role = "manager"
        
        if not user_telegram_id:
            raise HTTPException(status_code=401, detail="Пользователь не найден")
        
        logger.info(f"Using user_telegram_id={user_telegram_id}, user_role={user_role}")
        
        # Получаем данные календаря через универсальный сервис
        calendar_service = CalendarFilterService(db)
        logger.info(f"Calling CalendarFilterService with user_telegram_id={user_telegram_id}, user_role={user_role}")
        
        calendar_data = await calendar_service.get_calendar_data(
            user_telegram_id=user_telegram_id,
            user_role=user_role,
            date_range_start=start_date_obj,
            date_range_end=end_date_obj,
            object_filter=object_filter_list
        )
        
        logger.info(f"CalendarFilterService returned: {len(calendar_data.timeslots)} timeslots, {len(calendar_data.shifts)} shifts")
        
        # Преобразуем в формат, совместимый с существующим JavaScript
        logger.info("Converting timeslots to API format")
        timeslots_data = []
        for ts in calendar_data.timeslots:
                        timeslots_data.append({
                "id": ts.id,
                "object_id": ts.object_id,
                "object_name": ts.object_name,
                "date": ts.date.isoformat(),
                "start_time": ts.start_time.strftime("%H:%M"),
                "end_time": ts.end_time.strftime("%H:%M"),
                "hourly_rate": ts.hourly_rate,
                "max_employees": ts.max_employees,
                "current_employees": ts.current_employees,
                "available_slots": ts.available_slots,
                "status": ts.status.value,
                "is_active": ts.is_active,
                "notes": ts.notes,
                "work_conditions": ts.work_conditions,
                "shift_tasks": ts.shift_tasks,
                "coordinates": ts.coordinates,
                "can_edit": ts.can_edit,
                "can_plan": ts.can_plan,
                "can_view": ts.can_view
            })
        
        logger.info("Converting shifts to API format")
        shifts_data = []
        for s in calendar_data.shifts:
            # Получаем часовой пояс объекта
            object_timezone = s.timezone if hasattr(s, 'timezone') and s.timezone else 'Europe/Moscow'
            import pytz
            tz = pytz.timezone(object_timezone)
            
            # Конвертируем время в локальное время объекта
            def convert_to_local_time(utc_time):
                if utc_time:
                    # Если время уже имеет timezone info, конвертируем
                    if utc_time.tzinfo:
                        return utc_time.astimezone(tz).replace(tzinfo=None)
                    else:
                        # Если время naive, считаем его UTC и конвертируем
                        utc_aware = pytz.UTC.localize(utc_time)
                        return utc_aware.astimezone(tz).replace(tzinfo=None)
                return None
            
            shifts_data.append({
                "id": s.id,
                "user_id": s.user_id,
                "user_name": s.user_name,
                "object_id": s.object_id,
                "object_name": s.object_name,
                "time_slot_id": s.time_slot_id,
                "start_time": convert_to_local_time(s.start_time).isoformat() if s.start_time else None,
                "end_time": convert_to_local_time(s.end_time).isoformat() if s.end_time else None,
                "planned_start": convert_to_local_time(s.planned_start).isoformat() if s.planned_start else None,
                "planned_end": convert_to_local_time(s.planned_end).isoformat() if s.planned_end else None,
                "shift_type": s.shift_type.value,
                "status": s.status.value,
                "hourly_rate": s.hourly_rate,
                "total_hours": s.total_hours,
                "total_payment": s.total_payment,
                "notes": s.notes,
                "is_planned": s.is_planned,
                "schedule_id": s.schedule_id,
                "actual_shift_id": s.actual_shift_id,
                "start_coordinates": s.start_coordinates,
                "end_coordinates": s.end_coordinates,
                "can_edit": s.can_edit,
                "can_cancel": s.can_cancel,
                "can_view": s.can_view
            })
        
        logger.info("Preparing response")
        response_data = {
            "timeslots": timeslots_data,
            "shifts": shifts_data,
            "objects": [
                {
                    "id": obj['id'],
                    "name": obj['name'],
                    "timezone": obj.get('timezone', 'Europe/Moscow')
                }
                for obj in calendar_data.accessible_objects
            ],
            "date_range": {
                "start": start_date_obj.isoformat(),
                "end": end_date_obj.isoformat()
            },
            "user_role": user_role,
            "total_timeslots": len(timeslots_data),
            "total_shifts": len(shifts_data)
        }
        
        # Сохраняем в кэш (TTL 2 минуты)
        await cache.set(f"api_response:{cache_key}", response_data, ttl=120, serialize="json")
        logger.info(f"Manager calendar API: response cached with key {cache_key}")
        
        return response_data
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting manager calendar data: {e}")
        raise HTTPException(status_code=500, detail="Ошибка получения данных календаря")


@router.get("/calendar/api/timeslots-status")
async def get_timeslots_status_manager(
    year: int = Query(...),
    month: int = Query(...),
    object_id: Optional[int] = Query(None),
    current_user: dict = Depends(get_current_user_dependency()),
    db: AsyncSession = Depends(get_db_session)
):
    """Получение статуса тайм-слотов для календаря управляющего"""
    try:
        logger.info(f"Getting timeslots status for manager: {year}-{month}, object_id: {object_id}")
        
        # Получаем реальные тайм-слоты из базы
        from sqlalchemy import select, and_
        from sqlalchemy.orm import selectinload
        from domain.entities.time_slot import TimeSlot
        from domain.entities.object import Object
        from domain.entities.user import User
        
        # Получаем управляющего из текущего пользователя (по telegram_id)
        if not current_user or not getattr(current_user, "telegram_id", None):
            return []
        manager_query = select(User).where(User.telegram_id == current_user.telegram_id)
        manager_result = await db.execute(manager_query)
        manager = manager_result.scalar_one_or_none()
        
        if not manager:
            return []
        
        # Получаем доступные объекты управляющего
        permission_service = ManagerPermissionService(db)
        accessible_objects = await permission_service.get_user_accessible_objects(manager.id)
        object_ids = [obj.id for obj in accessible_objects]
        
        if not object_ids:
            return []
        
        # Фильтруем по выбранному объекту, если указан
        if object_id:
            if object_id not in object_ids:
                raise HTTPException(status_code=403, detail="Нет доступа к объекту")
            object_ids = [object_id]
        
        # Получаем тайм-слоты с текущего месяца до конца года (как у владельца)
        start_date = date(year, month, 1)
        end_date = date(year, 12, 31)
        
        timeslots_query = select(TimeSlot).options(
            selectinload(TimeSlot.object)
        ).where(
            and_(
                TimeSlot.object_id.in_(object_ids),
                TimeSlot.slot_date >= start_date,
                TimeSlot.slot_date < end_date,
                TimeSlot.is_active == True
            )
        ).order_by(TimeSlot.slot_date, TimeSlot.start_time)
        
        timeslots_result = await db.execute(timeslots_query)
        timeslots = timeslots_result.scalars().all()
        
        logger.info(f"Found {len(timeslots)} real timeslots for manager")
        
        # Получаем запланированные смены за месяц
        from domain.entities.shift_schedule import ShiftSchedule
        
        scheduled_shifts_query = select(ShiftSchedule).where(
            and_(
                ShiftSchedule.object_id.in_(object_ids),
                ShiftSchedule.planned_start >= start_date,
                ShiftSchedule.planned_start < end_date
            )
        ).order_by(ShiftSchedule.planned_start)
        
        scheduled_shifts_result = await db.execute(scheduled_shifts_query)
        scheduled_shifts = scheduled_shifts_result.scalars().all()
        
        logger.info(f"Found {len(scheduled_shifts)} scheduled shifts for manager")
        
        # Получаем отработанные смены за месяц
        from domain.entities.shift import Shift
        
        actual_shifts_query = select(Shift).options(
            selectinload(Shift.user)
        ).where(
            and_(
                Shift.object_id.in_(object_ids),
                Shift.start_time >= start_date,
                Shift.start_time < end_date
            )
        ).order_by(Shift.start_time)
        
        actual_shifts_result = await db.execute(actual_shifts_query)
        actual_shifts = actual_shifts_result.scalars().all()
        
        logger.info(f"Found {len(actual_shifts)} actual shifts for manager")
        
        # Создаем простой ответ для тестирования
        test_data = []
        for slot in timeslots:
            # Считаем смены для этого тайм-слота
            scheduled_shifts_for_slot = [s for s in scheduled_shifts if s.time_slot_id == slot.id]
            actual_shifts_for_slot = [s for s in actual_shifts if s.time_slot_id == slot.id]
            
            max_slots = slot.max_employees if slot.max_employees is not None else 1
            total_shifts = len(scheduled_shifts_for_slot) + len(actual_shifts_for_slot)
            
            if total_shifts < max_slots:
                status = "available"
                availability = f"{total_shifts}/{max_slots}"
            elif total_shifts == max_slots:
                status = "full"
                availability = f"{total_shifts}/{max_slots}"
            else:
                status = "overbooked"
                availability = f"{total_shifts}/{max_slots}"
            
            test_data.append({
                "slot_id": slot.id,
                "object_id": slot.object_id,
                "object_name": slot.object.name,
                "date": slot.slot_date.isoformat(),
                "start_time": slot.start_time.strftime("%H:%M"),
                "end_time": slot.end_time.strftime("%H:%M"),
                "hourly_rate": float(slot.hourly_rate) if slot.hourly_rate else 0,
                "status": status,
                "scheduled_shifts": [{"id": s.id, "status": s.status} for s in scheduled_shifts_for_slot],
                "actual_shifts": [{"id": s.id, "status": s.status} for s in actual_shifts_for_slot],
                "availability": availability,
                "occupied_slots": total_shifts,
                "max_slots": max_slots,
                "max_employees": max_slots
            })
        
        return test_data
        
    except Exception as e:
        logger.error(f"Error getting timeslots status for manager: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки статуса тайм-слотов")


# Удалена сломанная функция - используется корректная реализация ниже


@router.get("/calendar/api/timeslot/{timeslot_id}")
async def get_timeslot_details_manager(
    timeslot_id: int,
    current_user: dict = Depends(require_manager_or_owner)
):
    """Детали конкретного тайм-слота для управляющего"""
    try:
        from sqlalchemy import select, and_
        from sqlalchemy.orm import selectinload
        from domain.entities.time_slot import TimeSlot
        from domain.entities.object import Object
        from domain.entities.user import User
        from domain.entities.shift_schedule import ShiftSchedule
        from domain.entities.shift import Shift

        # Проверяем, что current_user не является RedirectResponse
        if isinstance(current_user, RedirectResponse):
            raise HTTPException(status_code=401, detail="Необходима авторизация")
        
        async with get_async_session() as db:
            # Получаем user_id
            user_id = await get_user_id_from_current_user(current_user, db)
            if not user_id:
                raise HTTPException(status_code=401, detail="Пользователь не найден")
            
            # Получаем доступные объекты управляющего
            permission_service = ManagerPermissionService(db)
            accessible_objects = await permission_service.get_user_accessible_objects(user_id)
            accessible_object_ids = [obj.id for obj in accessible_objects]
            
            if not accessible_object_ids:
                raise HTTPException(status_code=403, detail="Нет доступных объектов")

            # Получаем слот и проверяем доступ
            slot_q = select(TimeSlot).options(selectinload(TimeSlot.object)).where(TimeSlot.id == timeslot_id)
            slot = (await db.execute(slot_q)).scalar_one_or_none()
            if not slot:
                raise HTTPException(status_code=404, detail="Тайм-слот не найден")
            
            if slot.object_id not in accessible_object_ids:
                raise HTTPException(status_code=403, detail="Нет доступа к тайм-слоту")

            # Запланированные смены по тайм-слоту
            sched_q = select(ShiftSchedule).options(selectinload(ShiftSchedule.user)).where(ShiftSchedule.time_slot_id == timeslot_id).order_by(ShiftSchedule.planned_start)
            sched = (await db.execute(sched_q)).scalars().all()
            scheduled = [
                {
                    "id": s.id,
                    "user_id": s.user_id,
                    "user_name": f"{s.user.first_name} {s.user.last_name or ''}".strip() if s.user else None,
                    "status": s.status,
                    "start_time": s.planned_start.time().strftime("%H:%M"),
                    "end_time": s.planned_end.time().strftime("%H:%M"),
                    "notes": s.notes,
                }
                for s in sched
            ]

            # Фактические смены
            act_q = select(Shift).options(selectinload(Shift.user)).where(Shift.time_slot_id == timeslot_id).order_by(Shift.start_time)
            acts_linked = (await db.execute(act_q)).scalars().all()

            # Плюс спонтанные/несвязанные: по объекту и дате слота, с пересечением времени
            from datetime import datetime, time
            day_start = datetime.combine(slot.slot_date, time.min)
            day_end = datetime.combine(slot.slot_date, time.max)
            act_day_q = select(Shift).options(selectinload(Shift.user)).where(
                and_(
                    Shift.object_id == slot.object_id,
                    Shift.start_time >= day_start,
                    Shift.start_time <= day_end,
                )
            ).order_by(Shift.start_time)
            acts_day = (await db.execute(act_day_q)).scalars().all()

            # Отбираем пересекающиеся по времени и не дублируем
            def is_overlap(sh: Shift) -> bool:
                sh_start = sh.start_time.time()
                sh_end = sh.end_time.time() if sh.end_time else None
                if sh_end is None:
                    return sh_start < slot.end_time
                return (sh_start < slot.end_time) and (slot.start_time < sh_end)

            linked_ids = {sh.id for sh in acts_linked}
            acts_overlap = [sh for sh in acts_day if (sh.id not in linked_ids) and is_overlap(sh)]

            acts_all = acts_linked + acts_overlap
            actual = [
                {
                    "id": sh.id,
                    "user_id": sh.user_id,
                    "user_name": f"{sh.user.first_name} {sh.user.last_name or ''}".strip() if sh.user else None,
                    "status": sh.status,
                    "start_time": sh.start_time.time().strftime("%H:%M"),
                    "end_time": sh.end_time.time().strftime("%H:%M") if sh.end_time else None,
                    "total_hours": float(sh.total_hours) if sh.total_hours else None,
                    "total_payment": float(sh.total_payment) if sh.total_payment else None,
                    "is_planned": sh.is_planned,
                    "notes": sh.notes,
                }
                for sh in acts_all
            ]

            return {
                "slot": {
                    "id": slot.id,
                    "object_id": slot.object_id,
                    "object_name": slot.object.name if slot.object else None,
                    "date": slot.slot_date.strftime("%Y-%m-%d"),
                    "start_time": slot.start_time.strftime("%H:%M"),
                    "end_time": slot.end_time.strftime("%H:%M"),
                    "hourly_rate": float(slot.hourly_rate) if slot.hourly_rate else None,
                    "max_employees": slot.max_employees or 1,
                    "is_active": slot.is_active,
                    "notes": slot.notes or "",
                },
                "scheduled": scheduled,
                "actual": actual,
            }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting timeslot details for manager: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки деталей тайм-слота")


@router.get("/timeslots/{timeslot_id}", response_class=HTMLResponse)
async def manager_timeslot_detail(
    request: Request,
    timeslot_id: int,
    current_user: dict = Depends(require_manager_or_owner),
    db: AsyncSession = Depends(get_db_session)
):
    """Детали тайм-слота управляющего"""
    try:
        # Проверяем, что current_user - это словарь, а не RedirectResponse
        if isinstance(current_user, RedirectResponse):
            return current_user
            
        # Получаем внутренний ID пользователя
        user_id = await get_user_id_from_current_user(current_user, db)
        if not user_id:
            raise HTTPException(status_code=400, detail="Пользователь не найден")
        
        # Получаем доступные объекты управляющего
        permission_service = ManagerPermissionService(db)
        accessible_objects = await permission_service.get_user_accessible_objects(user_id)
        accessible_object_ids = [obj.id for obj in accessible_objects]
        
        if not accessible_object_ids:
            raise HTTPException(status_code=403, detail="Нет доступных объектов")
        
        # Получаем тайм-слот
        timeslot_query = select(TimeSlot).options(
            selectinload(TimeSlot.object)
        ).where(TimeSlot.id == timeslot_id)
        
        timeslot_result = await db.execute(timeslot_query)
        timeslot = timeslot_result.scalar_one_or_none()
        
        if not timeslot:
            raise HTTPException(status_code=404, detail="Тайм-слот не найден")
        
        # Проверяем доступ к объекту
        if timeslot.object_id not in accessible_object_ids:
            raise HTTPException(status_code=403, detail="Нет доступа к тайм-слоту")
        
        # Получаем связанные смены и расписания
        from sqlalchemy import and_
        # Запланированные смены (исключаем отмененные)
        scheduled_query = select(ShiftSchedule).options(
            selectinload(ShiftSchedule.user)
        ).where(
            and_(
                ShiftSchedule.time_slot_id == timeslot_id,
                ShiftSchedule.status != "cancelled"
            )
        ).order_by(ShiftSchedule.planned_start)
        
        scheduled_result = await db.execute(scheduled_query)
        scheduled_shifts = scheduled_result.scalars().all()
        
        # Фактические смены (исключаем отмененные)
        actual_query = select(Shift).options(
            selectinload(Shift.user)
        ).where(
            and_(
                Shift.time_slot_id == timeslot_id,
                Shift.status != "cancelled"
            )
        ).order_by(Shift.start_time)
        
        actual_result = await db.execute(actual_query)
        actual_shifts = actual_result.scalars().all()
        
        # Получаем available_interfaces
        login_service = RoleBasedLoginService(db)
        available_interfaces = await login_service.get_available_interfaces(user_id)
        
        return templates.TemplateResponse(
            "manager/timeslot_detail.html",
            {
                "request": request,
                "title": f"Тайм-слот {timeslot.object.name if timeslot.object else 'Неизвестный объект'}",
                "timeslot": timeslot,
                "scheduled_shifts": scheduled_shifts,
                "actual_shifts": actual_shifts,
                "current_user": current_user,
                "available_interfaces": available_interfaces
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting timeslot detail for manager: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки деталей тайм-слота")


@router.post("/timeslots/{timeslot_id}/edit")
async def manager_timeslot_edit(
    timeslot_id: int,
    request: Request,
    current_user: dict = Depends(require_manager_or_owner),
    db: AsyncSession = Depends(get_db_session)
):
    """Обновление тайм-слота управляющим"""
    try:
        # Проверяем, что current_user - это словарь, а не RedirectResponse
        if isinstance(current_user, RedirectResponse):
            return current_user
            
        # Получаем внутренний ID пользователя
        user_id = await get_user_id_from_current_user(current_user, db)
        if not user_id:
            raise HTTPException(status_code=400, detail="Пользователь не найден")
        
        # Получаем доступные объекты управляющего
        permission_service = ManagerPermissionService(db)
        accessible_objects = await permission_service.get_user_accessible_objects(user_id)
        accessible_object_ids = [obj.id for obj in accessible_objects]
        
        if not accessible_object_ids:
            raise HTTPException(status_code=403, detail="Нет доступных объектов")
        
        # Получаем тайм-слот
        timeslot_query = select(TimeSlot).where(TimeSlot.id == timeslot_id)
        timeslot_result = await db.execute(timeslot_query)
        timeslot = timeslot_result.scalar_one_or_none()
        
        if not timeslot:
            raise HTTPException(status_code=404, detail="Тайм-слот не найден")
        
        # Проверяем доступ к объекту
        if timeslot.object_id not in accessible_object_ids:
            raise HTTPException(status_code=403, detail="Нет доступа к тайм-слоту")
        
        # Получаем данные из формы
        form_data = await request.form()
        
        # Обновляем поля тайм-слота
        if 'slot_date' in form_data:
            from datetime import datetime
            timeslot.slot_date = datetime.strptime(form_data['slot_date'], "%Y-%m-%d").date()
        
        if 'start_time' in form_data:
            from datetime import time
            start_parts = form_data['start_time'].split(':')
            timeslot.start_time = time(int(start_parts[0]), int(start_parts[1]))
        
        if 'end_time' in form_data:
            from datetime import time
            end_parts = form_data['end_time'].split(':')
            timeslot.end_time = time(int(end_parts[0]), int(end_parts[1]))
        
        if 'max_employees' in form_data:
            timeslot.max_employees = int(form_data['max_employees'])
        
        if 'hourly_rate' in form_data:
            timeslot.hourly_rate = float(form_data['hourly_rate'])
        
        if 'notes' in form_data:
            timeslot.notes = form_data['notes']
        
        if 'is_active' in form_data:
            timeslot.is_active = form_data['is_active'].lower() == 'true'
        
        # Сохраняем изменения
        await db.commit()
        await db.refresh(timeslot)
        
        logger.info(f"Timeslot {timeslot_id} updated by manager {user_id}")
        
        return RedirectResponse(
            url=f"/manager/timeslots/object/{timeslot.object_id}",
            status_code=303
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating timeslot for manager: {e}")
        await db.rollback()
        raise HTTPException(status_code=500, detail="Ошибка обновления тайм-слота")


@router.get("/reports", response_class=HTMLResponse)
async def manager_reports(
    request: Request,
    current_user: dict = Depends(require_manager_or_owner),
):
    """Отчеты управляющего."""
    try:
        # Проверяем, что current_user не является RedirectResponse
        if isinstance(current_user, RedirectResponse):
            return current_user
        
        async with get_async_session() as db:
            user_id = await get_user_id_from_current_user(current_user, db)
            if not user_id:
                raise HTTPException(status_code=401, detail="Пользователь не найден")
            
            # Получаем доступные объекты управляющего
            permission_service = ManagerPermissionService(db)
            accessible_objects = await permission_service.get_user_accessible_objects(user_id)
            accessible_object_ids = [obj.id for obj in accessible_objects]
            
            if not accessible_object_ids:
                return templates.TemplateResponse("manager/reports/index.html", {
                    "request": request,
                    "current_user": current_user,
                    "available_interfaces": [],
                    "objects": [],
                    "employees": [],
                    "stats": {"total_shifts": 0, "total_hours": 0, "total_payment": 0, "active_objects": 0, "employees": 0}
                })
            
            # Получаем всех пользователей, которые работали на доступных объектах
            from sqlalchemy import select, and_
            from domain.entities.shift import Shift
            from domain.entities.user import User
            from datetime import datetime, timedelta
            
            employees_query = select(User.id, User.telegram_id, User.username, User.first_name, User.last_name, User.phone, User.role, User.is_active, User.created_at, User.updated_at).distinct().join(Shift, User.id == Shift.user_id).where(
                Shift.object_id.in_(accessible_object_ids)
            )
            employees_result = await db.execute(employees_query)
            employees = employees_result.all()
            
            # Если нет сотрудников из смен, показываем всех пользователей с ролью employee
            if not employees:
                all_employees_query = select(User.id, User.telegram_id, User.username, User.first_name, User.last_name, User.phone, User.role, User.is_active, User.created_at, User.updated_at).where(User.role == "employee")
                all_employees_result = await db.execute(all_employees_query)
                employees = all_employees_result.all()
            
            # Статистика за последний месяц
            month_ago = datetime.now() - timedelta(days=30)
            
            shifts_query = select(Shift).options(
                selectinload(Shift.object),
                selectinload(Shift.user)
            ).where(
                and_(
                    Shift.object_id.in_(accessible_object_ids),
                    Shift.start_time >= month_ago
                )
            )
            shifts_result = await db.execute(shifts_query)
            recent_shifts = shifts_result.scalars().all()
            
            stats = {
                "total_shifts": len(recent_shifts),
                "total_hours": sum(s.total_hours or 0 for s in recent_shifts if s.total_hours),
                "total_payment": sum(s.total_payment or 0 for s in recent_shifts if s.total_payment),
                "active_objects": len(accessible_objects),
                "employees": len(employees)
            }
            
            # Получаем данные для переключения интерфейсов
            login_service = RoleBasedLoginService(db)
            available_interfaces = await login_service.get_available_interfaces(user_id)
            
            return templates.TemplateResponse("manager/reports/index.html", {
                "request": request,
                "current_user": current_user,
                "objects": accessible_objects,
                "employees": employees,
                "available_interfaces": available_interfaces,
                "stats": stats
            })
        
    except Exception as e:
        logger.error(f"Error in manager reports: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки отчетов")


@router.post("/reports/generate")
async def manager_generate_report(
    request: Request,
    report_type: str = Form(...),
    date_from: str = Form(...),
    date_to: str = Form(...),
    object_id: Optional[int] = Form(None),
    employee_id: Optional[int] = Form(None),
    format: str = Form("excel"),
    current_user: dict = Depends(require_manager_or_owner),
):
    """Генерация отчета управляющего."""
    try:
        if isinstance(current_user, RedirectResponse):
            return current_user
        
        async with get_async_session() as db:
            user_id = await get_user_id_from_current_user(current_user, db)
            if not user_id:
                raise HTTPException(status_code=401, detail="Пользователь не найден")
            
            # Получаем доступные объекты управляющего
            permission_service = ManagerPermissionService(db)
            accessible_objects = await permission_service.get_user_accessible_objects(user_id)
            accessible_object_ids = [obj.id for obj in accessible_objects]
            
            if not accessible_object_ids:
                raise HTTPException(status_code=403, detail="Нет доступа к объектам")
            
            # Проверяем доступ к конкретному объекту, если указан
            if object_id and object_id not in accessible_object_ids:
                raise HTTPException(status_code=403, detail="Нет доступа к указанному объекту")
            
            # Парсим даты
            from datetime import datetime
            try:
                start_date = datetime.strptime(date_from, "%Y-%m-%d")
                end_date = datetime.strptime(date_to, "%Y-%m-%d")
            except ValueError:
                raise HTTPException(status_code=400, detail="Неверный формат даты")
            
            # Получаем смены за период
            from sqlalchemy import select, and_, desc
            from domain.entities.shift import Shift
            
            shifts_query = select(Shift).options(
                selectinload(Shift.object),
                selectinload(Shift.user)
            ).where(
                and_(
                    Shift.object_id.in_(accessible_object_ids),
                    Shift.start_time >= start_date,
                    Shift.start_time <= end_date + timedelta(days=1)
                )
            )
            
            # Применение фильтров
            if object_id and object_id in accessible_object_ids:
                shifts_query = shifts_query.where(Shift.object_id == object_id)
            
            if employee_id:
                shifts_query = shifts_query.where(Shift.user_id == employee_id)
            
            # Выполнение запроса
            shifts_result = await db.execute(shifts_query.order_by(desc(Shift.start_time)))
            shifts = shifts_result.scalars().all()
            
            # Генерация отчета в зависимости от типа
            if report_type == "shifts":
                return await _generate_shifts_report_manager(shifts, format, start_date, end_date)
            elif report_type == "employees":
                return await _generate_employees_report_manager(shifts, format, start_date, end_date)
            elif report_type == "objects":
                return await _generate_objects_report_manager(shifts, format, start_date, end_date)
            else:
                return {"error": "Неизвестный тип отчета"}
        
    except Exception as e:
        logger.error(f"Error generating manager report: {e}")
        raise HTTPException(status_code=500, detail="Ошибка генерации отчета")


@router.get("/reports/stats/period")
async def manager_reports_stats_period(
    request: Request,
    date_from: str = Query(...),
    date_to: str = Query(...),
    object_id: Optional[int] = Query(None),
    current_user: dict = Depends(require_manager_or_owner),
):
    """Статистика за период для управляющего."""
    try:
        if isinstance(current_user, RedirectResponse):
            return current_user
        
        async with get_async_session() as db:
            user_id = await get_user_id_from_current_user(current_user, db)
            if not user_id:
                raise HTTPException(status_code=401, detail="Пользователь не найден")
            
            # Получаем доступные объекты управляющего
            permission_service = ManagerPermissionService(db)
            accessible_objects = await permission_service.get_user_accessible_objects(user_id)
            accessible_object_ids = [obj.id for obj in accessible_objects]
            
            if not accessible_object_ids:
                return {"stats": {"total_shifts": 0, "total_hours": 0, "total_payment": 0}}
            
            # Проверяем доступ к конкретному объекту, если указан
            if object_id and object_id not in accessible_object_ids:
                raise HTTPException(status_code=403, detail="Нет доступа к указанному объекту")
            
            # Парсим даты
            from datetime import datetime
            try:
                start_date = datetime.strptime(date_from, "%Y-%m-%d")
                end_date = datetime.strptime(date_to, "%Y-%m-%d")
            except ValueError:
                raise HTTPException(status_code=400, detail="Неверный формат даты")
            
            # Получаем смены за период
            from sqlalchemy import select, and_
            from domain.entities.shift import Shift
            
            shifts_query = select(Shift).where(
                and_(
                    Shift.object_id.in_(accessible_object_ids),
                    Shift.start_time >= start_date,
                    Shift.start_time <= end_date
                )
            )
            
            if object_id:
                shifts_query = shifts_query.where(Shift.object_id == object_id)
            
            shifts_result = await db.execute(shifts_query)
            shifts = shifts_result.scalars().all()
            
            stats = {
                "total_shifts": len(shifts),
                "total_hours": sum(s.total_hours or 0 for s in shifts if s.total_hours),
                "total_payment": sum(s.total_payment or 0 for s in shifts if s.total_payment)
            }
            
            return {"stats": stats}
        
    except Exception as e:
        logger.error(f"Error getting manager stats: {e}")
        raise HTTPException(status_code=500, detail="Ошибка получения статистики")


async def _generate_shifts_report_manager(shifts: List[Shift], format: str, start_date: datetime, end_date: datetime):
    """Генерация отчета по сменам для управляющего"""
    data = []
    
    for shift in shifts:
        data.append({
            "ID": shift.id,
            "Сотрудник": f"{shift.user.first_name} {shift.user.last_name or ''}".strip(),
            "Объект": shift.object.name,
            "Дата начала": web_timezone_helper.format_datetime_with_timezone(shift.start_time, shift.object.timezone if shift.object else 'Europe/Moscow'),
            "Дата окончания": web_timezone_helper.format_datetime_with_timezone(shift.end_time, shift.object.timezone if shift.object else 'Europe/Moscow') if shift.end_time else "Не завершена",
            "Статус": shift.status,
            "Часов": shift.total_hours or 0,
            "Ставка": shift.hourly_rate or 0,
            "Сумма": shift.total_payment or 0,
            "Заметки": shift.notes or ""
        })
    
    if format == "excel":
        return await _create_excel_file_manager(data, f"manager_shifts_report_{start_date.strftime('%Y-%m-%d')}_{end_date.strftime('%Y-%m-%d')}")
    else:
        return {"data": data, "total": len(data)}


async def _generate_employees_report_manager(shifts: List[Shift], format: str, start_date: datetime, end_date: datetime):
    """Генерация отчета по сотрудникам для управляющего"""
    # Группировка по сотрудникам
    employees_data = {}
    
    for shift in shifts:
        employee_id = shift.user_id
        if employee_id not in employees_data:
            employees_data[employee_id] = {
                "employee": shift.user,
                "shifts": [],
                "total_hours": 0,
                "total_payment": 0
            }
        
        employees_data[employee_id]["shifts"].append(shift)
        employees_data[employee_id]["total_hours"] += shift.total_hours or 0
        employees_data[employee_id]["total_payment"] += shift.total_payment or 0
    
    data = []
    for emp_data in employees_data.values():
        data.append({
            "Сотрудник": f"{emp_data['employee'].first_name} {emp_data['employee'].last_name or ''}".strip(),
            "Количество смен": len(emp_data["shifts"]),
            "Общее время": emp_data["total_hours"],
            "Общая сумма": emp_data["total_payment"],
            "Средняя ставка": emp_data["total_payment"] / emp_data["total_hours"] if emp_data["total_hours"] > 0 else 0
        })
    
    if format == "excel":
        return await _create_excel_file_manager(data, f"manager_employees_report_{start_date.strftime('%Y-%m-%d')}_{end_date.strftime('%Y-%m-%d')}")
    else:
        return {"data": data, "total": len(data)}


async def _generate_objects_report_manager(shifts: List[Shift], format: str, start_date: datetime, end_date: datetime):
    """Генерация отчета по объектам для управляющего"""
    # Группировка по объектам
    objects_data = {}
    
    for shift in shifts:
        object_id = shift.object_id
        if object_id not in objects_data:
            objects_data[object_id] = {
                "object": shift.object,
                "shifts": [],
                "total_hours": 0,
                "total_payment": 0,
                "employees": set()
            }
        
        objects_data[object_id]["shifts"].append(shift)
        objects_data[object_id]["total_hours"] += shift.total_hours or 0
        objects_data[object_id]["total_payment"] += shift.total_payment or 0
        objects_data[object_id]["employees"].add(shift.user_id)
    
    data = []
    for obj_data in objects_data.values():
        data.append({
            "Объект": obj_data["object"].name,
            "Адрес": obj_data["object"].address or "",
            "Количество смен": len(obj_data["shifts"]),
            "Количество сотрудников": len(obj_data["employees"]),
            "Общее время": obj_data["total_hours"],
            "Общая сумма": obj_data["total_payment"],
            "Средняя ставка": obj_data["total_payment"] / obj_data["total_hours"] if obj_data["total_hours"] > 0 else 0
        })
    
    if format == "excel":
        return await _create_excel_file_manager(data, f"manager_objects_report_{start_date.strftime('%Y-%m-%d')}_{end_date.strftime('%Y-%m-%d')}")
    else:
        return {"data": data, "total": len(data)}


async def _create_excel_file_manager(data: List[dict], filename: str):
    """Создание Excel файла для управляющего"""
    if not data:
        return {"error": "Нет данных для отчета"}
    
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import Font, PatternFill, Alignment
        from fastapi.responses import StreamingResponse
        import io
        
        # Создание DataFrame
        df = pd.DataFrame(data)
        
        # Создание Excel файла
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчет управляющего"
        
        # Добавление данных
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Стилизация
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохранение в байты
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Возврат файла
        return StreamingResponse(
            io.BytesIO(excel_buffer.getvalue()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"}
        )
        
    except Exception as e:
        logger.error(f"Error creating Excel file for manager: {e}")
        return {"error": f"Ошибка создания Excel файла: {str(e)}"}


@router.get("/api/employees")
async def get_employees_for_manager(
    current_user: dict = Depends(require_manager_or_owner)
):
    """Получение списка сотрудников для управляющего"""
    try:
        if isinstance(current_user, RedirectResponse):
            raise HTTPException(status_code=401, detail="Необходима авторизация")
        
        # Проверяем кэш
        user_id_key = current_user.get("id") if isinstance(current_user, dict) else current_user.telegram_id
        from core.cache.redis_cache import cache
        cache_key = f"api_employees:{user_id_key}"
        cached_data = await cache.get(cache_key, serialize="json")
        if cached_data:
            logger.info(f"Manager employees API: cache HIT for user {user_id_key}")
            return cached_data
        
        logger.info(f"Manager employees API: cache MISS for user {user_id_key}")
        
        async with get_async_session() as db:
            user_id = await get_user_id_from_current_user(current_user, db)
            if not user_id:
                raise HTTPException(status_code=401, detail="Пользователь не найден")
            
            # Получаем доступные объекты управляющего
            permission_service = ManagerPermissionService(db)
            accessible_objects = await permission_service.get_user_accessible_objects(user_id)
            object_ids = [obj.id for obj in accessible_objects]
            
            if not object_ids:
                return []
            
            # Получаем сотрудников, которые имеют доступ к тем же объектам, что и управляющий
            from sqlalchemy import select, and_, or_
            from domain.entities.contract import Contract
            from domain.entities.user import User
            from sqlalchemy.dialects.postgresql import JSONB
            
            # Получаем договоры управляющего, чтобы узнать владельцев
            manager_contracts = await permission_service.get_manager_contracts_for_user(user_id)
            owner_ids = [contract.owner_id for contract in manager_contracts]
            
            if not owner_ids:
                return []
            
            # Получаем сотрудников, которые работают по договорам с этими владельцами
            # И имеют доступ к объектам, доступным управляющему
            employees_query = select(User).join(
                Contract, User.id == Contract.employee_id
            ).where(
                Contract.owner_id.in_(owner_ids),  # Договоры с владельцами управляющего
                Contract.is_active == True,
                Contract.status == "active"
            ).distinct()
            
            result = await db.execute(employees_query)
            all_employees = result.scalars().all()
            
            # Фильтруем сотрудников по доступу к объектам управляющего
            employees = []
            for emp in all_employees:
                # Получаем договоры сотрудника с владельцами управляющего
                emp_contracts_query = select(Contract).where(
                    Contract.employee_id == emp.id,
                    Contract.owner_id.in_(owner_ids),
                    Contract.is_active == True,
                    Contract.status == "active"
                )
                emp_contracts_result = await db.execute(emp_contracts_query)
                emp_contracts = emp_contracts_result.scalars().all()
                
                # Проверяем, есть ли у сотрудника доступ к объектам управляющего
                has_access = False
                for contract in emp_contracts:
                    if contract.allowed_objects:
                        # Проверяем пересечение объектов
                        for allowed_obj_id in contract.allowed_objects:
                            if allowed_obj_id in object_ids:
                                has_access = True
                                break
                    if has_access:
                        break
                
                if has_access:
                    employees.append(emp)
            
            employees_data = []
            for emp in employees:
                employees_data.append({
                    "id": emp.id,
                    "telegram_id": emp.telegram_id,
                    "first_name": emp.first_name,
                    "last_name": emp.last_name,
                    "username": emp.username,
                    "phone": emp.phone,
                    "is_active": emp.is_active,
                    "name": f"{emp.first_name} {emp.last_name or ''}".strip() or emp.username or f"ID {emp.id}"
                })
            
            # Сохраняем в кэш (TTL 2 минуты)
            await cache.set(cache_key, employees_data, ttl=120, serialize="json")
            logger.info(f"Manager employees API: cached {len(employees_data)} employees")
            
            return employees_data
            
    except Exception as e:
        logger.error(f"Error getting employees for manager: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки сотрудников")


# Удалено дублирование root-редиректа. Канонический редирект на /manager/dashboard определён выше.


@router.get("/calendar/api/objects")
async def get_objects_for_manager_calendar(
    current_user: dict = Depends(require_manager_or_owner)
):
    """Получение списка объектов для календаря управляющего"""
    try:
        if isinstance(current_user, RedirectResponse):
            raise HTTPException(status_code=401, detail="Необходима авторизация")
        
        # Проверяем кэш
        user_id_key = current_user.get("id") if isinstance(current_user, dict) else current_user.telegram_id
        from core.cache.redis_cache import cache
        cache_key = f"api_objects:{user_id_key}"
        cached_data = await cache.get(cache_key, serialize="json")
        if cached_data:
            logger.info(f"Manager objects API: cache HIT for user {user_id_key}")
            return cached_data
        
        logger.info(f"Manager objects API: cache MISS for user {user_id_key}")
        
        async with get_async_session() as db:
            user_id = await get_user_id_from_current_user(current_user, db)
            if not user_id:
                raise HTTPException(status_code=401, detail="Пользователь не найден")
            
            # Получаем доступные объекты управляющего
            permission_service = ManagerPermissionService(db)
            accessible_objects = await permission_service.get_user_accessible_objects(user_id)
            
            objects_data = []
            for obj in accessible_objects:
                # Формируем время работы
                working_hours = "Не указано"
                if obj.opening_time and obj.closing_time:
                    working_hours = f"{obj.opening_time.strftime('%H:%M')} - {obj.closing_time.strftime('%H:%M')}"
                elif obj.opening_time:
                    working_hours = f"с {obj.opening_time.strftime('%H:%M')}"
                elif obj.closing_time:
                    working_hours = f"до {obj.closing_time.strftime('%H:%M')}"
                
                objects_data.append({
                    "id": obj.id,
                    "name": obj.name,
                    "address": obj.address,
                    "hourly_rate": float(obj.hourly_rate) if obj.hourly_rate else 0,
                    "is_active": obj.is_active,
                    "opening_time": obj.opening_time.strftime('%H:%M') if obj.opening_time else None,
                    "closing_time": obj.closing_time.strftime('%H:%M') if obj.closing_time else None,
                    "working_hours": working_hours
                })
            
            # Сохраняем в кэш (TTL 2 минуты)
            await cache.set(cache_key, objects_data, ttl=120, serialize="json")
            logger.info(f"Manager objects API: cached {len(objects_data)} objects")
            
            return objects_data
            
    except Exception as e:
        logger.error(f"Error getting objects for manager calendar: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки объектов")


@router.get("/calendar/api/employees")
async def manager_calendar_api_employees(
    current_user: dict = Depends(require_manager_or_owner)
):
    """API для получения сотрудников для drag&drop панели"""
    try:
        if isinstance(current_user, RedirectResponse):
            raise HTTPException(status_code=401, detail="Необходима авторизация")
        
        async with get_async_session() as db:
            user_id = await get_user_id_from_current_user(current_user, db)
            if not user_id:
                raise HTTPException(status_code=401, detail="Пользователь не найден")
            
            # Получаем всех сотрудников
            employees_query = select(User).where(User.is_active == True)
            employees_result = await db.execute(employees_query)
            employees = employees_result.scalars().all()
            
            # Формируем список сотрудников для панели
            employees_data = []
            for emp in employees:
                # Исключаем владельца и неактивных пользователей
                if emp.is_owner or not emp.is_active:
                    continue
                    
                employees_data.append({
                    "id": emp.id,
                    "name": emp.name or emp.username,
                    "username": emp.username,
                    "role": emp.role,
                    "is_active": emp.is_active,
                    "is_owner": emp.is_owner
                })
            
            return employees_data
            
    except Exception as e:
        logger.error(f"Error getting employees for calendar: {e}")
        raise HTTPException(status_code=500, detail="Ошибка получения сотрудников")


# Удален тестовый endpoint


# Удален проблемный endpoint - используем рабочий /api/calendar/plan-shift


@router.get("/api/employees/for-object/{object_id}")
async def get_employees_for_object_manager(
    object_id: int,
    current_user: dict = Depends(require_manager_or_owner)
):
    """Получение сотрудников для конкретного объекта управляющим"""
    try:
        if isinstance(current_user, RedirectResponse):
            raise HTTPException(status_code=401, detail="Необходима авторизация")
        
        async with get_async_session() as db:
            user_id = await get_user_id_from_current_user(current_user, db)
            if not user_id:
                raise HTTPException(status_code=401, detail="Пользователь не найден")
            
            # Проверяем доступ к объекту
            permission_service = ManagerPermissionService(db)
            accessible_objects = await permission_service.get_user_accessible_objects(user_id)
            accessible_object_ids = [obj.id for obj in accessible_objects]
            
            if object_id not in accessible_object_ids:
                raise HTTPException(status_code=403, detail="Нет доступа к объекту")
            
            # Получаем сотрудников для объекта (копируем логику из календаря владельца)
            from sqlalchemy import select
            from domain.entities.contract import Contract
            from domain.entities.user import User
            import json
            
            # Получаем всех сотрудников с активными договорами
            employees_query = select(User).join(Contract, User.id == Contract.employee_id).where(
                Contract.is_active == True
            )
            employees_result = await db.execute(employees_query)
            employees = employees_result.scalars().all()
            
            employees_with_access = []
            added_employee_ids = set()  # Для отслеживания уже добавленных сотрудников
            
            for emp in employees:
                # Пропускаем, если сотрудник уже добавлен
                if emp.id in added_employee_ids:
                    continue
                
                # Проверяем, что пользователь имеет роль employee
                if "employee" not in (emp.roles if isinstance(emp.roles, list) else [emp.role]):
                    continue
                    
                # Получаем договоры сотрудника
                contract_query = select(Contract).where(
                    Contract.employee_id == emp.id,
                    Contract.is_active == True
                )
                contract_result = await db.execute(contract_query)
                contracts = contract_result.scalars().all()
                
                # Проверяем все договоры сотрудника
                for contract in contracts:
                    if contract and contract.allowed_objects:
                        allowed_objects = contract.allowed_objects if isinstance(contract.allowed_objects, list) else json.loads(contract.allowed_objects)
                        if object_id in allowed_objects:
                            employee_data = {
                                "id": int(emp.id),
                                "name": str(f"{emp.first_name or ''} {emp.last_name or ''}".strip() or emp.username or f"ID {emp.id}"),
                                "username": str(emp.username or ""),
                                "role": str(emp.role),
                                "is_active": bool(emp.is_active),
                                "telegram_id": int(emp.telegram_id) if emp.telegram_id else None
                            }
                            employees_with_access.append(employee_data)
                            added_employee_ids.add(emp.id)  # Помечаем сотрудника как добавленного
                            break  # Если нашли доступ, выходим из цикла по договорам
            
            employees_data = employees_with_access
            
            return employees_data
            
    except Exception as e:
        logger.error(f"Error getting employees for object {object_id}: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки сотрудников для объекта")


@router.post("/api/calendar/plan-shift")
async def plan_shift_manager(
    request: Request,
    current_user: dict = Depends(require_manager_or_owner)
):
    """Планирование смены управляющим"""
    logger.info("=== STARTING plan_shift_manager ===")
    logger.info("=== ENDPOINT HIT: /manager/api/calendar/plan-shift ===")
    try:
        if isinstance(current_user, RedirectResponse):
            raise HTTPException(status_code=401, detail="Необходима авторизация")
        
        data = await request.json()
        logger.info(f"Planning shift with data: {data}")
        
        timeslot_id = data.get('timeslot_id')
        employee_id = data.get('employee_id')
        
        logger.info(f"Processing: timeslot_id={timeslot_id}, employee_id={employee_id}")
        
        if not timeslot_id or not employee_id:
            raise HTTPException(status_code=400, detail="Не указан тайм-слот или сотрудник")
        
        async with get_async_session() as db:
            user_id = await get_user_id_from_current_user(current_user, db)
            if not user_id:
                raise HTTPException(status_code=401, detail="Пользователь не найден")
            
            # Получаем тайм-слот и проверяем доступ к объекту
            from sqlalchemy import select
            from domain.entities.time_slot import TimeSlot
            
            timeslot_query = select(TimeSlot).options(selectinload(TimeSlot.object)).where(TimeSlot.id == timeslot_id)
            timeslot = (await db.execute(timeslot_query)).scalar_one_or_none()
            if not timeslot:
                raise HTTPException(status_code=404, detail="Тайм-слот не найден")
            
            object_id = timeslot.object_id
            
            permission_service = ManagerPermissionService(db)
            accessible_objects = await permission_service.get_user_accessible_objects(user_id)
            accessible_object_ids = [obj.id for obj in accessible_objects]
            
            if object_id not in accessible_object_ids:
                raise HTTPException(status_code=403, detail="Нет доступа к объекту")
            
            # Проверяем, что у сотрудника есть договор с управляющим
            from domain.entities.contract import Contract
            from domain.entities.user import User
            
            # Получаем сотрудника
            employee_query = select(User).where(User.id == employee_id)
            employee = (await db.execute(employee_query)).scalar_one_or_none()
            if not employee:
                raise HTTPException(status_code=404, detail="Сотрудник не найден")
            
            # Получаем договоры управляющего, чтобы узнать владельцев
            manager_contracts = await permission_service.get_manager_contracts_for_user(user_id)
            owner_ids = [contract.owner_id for contract in manager_contracts]
            
            if not owner_ids:
                raise HTTPException(status_code=403, detail="У управляющего нет активных договоров")
            
            # Проверяем, что сотрудник работает по договорам с владельцами управляющего
            contract_query = select(Contract).where(
                Contract.employee_id == employee_id,
                Contract.owner_id.in_(owner_ids),  # Договоры с владельцами управляющего
                Contract.is_active == True,
                Contract.status == "active"
            )
            contracts = (await db.execute(contract_query)).scalars().all()
            if not contracts:
                raise HTTPException(status_code=400, detail="У сотрудника нет договора с владельцами управляющего")
            
            # Проверяем, что у сотрудника есть доступ к объекту тайм-слота
            has_object_access = False
            employee_contract = None  # Договор сотрудника с доступом к объекту
            for contract in contracts:
                if contract.allowed_objects and object_id in contract.allowed_objects:
                    has_object_access = True
                    employee_contract = contract  # Сохраняем договор для определения ставки
                    break
            
            if not has_object_access:
                raise HTTPException(status_code=400, detail=f"У сотрудника нет доступа к объекту ID {object_id}")
            
            # Проверяем пересечение смен сотрудника
            from domain.entities.shift import Shift
            from domain.entities.shift_schedule import ShiftSchedule
            import pytz
            
            # Получаем временную зону объекта для корректного сравнения времени
            object_timezone = timeslot.object.timezone if timeslot.object and timeslot.object.timezone else 'Europe/Moscow'
            tz = pytz.timezone(object_timezone)
            
            # Создаем naive datetime в локальной временной зоне объекта
            slot_datetime_naive = datetime.combine(timeslot.slot_date, timeslot.start_time)
            end_datetime_naive = datetime.combine(timeslot.slot_date, timeslot.end_time)
            
            # Локализуем в временную зону объекта, затем конвертируем в UTC для сравнения
            slot_datetime_utc = tz.localize(slot_datetime_naive).astimezone(pytz.UTC)
            end_datetime_utc = tz.localize(end_datetime_naive).astimezone(pytz.UTC)
            
            logger.info(f"Timezone conversion: {object_timezone} -> UTC")
            logger.info(f"Slot time: {slot_datetime_naive} -> {slot_datetime_utc}")
            logger.info(f"End time: {end_datetime_naive} -> {end_datetime_utc}")
            
            # Проверяем пересечение с активными сменами
            active_shifts_query = select(Shift).where(
                Shift.user_id == employee_id,
                Shift.status == "active",
                Shift.start_time < end_datetime_utc,
                Shift.end_time > slot_datetime_utc
            )
            active_shifts = (await db.execute(active_shifts_query)).scalars().all()
            
            if active_shifts:
                shift_times = []
                for shift in active_shifts:
                    # Конвертируем обратно в локальное время для отображения
                    local_start = shift.start_time.astimezone(tz).strftime('%H:%M')
                    local_end = shift.end_time.astimezone(tz).strftime('%H:%M')
                    shift_times.append(f"{local_start}-{local_end}")
                
                raise HTTPException(
                    status_code=400, 
                    detail=f"У сотрудника уже есть активная смена в это время: {', '.join(shift_times)}"
                )
            
            # Проверяем пересечение с запланированными сменами
            planned_shifts_query = select(ShiftSchedule).where(
                ShiftSchedule.user_id == employee_id,
                ShiftSchedule.status == "planned",
                ShiftSchedule.planned_start < end_datetime_utc,
                ShiftSchedule.planned_end > slot_datetime_utc
            )
            planned_shifts = (await db.execute(planned_shifts_query)).scalars().all()
            
            logger.info(f"Checking planned shifts for employee {employee_id}")
            logger.info(f"Slot time range: {slot_datetime_utc} - {end_datetime_utc}")
            logger.info(f"Found {len(planned_shifts)} planned shifts")
            
            if planned_shifts:
                shift_times = []
                for shift in planned_shifts:
                    logger.info(f"Conflicting shift: {shift.id}, {shift.planned_start} - {shift.planned_end}")
                    # Конвертируем обратно в локальное время для отображения
                    local_start = shift.planned_start.astimezone(tz).strftime('%H:%M')
                    local_end = shift.planned_end.astimezone(tz).strftime('%H:%M')
                    shift_times.append(f"{local_start}-{local_end}")
                
                error_msg = f"У сотрудника уже есть запланированная смена в это время: {', '.join(shift_times)}"
                logger.info(f"Raising HTTPException: {error_msg}")
                raise HTTPException(
                    status_code=400, 
                    detail=error_msg
                )
            
            # Создаем запланированную смену
            
            # Используем уже вычисленные времена в UTC
            slot_datetime = slot_datetime_utc
            end_datetime = end_datetime_utc
            
            # Вычисляем ставку: приоритет входного значения > contract.use_contract_rate > тайм-слот > объект
            effective_rate = None
            
            # Приоритет 1: входное значение hourly_rate
            try:
                raw_rate = data.get('hourly_rate')
                if isinstance(raw_rate, str):
                    raw_rate = raw_rate.strip()
                if raw_rate not in (None, ""):
                    normalized_rate = str(raw_rate).replace(",", ".")
                    candidate_rate = float(normalized_rate)
                    if candidate_rate > 0:
                        effective_rate = candidate_rate
            except Exception:
                # Если введено некорректно — игнорируем и перейдем к источникам ниже
                effective_rate = None

            # Если ставка не задана вручную, используем логику с учетом use_contract_rate
            if effective_rate is None and employee_contract:
                # Используем метод модели Contract для определения эффективной ставки
                timeslot_rate = None
                if getattr(timeslot, 'hourly_rate', None):
                    try:
                        ts_rate = float(timeslot.hourly_rate)
                        if ts_rate > 0:
                            timeslot_rate = ts_rate
                    except Exception:
                        pass
                
                object_rate = None
                try:
                    if timeslot.object and timeslot.object.hourly_rate:
                        object_rate = float(timeslot.object.hourly_rate)
                except Exception:
                    pass
                
                effective_rate = employee_contract.get_effective_hourly_rate(
                    timeslot_rate=timeslot_rate,
                    object_rate=object_rate
                )
            elif effective_rate is None:
                # Фолбэк для случаев без договора: тайм-слот > объект
                if getattr(timeslot, 'hourly_rate', None):
                    try:
                        ts_rate = float(timeslot.hourly_rate)
                        if ts_rate > 0:
                            effective_rate = ts_rate
                    except Exception:
                        pass
                
                if effective_rate is None:
                    try:
                        obj_rate = float(timeslot.object.hourly_rate) if timeslot.object and timeslot.object.hourly_rate else 0.0
                    except Exception:
                        obj_rate = 0.0
                    effective_rate = obj_rate

            shift_schedule = ShiftSchedule(
                user_id=int(employee_id),
                object_id=int(object_id),
                time_slot_id=int(timeslot_id),
                planned_start=slot_datetime,
                planned_end=end_datetime,
                status='planned',
                hourly_rate=effective_rate,
                notes=data.get('notes', '')
            )
            
            db.add(shift_schedule)
            await db.commit()
            await db.refresh(shift_schedule)
            
            # Очищаем кэш календаря для немедленного отображения
            from core.cache.redis_cache import cache
            await cache.clear_pattern("calendar_shifts:*")
            await cache.clear_pattern("api_response:*")
            logger.info(f"Calendar cache cleared after planning shift {shift_schedule.id}")
            
            logger.info(f"Successfully planned shift {shift_schedule.id}")
            return {
                "success": True,
                "message": "Смена успешно запланирована",
                "shift_id": shift_schedule.id
            }
            
    except HTTPException as e:
        # Передаем HTTPException как есть, чтобы сохранить детали ошибки
        logger.error(f"HTTPException in planning shift: {e.detail}")
        raise e
    except Exception as e:
        logger.error(f"Error planning shift: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка планирования смены: {str(e)}")


@router.post("/api/calendar/check-availability")
async def check_employee_availability(
    request: Request,
    current_user: dict = Depends(require_manager_or_owner)
):
    """Проверка доступности сотрудника для планирования смены"""
    try:
        if isinstance(current_user, RedirectResponse):
            raise HTTPException(status_code=401, detail="Необходима авторизация")
        
        data = await request.json()
        timeslot_id = data.get('timeslot_id')
        employee_id = data.get('employee_id')
        
        if not timeslot_id or not employee_id:
            raise HTTPException(status_code=400, detail="Не указан тайм-слот или сотрудник")
        
        async with get_async_session() as db:
            user_id = await get_user_id_from_current_user(current_user, db)
            if not user_id:
                raise HTTPException(status_code=401, detail="Пользователь не найден")
            
            # Получаем тайм-слот и проверяем доступ к объекту
            from sqlalchemy import select
            from domain.entities.time_slot import TimeSlot
            
            timeslot_query = select(TimeSlot).options(selectinload(TimeSlot.object)).where(TimeSlot.id == timeslot_id)
            timeslot = (await db.execute(timeslot_query)).scalar_one_or_none()
            if not timeslot:
                raise HTTPException(status_code=404, detail="Тайм-слот не найден")
            
            object_id = timeslot.object_id
            
            permission_service = ManagerPermissionService(db)
            accessible_objects = await permission_service.get_user_accessible_objects(user_id)
            accessible_object_ids = [obj.id for obj in accessible_objects]
            
            if object_id not in accessible_object_ids:
                raise HTTPException(status_code=403, detail="Нет доступа к объекту")
            
            # Проверяем, что у сотрудника есть договор с управляющим
            from domain.entities.contract import Contract
            from domain.entities.user import User
            
            # Получаем сотрудника
            employee_query = select(User).where(User.id == employee_id)
            employee = (await db.execute(employee_query)).scalar_one_or_none()
            if not employee:
                raise HTTPException(status_code=404, detail="Сотрудник не найден")
            
            # Получаем договоры управляющего, чтобы узнать владельцев
            manager_contracts = await permission_service.get_manager_contracts_for_user(user_id)
            owner_ids = [contract.owner_id for contract in manager_contracts]
            
            if not owner_ids:
                raise HTTPException(status_code=403, detail="У управляющего нет активных договоров")
            
            # Проверяем, что сотрудник работает по договорам с владельцами управляющего
            contract_query = select(Contract).where(
                Contract.employee_id == employee_id,
                Contract.owner_id.in_(owner_ids),
                Contract.is_active == True,
                Contract.status == "active"
            )
            contracts = (await db.execute(contract_query)).scalars().all()
            if not contracts:
                return {
                    "available": False,
                    "message": "У сотрудника нет договора с владельцами управляющего"
                }
            
            # Проверяем, что у сотрудника есть доступ к объекту тайм-слота
            has_object_access = False
            for contract in contracts:
                if contract.allowed_objects and object_id in contract.allowed_objects:
                    has_object_access = True
                    break
            
            if not has_object_access:
                return {
                    "available": False,
                    "message": f"У сотрудника нет доступа к объекту"
                }
            
            # Проверяем пересечение с активными сменами
            from domain.entities.shift import Shift
            from domain.entities.shift_schedule import ShiftSchedule
            import pytz
            
            # Получаем временную зону объекта для корректного сравнения времени
            object_timezone = timeslot.object.timezone if timeslot.object and timeslot.object.timezone else 'Europe/Moscow'
            tz = pytz.timezone(object_timezone)
            
            # Создаем naive datetime в локальной временной зоне объекта
            slot_datetime_naive = datetime.combine(timeslot.slot_date, timeslot.start_time)
            end_datetime_naive = datetime.combine(timeslot.slot_date, timeslot.end_time)
            
            # Локализуем в временную зону объекта, затем конвертируем в UTC для сравнения
            slot_datetime_utc = tz.localize(slot_datetime_naive).astimezone(pytz.UTC)
            end_datetime_utc = tz.localize(end_datetime_naive).astimezone(pytz.UTC)
            
            # Проверяем пересечение с активными сменами
            active_shifts_query = select(Shift).where(
                Shift.user_id == employee_id,
                Shift.status == "active",
                Shift.start_time < end_datetime_utc,
                Shift.end_time > slot_datetime_utc
            )
            active_shifts = (await db.execute(active_shifts_query)).scalars().all()
            
            if active_shifts:
                shift_times = []
                for shift in active_shifts:
                    local_start = shift.start_time.astimezone(tz).strftime('%H:%M')
                    local_end = shift.end_time.astimezone(tz).strftime('%H:%M')
                    shift_times.append(f"{local_start}-{local_end}")
                
                return {
                    "available": False,
                    "message": f"У сотрудника уже есть активная смена в это время: {', '.join(shift_times)}"
                }
            
            # Проверяем пересечение с запланированными сменами
            planned_shifts_query = select(ShiftSchedule).where(
                ShiftSchedule.user_id == employee_id,
                ShiftSchedule.status == "planned",
                ShiftSchedule.planned_start < end_datetime_utc,
                ShiftSchedule.planned_end > slot_datetime_utc
            )
            planned_shifts = (await db.execute(planned_shifts_query)).scalars().all()
            
            if planned_shifts:
                shift_times = []
                for shift in planned_shifts:
                    local_start = shift.planned_start.astimezone(tz).strftime('%H:%M')
                    local_end = shift.planned_end.astimezone(tz).strftime('%H:%M')
                    shift_times.append(f"{local_start}-{local_end}")
                
                return {
                    "available": False,
                    "message": f"У сотрудника уже есть запланированная смена в это время: {', '.join(shift_times)}"
                }
            
            return {
                "available": True,
                "message": "Сотрудник доступен"
            }
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error checking employee availability: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка проверки доступности: {str(e)}")


@router.get("/shifts/api/schedule/{schedule_id}/object-id", response_class=JSONResponse)
async def get_manager_schedule_object_id(
    schedule_id: int,
    current_user: dict = Depends(require_manager_or_owner),
    db: AsyncSession = Depends(get_db_session)
):
    """Получить object_id из запланированной смены для управляющего."""
    try:
        if isinstance(current_user, RedirectResponse):
            return JSONResponse({"error": "Unauthorized"}, status_code=401)
        
        user_id = await get_user_id_from_current_user(current_user, db)
        if not user_id:
            return JSONResponse({"error": "User not found"}, status_code=401)
        
        # Получаем запланированную смену
        from sqlalchemy import select
        from domain.entities.shift_schedule import ShiftSchedule
        from sqlalchemy.orm import selectinload
        
        schedule_query = select(ShiftSchedule).options(
            selectinload(ShiftSchedule.object)
        ).where(ShiftSchedule.id == schedule_id)
        
        result = await db.execute(schedule_query)
        schedule = result.scalar_one_or_none()
        
        if not schedule:
            return JSONResponse({"error": "Schedule not found"}, status_code=404)
        
        # Проверяем доступ управляющего к объекту
        permission_service = ManagerPermissionService(db)
        accessible_objects = await permission_service.get_user_accessible_objects(user_id)
        accessible_object_ids = [obj.id for obj in accessible_objects]
        
        if schedule.object_id not in accessible_object_ids:
            return JSONResponse({"error": "Access denied"}, status_code=403)
        
        return JSONResponse({"object_id": schedule.object_id})
        
    except Exception as e:
        logger.error(f"Error getting manager schedule object_id: {e}")
        return JSONResponse({"error": str(e)}, status_code=500)


@router.get("/shifts/plan", response_class=HTMLResponse)
async def manager_shifts_plan(
    request: Request,
    object_id: Optional[int] = Query(None, description="ID объекта для предзаполнения"),
    return_to: Optional[str] = Query(None, description="URL для возврата после планирования"),
    current_user: dict = Depends(require_manager_or_owner),
    db: AsyncSession = Depends(get_db_session)
):
    """Страница планирования смен для управляющего."""
    try:
        if isinstance(current_user, RedirectResponse):
            return current_user
        
        user_id = await get_user_id_from_current_user(current_user, db)
        if not user_id:
            raise HTTPException(status_code=401, detail="Пользователь не найден")
        
        # Получаем доступные объекты управляющего
        permission_service = ManagerPermissionService(db)
        accessible_objects = await permission_service.get_user_accessible_objects(user_id)
        
        objects_list = [{"id": obj.id, "name": obj.name} for obj in accessible_objects]
        
        selected_object_id = None
        if object_id:
            for obj in accessible_objects:
                if obj.id == object_id:
                    selected_object_id = object_id
                    break
        
        login_service = RoleBasedLoginService(db)
        available_interfaces = await login_service.get_available_interfaces(user_id)
        
        return templates.TemplateResponse("manager/shifts/plan.html", {
            "request": request,
            "current_user": current_user,
            "objects": objects_list,
            "selected_object_id": selected_object_id,
            "return_to": return_to or "/manager/shifts",
            "available_interfaces": available_interfaces
        })
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error loading manager shifts plan page: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки страницы планирования")


@router.get("/shifts", response_class=HTMLResponse, name="manager_shifts")
async def manager_shifts_list(
    request: Request,
    status: Optional[str] = Query(None, description="Фильтр по статусу: active, planned, completed"),
    date_from: Optional[str] = Query(None, description="Дата начала (YYYY-MM-DD)"),
    date_to: Optional[str] = Query(None, description="Дата окончания (YYYY-MM-DD)"),
    object_id: Optional[str] = Query(None, description="ID объекта"),
    sort: Optional[str] = Query(None, description="Поле для сортировки"),
    order: Optional[str] = Query("asc", description="Направление сортировки: asc, desc"),
    page: int = Query(1, ge=1, description="Номер страницы"),
    per_page: int = Query(20, ge=1, le=100, description="Количество на странице"),
    current_user: dict = Depends(require_manager_or_owner),
):
    """Список смен управляющего."""
    try:
        # Проверяем, что current_user - это словарь, а не RedirectResponse
        if isinstance(current_user, RedirectResponse):
            return current_user
        
        async with get_async_session() as db:
            # Получаем внутренний ID пользователя
            user_id = await get_user_id_from_current_user(current_user, db)
            if not user_id:
                raise HTTPException(status_code=401, detail="Пользователь не найден")
            
            # Получаем доступные объекты управляющего
            permission_service = ManagerPermissionService(db)
            accessible_objects = await permission_service.get_user_accessible_objects(user_id)
            accessible_object_ids = [obj.id for obj in accessible_objects]
            
            if not accessible_object_ids:
                return templates.TemplateResponse("manager/shifts/list.html", {
                    "request": request,
                    "current_user": current_user,
                    "available_interfaces": [],
                    "shifts": [],
                    "objects": [],
                    "stats": {"total": 0, "active": 0, "planned": 0, "completed": 0},
                    "filters": {"status": status, "date_from": date_from, "date_to": date_to, "object_id": object_id},
                    "pagination": {"page": 1, "per_page": 20, "total": 0, "pages": 0}
                })
            
            # Базовый запрос для смен
            from sqlalchemy import select, desc, asc
            from domain.entities.user import User
            from domain.entities.object import Object
            shifts_query = select(Shift).options(
                selectinload(Shift.object),
                selectinload(Shift.user)
            ).where(Shift.object_id.in_(accessible_object_ids))
            
            # Базовый запрос для запланированных смен
            schedules_query = select(ShiftSchedule).options(
                selectinload(ShiftSchedule.object),
                selectinload(ShiftSchedule.user)
            ).where(ShiftSchedule.object_id.in_(accessible_object_ids))
            
            # Применение фильтров
            if status:
                if status == "active":
                    # Только активные смены из таблицы shifts
                    shifts_query = shifts_query.where(Shift.status == "active")
                    schedules_query = schedules_query.where(False)  # Исключаем все запланированные
                elif status == "planned":
                    # Только запланированные смены из таблицы shift_schedules
                    shifts_query = shifts_query.where(False)  # Исключаем все обычные смены
                    schedules_query = schedules_query.where(ShiftSchedule.status == "planned")
                elif status == "completed":
                    # Только завершенные смены из таблицы shifts
                    shifts_query = shifts_query.where(Shift.status == "completed")
                    schedules_query = schedules_query.where(False)  # Исключаем все запланированные
                elif status == "cancelled":
                    # Отмененные смены из обеих таблиц
                    shifts_query = shifts_query.where(Shift.status == "cancelled")
                    schedules_query = schedules_query.where(ShiftSchedule.status == "cancelled")
            
            if date_from:
                from datetime import datetime
                date_from_obj = datetime.strptime(date_from, "%Y-%m-%d").date()
                shifts_query = shifts_query.where(Shift.start_time >= date_from_obj)
                schedules_query = schedules_query.where(ShiftSchedule.planned_start >= date_from_obj)
            
            if date_to:
                from datetime import datetime
                date_to_obj = datetime.strptime(date_to, "%Y-%m-%d").date()
                shifts_query = shifts_query.where(Shift.start_time <= date_to_obj)
                schedules_query = schedules_query.where(ShiftSchedule.planned_start <= date_to_obj)
            
            if object_id:
                if int(object_id) in accessible_object_ids:
                    shifts_query = shifts_query.where(Shift.object_id == int(object_id))
                    schedules_query = schedules_query.where(ShiftSchedule.object_id == int(object_id))
                else:
                    # Нет доступа к объекту
                    return templates.TemplateResponse("manager/shifts/list.html", {
                        "request": request,
                        "current_user": current_user,
                        "available_interfaces": [],
                        "shifts": [],
                        "objects": accessible_objects,
                        "stats": {"total": 0, "active": 0, "planned": 0, "completed": 0},
                        "filters": {"status": status, "date_from": date_from, "date_to": date_to, "object_id": object_id},
                        "pagination": {"page": 1, "per_page": 20, "total": 0, "pages": 0}
                    })
            
            # Получение данных (без сортировки на уровне БД)
            shifts_result = await db.execute(shifts_query.order_by(desc(Shift.created_at)))
            shifts = shifts_result.scalars().all()
            
            schedules_result = await db.execute(schedules_query.order_by(desc(ShiftSchedule.created_at)))
            schedules = schedules_result.scalars().all()
            
            # Объединение и форматирование данных
            all_shifts = []
            
            # Добавляем обычные смены
            for shift in shifts:
                # Корректные часы/оплата только для завершённых смен
                total_hours_val = None
                total_payment_val = None
                try:
                    if str(shift.status) == 'completed':
                        if getattr(shift, 'total_hours', None) is not None:
                            total_hours_val = float(shift.total_hours)
                        elif shift.end_time and shift.start_time:
                            duration = shift.end_time - shift.start_time
                            total_hours_val = round(duration.total_seconds() / 3600, 2)
                        if total_hours_val is not None:
                            if getattr(shift, 'total_payment', None) is not None:
                                total_payment_val = float(shift.total_payment)
                            elif getattr(shift, 'hourly_rate', None) is not None:
                                total_payment_val = round(total_hours_val * float(shift.hourly_rate), 2)
                except Exception:
                    total_hours_val = None
                    total_payment_val = None

                all_shifts.append({
                    'id': shift.id,
                    'type': 'shift',
                    'object_id': shift.object_id,
                    'object_name': shift.object.name if shift.object else 'Неизвестный объект',
                    'user_id': shift.user_id,
                    'user_first_name': (shift.user.first_name if shift.user else '') or '',
                    'user_last_name': (shift.user.last_name if shift.user else '') or '',
                    'user_name': (f"{(shift.user.last_name or '').strip()} {(shift.user.first_name or '').strip()}".strip() if shift.user else 'Неизвестный пользователь'),
                    'start_time': web_timezone_helper.format_datetime_with_timezone(shift.start_time, shift.object.timezone if shift.object else 'Europe/Moscow', '%Y-%m-%d %H:%M') if shift.start_time else '-',
                    'end_time': web_timezone_helper.format_datetime_with_timezone(shift.end_time, shift.object.timezone if shift.object else 'Europe/Moscow', '%Y-%m-%d %H:%M') if shift.end_time else '-',
                    'status': shift.status,
                    'is_planned': getattr(shift, 'is_planned', False),
                    'total_hours': total_hours_val,
                    'total_payment': total_payment_val,
                    'created_at': shift.created_at
                })
            
            # Добавляем запланированные смены
            for schedule in schedules:
                # Плановые часы/оплата для planned/confirmed
                planned_hours = None
                planned_payment = None
                try:
                    if str(schedule.status) in ['planned', 'confirmed']:
                        planned_hours = float(schedule.planned_duration_hours)
                        if getattr(schedule, 'hourly_rate', None) is not None:
                            planned_payment = round(planned_hours * float(schedule.hourly_rate), 2)
                except Exception:
                    planned_hours = None
                    planned_payment = None

                all_shifts.append({
                    'id': schedule.id,
                    'type': 'schedule',
                    'object_id': schedule.object_id,
                    'object_name': schedule.object.name if schedule.object else 'Неизвестный объект',
                    'user_id': schedule.user_id,
                    'user_first_name': (schedule.user.first_name if schedule.user else '') or '',
                    'user_last_name': (schedule.user.last_name if schedule.user else '') or '',
                    'user_name': (f"{(schedule.user.last_name or '').strip()} {(schedule.user.first_name or '').strip()}".strip() if schedule.user else 'Неизвестный пользователь'),
                    'start_time': web_timezone_helper.format_datetime_with_timezone(schedule.planned_start, schedule.object.timezone if schedule.object else 'Europe/Moscow', '%Y-%m-%d %H:%M') if schedule.planned_start else '-',
                    'end_time': web_timezone_helper.format_datetime_with_timezone(schedule.planned_end, schedule.object.timezone if schedule.object else 'Europe/Moscow', '%Y-%m-%d %H:%M') if schedule.planned_end else '-',
                    'status': schedule.status,
                    'is_planned': True,
                    'total_hours': planned_hours,
                    'total_payment': planned_payment,
                    'created_at': schedule.created_at
                })
            
            # Сортировка данных
            if sort:
                reverse = order == 'desc'
                if sort == "user_name":
                    all_shifts.sort(key=lambda x: ((x.get('user_last_name') or '').lower(), (x.get('user_first_name') or '').lower()), reverse=reverse)
                elif sort == "object_name":
                    all_shifts.sort(key=lambda x: x['object_name'].lower(), reverse=reverse)
                elif sort == "start_time":
                    all_shifts.sort(key=lambda x: x['start_time'], reverse=reverse)
                elif sort == "status":
                    all_shifts.sort(key=lambda x: x['status'], reverse=reverse)
                elif sort == "created_at":
                    all_shifts.sort(key=lambda x: x['created_at'], reverse=reverse)
            else:
                # Сортировка по умолчанию
                all_shifts.sort(key=lambda x: x['created_at'], reverse=True)
            
            # Пагинация
            total_shifts = len(all_shifts)
            start_idx = (page - 1) * per_page
            end_idx = start_idx + per_page
            paginated_shifts = all_shifts[start_idx:end_idx]
            
            # Статистика
            stats = {
                'total': total_shifts,
                'active': len([s for s in all_shifts if s['status'] == 'active']),
                'planned': len([s for s in all_shifts if s['type'] == 'schedule']),
                'completed': len([s for s in all_shifts if s['status'] == 'completed'])
            }
            
            # Получаем данные для переключения интерфейсов
            login_service = RoleBasedLoginService(db)
            available_interfaces = await login_service.get_available_interfaces(user_id)
            
            return templates.TemplateResponse("manager/shifts/list.html", {
                "request": request,
                "current_user": current_user,
                "available_interfaces": available_interfaces,
                "shifts": paginated_shifts,
                "objects": accessible_objects,
                "stats": stats,
                "filters": {
                    "status": status,
                    "date_from": date_from,
                    "date_to": date_to,
                    "object_id": object_id
                },
                "sort": {
                    "field": sort,
                    "order": order
                },
                "pagination": {
                    "page": page,
                    "per_page": per_page,
                    "total": total_shifts,
                    "pages": (total_shifts + per_page - 1) // per_page
                }
            })
            
    except Exception as e:
        logger.error(f"Error loading manager shifts: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки смен")


@router.get("/shifts/{shift_id}", response_class=HTMLResponse)
async def manager_shift_detail(
    request: Request, 
    shift_id: str,  # Изменено на str для поддержки префикса schedule_
    shift_type: Optional[str] = Query("shift"),
    current_user: dict = Depends(require_manager_or_owner),
):
    """Детали смены управляющего"""
    try:
        # Проверяем, что current_user - это словарь, а не RedirectResponse
        if isinstance(current_user, RedirectResponse):
            return current_user
        
        # Определяем тип смены по ID
        if shift_id.startswith('schedule_'):
            actual_shift_id = int(shift_id.replace('schedule_', ''))
            actual_shift_type = "schedule"
        else:
            actual_shift_id = int(shift_id)
            actual_shift_type = shift_type or "shift"
        
        async with get_async_session() as db:
            # Получаем внутренний ID пользователя
            user_id = await get_user_id_from_current_user(current_user, db)
            if not user_id:
                raise HTTPException(status_code=401, detail="Пользователь не найден")
            
            # Получаем доступные объекты управляющего
            permission_service = ManagerPermissionService(db)
            accessible_objects = await permission_service.get_user_accessible_objects(user_id)
            accessible_object_ids = [obj.id for obj in accessible_objects]
            
            if not accessible_object_ids:
                raise HTTPException(status_code=403, detail="Нет доступа к объектам")
            
            shift_data = None
            
            # Импортируем select для запросов
            from sqlalchemy import select
            
            if actual_shift_type == "schedule":
                # Запланированная смена
                query = select(ShiftSchedule).options(
                    selectinload(ShiftSchedule.object),
                    selectinload(ShiftSchedule.user)
                ).where(ShiftSchedule.id == actual_shift_id)
                
                result = await db.execute(query)
                schedule = result.scalar_one_or_none()
                
                if not schedule:
                    raise HTTPException(status_code=404, detail="Запланированная смена не найдена")
                
                # Проверяем доступ к объекту
                if schedule.object_id not in accessible_object_ids:
                    raise HTTPException(status_code=403, detail="Нет доступа к объекту")
                
                # Получаем информацию о тайм-слоте
                timeslot_info = None
                if schedule.time_slot_id:
                    timeslot_query = select(TimeSlot).where(TimeSlot.id == schedule.time_slot_id)
                    timeslot_result = await db.execute(timeslot_query)
                    timeslot = timeslot_result.scalar_one_or_none()
                    if timeslot:
                        # Подсчитываем количество запланированных смен для этого тайм-слота
                        scheduled_count_query = select(ShiftSchedule).where(
                            ShiftSchedule.time_slot_id == timeslot.id,
                            ShiftSchedule.status == 'planned'
                        )
                        scheduled_count_result = await db.execute(scheduled_count_query)
                        scheduled_count = len(scheduled_count_result.scalars().all())
                        
                        timeslot_info = {
                            'id': timeslot.id,
                            'start_time': timeslot.start_time.strftime('%H:%M') if timeslot.start_time else '-',
                            'end_time': timeslot.end_time.strftime('%H:%M') if timeslot.end_time else '-',
                            'max_employees': timeslot.max_employees,
                            'scheduled_count': scheduled_count
                        }
                
                shift_data = {
                    'id': schedule.id,
                    'type': 'schedule',
                    'object_id': schedule.object_id,
                    'object_name': schedule.object.name if schedule.object else 'Неизвестный объект',
                    'user_id': schedule.user_id,
                    'user_name': f"{schedule.user.first_name} {schedule.user.last_name or ''}".strip() if schedule.user else 'Неизвестный пользователь',
                    'start_time': schedule.planned_start.strftime('%Y-%m-%d %H:%M') if schedule.planned_start else '-',
                    'end_time': schedule.planned_end.strftime('%Y-%m-%d %H:%M') if schedule.planned_end else '-',
                    'status': schedule.status,
                    'hourly_rate': schedule.hourly_rate,
                    'notes': schedule.notes,
                    'created_at': schedule.created_at,
                    'timeslot_info': timeslot_info
                }
            else:
                # Обычная смена
                query = select(Shift).options(
                    selectinload(Shift.object),
                    selectinload(Shift.user)
                ).where(Shift.id == actual_shift_id)
                
                result = await db.execute(query)
                shift = result.scalar_one_or_none()
                
                if not shift:
                    raise HTTPException(status_code=404, detail="Смена не найдена")
                
                # Проверяем доступ к объекту
                if shift.object_id not in accessible_object_ids:
                    raise HTTPException(status_code=403, detail="Нет доступа к объекту")

                # Безопасное форматирование времени с учётом часового пояса объекта
                tz_name = 'Europe/Moscow'
                try:
                    if getattr(shift, 'object', None) and getattr(shift.object, 'timezone', None):
                        tz_name = shift.object.timezone
                except Exception:
                    tz_name = 'Europe/Moscow'

                def safe_format(dt):
                    if not dt:
                        return '-'
                    try:
                        return web_timezone_helper.format_datetime_with_timezone(dt, tz_name, '%Y-%m-%d %H:%M')
                    except Exception:
                        try:
                            return dt.strftime('%Y-%m-%d %H:%M')
                        except Exception:
                            return str(dt)

                # Приводим статус к строке (если Enum)
                status_value = getattr(shift.status, 'value', shift.status)
                
                # Получаем contract для hourly_rate
                hourly_rate = None
                contract_id = getattr(shift, 'contract_id', None)
                if contract_id:
                    contract_query = select(Contract).where(Contract.id == contract_id)
                    contract_result = await db.execute(contract_query)
                    contract = contract_result.scalar_one_or_none()
                    if contract:
                        hourly_rate = contract.hourly_rate
                
                # Получаем задачи смены из журнала
                shift_tasks = []
                from shared.services.shift_task_journal import ShiftTaskJournal
                from domain.entities.timeslot_task_template import TimeslotTaskTemplate
                
                journal = ShiftTaskJournal(db)
                tasks_entities = await journal.get_by_shift(shift.id)
                
                # Если журнал пуст - синхронизировать из конфигурации
                if not tasks_entities:
                    tasks_entities = await journal.sync_from_config(
                        shift_id=shift.id,
                        time_slot_id=shift.time_slot_id,
                        object_id=shift.object_id,
                        created_by_id=user_id
                    )
                
                # Преобразовать в dict для шаблона
                shift_tasks = [{
                    'id': t.id,
                    'task_text': t.task_text,
                    'source': t.source,
                    'source_id': t.source_id,
                    'is_mandatory': t.is_mandatory,
                    'requires_media': t.requires_media,
                    'deduction_amount': float(t.deduction_amount) if t.deduction_amount else 0,
                    'is_completed': t.is_completed,
                    'completed_at': t.completed_at,
                    'media_refs': t.media_refs,
                    'cost': float(t.cost) if t.cost else None
                } for t in tasks_entities]
                
                shift_data = {
                    'id': shift.id,
                    'type': 'shift',
                    'object_id': shift.object_id,
                    'object_name': shift.object.name if shift.object else 'Неизвестный объект',
                    'user_id': shift.user_id,
                    'user_name': f"{shift.user.first_name} {shift.user.last_name or ''}".strip() if shift.user else 'Неизвестный пользователь',
                    'start_time': safe_format(shift.start_time),
                    'end_time': safe_format(shift.end_time),
                    'status': status_value,
                    'total_hours': shift.total_hours,
                    'total_payment': shift.total_payment,
                    'hourly_rate': hourly_rate,
                    'shift_tasks': shift_tasks,
                    'notes': shift.notes,
                    'created_at': shift.created_at
                }
            
            # Получаем данные для переключения интерфейсов
            login_service = RoleBasedLoginService(db)
            available_interfaces = await login_service.get_available_interfaces(user_id)
            
            return templates.TemplateResponse("manager/shifts/detail.html", {
                "request": request,
                "current_user": current_user,
                "available_interfaces": available_interfaces,
                "shift": shift_data,
                "shift_type": actual_shift_type
            })
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error loading manager shift detail: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки деталей смены")


@router.post("/calendar/api/quick-create-timeslot")
async def quick_create_timeslot_manager(
    request: Request,
    object_id: int = Form(...),
    slot_date: str = Form(...),
    start_time: str = Form(...),
    end_time: str = Form(...),
    hourly_rate: int = Form(...),
    current_user: dict = Depends(require_manager_or_owner)
):
    """Быстрое создание тайм-слота управляющим"""
    try:
        if isinstance(current_user, RedirectResponse):
            raise HTTPException(status_code=401, detail="Необходима авторизация")
        
        logger.info(f"Quick create timeslot: object_id={object_id}, date={slot_date}, start={start_time}, end={end_time}, rate={hourly_rate}")
        
        async with get_async_session() as db:
            user_id = await get_user_id_from_current_user(current_user, db)
            if not user_id:
                raise HTTPException(status_code=401, detail="Пользователь не найден")
            
            logger.info(f"User ID: {user_id}")
            
            # Проверяем доступ к объекту
            logger.info(f"Object ID: {object_id}")
            
            permission_service = ManagerPermissionService(db)
            accessible_objects = await permission_service.get_user_accessible_objects(user_id)
            accessible_object_ids = [obj.id for obj in accessible_objects]
            logger.info(f"Accessible object IDs: {accessible_object_ids}")
            
            if object_id not in accessible_object_ids:
                raise HTTPException(status_code=403, detail="Нет доступа к объекту")
            
            # Валидация данных
            try:
                slot_date_obj = datetime.strptime(slot_date, "%Y-%m-%d").date()
                start_time_obj = time.fromisoformat(start_time)
                end_time_obj = time.fromisoformat(end_time)
            except ValueError as e:
                raise HTTPException(status_code=400, detail=f"Неверный формат данных: {str(e)}")
            
            if start_time_obj >= end_time_obj:
                raise HTTPException(status_code=400, detail="Время начала должно быть меньше времени окончания")
            if hourly_rate <= 0:
                raise HTTPException(status_code=400, detail="Ставка должна быть больше 0")
            
            # Создаем тайм-слот
            from domain.entities.time_slot import TimeSlot
            
            logger.info(f"Creating timeslot: object_id={object_id}, date={slot_date_obj}, start={start_time_obj}, end={end_time_obj}, rate={hourly_rate}")
            
            timeslot = TimeSlot(
                object_id=object_id,
                slot_date=slot_date_obj,
                start_time=start_time_obj,
                end_time=end_time_obj,
                hourly_rate=float(hourly_rate),
                max_employees=1,
                is_additional=False,
                is_active=True,
                notes=""
            )
            
            db.add(timeslot)
            await db.commit()
            await db.refresh(timeslot)
            
            # Очищаем кэш календаря для немедленного отображения
            from core.cache.redis_cache import cache
            await cache.clear_pattern("calendar_shifts:*")
            await cache.clear_pattern("api_response:*")
            logger.info(f"Calendar cache cleared after creating timeslot {timeslot.id}")
            
            logger.info(f"Timeslot created successfully with ID: {timeslot.id}")
            
            # Возвращаем результат до закрытия сессии
            result = {
                "success": True,
                "message": "Тайм-слот успешно создан",
                "timeslot_id": timeslot.id
            }
            
            logger.info(f"Returning success result: {result}")
            return result
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating timeslot: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка создания тайм-слота: {str(e)}")


@router.get("/profile", response_class=HTMLResponse)
async def manager_profile(
    request: Request,
    current_user: dict = Depends(require_manager_or_owner),
    db: AsyncSession = Depends(get_db_session)
):
    """Страница профиля управляющего"""
    try:
        if isinstance(current_user, RedirectResponse):
            return current_user
            
        user_id = await get_user_id_from_current_user(current_user, db)
        if not user_id:
            raise HTTPException(status_code=401, detail="Пользователь не найден")
        
        # Получаем данные пользователя
        user_query = select(User).where(User.id == user_id)
        user_result = await db.execute(user_query)
        user = user_result.scalar_one_or_none()
        
        if not user:
            raise HTTPException(status_code=404, detail="Пользователь не найден")
        
        # Получаем доступные объекты для статистики
        permission_service = ManagerPermissionService(db)
        accessible_objects = await permission_service.get_user_accessible_objects(user_id)
        
        # Получаем данные для переключения интерфейсов
        login_service = RoleBasedLoginService(db)
        available_interfaces = await login_service.get_available_interfaces(user_id)
        
        # Статистика профиля
        profile_stats = {
            'accessible_objects_count': len(accessible_objects),
            'total_contracts': 0,  # TODO: Добавить подсчет договоров
            'active_shifts_count': 0,  # TODO: Добавить подсчет активных смен
            'total_employees_count': 0  # TODO: Добавить подсчет сотрудников
        }
        
        return templates.TemplateResponse("manager/profile.html", {
            "request": request,
            "current_user": current_user,
            "user": user,
            "profile_stats": profile_stats,
            "accessible_objects": accessible_objects,
            "available_interfaces": available_interfaces,
            "success_message": request.query_params.get("success"),
            "error_message": request.query_params.get("error"),
        })
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in manager profile: {e}")
        raise HTTPException(status_code=500, detail="Ошибка загрузки профиля")


@router.post("/profile")
async def manager_profile_update(
    request: Request,
    first_name: Optional[str] = Form(None),
    last_name: Optional[str] = Form(None),
    username: Optional[str] = Form(None),
    phone: Optional[str] = Form(None),
    email: Optional[str] = Form(None),
    about: Optional[str] = Form(None),
    current_user: dict = Depends(require_manager_or_owner),
    db: AsyncSession = Depends(get_db_session)
):
    """Обновление профиля управляющего."""

    if isinstance(current_user, RedirectResponse):
        return current_user
        
    user_id = await get_user_id_from_current_user(current_user, db)
    if not user_id:
        raise HTTPException(status_code=401, detail="Пользователь не найден")
    
    user_query = select(User).where(User.id == user_id)
    user_result = await db.execute(user_query)
    user = user_result.scalar_one_or_none()
    
    if not user:
        raise HTTPException(status_code=404, detail="Пользователь не найден")
        
    try:
        user.first_name = (first_name or "").strip() or None
        user.last_name = (last_name or "").strip() or None
        user.username = (username or "").strip() or None
        user.phone = (phone or "").strip() or None
        user.email = (email or "").strip() or None
        user.about = (about or "").strip() or None

        await db.commit()
        success_message = quote("Профиль успешно обновлён")
        return RedirectResponse(
            url=f"/manager/profile?success={success_message}",
            status_code=status.HTTP_303_SEE_OTHER
        )
    except Exception as exc:
        await db.rollback()
        error_message = quote("Не удалось сохранить изменения")
        return RedirectResponse(
            url=f"/manager/profile?error={error_message}",
            status_code=status.HTTP_303_SEE_OTHER
        )

async def get_manager_context(user_id: int, session: AsyncSession):
    """Получает общий контекст для страниц управляющего"""
    
    # Получаем доступные интерфейсы
    login_service = RoleBasedLoginService(session)
    available_interfaces = await login_service.get_available_interfaces(user_id)
    
    # Получаем количество новых заявок
    new_applications_count = await get_new_applications_count(user_id, session, "manager")
    
    # Получаем права управляющего
    # Проверяем наличие права can_manage_payroll в manager_permissions (JSON поле договора)
    manager_contracts_query = select(Contract).where(
        Contract.employee_id == user_id,
        Contract.is_manager == True,
        Contract.is_active == True,
        Contract.status == 'active'
    )
    manager_contracts_result = await session.execute(manager_contracts_query)
    manager_contracts = manager_contracts_result.scalars().all()
    
    can_manage_payroll = False
    for contract in manager_contracts:
        permissions = contract.manager_permissions or {}
        if permissions.get("can_manage_payroll", False):
            can_manage_payroll = True
            break
    
    logger.info(
        f"Manager context for user {user_id}: can_manage_payroll={can_manage_payroll}, contracts={len(manager_contracts)}"
    )
    
    return {
        "available_interfaces": available_interfaces,
        "new_applications_count": new_applications_count,
        "can_manage_payroll": can_manage_payroll
    }

@router.get("/applications", response_class=HTMLResponse)
async def manager_applications(
    request: Request,
    current_user: dict = Depends(require_manager_or_owner),
    db: AsyncSession = Depends(get_db_session)
):
    """Страница заявок управляющего"""
    try:
        if isinstance(current_user, RedirectResponse):
            return current_user
            
        user_id = await get_user_id_from_current_user(current_user, db)
        if not user_id:
            raise HTTPException(status_code=401, detail="Пользователь не найден")
        
        # Получаем заявки по объектам, к которым у управляющего есть доступ
        from domain.entities.manager_object_permission import ManagerObjectPermission
        from domain.entities.contract import Contract
        
        # Сначала получаем объекты, к которым у управляющего есть доступ
        objects_query = select(Object.id).join(
            ManagerObjectPermission, Object.id == ManagerObjectPermission.object_id
        ).join(Contract, ManagerObjectPermission.contract_id == Contract.id).where(
            and_(
                Contract.employee_id == user_id,
                Contract.is_manager == True
            )
        )
        
        objects_result = await db.execute(objects_query)
        accessible_object_ids = [obj for obj in objects_result.scalars()]
        
        if not accessible_object_ids:
            # Если нет доступных объектов, возвращаем пустой список
            applications = []
        else:
            # Получаем заявки по доступным объектам
            applications_query = select(Application).where(
                Application.object_id.in_(accessible_object_ids)
            ).options(
                selectinload(Application.applicant),
                selectinload(Application.object)
            ).order_by(desc(Application.created_at))
            
        applications_result = await db.execute(applications_query)
        applications = applications_result.scalars().all()
    
        # Получаем общий контекст управляющего
        manager_context = await get_manager_context(user_id, db)
    
        return templates.TemplateResponse("manager/applications.html", {
            "request": request,
            "current_user": current_user,
            "applications": applications,
            "show_actions": True,
            "current_role": "manager",
            **manager_context
        })
        
    except Exception as e:
        logger.error(f"Ошибка загрузки заявок управляющего: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка загрузки заявок: {e}")

@router.post("/api/applications/approve")
async def manager_approve_application(
    request: Request,
    current_user: dict = Depends(require_manager_or_owner),
    db: AsyncSession = Depends(get_db_session)
):
    """Одобрение заявки управляющим с назначением собеседования"""
    try:
        if isinstance(current_user, RedirectResponse):
            return current_user
            
        user_id = await get_user_id_from_current_user(current_user, db)
        if not user_id:
            raise HTTPException(status_code=401, detail="Пользователь не найден")
        
        form_data = await request.form()
        application_id = form_data.get("application_id")
        interview_datetime = form_data.get("interview_datetime")
        interview_type = form_data.get("interview_type")
        interview_notes = form_data.get("interview_notes", "").strip()
        
        if not application_id or not interview_datetime or not interview_type:
            raise HTTPException(status_code=400, detail="Не все поля заполнены")
        
        # Получаем заявку с проверкой прав управляющего
        from domain.entities.manager_object_permission import ManagerObjectPermission
        from domain.entities.contract import Contract
        
        application_query = select(Application).join(Object).join(
            ManagerObjectPermission, Object.id == ManagerObjectPermission.object_id
        ).join(Contract, ManagerObjectPermission.contract_id == Contract.id).where(
            and_(
                Application.id == int(application_id),
                Contract.employee_id == user_id,
                Contract.is_manager == True
            )
        )
        application_result = await db.execute(application_query)
        application = application_result.scalar_one_or_none()
        
        if not application:
            raise HTTPException(status_code=404, detail="Заявка не найдена или нет доступа")
        
        # Обновляем заявку
        application.status = ApplicationStatus.INTERVIEW
        application.interview_scheduled_at = datetime.fromisoformat(interview_datetime.replace('T', ' '))
        application.interview_type = interview_type
        application.interview_result = interview_notes
        
        await db.commit()
        
        # Отправляем уведомления
        try:
            from core.database.session import get_sync_session
            from shared.services.notification_service import NotificationService
            from core.config.settings import settings
            from domain.entities.user import User
            
            session_factory = get_sync_session
            with session_factory() as session:
                notification_service = NotificationService(
                    session=session,
                    telegram_token=settings.telegram_bot_token
                )
                
                # Получаем информацию об управляющем для имени в уведомлении
                manager_query = select(User).where(User.id == user_id)
                manager_result = session.execute(manager_query)
                manager_user = manager_result.scalar_one_or_none()
                
                manager_name = "Управляющий"
                if manager_user:
                    if manager_user.first_name or manager_user.last_name:
                        parts = []
                        if manager_user.first_name:
                            parts.append(manager_user.first_name.strip())
                        if manager_user.last_name:
                            parts.append(manager_user.last_name.strip())
                        manager_name = " ".join(parts) if parts else manager_user.username
                    elif manager_user.username:
                        manager_name = manager_user.username
                
                # Уведомляем соискателя о назначении собеседования
                notification_payload = {
                    "application_id": application.id,
                    "object_name": application.object.name if hasattr(application, 'object') and application.object else "Объект",
                    "object_address": application.object.address if hasattr(application, 'object') and application.object else "—",
                    "employee_position": application.object.employee_position if hasattr(application, 'object') and application.object and hasattr(application.object, 'employee_position') else "Должность не указана",
                    "scheduled_at": application.interview_scheduled_at.isoformat(),
                    "interview_type": interview_type,
                    "owner_name": manager_name
                }
                
                notification_service.create(
                    [application.applicant_id],
                    "interview_assigned",
                    notification_payload,
                    send_telegram=True
                )
                session.commit()
        except Exception as notification_error:
            logger.error(f"Ошибка отправки уведомлений управляющим: {notification_error}")
        
        logger.info(f"Заявка {application_id} одобрена управляющим {user_id}, собеседование назначено на {interview_datetime}")
        
        return {
            "message": "Заявка одобрена и собеседование назначено",
            "interview_datetime": interview_datetime,
            "interview_type": interview_type
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка одобрения заявки управляющим: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка одобрения заявки: {e}")

@router.post("/api/applications/reject")
async def manager_reject_application(
    application_id: int = Form(...),
    reject_reason: str = Form(""),
    current_user: dict = Depends(require_manager_or_owner),
    db: AsyncSession = Depends(get_db_session)
):
    try:
        logger.info(f"*** MANAGER REJECT NOT OWNER! application_id={application_id}, reject_reason={reject_reason} ***")
        if isinstance(current_user, RedirectResponse):
            return current_user
        user_id = await get_user_id_from_current_user(current_user, db)
        if not user_id:
            raise HTTPException(status_code=401, detail="Пользователь не найден")
        query = select(Application).select_from(Application).join(
            Object, Application.object_id == Object.id
        ).join(
            ManagerObjectPermission, Object.id == ManagerObjectPermission.object_id
        ).join(
            Contract, and_(ManagerObjectPermission.contract_id == Contract.id, Contract.employee_id == user_id, Contract.is_manager == True)
        ).where(Application.id == application_id)
        result = await db.execute(query)
        application = result.scalar_one_or_none()
        if not application:
            raise HTTPException(status_code=404, detail="Заявка не найдена или нет доступа")
        application.status = ApplicationStatus.REJECTED
        application.interview_result = reject_reason
        await db.commit()
        
        # Отправляем уведомления
        try:
            from core.database.session import get_sync_session
            from shared.services.notification_service import NotificationService
            from core.config.settings import settings
            from domain.entities.user import User
            
            session_factory = get_sync_session
            with session_factory() as session:
                notification_service = NotificationService(
                    session=session,
                    telegram_token=settings.telegram_bot_token
                )
                
                # Получаем информацию об управляющем для имени в уведомлении
                manager_query = select(User).where(User.id == user_id)
                manager_result = session.execute(manager_query)
                manager_user = manager_result.scalar_one_or_none()
                
                manager_name = "Управляющий"
                if manager_user:
                    if manager_user.first_name or manager_user.last_name:
                        parts = []
                        if manager_user.first_name:
                            parts.append(manager_user.first_name.strip())
                        if manager_user.last_name:
                            parts.append(manager_user.last_name.strip())
                        manager_name = " ".join(parts) if parts else manager_user.username
                    elif manager_user.username:
                        manager_name = manager_user.username
                
                # Уведомляем соискателя об отклонении
                notification_payload = {
                    "application_id": application.id,
                    "object_name": application.object.name if hasattr(application, 'object') and application.object else "Объект",
                    "object_address": application.object.address if hasattr(application, 'object') and application.object else "—",
                    "employee_position": application.object.employee_position if hasattr(application, 'object') and application.object and hasattr(application.object, 'employee_position') else "Должность не указана",
                    "reason": reject_reason,
                    "owner_name": manager_name
                }
                
                notification_service.create(
                    [application.applicant_id],
                    "application_rejected",
                    notification_payload,
                    send_telegram=True
                )
                session.commit()
        except Exception as notification_error:
            logger.error(f"Ошибка отправки уведомлений об отклонении управляющим: {notification_error}")
        
        logger.info(f"Заявка {application_id} отклонена управляющим {user_id}")
        return {"message": "Заявка отклонена", "status": application.status.value}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка отклонения заявки управляющим: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка отклонения заявки: {e}")

@router.get("/api/applications/count")
async def manager_applications_count(
    current_user: dict = Depends(require_manager_or_owner),
    db: AsyncSession = Depends(get_db_session)
) -> dict[str, int]:
    """Количество новых заявок для управляющего."""
    if isinstance(current_user, RedirectResponse):
        return current_user
    user_id = await get_user_id_from_current_user(current_user, db)
    if not user_id:
        raise HTTPException(status_code=401, detail="Пользователь не найден")
    from apps.web.utils.applications_utils import get_new_applications_count
    count = await get_new_applications_count(user_id, db, "manager")
    return {"count": count}

@router.get("/api/applications/{application_id}")
async def manager_application_details_api(
    application_id: int,
    current_user: dict = Depends(require_manager_or_owner),
    db: AsyncSession = Depends(get_db_session)
):
    if isinstance(current_user, RedirectResponse):
        raise HTTPException(status_code=401, detail="Необходима авторизация")

    user_id = await get_user_id_from_current_user(current_user, db)
    if not user_id:
        raise HTTPException(status_code=401, detail="Пользователь не найден")

    from domain.entities.manager_object_permission import ManagerObjectPermission
    from domain.entities.contract import Contract

    query = select(Application, Object, User).select_from(Application).join(Object, Application.object_id == Object.id).join(User, Application.applicant_id == User.id).join(
        ManagerObjectPermission, and_(Object.id == ManagerObjectPermission.object_id)
    ).join(Contract, and_(ManagerObjectPermission.contract_id == Contract.id, Contract.employee_id == user_id, Contract.is_manager == True))
    result = await db.execute(query.where(Application.id == application_id))
    row = result.first()
    if not row:
        raise HTTPException(status_code=404, detail="Заявка не найдена или нет доступа")

    application, obj, applicant = row

    return {
        "id": application.id,
        "object_name": obj.name,
        "object_address": obj.address,
        "status": application.status.value,
        "message": application.message,
        "created_at": application.created_at.isoformat() if application.created_at else None,
        "interview_scheduled_at": application.interview_scheduled_at.isoformat() if application.interview_scheduled_at else None,
        "interview_type": application.interview_type,
        "applicant": {
            "full_name": applicant.full_name,
            "first_name": applicant.first_name,
            "last_name": applicant.last_name,
            "username": applicant.username,
            "email": applicant.email,
            "phone": applicant.phone,
            "skills": applicant.skills,
            "about": applicant.about,
            "work_experience": applicant.work_experience,
            "preferred_schedule": applicant.preferred_schedule,
            "education": applicant.education,
        }
    }

@router.post("/api/applications/finalize-contract")
async def manager_finalize_contract(
    application_id: int = Form(...),
    current_user: dict = Depends(require_manager_or_owner),
    db: AsyncSession = Depends(get_db_session)
):
    try:
        if isinstance(current_user, RedirectResponse):
            return current_user

        user_id = await get_user_id_from_current_user(current_user, db)
        if not user_id:
            raise HTTPException(status_code=401, detail="Пользователь не найден")

        query = select(Application).select_from(Application).join(
            Object, Application.object_id == Object.id
        ).join(
            ManagerObjectPermission, Object.id == ManagerObjectPermission.object_id
        ).join(
            Contract, and_(ManagerObjectPermission.contract_id == Contract.id, Contract.employee_id == user_id, Contract.is_manager == True)
        ).where(Application.id == application_id)

        result = await db.execute(query)
        application = result.scalar_one_or_none()
        if not application:
            raise HTTPException(status_code=404, detail="Заявка не найдена или нет доступа")

        application.status = ApplicationStatus.APPROVED
        await db.commit()

        logger.info(f"Заявка {application_id} переведена в статус APPROVED управляющим {user_id}")
        return {"message": "Заявка одобрена", "status": application.status.value}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка завершения заявки управляющим: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка обработки заявки: {e}")


@router.post("/shifts/{shift_id}/cancel")
async def manager_cancel_shift(
    request: Request, 
    shift_id: str,  # Поддержка префикса schedule_
    shift_type: Optional[str] = Query("shift"),
    current_user: dict = Depends(require_manager_or_owner),
    db: AsyncSession = Depends(get_db_session)
):
    """Отмена смены управляющим"""
    from fastapi.responses import JSONResponse
    from datetime import datetime
    
    try:
        # Определяем тип смены по ID
        if shift_id.startswith('schedule_'):
            actual_shift_id = int(shift_id.replace('schedule_', ''))
            actual_shift_type = "schedule"
        else:
            actual_shift_id = int(shift_id)
            actual_shift_type = shift_type or "shift"
        
        # Получаем внутренний ID пользователя
        user_id = await get_user_id_from_current_user(current_user, db)
        if not user_id:
            return JSONResponse(
                status_code=401,
                content={"success": False, "message": "Пользователь не найден"}
            )
        
        # Получаем доступные объекты управляющего
        permission_service = ManagerPermissionService(db)
        accessible_objects = await permission_service.get_user_accessible_objects(user_id)
        accessible_object_ids = [obj.id for obj in accessible_objects]
        
        if not accessible_object_ids:
            return JSONResponse(
                status_code=403,
                content={"success": False, "message": "Нет доступа к объектам"}
            )
        
        if actual_shift_type == "schedule":
            # Отмена запланированной смены
            query = select(ShiftSchedule).options(
                selectinload(ShiftSchedule.object)
            ).where(ShiftSchedule.id == actual_shift_id)
            result = await db.execute(query)
            shift = result.scalar_one_or_none()
            
            if shift and shift.status == "planned":
                # Проверяем доступ к объекту
                if shift.object_id not in accessible_object_ids:
                    return JSONResponse(
                        status_code=403,
                        content={"success": False, "message": "Нет доступа к объекту"}
                    )
                
                # Отменяем смену
                shift.status = "cancelled"
                shift.updated_at = datetime.utcnow()
                await db.commit()
                
                return JSONResponse(
                    status_code=200,
                    content={"success": True, "message": "Смена отменена"}
                )
            else:
                return JSONResponse(
                    status_code=400,
                    content={"success": False, "message": "Смена не найдена или уже отменена"}
                )
        else:
            # Отмена реальной смены
            query = select(Shift).options(
                selectinload(Shift.object)
            ).where(Shift.id == actual_shift_id)
            result = await db.execute(query)
            shift = result.scalar_one_or_none()
            
            if shift and shift.status == "active":
                # Проверяем доступ к объекту
                if shift.object_id not in accessible_object_ids:
                    return JSONResponse(
                        status_code=403,
                        content={"success": False, "message": "Нет доступа к объекту"}
                    )
                
                # Отменяем смену
                shift.status = "cancelled"
                shift.updated_at = datetime.utcnow()
                await db.commit()
                
                return JSONResponse(
                    status_code=200,
                    content={"success": True, "message": "Смена отменена"}
                )
            else:
                return JSONResponse(
                    status_code=400,
                    content={"success": False, "message": "Смена не найдена или уже завершена"}
                )
                
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка отмены смены управляющим: {e}")
        return JSONResponse(
            status_code=500,
            content={"success": False, "message": "Ошибка отмены смены"}
        )


# Роут /payroll-adjustments перенесен в apps/web/routes/manager_payroll_adjustments.py

# ==================== CONTRACT MANAGEMENT ====================


@router.get("/contracts/{contract_id}/view", response_class=HTMLResponse)
async def manager_contract_view(
    contract_id: int,
    request: Request,
    current_user: dict = Depends(require_manager_or_owner),
    db: AsyncSession = Depends(get_db_session)
):
    """Просмотр договора (доступен всегда)."""
    try:
        user_id = await get_user_id_from_current_user(current_user, db)
        if not user_id:
            raise HTTPException(status_code=401, detail="Пользователь не найден")
        
        # Получаем договор
        query = select(Contract).where(Contract.id == contract_id).options(
            selectinload(Contract.employee),
            selectinload(Contract.owner),
            selectinload(Contract.template)
        )
        result = await db.execute(query)
        contract = result.scalar_one_or_none()
        
        if not contract:
            raise HTTPException(status_code=404, detail="Договор не найден")
        
        # Проверяем доступ: управляющий должен иметь права на объекты из allowed_objects
        from shared.services.manager_permission_service import ManagerPermissionService
        permission_service = ManagerPermissionService(db)
        accessible_objects = await permission_service.get_user_accessible_objects(user_id)
        accessible_object_ids = [obj.id for obj in accessible_objects]
        
        # Проверяем, что хотя бы один объект из договора доступен управляющему
        contract_objects = contract.allowed_objects or []
        has_access = any(obj_id in accessible_object_ids for obj_id in contract_objects)
        
        if not has_access:
            raise HTTPException(status_code=403, detail="Доступ запрещён")
        
        # Получаем объекты для отображения
        objects_info = []
        if contract_objects:
            objects_query = select(Object).where(Object.id.in_(contract_objects))
            objects_result = await db.execute(objects_query)
            objects_info = objects_result.scalars().all()
        
        return templates.TemplateResponse(
            "manager/contracts/view.html",
            {
                "request": request,
                "current_user": current_user,
                "contract": contract,
                "objects": objects_info
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error viewing contract: {e}")
        raise HTTPException(status_code=500, detail="Ошибка просмотра договора")


@router.get("/contracts/{contract_id}/edit", response_class=HTMLResponse)
async def manager_contract_edit(
    contract_id: int,
    request: Request,
    current_user: dict = Depends(require_manager_or_owner),
    db: AsyncSession = Depends(get_db_session)
):
    """Редактирование договора (только с правом can_manage_employees)."""
    try:
        user_id = await get_user_id_from_current_user(current_user, db)
        if not user_id:
            raise HTTPException(status_code=401, detail="Пользователь не найден")
        
        # Получаем договор
        query = select(Contract).where(Contract.id == contract_id).options(
            selectinload(Contract.employee),
            selectinload(Contract.owner),
            selectinload(Contract.template)
        )
        result = await db.execute(query)
        contract = result.scalar_one_or_none()
        
        if not contract:
            raise HTTPException(status_code=404, detail="Договор не найден")
        
        # Запрещаем управляющему редактировать свой собственный договор
        if contract.employee_id == user_id:
            raise HTTPException(status_code=403, detail="Вы не можете редактировать свой собственный договор через этот интерфейс")
        
        # Проверяем право can_manage_employees
        # Получаем договоры управляющего
        manager_contracts_query = select(Contract).where(
            Contract.employee_id == user_id,
            Contract.is_manager == True,
            Contract.is_active == True
        )
        manager_contracts_result = await db.execute(manager_contracts_query)
        manager_contracts = manager_contracts_result.scalars().all()
        manager_contract_ids = [mc.id for mc in manager_contracts]
        
        if not manager_contract_ids:
            raise HTTPException(status_code=403, detail="Вы не являетесь управляющим")
        
        # Проверяем права на объекты договора сотрудника
        contract_objects = contract.allowed_objects or []
        
        from domain.entities.manager_object_permission import ManagerObjectPermission
        permissions_query = select(ManagerObjectPermission).where(
            ManagerObjectPermission.contract_id.in_(manager_contract_ids),
            ManagerObjectPermission.object_id.in_(contract_objects),
            ManagerObjectPermission.can_manage_employees == True
        )
        permissions_result = await db.execute(permissions_query)
        permissions = permissions_result.scalars().all()
        
        if not permissions:
            raise HTTPException(status_code=403, detail="У вас нет прав на редактирование этого договора")
        
        # Получаем объекты для отображения
        # Нужно показать все объекты, доступные управляющему по договору, в котором есть объект из редактируемого договора сотрудника
        contract_objects = contract.allowed_objects or []
        
        # Получаем владельцев объектов из редактируемого договора
        owner_ids = set()
        if contract_objects:
            objects_query = select(Object).where(Object.id.in_(contract_objects))
            objects_result = await db.execute(objects_query)
            contract_objects_list = objects_result.scalars().all()
            owner_ids = {obj.owner_id for obj in contract_objects_list if obj.owner_id}
        
        # Получаем все объекты из договоров управляющего с этими владельцами
        accessible_objects = []
        if owner_ids:
            # Находим договоры управляющего с нужными владельцами
            manager_contracts_with_owners = [
                mc for mc in manager_contracts 
                if mc.owner_id in owner_ids
            ]
            
            # Получаем все объекты из allowed_objects этих договоров
            accessible_object_ids = set()
            for mc in manager_contracts_with_owners:
                if mc.allowed_objects:
                    accessible_object_ids.update(mc.allowed_objects)
            
            # Получаем объекты из БД
            if accessible_object_ids:
                objects_query = select(Object).where(Object.id.in_(accessible_object_ids))
                objects_result = await db.execute(objects_query)
                accessible_objects = objects_result.scalars().all()
        
        # Если нет объектов через owner_id, используем все доступные объекты управляющего
        if not accessible_objects:
            from shared.services.manager_permission_service import ManagerPermissionService
            permission_service = ManagerPermissionService(db)
            accessible_objects = await permission_service.get_user_accessible_objects(user_id)
        
        return templates.TemplateResponse(
            "manager/contracts/edit.html",
            {
                "request": request,
                "current_user": current_user,
                "contract": contract,
                "objects": accessible_objects,
                "accessible_objects": accessible_objects  # Для совместимости с шаблоном
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error editing contract: {e}")
        raise HTTPException(status_code=500, detail="Ошибка редактирования договора")


@router.post("/contracts/{contract_id}/terminate")
async def manager_contract_terminate(
    contract_id: int,
    request: Request,
    reason: str = Form(...),
    reason_category: str = Form(...),
    termination_date: Optional[str] = Form(None),
    payout_mode: str = Form("schedule"),
    current_user: dict = Depends(require_manager_or_owner),
    db: AsyncSession = Depends(get_db_session)
):
    """Расторжение договора (только с правом can_manage_employees)."""
    try:
        user_id = await get_user_id_from_current_user(current_user, db)
        if not user_id:
            raise HTTPException(status_code=401, detail="Пользователь не найден")
        
        # Получаем договор
        query = select(Contract).where(Contract.id == contract_id)
        result = await db.execute(query)
        contract = result.scalar_one_or_none()
        
        if not contract:
            raise HTTPException(status_code=404, detail="Договор не найден")
        
        # Запрещаем управляющему расторгать свой собственный договор
        if contract.employee_id == user_id:
            raise HTTPException(status_code=403, detail="Вы не можете расторгать свой собственный договор через этот интерфейс")
        
        # Проверяем право can_manage_employees
        # Получаем договоры управляющего
        manager_contracts_query = select(Contract).where(
            Contract.employee_id == user_id,
            Contract.is_manager == True,
            Contract.is_active == True
        )
        manager_contracts_result = await db.execute(manager_contracts_query)
        manager_contracts = manager_contracts_result.scalars().all()
        manager_contract_ids = [mc.id for mc in manager_contracts]
        
        if not manager_contract_ids:
            raise HTTPException(status_code=403, detail="Вы не являетесь управляющим")
        
        # Проверяем права на объекты договора сотрудника
        contract_objects = contract.allowed_objects or []
        
        from domain.entities.manager_object_permission import ManagerObjectPermission
        permissions_query = select(ManagerObjectPermission).where(
            ManagerObjectPermission.contract_id.in_(manager_contract_ids),
            ManagerObjectPermission.object_id.in_(contract_objects),
            ManagerObjectPermission.can_manage_employees == True
        )
        permissions_result = await db.execute(permissions_query)
        permissions = permissions_result.scalars().all()
        
        if not permissions:
            raise HTTPException(status_code=403, detail="У вас нет прав на расторжение этого договора")
        
        # Парсим дату увольнения
        term_date = None
        if termination_date:
            try:
                term_date = datetime.strptime(termination_date, "%Y-%m-%d").date()
            except ValueError:
                raise HTTPException(status_code=400, detail="Неверный формат даты увольнения")
        
        # Определяем политику финрасчёта
        settlement_policy = "termination_date" if payout_mode == "termination_date" else "schedule"
        
        # Расторгаем договор
        terminated_at_now = datetime.now()
        contract.status = 'terminated'
        contract.is_active = False
        contract.terminated_at = terminated_at_now
        contract.termination_date = term_date
        contract.settlement_policy = settlement_policy
        
        # Добавляем причину в values или notes
        if not contract.values:
            contract.values = {}
        contract.values['termination_reason'] = reason
        contract.values['termination_reason_category'] = reason_category
        contract.values['terminated_by'] = user_id
        
        # Создаём запись о расторжении для аналитики
        from domain.entities.contract_termination import ContractTermination
        termination_record = ContractTermination(
            contract_id=contract_id,
            employee_id=contract.employee_id,
            owner_id=contract.owner_id,
            terminated_by_id=user_id,
            terminated_by_type='manager',
            reason_category=reason_category,
            reason=reason,
            termination_date=term_date,
            settlement_policy=settlement_policy,
            terminated_at=terminated_at_now
        )
        db.add(termination_record)
        
        # Отменяем плановые смены после termination_date
        if term_date:
            from domain.entities.shift_schedule import ShiftSchedule
            from domain.entities.shift_cancellation import ShiftCancellation
            
            shifts_query = select(ShiftSchedule).where(
                and_(
                    ShiftSchedule.user_id == contract.employee_id,
                    ShiftSchedule.status == 'planned',
                    func.date(ShiftSchedule.planned_start) > term_date
                )
            )
            shifts_result = await db.execute(shifts_query)
            shifts_to_cancel = shifts_result.scalars().all()
            
            for shift in shifts_to_cancel:
                shift.status = 'cancelled'
                cancellation = ShiftCancellation(
                    shift_schedule_id=shift.id,
                    employee_id=contract.employee_id,
                    object_id=shift.object_id,
                    cancelled_by_id=user_id,
                    cancelled_by_type='manager',
                    cancellation_reason='contract_termination',
                    reason_notes=f"Расторгнут договор (дата увольнения: {term_date}). Причина: {reason}",
                    contract_id=contract.id,
                    hours_before_shift=None,
                    fine_amount=None,
                    fine_reason=None,
                    fine_applied=False
                )
                db.add(cancellation)
        
        # Сохраняем employee_id перед коммитом
        employee_id = contract.employee_id
        
        await db.commit()
        
        # Проверяем, есть ли у сотрудника другие активные договоры
        from domain.entities.user import User
        remaining_contracts_query = select(Contract).where(
            Contract.employee_id == employee_id,
            Contract.is_active == True,
            Contract.status == 'active'
        )
        remaining_result = await db.execute(remaining_contracts_query)
        remaining_contracts = remaining_result.scalars().all()
        
        # Если это был последний активный договор - делаем сотрудника неактивным
        if not remaining_contracts:
            user_query = select(User).where(User.id == employee_id)
            user_result = await db.execute(user_query)
            user = user_result.scalar_one_or_none()
            if user:
                user.is_active = False
                await db.commit()
                logger.info(f"User {employee_id} marked as inactive (no active contracts)")
        
        logger.info(
            f"Contract terminated by manager",
            contract_id=contract_id,
            employee_id=employee_id,
            manager_id=user_id,
            reason_category=reason_category
        )
        
        return RedirectResponse(
            url=f"/manager/employees/{employee_id}",
            status_code=302
        )
        
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        logger.error(f"Error terminating contract: {e}\n{traceback.format_exc()}")
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка расторжения договора: {str(e)}")
